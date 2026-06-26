import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import sqlite3
import os
import sys
import shutil
import re
import threading
import subprocess
import socket
from datetime import date, datetime
import json
from PIL import Image, ImageTk
import requests
import database


def center_dialog(w, pw):
    w.update_idletasks()
    x = pw.winfo_x() + (pw.winfo_width() - w.winfo_width()) // 2
    y = pw.winfo_y() + (pw.winfo_height() - w.winfo_height()) // 2
    w.geometry("+%d+%d" % (x, y))


def add_keyboard_shortcuts(dialog, save_cb):
    dialog.bind("<Escape>", lambda e: dialog.destroy())
    dialog.bind("<Control-Return>", lambda e: save_cb())


def validate_required(entries, labels):
    for e, lbl in zip(entries, labels):
        if not e.get().strip():
            messagebox.showerror("Error", f"{lbl} es obligatorio")
            return False
    return True


DB_PATH = database.DB_PATH
PHOTOS_DIR = database.PHOTOS_DIR
EXAMENES_DIR = database.EXAMENES_DIR
DOCS_DIR = database.DOCS_DIR
class AutocompleteEntry(tk.Frame):
    def __init__(self, master=None, full_values=None, width=37, font=None, **kwargs):
        super().__init__(master, **kwargs)
        self._full_values = full_values or []
        self._selected_value = ""
        self._font = font or ("Segoe UI", 10)
        self._top = None

        self.entry = tk.Entry(self, width=width, font=self._font)
        self.entry.pack(fill="x")
        self.entry.bind("<KeyRelease>", self._on_keyrelease)
        self.entry.bind("<FocusOut>", self._on_focusout)

    def get(self):
        return self.entry.get()

    def set(self, value):
        self.entry.delete(0, "end")
        self.entry.insert(0, value)
        self._selected_value = value

    def set_full_values(self, values):
        self._full_values = list(values)

    def _hide_top(self):
        if self._top:
            self._top.destroy()
            self._top = None

    def _on_keyrelease(self, event):
        if event.keysym in ("Up", "Down", "Return", "Tab", "Shift_L", "Shift_R", "Control_L", "Control_R", "Escape"):
            return
        typed = self.entry.get().lower()
        self._hide_top()
        if not typed:
            matches = list(self._full_values)
        else:
            matches = [v for v in self._full_values if typed in v.lower()]
        if not matches:
            return
        self._top = tk.Toplevel(self.master)
        self._top.wm_overrideredirect(True)
        # Calculate dropdown width based on longest match
        max_len = max(len(m) for m in matches) + 2
        listbox_width = max(max_len, self.entry.cget("width"))
        listbox = tk.Listbox(self._top, height=min(len(matches), 8), font=self._font,
                             width=listbox_width,
                             exportselection=False, activestyle="none")
        listbox.pack()
        for v in matches:
            listbox.insert("end", v)
        # Position dropdown aligned with entry
        x = self.winfo_rootx()
        y = self.winfo_rooty() + self.entry.winfo_height()
        self._top.wm_geometry("+%d+%d" % (x, y))

        def on_select(ev=None):
            sel = listbox.curselection()
            if sel:
                val = listbox.get(sel[0])
                self.entry.delete(0, "end")
                self.entry.insert(0, val)
                self._selected_value = val
            self._hide_top()
            self.entry.focus_set()

        listbox.bind("<ButtonRelease-1>", on_select)
        listbox.bind("<Return>", on_select)
        self._top.bind("<Escape>", lambda e: self._hide_top())

    def _on_focusout(self, event):
        self.after(400, self._hide_top)


# DB_PATH, PHOTOS_DIR, EXAMENES_DIR imported from database module

class App:
    def __init__(self, root):
        self.root = root
        self.root.title('SCRIPTYFY V.02 - Historial Clinico')
        self.root.state('zoomed')
        self.root.configure(bg='#e8edf2')
        self._dark_mode = False
        self._user_logged_in = False
        self._user_id = None
        self._user_name = 'Invitado'
        self._user_role = 'empleado'
        self._theme_bg = '#e8edf2'
        self._caja_abierta = False
        self._theme = self._get_theme('light')
        self._current_view = ''
        self._saldo_inicial = 0.0
        os.makedirs(PHOTOS_DIR, exist_ok=True)
        os.makedirs(EXAMENES_DIR, exist_ok=True)
        os.makedirs(DOCS_DIR, exist_ok=True)
        self._init_tables()
        self._web_process = None

        # Check if caja was left open from yesterday/today
        conn = self.get_db()
        today = str(date.today())
        apertura = conn.execute("SELECT monto FROM caja WHERE fecha=? AND tipo='apertura'", (today,)).fetchone()
        cierre = conn.execute("SELECT id FROM caja WHERE fecha=? AND tipo='cierre'", (today,)).fetchone()
        if apertura and not cierre:
            self._caja_abierta = True
            self._saldo_inicial = apertura['monto']
        conn.close()

        style = ttk.Style()
        style.theme_use('clam')
        style.configure('Treeview', rowheight=28, font=('Segoe UI', 10))
        style.configure('Treeview.Heading', font=('Segoe UI', 10, 'bold'))

        self._show_login_screen()

    def _get_theme(self, mode='light'):
        if mode == 'dark':
            return {
                'sidebar_bg': '#121220', 'sidebar_fg': '#e0e0e0',
                'sidebar_btn': '#1e1e32', 'sidebar_btn_active': '#2a2a45',
                'main_bg': '#0d0d1a', 'card_bg': '#1a1a30',
                'fg': '#e0e0e0', 'fg_secondary': '#999',
                'accent': '#e94560', 'accent2': '#2a2a45',
                'highlight': '#533483',
                'label_bg': '#1a1a30', 'label_fg': '#e0e0e0',
                'entry_bg': '#252540', 'entry_fg': '#e0e0e0',
                'tree_bg': '#1a1a30', 'tree_fg': '#e0e0e0', 'tree_sel': '#533483',
                'button_bg': '#252540', 'button_fg': '#e0e0e0',
                'title_fg': '#e94560',
            }
        else:
            return {
                'sidebar_bg': '#2c3e50', 'sidebar_fg': 'white',
                'sidebar_btn': '#34495e', 'sidebar_btn_active': '#1abc9c',
                'main_bg': '#f0f4f8', 'card_bg': 'white',
                'fg': '#2c3e50', 'fg_secondary': '#555',
                'accent': '#1abc9c', 'accent2': '#3498db',
                'highlight': '#2c3e50',
                'label_bg': '#f0f4f8', 'label_fg': '#2c3e50',
                'entry_bg': 'white', 'entry_fg': 'black',
                'tree_bg': 'white', 'tree_fg': 'black', 'tree_sel': '#3498db',
                'button_bg': '#ecf0f1', 'button_fg': '#2c3e50',
                'title_fg': '#2c3e50',
            }

    def _toggle_theme(self):
        self._dark_mode = not self._dark_mode
        mode = 'dark' if self._dark_mode else 'light'
        self._theme = self._get_theme(mode)
        # Rebuild current view first so new widgets are created
        if hasattr(self, '_current_view'):
            view = self._current_view
            self._current_view = ''
            if view == 'dashboard':
                self.show_dashboard()
            elif view == 'ventas':
                self.show_ventas()
            elif view == 'servicios_productos':
                self.show_servicios_productos()
            elif view == 'medical':
                self.show_medical_history()
            elif view == 'creditos':
                self.show_creditos()
        self._apply_theme()
        self._reapply_ttk_styles()
        if hasattr(self, 'menu_frame'):
            for w in self.menu_frame.winfo_children():
                if isinstance(w, tk.Button) and ('Tema' in w.cget('text') or 'Tema' in str(w.cget('text'))):
                    w.configure(text='☀️ Tema Claro' if self._dark_mode else '🌙 Tema Oscuro')

    def _get_local_ip(self):
        try:
            import subprocess
            r = subprocess.run(['ipconfig'], capture_output=True, text=True, creationflags=subprocess.CREATE_NO_WINDOW)
            lines = r.stdout.split('\n')
            for line in lines:
                line = line.strip()
                if 'IPv4' in line or 'Direcci' in line:
                    parts = line.split(':')
                    if len(parts) >= 2:
                        ip = parts[-1].strip()
                        if ip.startswith('192.') or ip.startswith('10.') or ip.startswith('172.'):
                            return ip
        except:
            pass
        try:
            s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
            s.settimeout(2)
            s.connect(("8.8.8.8", 80))
            ip = s.getsockname()[0]
            s.close()
            return ip
        except:
            return "127.0.0.1"

    def _toggle_web_server(self):
        if hasattr(self, '_web_process') and self._web_process and self._web_process.poll() is None:
            self._web_process.terminate()
            self._web_process = None
            if hasattr(self, '_web_btn'):
                self._web_btn.configure(text='\U0001f310 Iniciar Web', bg='#27ae60')
            if hasattr(self, '_web_status'):
                self._web_status.configure(text='')
            if hasattr(self, '_web_info_frame') and self._web_info_frame:
                self._web_info_frame.destroy()
                self._web_info_frame = None
            return
        ip = self._get_local_ip()
        port = 5000
        script = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'servidor_web.py')
        logfile = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'server_log.txt')
        try:
            flog = open(logfile, 'w')
            self._web_process = subprocess.Popen([sys.executable, script],
                                                  stdout=flog, stderr=flog)
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo iniciar el servidor:\n{e}")
            return
        self.root.after(2000, lambda: self._check_web_started(ip, port, logfile))

    def _check_web_started(self, ip, port, logfile):
        if not hasattr(self, '_web_process') or not self._web_process:
            return
        rc = self._web_process.poll()
        if rc is not None:
            try:
                with open(logfile, 'r') as f:
                    log = f.read()
            except:
                log = "(no se pudo leer log)"
            messagebox.showerror("Error del servidor",
                                 f"El servidor se detuvo (código {rc}).\n\nLog:\n{log[:500]}")
            if hasattr(self, '_web_btn'):
                self._web_btn.configure(text='🌐 Iniciar Web', bg='#27ae60')
            if hasattr(self, '_web_status'):
                self._web_status.configure(text='')
            return
        if hasattr(self, '_web_btn'):
            self._web_btn.configure(text='\U0001f6d1 Detener Web', bg='#c0392b')
        if hasattr(self, '_web_status'):
            self._web_status.configure(text=f'http://{ip}:{port}')
        self._show_web_info(ip, port)

    def _show_web_info(self, ip, port):
        if hasattr(self, '_web_info_frame') and self._web_info_frame:
            self._web_info_frame.destroy()
            self._web_info_frame = None
        url = f'http://{ip}:{port}'
        win = tk.Toplevel(self.root)
        win.title('Acceso Doctores')
        win.configure(bg='white')
        win.resizable(False, False)
        try:
            import qrcode
            from PIL import ImageTk
            qr_img = qrcode.make(url, box_size=10, border=2)
            self._qr_popup = ImageTk.PhotoImage(qr_img)
            tk.Label(win, image=self._qr_popup, bg='white').pack(padx=25, pady=20)
        except Exception as ex:
            print('QR error:', ex)
        tk.Label(win, text=url, font=('Consolas', 13, 'bold'),
                 bg='white', fg='#2c3e50').pack(pady=(0, 5))
        tk.Label(win, text='Escanee con su celular (misma red WiFi)',
                 font=('Segoe UI', 10), bg='white', fg='#7f8c8d').pack(pady=(0, 15))
        tk.Button(win, text='Cerrar', bg='#e74c3c', fg='white',
                  font=('Segoe UI', 10), relief='flat', padx=20, pady=4,
                  command=win.destroy, cursor='hand2').pack(pady=(0, 15))
        win.update_idletasks()
        sw = win.winfo_screenwidth()
        sh = win.winfo_screenheight()
        x = (sw - win.winfo_width()) // 2
        y = (sh - win.winfo_height()) // 2
        win.geometry('+{}+{}'.format(x, y))
        win.transient(self.root)
        win.grab_set()

    def _test_web(self, ip, port):
        try:
            import urllib.request
            r = urllib.request.urlopen(f'http://{ip}:{port}/', timeout=5)
            if r.status == 200:
                messagebox.showinfo("OK", f"Servidor responde OK (HTTP {r.status})\n\nhttp://{ip}:{port}")
            else:
                messagebox.showwarning("Atencion", f"Servidor responde con codigo {r.status}")
        except Exception as e:
            if 'certificate' in str(e).lower() or 'ssl' in str(e).lower():
                messagebox.showerror("Error SSL", "Error de certificado SSL")
            elif 'refused' in str(e).lower():
                messagebox.showerror("No conecta",
                                     f"No se puede conectar al servidor.\n\n"
                                     f"1. Verifica que el servidor este iniciado (boton Detener Web)\n"
                                     f"2. Prueba desactivar el firewall de Windows temporalmente\n"
                                     f"3. Verifica que ambos esten en la misma red WiFi\n\n"
                                     f"Detalle: {str(e)[:100]}")
            else:
                messagebox.showerror("Error", f"No se pudo conectar:\n{str(e)[:200]}")

    def _reapply_ttk_styles(self):
        style = ttk.Style()
        style.theme_use('clam')
        style.configure('Treeview', background=self._theme['tree_bg'],
                        fieldbackground=self._theme['tree_bg'],
                        foreground=self._theme['tree_fg'],
                        rowheight=28, font=('Segoe UI', 10))
        style.map('Treeview', background=[('selected', self._theme['tree_sel'])],
                  foreground=[('selected', 'white')])
        style.configure('Treeview.Heading', font=('Segoe UI', 10, 'bold'),
                        background=self._theme['sidebar_btn'],
                        foreground=self._theme['sidebar_fg'])
        style.map('Treeview.Heading', background=[('active', self._theme['accent'])])
        style.configure('TCombobox', fieldbackground=self._theme['entry_bg'],
                        foreground=self._theme['entry_fg'],
                        arrowcolor=self._theme['entry_fg'])
        style.map('TCombobox', fieldbackground=[('readonly', self._theme['entry_bg'])])
        style.configure('TSpinbox', fieldbackground=self._theme['entry_bg'],
                        foreground=self._theme['entry_fg'])

    def _apply_theme(self):
        self.root.configure(bg=self._theme['main_bg'])
        # Sidebar
        if hasattr(self, 'menu_frame'):
            self.menu_frame.configure(bg=self._theme['sidebar_bg'])
            for w in self.menu_frame.winfo_children():
                if isinstance(w, tk.Label):
                    try: w.configure(bg=self._theme['sidebar_bg'], fg=self._theme['sidebar_fg'])
                    except: pass
                elif isinstance(w, tk.Button):
                    try: w.configure(bg=self._theme['sidebar_btn'], fg=self._theme['sidebar_fg'],
                                     activebackground=self._theme['sidebar_btn_active'])
                    except: pass
                elif isinstance(w, tk.Frame):
                    w.configure(bg=self._theme['sidebar_bg'])
                    for cw in w.winfo_children():
                        if isinstance(cw, tk.Entry):
                            try: cw.configure(bg=self._theme['entry_bg'], fg=self._theme['entry_fg'],
                                              insertbackground=self._theme['entry_fg'])
                            except: pass
        # Recursively update all widgets in the main area
        self._apply_theme_recursive(self.root)

    def _apply_theme_recursive(self, parent, depth=0):
        if depth > 20:
            return
        for w in parent.winfo_children():
            # Skip sidebar, it's handled separately
            if w is getattr(self, 'menu_frame', None):
                continue
            try:
                cls = w.__class__.__name__
                if cls == 'Frame':
                    try:
                        bg = w.cget('bg')
                        if bg in ('#f0f4f8', '#e8edf2', 'white', '#1a1a30', '#252540', '#121220'):
                            if self._dark_mode:
                                w.configure(bg=self._theme['card_bg'] if bg == 'white' else self._theme['main_bg'])
                            else:
                                w.configure(bg=self._theme['main_bg'] if bg in ('#1a1a30', '#252540', '#121220') else bg)
                    except: pass
                elif cls == 'Label':
                    try:
                        bg = w.cget('bg')
                        if self._dark_mode and bg in ('#f0f4f8', '#e8edf2', 'white', '#1a1a30'):
                            w.configure(bg=self._theme['label_bg'], fg=self._theme['label_fg'])
                        elif not self._dark_mode and bg in ('#1a1a30', '#252540'):
                            w.configure(bg=self._theme['label_bg'], fg=self._theme['label_fg'])
                    except: pass
                elif cls == 'Button':
                    try:
                        bg = w.cget('bg')
                        if self._dark_mode and bg not in ('#1a1a2e', '#16213e', '#121220', '#1e1e32'):
                            w.configure(bg=self._theme['button_bg'], fg=self._theme['button_fg'],
                                        activebackground=self._theme['accent'])
                        elif not self._dark_mode and bg in ('#252540', '#1e1e32', '#121220'):
                            w.configure(bg='#ecf0f1', fg='#2c3e50',
                                        activebackground='#3498db')
                    except: pass
                elif cls == 'LabelFrame':
                    try:
                        if self._dark_mode:
                            w.configure(bg=self._theme['main_bg'], fg=self._theme['fg'])
                    except: pass
                elif cls == 'Entry':
                    try:
                        if self._dark_mode:
                            w.configure(bg=self._theme['entry_bg'], fg=self._theme['entry_fg'],
                                        insertbackground=self._theme['entry_fg'])
                    except: pass
                elif cls in ('Treeview', 'Treeview.Heading'):
                    pass  # handled by ttk styles
                elif cls in ('Text', 'ScrolledText'):
                    try:
                        if self._dark_mode:
                            w.configure(bg=self._theme['entry_bg'], fg=self._theme['entry_fg'],
                                        insertbackground=self._theme['entry_fg'])
                    except: pass
            except:
                pass
            self._apply_theme_recursive(w, depth + 1)

    def _show_login_screen(self):
        LOGO_PATH = os.path.join(os.path.dirname(__file__), 'photos', 'logo.png')
        BG_PATH = os.path.expanduser(r'~\Downloads\82023-gato-raza_de_perro-cuidado_de_mascotas-refugio_de_animales-veterinario-1366x768.jpg')
        self._login_frame = tk.Frame(self.root)
        self._login_frame.pack(fill='both', expand=True)

        canvas = tk.Canvas(self._login_frame, highlightthickness=0)
        canvas.pack(fill='both', expand=True)

        entry_bg = '#ffffff'
        sw = self.root.winfo_screenwidth()
        sh = self.root.winfo_screenheight()
        try:
            pil_bg = Image.open(BG_PATH)
            if pil_bg.mode == 'RGBA':
                pil_bg = pil_bg.convert('RGB')
            cx, cy = pil_bg.width // 2, pil_bg.height // 2
            avg_color = pil_bg.getpixel((cx, cy))
            entry_bg = '#%02x%02x%02x' % avg_color
            ratio = max(sw / pil_bg.width, sh / pil_bg.height)
            new_w = int(pil_bg.width * ratio)
            new_h = int(pil_bg.height * ratio)
            pil_bg = pil_bg.resize((new_w, new_h), Image.Resampling.LANCZOS)
            left = (new_w - sw) // 2
            top = (new_h - sh) // 2
            pil_bg = pil_bg.crop((left, top, left + sw, top + sh))
            self._bg_tk = ImageTk.PhotoImage(pil_bg)
            canvas.create_image(0, 0, image=self._bg_tk, anchor='nw')
        except:
            canvas.configure(bg='#1a1a2e')

        logo_id = None
        try:
            pil_logo = Image.open(LOGO_PATH)
            max_h = int(sh * 0.35)
            ratio = max_h / pil_logo.height
            new_w = int(pil_logo.width * ratio)
            pil_logo = pil_logo.resize((new_w, max_h), Image.Resampling.LANCZOS)
            self._logo_tk = ImageTk.PhotoImage(pil_logo)
            logo_id = canvas.create_image(sw // 2, int(sh * 0.24), image=self._logo_tk)
        except:
            pass

        # Form fields directly on canvas - no dark card
        fy = int(sh * 0.47)
        lbl_font = ('Segoe UI', 10, 'bold')
        canvas.create_text(sw // 2 + 1, fy - 29, text='USUARIO', font=lbl_font, fill='#1a1a2e')
        canvas.create_text(sw // 2, fy - 30, text='USUARIO', font=lbl_font, fill='#1abc9c')
        e_user = tk.Entry(self._login_frame, width=26, font=('Segoe UI', 13), relief='flat',
                          bg=entry_bg, fg='#333', insertbackground='#333',
                          bd=0, highlightthickness=1, highlightbackground='#ccc',
                          highlightcolor='#1abc9c')
        e_user.insert(0, 'admin')
        e_user_window = canvas.create_window(sw // 2, fy, window=e_user, height=38)
        e_user.focus_set()

        fy2 = fy + 55
        canvas.create_text(sw // 2 + 1, fy2 - 29, text='CONTRASEÑA', font=lbl_font, fill='#1a1a2e')
        canvas.create_text(sw // 2, fy2 - 30, text='CONTRASEÑA', font=lbl_font, fill='#1abc9c')
        e_pass = tk.Entry(self._login_frame, width=26, font=('Segoe UI', 13),
                          show="*", relief="flat", bg=entry_bg, fg="#333",
                          insertbackground="#333", bd=0, highlightthickness=1,
                          highlightbackground="#ccc", highlightcolor="#1abc9c")
        e_pass.insert(0, 'admin123')
        e_pass_window = canvas.create_window(sw // 2, fy2, window=e_pass, height=38)

        error_var = tk.StringVar()
        lbl_error = tk.Label(self._login_frame, textvariable=error_var,
                             fg='#e94560', bg=entry_bg, font=('Segoe UI', 9))
        canvas.create_window(sw // 2, fy2 + 40, window=lbl_error)

        def do_login():
            user = e_user.get().strip()
            pwd = e_pass.get().strip()
            if not user or not pwd:
                error_var.set('Complete todos los campos')
                return
            conn = self.get_db()
            row = conn.execute(
                'SELECT id, nombre, rol FROM usuarios WHERE username=? AND password=? AND activo=1',
                (user, pwd)).fetchone()
            conn.close()
            if row:
                self._user_logged_in = True
                self._user_id = row['id']
                self._user_name = row['nombre']
                self._user_role = row['rol']
                self._login_frame.destroy()
                self.build_menu()
                self.show_dashboard()
                self._backup_database()
                self._setup_shortcuts()
                if hasattr(self, 'menu_frame'):
                    for w in self.menu_frame.winfo_children():
                        try:
                            if isinstance(w, tk.Label) and 'Usuario:' in str(w.cget('text')):
                                w.configure(text=f'Usuario: {self._user_name}')
                        except:
                            pass
            else:
                error_var.set('Usuario o contraseña incorrectos')
                e_pass.delete(0, 'end')
                e_pass.focus_set()

        e_pass.bind('<Return>', lambda e: do_login())
        btn = tk.Button(self._login_frame, text='INGRESAR', bg='#1abc9c', fg='white',
                        font=('Segoe UI', 11, 'bold'), relief='flat', padx=40, pady=8,
                        activebackground='#16a085', activeforeground='white',
                        command=do_login, cursor='hand2')
        canvas.create_window(sw // 2, fy2 + 75, window=btn)

    def _backup_database(self):
        database.backup_database()

    def _setup_shortcuts(self):
        self.root.bind('<F1>', lambda e: self.add_owner_dialog())
        self.root.bind('<F2>', lambda e: self.add_appointment_dialog())
        self.root.bind('<F3>', lambda e: self.add_medical_record_dialog())
        self.root.bind('<F4>', lambda e: self.show_dashboard())
        self.root.bind('<F5>', lambda e: self.show_ventas())
        self.root.bind('<Control-q>', lambda e: self.root.quit())
        self.root.bind('<Escape>', lambda e: self.show_dashboard())

    def get_db(self):
        return database.get_db()

    def _init_tables(self):
        database.init_db()

    # _init_doctores_data, _init_productos_default, _init_servicios_grooming_default moved to database module

    def get_doctores(self):
        conn = self.get_db()
        rows = conn.execute('SELECT nombre FROM doctores ORDER BY nombre').fetchall()
        conn.close()
        return [r['nombre'] for r in rows]

    def manage_doctores_dialog(self, callback=None):
        dialog = tk.Toplevel(self.root)
        dialog.title('Gestionar Doctores')
        dialog.geometry('360x280')
        dialog.resizable(False, False)
        dialog.configure(bg='#e8edf2')
        dialog.transient(self.root)
        dialog.grab_set()
        center_dialog(dialog, self.root)

        tk.Label(dialog, text='Doctores', font=('Segoe UI', 12, 'bold'),
                 bg='#e8edf2', fg='#2c3e50').pack(pady=(12, 5))

        lb = tk.Listbox(dialog, font=('Segoe UI', 10), height=8)
        lb.pack(fill='both', expand=True, padx=15, pady=5)

        def refresh():
            lb.delete(0, 'end')
            for d in self.get_doctores():
                lb.insert('end', d)

        refresh()

        btn_frame = tk.Frame(dialog, bg='#e8edf2')
        btn_frame.pack(fill='x', padx=15, pady=(0, 12))

        tk.Button(btn_frame, text='Agregar', bg='#2ecc71', fg='white',
                  font=('Segoe UI', 9), relief='flat', padx=12, pady=3,
                  cursor='hand2',
                  command=lambda: self._add_doctor_dialog(dialog, refresh, callback)
                  ).pack(side='left', padx=2)
        tk.Button(btn_frame, text='Eliminar', bg='#e74c3c', fg='white',
                  font=('Segoe UI', 9), relief='flat', padx=12, pady=3,
                  cursor='hand2',
                  command=lambda: self._delete_doctor(lb, dialog, refresh, callback)
                  ).pack(side='left', padx=2)
        tk.Button(btn_frame, text='Cerrar', bg='#95a5a6', fg='white',
                  font=('Segoe UI', 9), relief='flat', padx=12, pady=3,
                  cursor='hand2', command=dialog.destroy).pack(side='right', padx=2)

    def _add_doctor_dialog(self, parent_dialog, refresh_cb, final_cb=None):
        d2 = tk.Toplevel(parent_dialog)
        d2.title('Nuevo Doctor')
        d2.geometry('300x120')
        d2.resizable(False, False)
        d2.configure(bg='#e8edf2')
        d2.transient(parent_dialog)
        d2.grab_set()
        center_dialog(d2, parent_dialog)
        tk.Label(d2, text='Nombre del Doctor:', bg='#e8edf2', font=('Segoe UI', 10)).pack(pady=(15, 5))
        e_name = tk.Entry(d2, width=35, font=('Segoe UI', 10))
        e_name.pack()
        e_name.focus_set()
        def save():
            name = e_name.get().strip()
            if not name:
                messagebox.showwarning('Error', 'Ingrese un nombre')
                return
            conn = self.get_db()
            try:
                conn.execute('INSERT INTO doctores (nombre) VALUES (?)', (name,))
                conn.commit()
                conn.close()
                d2.destroy()
                refresh_cb()
                if final_cb:
                    final_cb()
            except sqlite3.IntegrityError:
                conn.close()
                messagebox.showwarning('Error', 'Ese doctor ya existe')
        add_keyboard_shortcuts(d2, save)
        tk.Button(d2, text='Guardar', bg='#2ecc71', fg='white',
                  font=('Segoe UI', 10), relief='flat', padx=15, pady=3,
                  command=save, cursor='hand2').pack(pady=(10, 5))

    def _delete_doctor(self, lb, parent_dialog, refresh_cb, final_cb=None):
        sel = lb.curselection()
        if not sel:
            messagebox.showwarning('Error', 'Seleccione un doctor')
            return
        name = lb.get(sel[0])
        if messagebox.askyesno('Confirmar', f'Eliminar a {name}?'):
            conn = self.get_db()
            conn.execute('DELETE FROM doctores WHERE nombre=?', (name,))
            conn.commit()
            conn.close()
            refresh_cb()
            if final_cb:
                final_cb()

    def clear_frame(self):
        for w in self.content_area.winfo_children():
            w.destroy()

    def build_menu(self):
        self.menu_frame = tk.Frame(self.root, bg='#2c3e50', width=200)
        self.menu_frame.pack(side='left', fill='y')
        self.menu_frame.pack_propagate(False)
        self.content_area = tk.Frame(self.root, bg='#f0f4f8')
        self.content_area.pack(side='left', fill='both', expand=True)

        tk.Label(self.menu_frame, text='SCRIPTYFY', font=('Segoe UI', 13, 'bold'),
                 bg='#2c3e50', fg='white').pack(pady=(15, 5))
        tk.Label(self.menu_frame, text='V.02', font=('Segoe UI', 9),
                 bg='#2c3e50', fg='#1abc9c').pack(pady=(0, 15))

        tk.Label(self.menu_frame, text=f'Usuario: {self._user_name}',
                 font=('Segoe UI', 8), bg='#2c3e50', fg='#888').pack(pady=(0, 10))

        # Global search
        search_frame = tk.Frame(self.menu_frame, bg='#2c3e50')
        search_frame.pack(fill='x', padx=8, pady=(0, 10))
        e_search = tk.Entry(search_frame, font=('Segoe UI', 9), bg='#34495e', fg='white',
                            insertbackground='white', relief='flat')
        e_search.pack(side='left', fill='x', expand=True, ipady=3, padx=(0, 3))
        e_search.insert(0, 'Buscar...')

        def on_search_focus_in(e):
            if e_search.get() == 'Buscar...':
                e_search.delete(0, 'end')
                e_search.config(fg='white')

        def on_search_focus_out(e):
            if not e_search.get().strip():
                e_search.insert(0, 'Buscar...')
                e_search.config(fg='#888')

        e_search.bind('<FocusIn>', on_search_focus_in)
        e_search.bind('<FocusOut>', on_search_focus_out)
        e_search.bind('<Return>', lambda e: self._global_search(e_search.get() if e_search.get() != 'Buscar...' else ''))

        tk.Button(search_frame, text='🔍', font=('Segoe UI', 11), bg='#1abc9c', fg='white',
                  relief='flat', padx=6, pady=0, cursor='hand2',
                  command=lambda: self._global_search(e_search.get() if e_search.get() != 'Buscar...' else '')).pack(side='left')

        role = self._user_role

        if role == 'recepcion':
            nav_buttons = [
                ('🏠 Inicio', self.show_dashboard),
                ('💰 Cobros Pendientes', self.show_cobros_pendientes),
                ('📅 Citas', self.add_appointment_dialog),
                ('💰 Ventas / Caja', self.show_ventas),
            ]
        elif role == 'doctor':
            nav_buttons = [
                ('🏠 Inicio', self.show_dashboard),
                ('📋 Historial Clinico', self.show_medical_history),
                ('📅 Citas', self.add_appointment_dialog),
            ]
        elif role == 'admin':
            nav_buttons = [
                ('🏠 Inicio', self.show_dashboard),
                ('📋 Historial Clinico', self.show_medical_history),
                ('📦 Servicios y Productos', self.show_servicios_productos),
                ('📋 Creditos', self.show_creditos),
                ('💰 Ventas / Caja', self.show_ventas),
                ('💰 Cobros Pendientes', self.show_cobros_pendientes),
            ]
        else:
            nav_buttons = [
                ('🏠 Inicio', self.show_dashboard),
                ('📋 Historial Clinico', self.show_medical_history),
                ('📦 Servicios y Productos', self.show_servicios_productos),
                ('📋 Creditos', self.show_creditos),
                ('💰 Ventas / Caja', self.show_ventas),
            ]
        for txt, cmd in nav_buttons:
            tk.Button(self.menu_frame, text=txt, font=('Segoe UI', 10),
                      bg='#34495e', fg='white', relief='flat', anchor='w', padx=15,
                      activebackground='#1abc9c', activeforeground='white',
                      command=cmd, cursor='hand2').pack(fill='x', padx=8, pady=3)

        tk.Frame(self.menu_frame, bg='#34495e', height=1).pack(fill='x', padx=15, pady=15)

        if role == 'admin':
            quick_buttons = [
                ('Doctores', lambda: self.manage_doctores_dialog(), '#3498db'),
                ('👥 Usuarios', lambda: self.manage_usuarios_dialog(), '#9b59b6'),
            ]
        else:
            quick_buttons = [
                ('Doctores', lambda: self.manage_doctores_dialog(), '#3498db'),
            ]
        for txt, cmd, color in quick_buttons:
            tk.Button(self.menu_frame, text=txt, font=('Segoe UI', 9),
                      bg=color, fg='white', relief='flat', anchor='w', padx=15,
                      command=cmd, cursor='hand2').pack(fill='x', padx=8, pady=2)

        tk.Frame(self.menu_frame, bg='#34495e', height=1).pack(fill='x', padx=15, pady=10)
        tk.Button(self.menu_frame, text='🌙 Tema Oscuro',
                  font=('Segoe UI', 9), bg='#2c3e50', fg='white', relief='flat', padx=15,
                  command=self._toggle_theme, cursor='hand2').pack(fill='x', padx=8, pady=2)
        if self._user_role in ('admin', 'recepcion'):
            self._web_btn = tk.Button(self.menu_frame, text='🌐 Iniciar Web',
                      font=('Segoe UI', 9), bg='#27ae60', fg='white', relief='flat', padx=15,
                      command=self._toggle_web_server, cursor='hand2')
            self._web_btn.pack(fill='x', padx=8, pady=(2, 0))
            self._web_status = tk.Label(self.menu_frame, text='', font=('Segoe UI', 8),
                                         bg='#2c3e50', fg='#aaa')
            self._web_status.pack(fill='x', padx=8, pady=(0, 2))
        tk.Button(self.menu_frame, text='🚪 Cerrar Sesion',
                  font=('Segoe UI', 9), bg='#c0392b', fg='white', relief='flat', padx=15,
                  command=self._logout, cursor='hand2').pack(fill='x', padx=8, pady=(2, 10))

    def toggle_theme(self):
        self._dark_mode = not self._dark_mode
        if self._dark_mode:
            self._theme_bg = '#2c3e50'
            self.root.configure(bg='#2c3e50')
        else:
            self._theme_bg = '#e8edf2'
            self.root.configure(bg='#e8edf2')

    def _logout(self):
        if messagebox.askyesno("Cerrar Sesion", "Seguro de cerrar sesion?"):
            self.menu_frame.destroy()
            self.clear_frame()
            self.content_area.destroy()
            self._user_logged_in = False
            self._user_id = None
            self._user_name = 'Invitado'
            self._user_role = 'empleado'
            self._current_view = ''
            self._show_login_screen()

    # ---------- INICIO ----------
    def show_dashboard(self):
        self.root.unbind_all('<MouseWheel>')
        self._current_view = 'dashboard'
        self.clear_frame()
        main = tk.Frame(self.content_area, bg='#f0f4f8')
        main.pack(fill='both', expand=True, padx=20, pady=10)

        dash = tk.Frame(main, bg='#f0f4f8')
        dash.pack(fill='both', expand=True)

        logo_path = os.path.join('photos', 'logo.png')
        self._dashboard_logo = None
        if os.path.exists(logo_path):
            pil_img = Image.open(logo_path).resize((70, 70), Image.LANCZOS)
            self._dashboard_logo = ImageTk.PhotoImage(pil_img)

        header = tk.Frame(dash, bg='#f0f4f8')
        header.pack(anchor='w', pady=(0, 8))
        if self._dashboard_logo:
            tk.Label(header, image=self._dashboard_logo, bg='#f0f4f8').pack(side='left', padx=(0, 10))
        tk.Label(header, text='Panel Principal', font=('Segoe UI', 14, 'bold'),
                 bg='#f0f4f8', fg='#2c3e50').pack(side='left')
        tk.Label(header, text='{} | {}'.format(self._user_name, self._user_role.capitalize()),
                 font=('Segoe UI', 9), bg='#f0f4f8', fg='#7f8c8d').pack(side='right', padx=(0, 5))

        import calendar as cal_mod
        hoy = date.today()

        conn = self.get_db()
        total_d = conn.execute('SELECT COUNT(*) FROM duenos').fetchone()[0]
        total_a = conn.execute('SELECT COUNT(*) FROM animales').fetchone()[0]
        citas_vet_hoy = conn.execute(
            "SELECT COUNT(*) FROM citas WHERE fecha = ? AND estado = 'pendiente' AND (tipo = 'veterinaria' OR tipo = '')",
            (str(date.today()),)).fetchone()[0]
        citas_grooming_hoy = conn.execute(
            "SELECT COUNT(*) FROM citas WHERE fecha = ? AND estado = 'pendiente' AND tipo = 'grooming'",
            (str(date.today()),)).fetchone()[0]
        ingresos_hoy = conn.execute(
            'SELECT COALESCE(SUM(total), 0) FROM ventas WHERE fecha = ?',
            (str(date.today()),)).fetchone()[0]
        stock_bajo = conn.execute(
            'SELECT COUNT(*) FROM productos WHERE stock <= 3 AND activo = 1').fetchone()[0]
        creditos_pend = conn.execute(
            "SELECT COUNT(*), COALESCE(SUM(saldo), 0) FROM creditos WHERE estado = 'pendiente'").fetchone()
        cumple_mes = conn.execute(
            "SELECT a.nombre, a.especie, a.fecha_nacimiento, d.nombre as dueno "
            "FROM animales a LEFT JOIN duenos d ON a.id_dueno = d.id "
            "WHERE a.fecha_nacimiento != '' AND a.fecha_nacimiento IS NOT NULL "
            "AND substr(a.fecha_nacimiento, 6, 2) = ? ORDER BY a.fecha_nacimiento",
            ('{:02d}'.format(hoy.month),)).fetchall()
        conn.close()

        # --- ALERTAS VISIBLES ---
        alerts = []
        if citas_vet_hoy + citas_grooming_hoy > 0:
            alerts.append(('info', '{} cita(s) pendiente(s) para hoy'.format(citas_vet_hoy + citas_grooming_hoy)))
        if creditos_pend[0] > 0:
            alerts.append(('warning', '{} credito(s) pendiente(s) por S/.{:.2f}'.format(creditos_pend[0], creditos_pend[1])))
        if cumple_mes:
            alerts.append(('info', '{} mascota(s) cumplen anos este mes'.format(len(cumple_mes))))
        if alerts:
            alert_frame = tk.Frame(dash, bg='#f0f4f8')
            alert_frame.pack(fill='x', pady=(0, 8))
            for level, msg in alerts:
                colors = {'info': ('#d6eaf8', '#2980b9'), 'danger': ('#fadbd8', '#c0392b'), 'warning': ('#fdebd0', '#d35400')}
                bg_c, fg_c = colors.get(level, ('#fff', '#333'))
                al = tk.Frame(alert_frame, bg=bg_c, highlightbackground=fg_c, highlightthickness=1, padx=12, pady=6)
                al.pack(fill='x', pady=2)
                tk.Label(al, text=msg, font=('Segoe UI', 10, 'bold'),
                         bg=bg_c, fg=fg_c, anchor='w').pack(side='left')

        # --- CARDS CLIQUEABLES ---
        cards = tk.Frame(dash, bg='#f0f4f8')
        cards.pack(fill='x')

        citas_hoy_total = citas_vet_hoy + citas_grooming_hoy
        card_info = [
            ('#3498db', total_d, 'Due\u00f1os registrados', lambda: self.add_owner_dialog()),
            ('#2ecc71', total_a, 'Animales registrados', self.show_medical_history),
            ('#f39c12', citas_hoy_total, 'Citas Hoy', self.add_appointment_dialog),
            ('#1abc9c', 'S/.{:.2f}'.format(ingresos_hoy), 'Ingresos Hoy', self.show_ventas),
            ('#e74c3c', stock_bajo, 'Stock Bajo', self.show_servicios_productos),
        ]
        for bg_c, val, lab, cmd in card_info:
            shadow = tk.Frame(cards, bg='#d5d8dc', padx=2, pady=2)
            shadow.pack(side='left', fill='x', expand=True, padx=4)
            card = tk.Frame(shadow, bg='white', padx=12, pady=5)
            card.pack(fill='x', expand=True)
            card.bind('<Button-1>', lambda e, c=cmd: c())
            lb_val = tk.Label(card, text=str(val), font=('Segoe UI', 22, 'bold'),
                              bg='white', fg=bg_c)
            lb_val.pack()
            lb_val.bind('<Button-1>', lambda e, c=cmd: c())
            lb_lab = tk.Label(card, text=lab, font=('Segoe UI', 8), bg='white', fg='#7f8c8d')
            lb_lab.pack()
            lb_lab.bind('<Button-1>', lambda e, c=cmd: c())
            tk.Frame(card, bg=bg_c, height=3).pack(fill='x', pady=(3, 0))
            shadow.configure(cursor='hand2')

        # --- BOTTOM SPLIT ---
        bottom_split = tk.Frame(dash, bg='#f0f4f8')
        bottom_split.pack(fill='both', expand=True, pady=(8, 0))

        left_col = tk.Frame(bottom_split, bg='#f0f4f8')
        left_col.pack(side='left', fill='both', expand=True, padx=(0, 6))

        # --- CITAS PENDIENTES + CALENDARIO ---

        citas_frame = tk.Frame(left_col, bg='#f0f4f8')
        citas_frame.pack(fill='x', pady=(0, 5))

        # Mini calendar filter (never destroyed)
        cal_filter_frame = tk.Frame(citas_frame, bg='white', highlightbackground='#ddd',
                                    highlightthickness=1)
        cal_filter_frame.pack(fill='x', pady=(0, 5))

        conn = self.get_db()
        dias_citas = set()
        for r in conn.execute("SELECT DISTINCT fecha FROM citas WHERE estado='pendiente'").fetchall():
            dias_citas.add(r['fecha'])
        conn.close()

        cal_inner = tk.Frame(cal_filter_frame, bg='white', padx=6, pady=4)
        cal_inner.pack(fill='x')
        mes_actual = cal_mod.month_abbr[hoy.month]
        tk.Label(cal_inner, text='{} {} (clic en fecha para filtrar)'.format(mes_actual, hoy.year),
                 font=('Segoe UI', 8, 'bold'), bg='white', fg='#2c3e50').pack()
        day_header = tk.Frame(cal_inner, bg='white')
        day_header.pack()
        for d in ['Lu', 'Ma', 'Mi', 'Ju', 'Vi', 'Sa', 'Do']:
            tk.Label(day_header, text=d, font=('Segoe UI', 6, 'bold'),
                     bg='#ecf0f1', fg='#555', width=3, relief='ridge').pack(side='left', padx=1)
        cal_grid = tk.Frame(cal_inner, bg='white')
        cal_grid.pack(pady=(2, 0))
        _, days_in_month = cal_mod.monthrange(hoy.year, hoy.month)
        first_weekday = cal_mod.monthrange(hoy.year, hoy.month)[0]
        selected_date = tk.StringVar(value='')
        row_fr = None
        for d in range(1, days_in_month + 1):
            wd = (first_weekday + d - 1) % 7
            if wd == 0 or d == 1:
                row_fr = tk.Frame(cal_grid, bg='white')
                row_fr.pack()
                if d == 1:
                    for _ in range(first_weekday):
                        tk.Label(row_fr, text='', width=3, bg='white').pack(side='left', padx=1)
            date_str = '{:04d}-{:02d}-{:02d}'.format(hoy.year, hoy.month, d)
            has_cita = date_str in dias_citas
            if d == hoy.day and has_cita:
                bg_day = '#1abc9c'
                fg_day = 'white'
            elif d == hoy.day:
                bg_day = '#3498db'
                fg_day = 'white'
            elif has_cita:
                bg_day = '#2ecc71'
                fg_day = 'white'
            else:
                bg_day = 'white'
                fg_day = '#555'
            lb = tk.Label(row_fr, text=str(d), width=3, bg=bg_day, fg=fg_day,
                          font=('Segoe UI', 6, 'bold') if has_cita else ('Segoe UI', 6),
                          cursor='hand2')
            lb.pack(side='left', padx=1)
            if has_cita:
                def on_click_date(e, fs=date_str):
                    try:
                        conn2 = self.get_db()
                        rows2 = conn2.execute(
                            "SELECT a.nombre as animal, d.nombre as dueno, d.telefono, c.motivo, c.tipo "
                            "FROM citas c JOIN animales a ON c.id_animal = a.id "
                            "JOIN duenos d ON c.id_dueno = d.id "
                            "WHERE c.fecha = ? AND c.estado = 'pendiente'",
                            (fs,)).fetchall()
                        conn2.close()
                        if not rows2:
                            messagebox.showinfo('Sin citas', 'No hay citas para esta fecha')
                            return
                        lines = ['Citas del {}:'.format(fs), '']
                        for r in rows2:
                            tipo = r['tipo'] or 'veterinaria'
                            tel = r['telefono'] or '---'
                            lines.append('  {} - Dueno: {} (Tel: {}) - {}'.format(r['animal'], r['dueno'], tel, r['motivo']))
                        messagebox.showinfo('Citas del dia', '\n'.join(lines))
                    except Exception as ex:
                        messagebox.showerror('Error', str(ex))
                lb.bind('<Button-1>', on_click_date)

        # --- LEFT COLUMN BOTTOM: TOP VENDIDO + CUMPLEANOS + FRECUENTES ---
        left_bottom = tk.Frame(left_col, bg='#f0f4f8')
        left_bottom.pack(fill='x')

        # Sub-row 1: Top Vendido + Cumpleaños
        lr1 = tk.Frame(left_bottom, bg='#f0f4f8')
        lr1.pack(fill='x', pady=(0, 5))
        conn = self.get_db()
        top_items = conn.execute(
            "SELECT vi.nombre, SUM(vi.cantidad) as total_q, SUM(vi.subtotal) as total_s "
            "FROM venta_items vi JOIN ventas v ON vi.id_venta = v.id "
            "WHERE v.fecha >= ? "
            "GROUP BY vi.nombre ORDER BY total_q DESC LIMIT 5",
            (str(date.today()),)).fetchall()
        conn.close()
        top_frame = tk.Frame(lr1, bg='white', highlightbackground='#ddd', highlightthickness=1)
        top_frame.pack(side='left', fill='x', expand=True, padx=(0, 3))
        tk.Label(top_frame, text='Top Vendido Hoy', font=('Segoe UI', 9, 'bold'),
                 bg='white', fg='#2c3e50', padx=6, pady=3).pack(anchor='w')
        top_inner = tk.Frame(top_frame, bg='white')
        top_inner.pack(fill='x', padx=6, pady=(0, 3))
        if top_items:
            for i, r in enumerate(top_items, 1):
                tk.Label(top_inner, text='  {}. {} ({} und.)'.format(i, r['nombre'], r['total_q']),
                    font=('Segoe UI', 7), bg='white', fg='#555', anchor='w').pack(fill='x', pady=1)
        else:
            tk.Label(top_inner, text='  Sin ventas hoy', font=('Segoe UI', 7), bg='white', fg='#999').pack(anchor='w')

        cumple_frame = tk.Frame(lr1, bg='white', highlightbackground='#ddd', highlightthickness=1)
        cumple_frame.pack(side='left', fill='x', expand=True, padx=(3, 0))
        tk.Label(cumple_frame, text='Cumpleaños del Mes', font=('Segoe UI', 9, 'bold'),
                 bg='white', fg='#e74c3c', padx=6, pady=3).pack(anchor='w')
        cumple_inner = tk.Frame(cumple_frame, bg='white')
        cumple_inner.pack(fill='x', padx=6, pady=(0, 3))
        if cumple_mes:
            for r in cumple_mes:
                tk.Label(cumple_inner, text='  {n} ({e}) - {d}'.format(n=r['nombre'], e=r['especie'], d=r['dueno'] or '---'),
                    font=('Segoe UI', 7), bg='white', fg='#555', anchor='w').pack(fill='x', pady=1)
        else:
            tk.Label(cumple_inner, text='  Sin cumpleaños este mes', font=('Segoe UI', 7), bg='white', fg='#999').pack(anchor='w')

        # Sub-row 2: Pacientes Frecuentes
        conn = self.get_db()
        frecuentes = conn.execute(
            "SELECT a.nombre, a.especie, COUNT(rm.id) as visitas "
            "FROM animales a JOIN registros_medicos rm ON a.id = rm.id_animal "
            "GROUP BY a.id ORDER BY visitas DESC LIMIT 5").fetchall()
        conn.close()
        freq_frame = tk.Frame(left_bottom, bg='white', highlightbackground='#ddd', highlightthickness=1)
        freq_frame.pack(fill='x', pady=(0, 5))
        tk.Label(freq_frame, text='Pacientes mas Frecuentes', font=('Segoe UI', 9, 'bold'),
                 bg='white', fg='#8e44ad', padx=6, pady=3).pack(anchor='w')
        freq_inner = tk.Frame(freq_frame, bg='white')
        freq_inner.pack(fill='x', padx=6, pady=(0, 3))
        if frecuentes:
            for i, r in enumerate(frecuentes, 1):
                tk.Label(freq_inner, text='  {}. {} ({}) - {} visitas'.format(i, r['nombre'], r['especie'], r['visitas']),
                    font=('Segoe UI', 7), bg='white', fg='#555', anchor='w').pack(fill='x', pady=1)
        else:
            tk.Label(freq_inner, text='  Sin consultas registradas', font=('Segoe UI', 7), bg='white', fg='#999').pack(anchor='w')

        # --- RIGHT COLUMN ---
        right_col = tk.Frame(bottom_split, bg='#f0f4f8')
        right_col.pack(side='left', fill='both', expand=True)

        # --- SEMANAL INCOME CHART ---
        chart_frame = tk.Frame(right_col, bg='white', highlightbackground='#ddd',
                               highlightthickness=1)
        chart_frame.pack(fill='x', pady=(0, 5))
        tk.Label(chart_frame, text='Ingresos Semanales', font=('Segoe UI', 10, 'bold'),
                 bg='white', fg='#2c3e50', padx=8, pady=4).pack(anchor='w')
        try:
            from matplotlib.figure import Figure
            from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
            conn_stats = self.get_db()
            from datetime import timedelta
            dias_semana = []
            montos = []
            for i in range(6, -1, -1):
                d = hoy - timedelta(days=i)
                dias_semana.append(d.strftime('%a'))
                r = conn_stats.execute(
                    'SELECT COALESCE(SUM(total), 0) FROM ventas WHERE fecha = ?',
                    (d.isoformat(),)).fetchone()[0]
                montos.append(r)
            conn_stats.close()

            fig = Figure(figsize=(4, 1.5), dpi=85, facecolor='white')
            ax = fig.add_subplot(111)
            ax.plot(dias_semana, montos, color='#1abc9c', marker='o', linewidth=2)
            ax.fill_between(range(len(montos)), montos, alpha=0.15, color='#1abc9c')
            ax.set_ylabel('S/.', fontsize=7)
            ax.tick_params(axis='both', labelsize=7)
            fig.tight_layout(pad=1)
            canvas_chart = FigureCanvasTkAgg(fig, master=chart_frame)
            canvas_chart.draw()
            canvas_chart.get_tk_widget().pack(fill='x', padx=5, pady=5)
        except Exception as ex:
            tk.Label(chart_frame, text='Grafico no disponible', bg='white',
                     fg='#999').pack(expand=True)

        # --- INGRESOS MENSUALES (barra de progreso) ---
        conn = self.get_db()
        ingreso_mes = conn.execute(
            "SELECT COALESCE(SUM(total), 0) FROM ventas WHERE strftime('%Y-%m', fecha) = ?",
            (hoy.strftime('%Y-%m'),)).fetchone()[0]
        conn.close()
        meta_mensual = 15000.0
        pct = min(ingreso_mes / meta_mensual * 100, 100) if meta_mensual > 0 else 0
        ingreso_frame = tk.Frame(right_col, bg='white', highlightbackground='#ddd',
                                 highlightthickness=1)
        ingreso_frame.pack(fill='x', pady=(0, 5))
        tk.Label(ingreso_frame, text='Ingresos del Mes', font=('Segoe UI', 10, 'bold'),
                 bg='white', fg='#2c3e50', padx=8, pady=4).pack(anchor='w')
        bar_bg = tk.Frame(ingreso_frame, bg='#ecf0f1', height=16, padx=0, pady=0)
        bar_bg.pack(fill='x', padx=8, pady=4)
        bar_px = int(pct / 100 * 400)
        bar_fill = tk.Frame(bar_bg, bg='#2ecc71', width=bar_px, height=16)
        bar_fill.pack(side='left')
        tk.Label(ingreso_frame, text='S/.{:.2f} de S/.{:.2f} ({:.0f}%)'.format(ingreso_mes, meta_mensual, pct),
                 font=('Segoe UI', 8), bg='white', fg='#7f8c8d', padx=8, pady=4).pack(anchor='w')

        # --- ACTIVIDAD RECIENTE (con fotos) ---
        feed_frame = tk.Frame(right_col, bg='white', highlightbackground='#ddd',
                              highlightthickness=1)
        feed_frame.pack(fill='x', pady=(0, 5))
        tk.Label(feed_frame, text='Actividad Reciente', font=('Segoe UI', 10, 'bold'),
                 bg='white', fg='#2c3e50', padx=8, pady=4).pack(anchor='w')
        conn3 = self.get_db()
        ultimas = conn3.execute(
            'SELECT rm.fecha, rm.hora, a.nombre as animal, rm.diagnostico '
            'FROM registros_medicos rm JOIN animales a ON rm.id_animal = a.id '
            'ORDER BY rm.fecha DESC, rm.hora DESC LIMIT 10').fetchall()
        nuevos = conn3.execute(
            'SELECT id, nombre, especie, foto FROM animales ORDER BY id DESC LIMIT 5').fetchall()
        conn3.close()
        feed_inner = tk.Frame(feed_frame, bg='white')
        feed_inner.pack(fill='both', expand=True, padx=8, pady=5)
        if ultimas:
            tk.Label(feed_inner, text='Ultimas Consultas', font=('Segoe UI', 9, 'bold'),
                     bg='white', fg='#7f8c8d').pack(anchor='w', pady=(0, 3))
            for r in ultimas[:5]:
                tk.Label(feed_inner,
                    text='{fecha} - {animal}: {diagnostico}'.format(**r),
                    font=('Segoe UI', 9), bg='white', fg='#555',
                    anchor='w', wraplength=300, justify='left').pack(fill='x', pady=1)
        if nuevos:
            tk.Label(feed_inner, text='Nuevos Pacientes', font=('Segoe UI', 9, 'bold'),
                     bg='white', fg='#7f8c8d').pack(anchor='w', pady=(5, 3))
            for r in nuevos:
                row_f = tk.Frame(feed_inner, bg='white')
                row_f.pack(fill='x', pady=2)
                # Show photo thumbnail if available
                photo_label = None
                if r['foto'] and os.path.exists(r['foto']):
                    try:
                        pimg = Image.open(r['foto']).resize((24, 24), Image.LANCZOS)
                        pphoto = ImageTk.PhotoImage(pimg)
                        photo_label = tk.Label(row_f, image=pphoto, bg='white')
                        photo_label.image = pphoto
                    except:
                        pass
                if photo_label:
                    photo_label.pack(side='left', padx=(0, 5))
                else:
                    tk.Label(row_f, text='  ', bg='white').pack(side='left')
                tk.Label(row_f, text='{nombre} ({especie})'.format(**r),
                         font=('Segoe UI', 9), bg='white', fg='#555',
                         anchor='w').pack(side='left', fill='x')

        # --- CREDITOS PENDIENTES (deudores) ---
        conn = self.get_db()
        deudores = conn.execute(
            "SELECT c.id, c.saldo, c.fecha_vencimiento, d.nombre as dueno "
            "FROM creditos c LEFT JOIN duenos d ON c.id_cliente = d.id "
            "WHERE c.estado = 'pendiente' ORDER BY c.fecha_vencimiento ASC LIMIT 5").fetchall()
        conn.close()
        if deudores:
            deud_frame = tk.Frame(right_col, bg='white', highlightbackground='#e74c3c',
                                  highlightthickness=1)
            deud_frame.pack(fill='x', pady=(0, 5))
            tk.Label(deud_frame, text='Deudores / Creditos Pendientes', font=('Segoe UI', 10, 'bold'),
                     bg='white', fg='#e74c3c', padx=8, pady=4).pack(anchor='w')
            deud_inner = tk.Frame(deud_frame, bg='white')
            deud_inner.pack(fill='x', padx=8, pady=5)
            for r in deudores:
                tk.Label(deud_inner,
                    text='{dueno} - S/.{saldo:.2f} (vence: {fecha_vencimiento})'.format(**r),
                    font=('Segoe UI', 9), bg='white', fg='#c0392b', anchor='w').pack(fill='x', pady=1)

    def show_medical_history(self):
        self._current_view = 'medical'
        self.clear_frame()
        main = tk.Frame(self.content_area, bg="#f0f4f8")
        main.pack(fill="both", expand=True, padx=20, pady=15)

        # Top bar
        top_bar = tk.Frame(main, bg="#f0f4f8")
        top_bar.pack(fill="x")
        tk.Label(top_bar, text="Historial Clinico", font=("Segoe UI", 16, "bold"),
                 bg="#f0f4f8", fg="#2c3e50").pack(side="left")

        for txt, cmd, color in [("+ NUEVO", self.add_animal_dialog, "#2ecc71")]:
            tk.Button(top_bar, text=txt, bg=color, fg="white", font=("Segoe UI", 9),
                      relief="flat", padx=10, pady=4, command=cmd,
                      cursor="hand2").pack(side="right", padx=2)

        content = tk.Frame(main, bg="#f0f4f8")
        content.pack(fill="both", expand=True, pady=(10, 0))

        # Left panel - patient list
        left = tk.Frame(content, bg="white", width=280,
                        highlightbackground="#ddd", highlightthickness=1)
        left.pack(side="left", fill="y", padx=(0, 10))
        left.pack_propagate(False)

        tk.Label(left, text="Pacientes", font=("Segoe UI", 12, "bold"),
                 bg="#2c3e50", fg="white", padx=10, pady=8).pack(fill="x")

        search_frame = tk.Frame(left, bg="white", padx=8, pady=5)
        search_frame.pack(fill="x")
        tk.Label(search_frame, text="Buscar:", bg="white", font=("Segoe UI", 9)).pack(anchor="w")
        search_entry = tk.Entry(search_frame, font=("Segoe UI", 10))
        search_entry.pack(fill="x")

        especie_frame = tk.Frame(left, bg="white")
        especie_frame.pack(fill="x", padx=8, pady=(0, 5))
        tk.Label(especie_frame, text="Filtrar:", bg="white", font=("Segoe UI", 9)).pack(side="left", padx=(0, 5))
        filtro_especie = ttk.Combobox(especie_frame, values=["Todos", "Perro", "Gato"],
                                      width=20, font=("Segoe UI", 9), state="readonly")
        filtro_especie.set("Todos")
        filtro_especie.pack(side="left")

        list_frame = tk.Frame(left, bg="white")
        list_frame.pack(fill="both", expand=True, padx=5, pady=5)

        canvas = tk.Canvas(list_frame, bg="white", highlightthickness=0)
        scroll = ttk.Scrollbar(list_frame, orient="vertical", command=canvas.yview)
        scrollable = tk.Frame(canvas, bg="white")
        scrollable.bind("<Configure>",
                        lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.create_window((0, 0), window=scrollable, anchor="nw")
        canvas.configure(yscrollcommand=scroll.set)
        canvas.pack(side="left", fill="both", expand=True)
        scroll.pack(side="right", fill="y")

        # Right panel - patient detail
        right = tk.Frame(content, bg="white",
                         highlightbackground="#ddd", highlightthickness=1)
        right.pack(side="left", fill="both", expand=True)

        detail_frame = tk.Frame(right, bg="#f0f4f8")
        detail_frame.pack(fill="both", expand=True, padx=10, pady=10)

        tk.Label(detail_frame, text="Seleccione un paciente de la lista",
                 font=("Segoe UI", 14), bg="#f0f4f8", fg="#999").pack(expand=True)

        self._selected_card = None

        def load_patients(filter_text="", especie_filtro="Todos"):
            self._selected_card = None
            for w in scrollable.winfo_children():
                w.destroy()
            conn = self.get_db()
            query = ("SELECT a.id, a.nombre, a.especie, a.raza, a.edad, a.foto, "
                     "d.nombre as dueno_nombre, d.telefono as dueno_telefono, d.dni as dueno_dni "
                     "FROM animales a JOIN duenos d ON a.id_dueno = d.id ")
            params = []
            if especie_filtro != "Todos":
                query += "WHERE a.especie = ? "
                params.append(especie_filtro)
            query += "ORDER BY a.nombre"
            rows = conn.execute(query, params).fetchall()
            conn.close()
            for a in rows:
                if filter_text:
                    fl = filter_text.lower()
                    if (fl not in a["nombre"].lower() and
                        fl not in a["dueno_nombre"].lower() and
                        fl not in (a["dueno_dni"] or "")):
                        continue
                card = tk.Frame(scrollable, bg="white",
                                highlightbackground="#e8e8e8",
                                highlightthickness=1, padx=8, pady=6, cursor="hand2")
                card.pack(fill="x", pady=2)
                if a["foto"] and os.path.exists(os.path.join(PHOTOS_DIR, a["foto"])):
                    try:
                        img = Image.open(os.path.join(PHOTOS_DIR, a["foto"]))
                        img.thumbnail((40, 40))
                        tk_img = ImageTk.PhotoImage(img)
                        foto_lbl = tk.Label(card, image=tk_img, bg="#ddd")
                        foto_lbl.image = tk_img
                    except:
                        foto_lbl = tk.Label(card, text="  ", bg="#ddd", width=4, height=2)
                else:
                    foto_lbl = tk.Label(card, text="  ", bg="#ddd", width=4, height=2)
                foto_lbl.pack(side="left", padx=(0, 8))
                text_frame = tk.Frame(card, bg="white")
                text_frame.pack(side="left", fill="x", expand=True)
                tk.Label(text_frame, text=a["nombre"],
                         font=("Segoe UI", 10, "bold"),
                         bg="white", fg="#2c3e50").pack(anchor="w")
                tk.Label(text_frame,
                         text=a["especie"] + " - " + a["dueno_nombre"] +
                               (" | Tel: " + a["dueno_telefono"] if a["dueno_telefono"] else ""),
                         font=("Segoe UI", 8), bg="white",
                         fg="#7f8c8d").pack(anchor="w")
                aid = a["id"]

                def select_card(e, c=card, aid=aid):
                    if self._selected_card:
                        try:
                            self._selected_card.configure(bg="white", highlightbackground="#e8e8e8", highlightthickness=1)
                        except tk.TclError:
                            pass
                    self._selected_card = c
                    c.configure(bg="#d6eaf8", highlightbackground="#2980b9", highlightthickness=2)
                    self.show_detail(detail_frame, aid, load_patients)

                card.bind("<Button-1>", select_card)
                def bind_all_children(w):
                    for ch in w.winfo_children():
                        ch.bind("<Button-1>", select_card)
                        bind_all_children(ch)
                bind_all_children(card)

        def on_search(*args):
            load_patients(search_entry.get(), filtro_especie.get())

        def on_especie_filter(*args):
            load_patients(search_entry.get(), filtro_especie.get())

        search_entry.bind("<KeyRelease>", on_search)
        filtro_especie.bind("<<ComboboxSelected>>", on_especie_filter)
        load_patients()
        self.root.bind("<F5>", lambda e: load_patients(search_entry.get(), filtro_especie.get()))

    def show_detail(self, detail_frame, animal_id, reload_callback):
        for w in detail_frame.winfo_children():
            w.destroy()

        conn = self.get_db()
        animal = conn.execute(
            "SELECT a.*, d.id as dueno_id, d.nombre as dueno_nombre, d.telefono as dueno_telefono, "
            "d.email as dueno_email, d.direccion as dueno_direccion "
            "FROM animales a JOIN duenos d ON a.id_dueno = d.id "
            "WHERE a.id=?", (animal_id,)).fetchone()
        historial = conn.execute(
            "SELECT * FROM registros_medicos WHERE id_animal=? ORDER BY fecha DESC",
            (animal_id,)).fetchall()
        conn.close()

        if not animal:
            tk.Label(detail_frame, text="Animal no encontrado",
                     bg="#f0f4f8", fg="red").pack()
            return

        # Scrollable container
        canvas_outer = tk.Frame(detail_frame, bg="#f0f4f8")
        canvas_outer.pack(fill="both", expand=True)

        canvas = tk.Canvas(canvas_outer, bg="#f0f4f8", highlightthickness=0)
        v_scroll = ttk.Scrollbar(canvas_outer, orient="vertical", command=canvas.yview)
        scrollable = tk.Frame(canvas, bg="#f0f4f8")
        scrollable.bind("<Configure>",
                        lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        window_id = canvas.create_window((0, 0), window=scrollable, anchor="nw")
        canvas.configure(yscrollcommand=v_scroll.set)
        canvas.pack(side="left", fill="both", expand=True)
        v_scroll.pack(side="right", fill="y")

        def _configure_canvas(event):
            canvas.itemconfig(window_id, width=event.width)
        canvas.bind("<Configure>", _configure_canvas)
        canvas.bind("<MouseWheel>", lambda e: canvas.yview_scroll(int(-1 * (e.delta / 120)), "units"))

        outer = scrollable

        # Header with photo
        header = tk.Frame(outer, bg="white", padx=15, pady=12,
                          highlightbackground="#ddd", highlightthickness=1)
        header.pack(fill="x")

        left_p = tk.Frame(header, bg="white")
        left_p.pack(side="left", padx=(0, 15))

        photo_frame = tk.Frame(left_p, bg="#eee", width=180, height=180,
                               highlightbackground="#ccc", highlightthickness=1)
        photo_frame.pack()
        photo_frame.pack_propagate(False)

        foto_path = os.path.join(PHOTOS_DIR, animal["foto"]) if animal["foto"] else ""
        if animal["foto"] and os.path.exists(foto_path):
            try:
                img = Image.open(foto_path)
                img.thumbnail((170, 170))
                tk_img = ImageTk.PhotoImage(img)
                lbl_img = tk.Label(photo_frame, image=tk_img, bg="#eee")
                lbl_img.image = tk_img
                lbl_img.pack(fill="both", expand=True)
            except:
                tk.Label(photo_frame, text="Sin foto", bg="#eee",
                         fg="#999").pack(expand=True)
        else:
            tk.Label(photo_frame, text="Sin foto", bg="#eee",
                     fg="#999").pack(expand=True)

        def change_photo():
            path = filedialog.askopenfilename(
                filetypes=[("Imagenes", "*.jpg *.jpeg *.png *.bmp *.gif")])
            if path:
                ext = os.path.splitext(path)[1]
                dest = os.path.join(PHOTOS_DIR, str(animal_id) + ext)
                shutil.copy2(path, dest)
                c = self.get_db()
                c.execute("UPDATE animales SET foto=? WHERE id=?",
                          (str(animal_id) + ext, animal_id))
                c.commit()
                c.close()
                self.show_detail(detail_frame, animal_id, reload_callback)

        tk.Button(left_p, text="Cambiar Foto", bg="#3498db", fg="white",
                  font=("Segoe UI", 8), relief="flat", padx=8, pady=2,
                  command=change_photo, cursor="hand2").pack(pady=(5, 0))

        info_p = tk.Frame(header, bg="white")
        info_p.pack(side="left", fill="x", expand=True)

        tk.Label(info_p, text=animal["nombre"] + " (" + animal["especie"] + ")",
                 font=("Segoe UI", 16, "bold"), bg="white", fg="#2c3e50").pack(anchor="w")
        det = ("Raza: " + (animal["raza"] or "N/A") +
               "  |  Edad: " + str(animal["edad"]) + " años" +
               "  |  Peso: " + str(animal["peso"]) + " kg" +
               "  |  Sexo: " + (animal["sexo"] or "N/A") +
               "  |  Color: " + (animal["color"] or "N/A") +
               "  |  Esterilizado: " + ("Si" if animal["esterilizado"] else "No"))
        tk.Label(info_p, text=det, bg="white", fg="#555",
                 font=("Segoe UI", 10)).pack(anchor="w")
        tk.Label(info_p, text="", bg="white").pack(anchor="w", pady=2)
        tk.Label(info_p, text="Due\u00f1o: ", bg="white", fg="#555",
                 font=("Segoe UI", 10, "bold")).pack(anchor="w")
        dueno_line = tk.Frame(info_p, bg="white")
        dueno_line.pack(anchor="w")
        lbl_dueno = tk.Label(dueno_line, text=animal["dueno_nombre"],
                             bg="white", fg="#2980b9", font=("Segoe UI", 10, "bold"),
                             cursor="hand2")
        lbl_dueno.pack(side="left")
        lbl_dueno.bind("<Button-1>", lambda e: self.show_owner_detail(animal["dueno_id"], detail_frame, reload_callback))
        tk.Button(dueno_line, text="Editar", bg="#f39c12", fg="white",
                  font=("Segoe UI", 8), relief="flat", padx=6, pady=0,
                  cursor="hand2",
                  command=lambda: self.edit_owner_dialog(animal["dueno_id"], animal_id, detail_frame, reload_callback)
                  ).pack(side="left", padx=(8, 0))
        tk.Label(info_p,
                 text="Telefono: " + animal["dueno_telefono"] +
                 "  |  Email: " + (animal["dueno_email"] or "N/A"),
                 bg="white", fg="#555", font=("Segoe UI", 10)).pack(anchor="w")
        tk.Label(info_p, text="Direccion: " + (animal["dueno_direccion"] or "N/A"),
                 bg="white", fg="#555", font=("Segoe UI", 10)).pack(anchor="w")

        # ---------- BOTONES DE ACCION (barra horizontal) ----------
        def delete_animal():
            if messagebox.askyesno("Confirmar", "\u00bfEliminar este animal y todos sus registros?"):
                c = self.get_db()
                c.execute("DELETE FROM signos_vitales WHERE id_registro IN (SELECT id FROM registros_medicos WHERE id_animal=?)", (animal_id,))
                c.execute("DELETE FROM registros_medicos WHERE id_animal=?", (animal_id,))
                c.execute("DELETE FROM vacunas WHERE id_animal=?", (animal_id,))
                c.execute("DELETE FROM alergias WHERE id_animal=?", (animal_id,))
                c.execute("DELETE FROM examenes_auxiliares WHERE id_animal=?", (animal_id,))
                c.execute("DELETE FROM medicacion WHERE id_animal=?", (animal_id,))
                c.execute("DELETE FROM citas WHERE id_animal=?", (animal_id,))
                c.execute("DELETE FROM historial_grooming WHERE id_animal=?", (animal_id,))
                c.execute("DELETE FROM animales WHERE id=?", (animal_id,))
                c.commit()
                c.close()
                reload_callback()
        action_bar = tk.Frame(outer, bg="white", highlightbackground="#ddd", highlightthickness=1)
        action_bar.pack(fill="x", pady=(6, 0))
        bf = tk.Frame(action_bar, bg="white")
        bf.pack(expand=True, pady=4)
        btn_style = {"font": ("Segoe UI", 10, "bold"), "relief": "raised",
                      "padx": 16, "pady": 6, "cursor": "hand2"}
        tk.Button(bf, text="Editar", bg="#f39c12", fg="white", width=10,
                  command=lambda: self.edit_animal_dialog(animal_id, reload_callback),
                  **btn_style).pack(side="left", padx=4)
        tk.Button(bf, text="Grooming", bg="#8e44ad", fg="white", width=10,
                  command=lambda: self.add_grooming_from_detail(animal_id, detail_frame, reload_callback),
                  **btn_style).pack(side="left", padx=4)
        tk.Button(bf, text="Eliminar", bg="#c0392b", fg="white", width=10,
                  command=delete_animal,
                  **btn_style).pack(side="left", padx=4)
        tk.Button(bf, text="PDF", bg="#8e44ad", fg="white", width=10,
                  command=lambda: self.export_pdf(animal_id),
                  **btn_style).pack(side="left", padx=4)
        hist_frame = tk.Frame(outer, bg="#e8edf2")
        hist_frame.pack(fill="both", expand=True, pady=(8, 0))

        # --- RESUMEN RAPIDO ---
        self._resumen_rapido(hist_frame, animal_id)

        # --- BUSQUEDA GLOBAL ---
        busq_frame = tk.Frame(hist_frame, bg="white", highlightbackground="#ddd", highlightthickness=1, padx=8, pady=4)
        busq_frame.pack(fill="x", pady=(0, 6))
        tk.Label(busq_frame, text="🔍 Buscar en todo el expediente:", bg="white",
                 font=("Segoe UI", 9, "bold"), fg="#2c3e50").pack(side="left")
        busq_entry = tk.Entry(busq_frame, font=("Segoe UI", 9), width=35)
        busq_entry.pack(side="left", padx=6)
        busq_results = tk.Label(busq_frame, text="", bg="white", fg="#555", font=("Segoe UI", 9))
        busq_results.pack(side="left", padx=4)

        def do_global_search(event=None):
            texto = busq_entry.get().strip().lower()
            if not texto:
                busq_results.configure(text="")
                return
            total = 0
            conn = self.get_db()
            for tbl, col in [("registros_medicos","diagnostico"),("registros_medicos","tratamiento"),
                             ("registros_medicos","observaciones"),("vacunas","nombre"),
                             ("alergias","alergeno"),("medicacion","medicamento"),
                             ("examenes_auxiliares","nombre"),("examenes_auxiliares","resultados"),
                             ("citas","motivo"),("historial_grooming","observaciones")]:
                c = conn.execute(f"SELECT COUNT(*) FROM {tbl} WHERE id_animal=? AND LOWER({col}) LIKE ?",
                                 (animal_id, f"%{texto}%")).fetchone()[0]
                total += c
            conn.close()
            busq_results.configure(text=f"{total} coincidencia(s)")

        busq_entry.bind("<KeyRelease>", do_global_search)

        # Notebook (tabs)
        nb = ttk.Notebook(hist_frame)
        nb.pack(fill="both", expand=True, pady=(6, 0))

        # ----- TAB 1: HISTORIAL -----
        th = tk.Frame(nb, bg="#e8edf2")
        nb.add(th, text="  📋 Consultas  ")

        hdr = tk.Frame(th, bg="#e8edf2")
        hdr.pack(fill="x", pady=(6, 4))
        tk.Label(hdr, text="Historial Medico",
                 font=("Segoe UI", 14, "bold"),
                 bg="#e8edf2", fg="#2c3e50").pack(side="left")
        tk.Button(hdr, text="+ Nueva Consulta", bg="#2ecc71",
                  fg="white", font=("Segoe UI", 9), relief="flat", padx=12, pady=3,
                  command=lambda: self.add_medical_record_for(
                      animal_id, callback=lambda: self.show_detail(
                          detail_frame, animal_id, reload_callback)),
                  cursor="hand2").pack(side="right", padx=2)
        tk.Button(hdr, text="+ Programar Cita", bg="#3498db",
                  fg="white", font=("Segoe UI", 9), relief="flat", padx=12, pady=3,
                  command=lambda: self.add_appointment_dialog(animal_id),
                  cursor="hand2").pack(side="right")

        # --- NOTAS RAPIDAS ---
        conn_nr = self.get_db()
        nr = conn_nr.execute("SELECT contenido FROM notas_rapidas WHERE id_animal=?", (animal_id,)).fetchone()
        conn_nr.close()
        nr_frame = tk.Frame(th, bg="white", highlightbackground="#f39c12", highlightthickness=1, padx=8, pady=4)
        nr_frame.pack(fill="x", pady=(0, 6))
        nr_header = tk.Frame(nr_frame, bg="white")
        nr_header.pack(fill="x")
        tk.Label(nr_header, text="📝 Notas rapidas", font=("Segoe UI", 9, "bold"),
                 bg="white", fg="#e67e22").pack(side="left")
        nr_count = tk.Label(nr_header, text="", font=("Segoe UI", 8),
                            bg="white", fg="#999")
        nr_count.pack(side="right")
        nr_text = tk.Text(nr_frame, font=("Segoe UI", 9), height=3, width=60,
                          wrap="word", relief="flat", highlightbackground="#ddd", highlightthickness=1)
        nr_text.pack(fill="x", pady=(4, 2))
        if nr and nr["contenido"]:
            nr_text.insert("1.0", nr["contenido"])
            nr_count.configure(text=f"{len(nr['contenido'].split())} palabras")
        def save_nota(event=None):
            contenido = nr_text.get("1.0", "end-1c").strip()
            conn = self.get_db()
            existing = conn.execute("SELECT id FROM notas_rapidas WHERE id_animal=?", (animal_id,)).fetchone()
            if existing:
                conn.execute("UPDATE notas_rapidas SET contenido=?, fecha_actualizacion=? WHERE id_animal=?",
                             (contenido, str(date.today()), animal_id))
            else:
                conn.execute("INSERT INTO notas_rapidas (id_animal, contenido, fecha_actualizacion) VALUES (?,?,?)",
                             (animal_id, contenido, str(date.today())))
            conn.commit()
            conn.close()
            nr_count.configure(text=f"{len(contenido.split())} palabras" if contenido else "")
        nr_text.bind("<FocusOut>", save_nota)

        total_consultas = len(historial)
        stats_bar = tk.Frame(th, bg="white", highlightbackground="#ddd", highlightthickness=1, padx=12, pady=4)
        stats_bar.pack(fill="x", pady=(0, 6))
        stats_data = [("Total", str(total_consultas), "#2c3e50")]
        if total_consultas > 0:
            stats_data.append(("Última", historial[0]["fecha"], "#3498db"))
            doctores = len(set(r["doctor"] for r in historial if r["doctor"]))
            stats_data.append(("Doctores", str(doctores), "#9b59b6"))
        for lbl_t, val, color in stats_data:
            cell = tk.Frame(stats_bar, bg="white", padx=10, pady=2)
            cell.pack(side="left", fill="x", expand=True)
            tk.Label(cell, text=lbl_t, font=("Segoe UI", 8, "bold"),
                     bg="white", fg=color).pack(anchor="w")
            tk.Label(cell, text=val, font=("Segoe UI", 9),
                     bg="white", fg="#333").pack(anchor="w")

        filtros_frame = tk.Frame(th, bg="#e8edf2")
        filtros_frame.pack(fill="x", pady=(0, 6))
        tk.Label(filtros_frame, text="🔍", bg="#e8edf2", font=("Segoe UI", 9)).pack(side="left", padx=(0, 2))
        filtro_entry = tk.Entry(filtros_frame, font=("Segoe UI", 9), width=28)
        filtro_entry.pack(side="left", padx=2)

        if not historial:
            tk.Label(th, text="No hay registros medicos para este paciente",
                     bg="white", fg="#999", font=("Segoe UI", 10),
                     padx=15, pady=20).pack(fill="x", pady=5)
        else:
            hist_container = tk.Frame(th, bg="#e8edf2")
            hist_container.pack(fill="both", expand=True)

            table_frame = tk.Frame(hist_container, bg="white",
                                   highlightbackground="#ddd", highlightthickness=1)
            table_frame.pack(fill="both", expand=True, pady=(0, 4))

            columns = ("fecha", "doctor", "diagnostico", "tratamiento", "examenes")
            tree = ttk.Treeview(table_frame, columns=columns, show="headings", height=8)
            style = ttk.Style()
            style.configure("Treeview", rowheight=30, font=("Segoe UI", 10))
            style.configure("Treeview.Heading", font=("Segoe UI", 10, "bold"))
            style.map("Treeview", background=[("selected", "#3498db")])
            tree.tag_configure("even", background="#f8f9fa")
            tree.tag_configure("odd", background="white")
            for col, txt, w in [("fecha","Fecha",160),("doctor","Doctor",140),
                                ("diagnostico","Diagnóstico",280),("tratamiento","Tratamiento",240),
                                ("examenes","Exámenes",280)]:
                tree.heading(col, text=txt)
                tree.column(col, width=w)
            tree.pack(fill="both", expand=True, side="left")
            t_scroll = ttk.Scrollbar(table_frame, orient="vertical", command=tree.yview)
            tree.configure(yscrollcommand=t_scroll.set)
            t_scroll.pack(side="right", fill="y")

            def build_exam_summary(r):
                labels = {"exam_ecografia":"Eco","exam_radiografia":"Rx","exam_hemograma":"Hemo",
                          "exam_bioquimico":"Bioq","exam_orina":"Orina","exam_otros":"Otros"}
                parts = []
                for key, short in labels.items():
                    v = r[key]
                    if not v: continue
                    if v.startswith("FILE:"): parts.append(f"{short}:📷")
                    elif "|||FILE:" in v:
                        t = v.split("|||FILE:",1)[0]
                        parts.append(f"{short}:{t[:25]}{'…' if len(t)>25 else ''}📷")
                    else: parts.append(f"{short}:{v[:30]}{'…' if len(v)>30 else ''}")
                return " | ".join(parts) if parts else ""

            reg_map = {}
            for idx, r in enumerate(historial):
                tag = "even" if idx % 2 == 0 else "odd"
                diag = r["diagnostico"] or r["diagnostico_definitivo"] or r["diagnostico_presuntivo"] or ""
                ex_resumen = build_exam_summary(r)
                iid = str(r["id"])
                tree.insert("", "end", iid=iid, tags=(tag,), values=(
                    f"{r['fecha']}  {r['hora'] or ''}", r["doctor"] or "", diag,
                    r["tratamiento"] or "", ex_resumen))
                reg_map[iid] = r

            detail_panel = tk.Frame(hist_container, bg="white",
                                    highlightbackground="#ddd", highlightthickness=1, padx=12, pady=8)
            detail_panel.pack_forget()

            def _show_detail_panel():
                sel = tree.selection()
                if not sel: detail_panel.pack_forget(); return
                r = reg_map.get(sel[0])
                if not r: detail_panel.pack_forget(); return
                for w in detail_panel.winfo_children(): w.destroy()
                detail_panel.pack(fill="x", pady=(0, 4))
                tk.Label(detail_panel, text=f"📅 {r['fecha']}  🕐 {r['hora'] or '--'}  ⚕ {r['doctor'] or 'N/A'}  ⚖ {str(r['peso'])+' kg' if r['peso'] else '--'}",
                         bg="white", fg="#2c3e50", font=("Segoe UI", 10, "bold")).pack(anchor="w")
                tk.Frame(detail_panel, bg="#eee", height=1).pack(fill="x", pady=4)
                for lbl, key in [("Anamnesis","anamnesis"),("Dx Presuntivo","diagnostico_presuntivo"),
                                 ("Dx Definitivo","diagnostico_definitivo"),("Diagnóstico","diagnostico"),
                                 ("Tratamiento","tratamiento"),("Observaciones","observaciones")]:
                    val = r[key]
                    if not val: continue
                    fr = tk.Frame(detail_panel, bg="white")
                    fr.pack(fill="x", pady=1)
                    tk.Label(fr, text=lbl+":", bg="white", fg="#2c3e50",
                             font=("Segoe UI", 9, "bold"), width=16, anchor="e").pack(side="left")
                    tk.Label(fr, text=val, bg="white", fg="#333",
                             font=("Segoe UI", 9), wraplength=550, justify="left").pack(side="left", padx=(6, 0))
                exams = [("Ecografía","exam_ecografia"),("Radiografía","exam_radiografia"),
                         ("Hemograma","exam_hemograma"),("Bioquímico","exam_bioquimico"),
                         ("Orina","exam_orina"),("Otros","exam_otros")]
                if any(r[k] for _,k in exams):
                    tk.Frame(detail_panel, bg="#eee", height=1).pack(fill="x", pady=4)
                    tk.Label(detail_panel, text="Exámenes Complementarios",bg="white",
                             fg="#7f8c8d", font=("Segoe UI", 9, "bold")).pack(anchor="w")
                    for lbl, key in exams:
                        val = r[key]
                        if not val: continue
                        txt, fpath = "", ""
                        if val.startswith("FILE:"): fpath = val[5:]
                        elif "|||FILE:" in val:
                            p = val.split("|||FILE:", 1); txt, fpath = p[0], p[1]
                        else: txt = val
                        if not txt and not fpath: continue
                        fr = tk.Frame(detail_panel, bg="white")
                        fr.pack(fill="x", pady=1)
                        tk.Label(fr, text=lbl+":", bg="white", fg="#555",
                                 font=("Segoe UI", 9), width=16, anchor="e").pack(side="left")
                        if txt:
                            tk.Label(fr, text=txt, bg="white", fg="#333",
                                     font=("Segoe UI", 9)).pack(side="left", padx=(6, 0))
                        if fpath and os.path.exists(fpath):
                            flbl = tk.Label(fr, text=f"📷 {os.path.basename(fpath)[:30]}", bg="white",
                                            fg="#2980b9", font=("Segoe UI", 9, "underline"), cursor="hand2")
                            flbl.pack(side="left", padx=4)
                            flbl.bind("<Button-1>", lambda e, p=fpath: os.startfile(p))

            def edit_selected():
                sel = tree.selection()
                if not sel: return
                self.edit_registro_dialog(int(sel[0]), animal_id,
                    lambda: self.show_detail(detail_frame, animal_id, reload_callback))
            def delete_selected():
                sel = tree.selection()
                if not sel: return
                self._delete_registro(int(sel[0]), animal_id, detail_frame, reload_callback)

            tree.bind("<<TreeviewSelect>>", lambda e: _show_detail_panel())
            tree.bind("<Double-1>", lambda e: edit_selected())

            btn_row = tk.Frame(th, bg="#e8edf2")
            btn_row.pack(fill="x", pady=(2, 0))
            tk.Button(btn_row, text="✏ Editar", bg="#f39c12", fg="white",
                      font=("Segoe UI", 9), relief="flat", padx=14, pady=3,
                      command=edit_selected, cursor="hand2").pack(side="left", padx=2)
            tk.Button(btn_row, text="🗑 Eliminar", bg="#e74c3c", fg="white",
                      font=("Segoe UI", 9), relief="flat", padx=14, pady=3,
                      command=delete_selected, cursor="hand2").pack(side="left", padx=2)
            tk.Button(btn_row, text="❤ Signos Vitales", bg="#e67e22", fg="white",
                      font=("Segoe UI", 9), relief="flat", padx=14, pady=3, cursor="hand2",
                      command=lambda: self._add_signos_vitales_dialog(
                          animal_id, detail_frame, reload_callback)).pack(side="left", padx=2)
            tk.Button(btn_row, text="📋 Receta", bg="#2c3e50", fg="white",
                      font=("Segoe UI", 9), relief="flat", padx=14, pady=3, cursor="hand2",
                      command=lambda: self.generar_receta(animal_id, tree)).pack(side="left", padx=2)
            if self._user_role in ('admin', 'doctor'):
                tk.Button(btn_row, text="🛒 Listo pa' cobrar", bg="#e74c3c", fg="white",
                          font=("Segoe UI", 9), relief="flat", padx=14, pady=3, cursor="hand2",
                          command=lambda: self._marcar_para_cobro(animal_id, tree, detail_frame, reload_callback)
                          ).pack(side="left", padx=2)

            def filtrar_historial():
                texto = filtro_entry.get().lower()
                for item in tree.get_children():
                    vals = tree.item(item, "values")
                    if not texto:
                        tree.reattach(item, "", "end")
                    elif any(texto in str(v).lower() for v in vals):
                        tree.reattach(item, "", "end")
                    else:
                        tree.detach(item)
                _show_detail_panel()

            filtro_entry.bind("<KeyRelease>", lambda e: filtrar_historial())
            tk.Button(filtros_frame, text="Filtrar", bg="#3498db", fg="white",
                      font=("Segoe UI", 8), relief="flat", padx=8, pady=1,
                      command=filtrar_historial, cursor="hand2").pack(side="left", padx=2)
            tk.Button(filtros_frame, text="Limpiar", bg="#95a5a6", fg="white",
                      font=("Segoe UI", 8), relief="flat", padx=8, pady=1,
                      command=lambda: (filtro_entry.delete(0, "end"), filtrar_historial()),
                      cursor="hand2").pack(side="left", padx=2)

        # --- GRAFICO PESO ---
        self._grafico_peso(th, historial)

        # --- DIAGNOSTICOS RECURRENTES ---
        self._diagnosticos_recurrentes(th, animal_id)

        # --- LINEA DE TIEMPO ---
        self._linea_tiempo(th, animal_id)

        # ----- TAB 2: ALERGIAS -----
        tab_alerg = tk.Frame(nb, bg="#e8edf2")
        nb.add(tab_alerg, text="  ⚠ Alergias  ")
        self._seccion_alergias(tab_alerg, animal_id, detail_frame, reload_callback)

        # ----- TAB 3: VACUNAS Y MEDICACIÓN -----
        tab_vac = tk.Frame(nb, bg="#e8edf2")
        nb.add(tab_vac, text="  💉 Vacunas  ")
        self._seccion_vacunas(tab_vac, animal_id, detail_frame, reload_callback)
        tk.Frame(tab_vac, bg="#e8edf2", height=8).pack()
        self._seccion_medicacion(tab_vac, animal_id, detail_frame, reload_callback)

        # ----- TAB 4: EXÁMENES -----
        tab_exam = tk.Frame(nb, bg="#e8edf2")
        nb.add(tab_exam, text="  🔬 Exámenes  ")
        self._seccion_signos_vitales(tab_exam, animal_id)
        tk.Frame(tab_exam, bg="#e8edf2", height=8).pack()
        self._seccion_examenes(tab_exam, animal_id, detail_frame, reload_callback)

        # ----- TAB 5: CITAS Y GROOMING -----
        tab_citas = tk.Frame(nb, bg="#e8edf2")
        nb.add(tab_citas, text="  📅 Citas  ")
        conn2 = self.get_db()
        citas_paciente = conn2.execute(
            "SELECT id, fecha, motivo, estado, tipo FROM citas WHERE id_animal=? ORDER BY fecha DESC",
            (animal_id,)).fetchall()
        conn2.close()
        if citas_paciente:
            sec, lbl = self._make_section(tab_citas, "Citas", "#3498db")
            c_frame = self._make_table_frame(tab_citas)
            cols_c = ("fecha_c", "motivo_c", "tipo_c", "estado_c")
            c_tree = ttk.Treeview(c_frame, columns=cols_c, show="headings", height=4)
            c_tree.heading("fecha_c", text="Fecha")
            c_tree.heading("motivo_c", text="Motivo")
            c_tree.heading("tipo_c", text="Tipo")
            c_tree.heading("estado_c", text="Estado")
            c_tree.column("fecha_c", width=100)
            c_tree.column("motivo_c", width=200)
            c_tree.column("tipo_c", width=80)
            c_tree.column("estado_c", width=80)
            c_tree.pack(fill="x")
            for r in citas_paciente:
                c_tree.insert("", "end", values=(r["fecha"], r["motivo"], r["tipo"] or "veterinaria", r["estado"] or "pendiente"))
        tk.Frame(tab_citas, bg="#e8edf2", height=8).pack()
        self._seccion_historial_grooming(tab_citas, animal_id)

        # ----- TAB 6: DOCUMENTOS -----
        tab_docs = tk.Frame(nb, bg="#e8edf2")
        nb.add(tab_docs, text="  📎 Documentos  ")
        self._seccion_documentos(tab_docs, animal_id, detail_frame, reload_callback)

        # Notificacion de citas proximas
        conn3 = self.get_db()
        prox = conn3.execute(
            "SELECT fecha, motivo FROM citas WHERE id_animal=? AND estado='pendiente' AND fecha >= ? ORDER BY fecha ASC LIMIT 1",
            (animal_id, str(date.today()))).fetchone()
        if prox:
            dias = (date.fromisoformat(prox["fecha"]) - date.today()).days
            if dias <= 3:
                msg = f"Proxima cita: {prox['fecha']} ({prox['motivo']})"
                if dias == 0: msg = f"Cita programada para HOY: {prox['motivo']}"
                elif dias == 1: msg = f"Cita programada para MA\u00d1ANA: {prox['motivo']}"
                lbl_notif = tk.Label(hist_frame, text=msg, font=("Segoe UI", 9, "bold"),
                                      bg="#fef9e7", fg="#e67e22", anchor="w", padx=10, pady=4)
                lbl_notif.pack(fill="x", pady=(6, 0))
        conn3.close()

    # ---------- NUEVAS SECCIONES ----------
    def _make_section(self, parent, title, color="#2c3e50"):
        f = tk.Frame(parent, bg="#e8edf2")
        f.pack(fill="x", pady=(12, 6))
        lbl = tk.Label(f, text=title, font=("Segoe UI", 12, "bold"),
                       bg="#e8edf2", fg=color)
        lbl.pack(side="left")
        return f, lbl

    def _make_table_frame(self, parent):
        f = tk.Frame(parent, bg="white", highlightbackground="#ddd", highlightthickness=1)
        f.pack(fill="x", pady=(0, 8))
        return f

    def _resumen_rapido(self, parent, animal_id):
        conn = self.get_db()
        total_consultas = conn.execute(
            "SELECT COUNT(*) FROM registros_medicos WHERE id_animal=?", (animal_id,)).fetchone()[0]
        ultima = conn.execute(
            "SELECT fecha, diagnostico FROM registros_medicos WHERE id_animal=? ORDER BY fecha DESC LIMIT 1",
            (animal_id,)).fetchone()
        ultima_vacuna = conn.execute(
            "SELECT fecha, nombre FROM vacunas WHERE id_animal=? AND tipo='vacuna' ORDER BY fecha DESC LIMIT 1",
            (animal_id,)).fetchone()
        prox_cita = conn.execute(
            "SELECT fecha, motivo FROM citas WHERE id_animal=? AND estado='pendiente' ORDER BY fecha ASC LIMIT 1",
            (animal_id,)).fetchone()
        vacunas_proximas = conn.execute(
            "SELECT nombre, proxima_dosis FROM vacunas WHERE id_animal=? AND proxima_dosis IS NOT NULL AND proxima_dosis != '' AND proxima_dosis <= ? ORDER BY proxima_dosis ASC",
            (animal_id, str(date.today()))).fetchall()
        pesos = conn.execute(
            "SELECT peso, fecha FROM registros_medicos WHERE id_animal=? AND peso IS NOT NULL AND peso > 0 ORDER BY fecha DESC LIMIT 2",
            (animal_id,)).fetchall()
        conn.close()

        card = tk.Frame(parent, bg="white", highlightbackground="#ddd",
                        highlightthickness=1, padx=16, pady=10)
        card.pack(fill="x", pady=(0, 8))
        items = [("Consultas", str(total_consultas), "#3498db")]
        if ultima:
            items.append(("Ultima Consulta", f"{ultima['fecha']} - {ultima['diagnostico'][:30]}", "#2ecc71"))
        if ultima_vacuna:
            items.append(("Ultima Vacuna", f"{ultima_vacuna['fecha']} - {ultima_vacuna['nombre']}", "#9b59b6"))
        if prox_cita:
            items.append(("Proxima Cita", f"{prox_cita['fecha']} - {prox_cita['motivo']}", "#e67e22"))
        inner = tk.Frame(card, bg="white")
        inner.pack(fill="x")
        for lbl_t, val, color in items:
            cell = tk.Frame(inner, bg="white", padx=10, pady=4)
            cell.pack(side="left", fill="x", expand=True)
            tk.Label(cell, text=lbl_t, font=("Segoe UI", 8, "bold"),
                     bg="white", fg=color).pack(anchor="w")
            tk.Label(cell, text=val, font=("Segoe UI", 9),
                     bg="white", fg="#555").pack(anchor="w")

        # Alertas inteligentes
        alertas = []
        for v in vacunas_proximas:
            dias = (date.today() - date.fromisoformat(v["proxima_dosis"])).days
            if dias > 0:
                alertas.append(("⚠", f"{v['nombre']} vencida hace {dias} dia(s)", "#e74c3c"))
            else:
                alertas.append(("⏳", f"{v['nombre']} vence en {abs(dias)} dia(s)", "#e67e22"))
        if len(pesos) == 2:
            p1, p2 = pesos[0]["peso"], pesos[1]["peso"]
            cambio = abs(p1 - p2) / max(p2, 0.01) * 100
            if cambio > 20:
                alertas.append(("⚖", f"Peso cambio {cambio:.0f}% ({pesos[1]['fecha']}: {p2}kg → {pesos[0]['fecha']}: {p1}kg)", "#e74c3c"))
        for icono, texto, color in alertas:
            af = tk.Frame(card, bg="white")
            af.pack(fill="x", pady=(2, 0))
            tk.Label(af, text=f"{icono} {texto}", bg="white", fg=color,
                     font=("Segoe UI", 9, "bold"), anchor="w").pack(fill="x")

    def _seccion_alergias(self, parent, animal_id, detail_frame, reload_callback):
        conn = self.get_db()
        alergias = conn.execute(
            "SELECT * FROM alergias WHERE id_animal=? ORDER BY alergeno", (animal_id,)).fetchall()
        conn.close()
        sec, lbl = self._make_section(parent, "Alergias", "#e74c3c")

        def add_alergia():
            dialog = tk.Toplevel(self.root)
            dialog.title("Agregar Alergia")
            dialog.geometry("420x280")
            dialog.resizable(False, False)
            dialog.configure(bg="#e8edf2")
            dialog.transient(self.root)
            dialog.grab_set()
            center_dialog(dialog, self.root)
            tk.Label(dialog, text="Alergeno", bg="#e8edf2", font=("Segoe UI", 10)).grid(row=0, column=0, sticky="e", padx=8, pady=3)
            e_alergeno = tk.Entry(dialog, width=40, font=("Segoe UI", 10))
            e_alergeno.grid(row=0, column=1, padx=8, pady=3)
            tk.Label(dialog, text="Tipo", bg="#e8edf2", font=("Segoe UI", 10)).grid(row=1, column=0, sticky="e", padx=8, pady=3)
            e_tipo = ttk.Combobox(dialog, values=["Alimentaria", "Ambiental", "Farmacologica", "Picadura", "Contacto", "Otra"],
                                  width=37, font=("Segoe UI", 10))
            e_tipo.grid(row=1, column=1, padx=8, pady=3)
            tk.Label(dialog, text="Severidad", bg="#e8edf2", font=("Segoe UI", 10)).grid(row=2, column=0, sticky="e", padx=8, pady=3)
            e_sev = ttk.Combobox(dialog, values=["Leve", "Moderada", "Grave"],
                                 width=37, font=("Segoe UI", 10))
            e_sev.set("Leve")
            e_sev.grid(row=2, column=1, padx=8, pady=3)
            tk.Label(dialog, text="Observaciones", bg="#e8edf2", font=("Segoe UI", 10)).grid(row=3, column=0, sticky="e", padx=8, pady=3)
            e_obs = tk.Entry(dialog, width=40, font=("Segoe UI", 10))
            e_obs.grid(row=3, column=1, padx=8, pady=3)
            def save():
                conn = self.get_db()
                conn.execute("INSERT INTO alergias (id_animal, alergeno, tipo, severidad, observaciones) VALUES (?,?,?,?,?)",
                             (animal_id, e_alergeno.get(), e_tipo.get(), e_sev.get(), e_obs.get()))
                conn.commit()
                conn.close()
                dialog.destroy()
                self.show_detail(detail_frame, animal_id, reload_callback)
            tk.Button(dialog, text="Guardar", bg="#2ecc71", fg="white",
                      font=("Segoe UI", 10), relief="flat", padx=20, pady=5,
                      command=save, cursor="hand2").grid(row=4, column=1, pady=15, sticky="e")

        tk.Button(sec, text="+", bg="#e74c3c", fg="white", font=("Segoe UI", 9, "bold"),
                  relief="flat", width=2, cursor="hand2", command=add_alergia).pack(side="right", padx=2)
        if alergias:
            f = self._make_table_frame(parent)
            for r in alergias:
                color = "#e74c3c" if r["severidad"] == "Grave" else "#f39c12" if r["severidad"] == "Moderada" else "#2ecc71"
                row_f = tk.Frame(f, bg="white", padx=12, pady=6)
                row_f.pack(fill="x")
                tk.Label(row_f, text=r["alergeno"], font=("Segoe UI", 10, "bold"),
                         bg="white", fg=color).pack(side="left", padx=5)
                tk.Label(row_f, text=f"({r['tipo']}) [{r['severidad']}]",
                         font=("Segoe UI", 9), bg="white", fg="#7f8c8d").pack(side="left")
                if r["observaciones"]:
                    tk.Label(row_f, text=r["observaciones"], font=("Segoe UI", 9),
                             bg="white", fg="#555").pack(side="left", padx=10)
        else:
            tk.Label(parent, text="  No se registraron alergias", font=("Segoe UI", 9),
                     bg="#e8edf2", fg="#999", anchor="w").pack(fill="x")

    def _delete_registro(self, reg_id, animal_id, detail_frame, reload_callback):
        if messagebox.askyesno("Confirmar", "\u00bfEliminar este registro medico?"):
            c = self.get_db()
            c.execute("DELETE FROM signos_vitales WHERE id_registro=?", (reg_id,))
            c.execute("DELETE FROM registros_medicos WHERE id=?", (reg_id,))
            c.commit()
            c.close()
            self.show_detail(detail_frame, animal_id, reload_callback)

    def _seccion_vacunas(self, parent, animal_id, detail_frame, reload_callback):
        conn = self.get_db()
        vacunas = conn.execute(
            "SELECT * FROM vacunas WHERE id_animal=? ORDER BY fecha DESC", (animal_id,)).fetchall()
        conn.close()
        sec, lbl = self._make_section(parent, "Vacunas y Desparasitaciones", "#9b59b6")

        def add_vacuna():
            from tkcalendar import DateEntry
            dialog = tk.Toplevel(self.root)
            dialog.title("Agregar Vacuna / Desparasitacion")
            dialog.geometry("450x340")
            dialog.resizable(False, False)
            dialog.configure(bg="#e8edf2")
            dialog.transient(self.root)
            dialog.grab_set()
            center_dialog(dialog, self.root)
            tk.Label(dialog, text="Tipo", bg="#e8edf2", font=("Segoe UI", 10)).grid(row=0, column=0, sticky="e", padx=8, pady=3)
            e_tipo = ttk.Combobox(dialog, values=["Vacuna", "Desparasitacion"], width=37, font=("Segoe UI", 10), state="readonly")
            e_tipo.set("Vacuna")
            e_tipo.grid(row=0, column=1, padx=8, pady=3)
            tk.Label(dialog, text="Nombre", bg="#e8edf2", font=("Segoe UI", 10)).grid(row=1, column=0, sticky="e", padx=8, pady=3)
            e_nombre = tk.Entry(dialog, width=40, font=("Segoe UI", 10))
            e_nombre.grid(row=1, column=1, padx=8, pady=3)
            tk.Label(dialog, text="Fecha aplicacion", bg="#e8edf2", font=("Segoe UI", 10)).grid(row=2, column=0, sticky="e", padx=8, pady=3)
            e_fecha = DateEntry(dialog, width=37, font=("Segoe UI", 10),
                                background="#2c3e50", foreground="white",
                                borderwidth=2, date_pattern="yyyy-mm-dd")
            e_fecha.grid(row=2, column=1, padx=8, pady=3)
            tk.Label(dialog, text="Lote", bg="#e8edf2", font=("Segoe UI", 10)).grid(row=3, column=0, sticky="e", padx=8, pady=3)
            e_lote = tk.Entry(dialog, width=40, font=("Segoe UI", 10))
            e_lote.grid(row=3, column=1, padx=8, pady=3)
            tk.Label(dialog, text="Proxima dosis", bg="#e8edf2", font=("Segoe UI", 10)).grid(row=4, column=0, sticky="e", padx=8, pady=3)
            e_prox = DateEntry(dialog, width=37, font=("Segoe UI", 10),
                               background="#2c3e50", foreground="white",
                               borderwidth=2, date_pattern="yyyy-mm-dd")
            e_prox.grid(row=4, column=1, padx=8, pady=3)
            def save():
                conn = self.get_db()
                conn.execute(
                    "INSERT INTO vacunas (id_animal, tipo, nombre, fecha, lote, proxima_dosis) VALUES (?,?,?,?,?,?)",
                    (animal_id, e_tipo.get().lower(), e_nombre.get(), e_fecha.get(), e_lote.get(), e_prox.get()))
                conn.commit()
                conn.close()
                dialog.destroy()
                self.show_detail(detail_frame, animal_id, reload_callback)
            tk.Button(dialog, text="Guardar", bg="#2ecc71", fg="white",
                      font=("Segoe UI", 10), relief="flat", padx=20, pady=5,
                      command=save, cursor="hand2").grid(row=5, column=1, pady=15, sticky="e")

        tk.Button(sec, text="+", bg="#9b59b6", fg="white", font=("Segoe UI", 9, "bold"),
                  relief="flat", width=2, cursor="hand2", command=add_vacuna).pack(side="right", padx=2)
        if vacunas:
            f = self._make_table_frame(parent)
            cols = ("f_v", "tipo_v", "nombre_v", "lote_v", "prox_v")
            tree = ttk.Treeview(f, columns=cols, show="headings", height=4)
            tree.heading("f_v", text="Fecha")
            tree.heading("tipo_v", text="Tipo")
            tree.heading("nombre_v", text="Nombre")
            tree.heading("lote_v", text="Lote")
            tree.heading("prox_v", text="Proxima Dosis")
            tree.column("f_v", width=90)
            tree.column("tipo_v", width=100)
            tree.column("nombre_v", width=180)
            tree.column("lote_v", width=100)
            tree.column("prox_v", width=100)
            tree.pack(fill="x")
            for r in vacunas:
                prox = r["proxima_dosis"] or ""
                tag = ""
                if prox:
                    try:
                        if date.fromisoformat(prox) < date.today():
                            tag = "vencida"
                        elif (date.fromisoformat(prox) - date.today()).days <= 15:
                            tag = "proxima"
                    except: pass
                tree.insert("", "end", values=(r["fecha"], r["tipo"].capitalize(),
                                               r["nombre"], r["lote"] or "", prox),
                            tags=(tag,) if tag else ())
            tree.tag_configure("vencida", background="#fadbd8")
            tree.tag_configure("proxima", background="#fef9e7")
        else:
            tk.Label(parent, text="  No hay vacunas registradas", font=("Segoe UI", 9),
                     bg="#e8edf2", fg="#999", anchor="w").pack(fill="x")

    def _seccion_medicacion(self, parent, animal_id, detail_frame, reload_callback):
        conn = self.get_db()
        medicacion = conn.execute(
            "SELECT * FROM medicacion WHERE id_animal=? ORDER BY activo DESC, fecha_inicio DESC",
            (animal_id,)).fetchall()
        conn.close()
        sec, lbl = self._make_section(parent, "Medicacion Actual", "#e67e22")

        def add_medicacion():
            from tkcalendar import DateEntry
            dialog = tk.Toplevel(self.root)
            dialog.title("Agregar Medicacion")
            dialog.geometry("480x350")
            dialog.resizable(False, False)
            dialog.configure(bg="#e8edf2")
            dialog.transient(self.root)
            dialog.grab_set()
            center_dialog(dialog, self.root)
            tk.Label(dialog, text="Medicamento", bg="#e8edf2", font=("Segoe UI", 10)).grid(row=0, column=0, sticky="e", padx=8, pady=3)
            e_med = tk.Entry(dialog, width=40, font=("Segoe UI", 10))
            e_med.grid(row=0, column=1, padx=8, pady=3)
            tk.Label(dialog, text="Dosis", bg="#e8edf2", font=("Segoe UI", 10)).grid(row=1, column=0, sticky="e", padx=8, pady=3)
            e_dosis = tk.Entry(dialog, width=40, font=("Segoe UI", 10))
            e_dosis.grid(row=1, column=1, padx=8, pady=3)
            tk.Label(dialog, text="Frecuencia", bg="#e8edf2", font=("Segoe UI", 10)).grid(row=2, column=0, sticky="e", padx=8, pady=3)
            e_freq = ttk.Combobox(dialog, values=["Cada 8h", "Cada 12h", "Cada 24h", "Cada 48h", "Semanal", "Mensual", "Segun indicacion"],
                                  width=37, font=("Segoe UI", 10))
            e_freq.grid(row=2, column=1, padx=8, pady=3)
            tk.Label(dialog, text="Via", bg="#e8edf2", font=("Segoe UI", 10)).grid(row=3, column=0, sticky="e", padx=8, pady=3)
            e_via = ttk.Combobox(dialog, values=["Oral", "Topica", "Intramuscular", "Subcutanea", "Intravenosa", "Otica", "Oftalmica", "Otra"],
                                 width=37, font=("Segoe UI", 10))
            e_via.grid(row=3, column=1, padx=8, pady=3)
            tk.Label(dialog, text="Inicio", bg="#e8edf2", font=("Segoe UI", 10)).grid(row=4, column=0, sticky="e", padx=8, pady=3)
            e_ini = DateEntry(dialog, width=37, font=("Segoe UI", 10),
                              background="#2c3e50", foreground="white", borderwidth=2, date_pattern="yyyy-mm-dd")
            e_ini.grid(row=4, column=1, padx=8, pady=3)
            tk.Label(dialog, text="Fin", bg="#e8edf2", font=("Segoe UI", 10)).grid(row=5, column=0, sticky="e", padx=8, pady=3)
            e_fin = DateEntry(dialog, width=37, font=("Segoe UI", 10),
                              background="#2c3e50", foreground="white", borderwidth=2, date_pattern="yyyy-mm-dd")
            e_fin.grid(row=5, column=1, padx=8, pady=3)
            def save():
                conn = self.get_db()
                conn.execute(
                    "INSERT INTO medicacion (id_animal, medicamento, dosis, frecuencia, via, fecha_inicio, fecha_fin) VALUES (?,?,?,?,?,?,?)",
                    (animal_id, e_med.get(), e_dosis.get(), e_freq.get(), e_via.get(), e_ini.get(), e_fin.get()))
                conn.commit()
                conn.close()
                dialog.destroy()
                self.show_detail(detail_frame, animal_id, reload_callback)
            tk.Button(dialog, text="Guardar", bg="#2ecc71", fg="white",
                      font=("Segoe UI", 10), relief="flat", padx=20, pady=5,
                      command=save, cursor="hand2").grid(row=6, column=1, pady=15, sticky="e")

        tk.Button(sec, text="+", bg="#e67e22", fg="white", font=("Segoe UI", 9, "bold"),
                  relief="flat", width=2, cursor="hand2", command=add_medicacion).pack(side="right", padx=2)
        if medicacion:
            f = self._make_table_frame(parent)
            cols = ("m_med", "m_dosis", "m_freq", "m_via", "m_ini", "m_fin", "m_act")
            tree = ttk.Treeview(f, columns=cols, show="headings", height=4)
            tree.heading("m_med", text="Medicamento")
            tree.heading("m_dosis", text="Dosis")
            tree.heading("m_freq", text="Frecuencia")
            tree.heading("m_via", text="Via")
            tree.heading("m_ini", text="Inicio")
            tree.heading("m_fin", text="Fin")
            tree.heading("m_act", text="Activo")
            for c in cols: tree.column(c, width=90)
            tree.column("m_med", width=130)
            tree.column("m_act", width=50)
            tree.pack(fill="x")
            for r in medicacion:
                tree.insert("", "end", values=(r["medicamento"], r["dosis"], r["frecuencia"],
                                               r["via"] or "", r["fecha_inicio"], r["fecha_fin"] or "",
                                               "Si" if r["activo"] else "No"))
        else:
            tk.Label(parent, text="  Sin medicacion activa", font=("Segoe UI", 9),
                     bg="#e8edf2", fg="#999", anchor="w").pack(fill="x")

    def _seccion_examenes(self, parent, animal_id, detail_frame, reload_callback):
        conn = self.get_db()
        examenes = conn.execute(
            "SELECT * FROM examenes_auxiliares WHERE id_animal=? ORDER BY fecha DESC",
            (animal_id,)).fetchall()
        conn.close()
        sec, lbl = self._make_section(parent, "Examenes Auxiliares", "#1abc9c")

        def add_examen():
            from tkcalendar import DateEntry
            dialog = tk.Toplevel(self.root)
            dialog.title("Agregar Examen Auxiliar")
            dialog.geometry("480x350")
            dialog.resizable(False, False)
            dialog.configure(bg="#e8edf2")
            dialog.transient(self.root)
            dialog.grab_set()
            center_dialog(dialog, self.root)
            tk.Label(dialog, text="Tipo", bg="#e8edf2", font=("Segoe UI", 10)).grid(row=0, column=0, sticky="e", padx=8, pady=3)
            e_tipo = ttk.Combobox(dialog, values=["Laboratorio", "Radiografia", "Ecografia", "Electrocardiograma", "Endoscopia", "Otro"],
                                  width=37, font=("Segoe UI", 10))
            e_tipo.grid(row=0, column=1, padx=8, pady=3)
            tk.Label(dialog, text="Nombre", bg="#e8edf2", font=("Segoe UI", 10)).grid(row=1, column=0, sticky="e", padx=8, pady=3)
            e_nombre = tk.Entry(dialog, width=40, font=("Segoe UI", 10))
            e_nombre.grid(row=1, column=1, padx=8, pady=3)
            tk.Label(dialog, text="Fecha", bg="#e8edf2", font=("Segoe UI", 10)).grid(row=2, column=0, sticky="e", padx=8, pady=3)
            e_fecha = DateEntry(dialog, width=37, font=("Segoe UI", 10),
                                background="#2c3e50", foreground="white", borderwidth=2, date_pattern="yyyy-mm-dd")
            e_fecha.grid(row=2, column=1, padx=8, pady=3)
            tk.Label(dialog, text="Resultados", bg="#e8edf2", font=("Segoe UI", 10)).grid(row=3, column=0, sticky="ne", padx=8, pady=3)
            e_result = tk.Text(dialog, width=30, height=3, font=("Segoe UI", 10))
            e_result.grid(row=3, column=1, padx=8, pady=3)
            selected_file = tk.StringVar(value="")
            tk.Label(dialog, text="Adjuntar archivo", bg="#e8edf2", font=("Segoe UI", 10)).grid(row=4, column=0, sticky="e", padx=8, pady=3)
            file_frame = tk.Frame(dialog, bg="#e8edf2")
            file_frame.grid(row=4, column=1, padx=10, pady=6, sticky="w")
            file_label = tk.Label(file_frame, text="Ninguno", bg="#e8edf2", fg="#999", font=("Segoe UI", 9))
            file_label.pack(side="left")
            def choose_file():
                path = filedialog.askopenfilename(filetypes=[("Todos los archivos", "*.*")])
                if path:
                    selected_file.set(path)
                    file_label.config(text=os.path.basename(path))
            tk.Button(file_frame, text="Examinar", bg="#3498db", fg="white",
                      font=("Segoe UI", 9), relief="flat", padx=8, pady=1,
                      command=choose_file, cursor="hand2").pack(side="left", padx=5)
            def save():
                conn = self.get_db()
                archivo = ""
                if selected_file.get():
                    ext = os.path.splitext(selected_file.get())[1]
                    dest = os.path.join(EXAMENES_DIR, f"{animal_id}_{datetime.now().strftime('%Y%m%d%H%M%S')}{ext}")
                    shutil.copy2(selected_file.get(), dest)
                    archivo = os.path.basename(dest)
                conn.execute(
                    "INSERT INTO examenes_auxiliares (id_animal, tipo, nombre, fecha, archivo, resultados) VALUES (?,?,?,?,?,?)",
                    (animal_id, e_tipo.get(), e_nombre.get(), e_fecha.get(), archivo, e_result.get("1.0", "end-1c")))
                conn.commit()
                conn.close()
                dialog.destroy()
                self.show_detail(detail_frame, animal_id, reload_callback)
            tk.Button(dialog, text="Guardar", bg="#2ecc71", fg="white",
                      font=("Segoe UI", 10), relief="flat", padx=20, pady=5,
                      command=save, cursor="hand2").grid(row=5, column=1, pady=15, sticky="e")

        tk.Button(sec, text="+", bg="#1abc9c", fg="white", font=("Segoe UI", 9, "bold"),
                  relief="flat", width=2, cursor="hand2", command=add_examen).pack(side="right", padx=2)
        if examenes:
            f = self._make_table_frame(parent)
            cols = ("e_fecha", "e_tipo", "e_nombre", "e_result", "e_archivo")
            tree = ttk.Treeview(f, columns=cols, show="headings", height=4)
            tree.heading("e_fecha", text="Fecha")
            tree.heading("e_tipo", text="Tipo")
            tree.heading("e_nombre", text="Nombre")
            tree.heading("e_result", text="Resultados")
            tree.heading("e_archivo", text="Archivo")
            tree.column("e_fecha", width=80)
            tree.column("e_tipo", width=90)
            tree.column("e_nombre", width=130)
            tree.column("e_result", width=170)
            tree.column("e_archivo", width=100)
            tree.pack(fill="x")

            def ver_examen():
                sel = tree.selection()
                if not sel: return
                vals = tree.item(sel[0], "values")
                archivo = vals[4] if len(vals) > 4 else ""
                if not archivo:
                    messagebox.showinfo("Info", "Este examen no tiene archivo adjunto")
                    return
                ruta = os.path.join(EXAMENES_DIR, archivo)
                if not os.path.exists(ruta):
                    messagebox.showerror("Error", "Archivo no encontrado")
                    return
                ext = os.path.splitext(archivo)[1].lower()
                if ext in (".jpg", ".jpeg", ".png", ".bmp", ".gif", ".tiff", ".webp"):
                    try:
                        viewer = tk.Toplevel(self.root)
                        viewer.title(f"Radiografia/Imagen - {vals[2]}")
                        viewer.geometry("900x700")
                        viewer.transient(self.root)
                        img = Image.open(ruta)
                        w, h = img.size
                        max_w, max_h = 850, 650
                        if w > max_w or h > max_h:
                            ratio = min(max_w/w, max_h/h)
                            img = img.resize((int(w*ratio), int(h*ratio)), Image.LANCZOS)
                        tk_img = ImageTk.PhotoImage(img)
                        lbl = tk.Label(viewer, image=tk_img)
                        lbl.image = tk_img
                        lbl.pack(expand=True)
                        info_f = tk.Frame(viewer, bg="#f0f4f8", padx=8, pady=3)
                        info_f.pack(fill="x")
                        tk.Label(info_f, text=f"{vals[2]} ({vals[1]}) - {vals[0]}",
                                 font=("Segoe UI", 10), bg="#f0f4f8").pack()
                        tk.Button(info_f, text="Abrir con visor externo", bg="#3498db", fg="white",
                                  font=("Segoe UI", 9), relief="flat", padx=10,
                                  command=lambda: os.startfile(ruta), cursor="hand2").pack(pady=5)
                    except Exception as ex:
                        messagebox.showerror("Error", f"No se pudo abrir imagen: {ex}")
                else:
                    try:
                        os.startfile(ruta)
                    except Exception as ex:
                        messagebox.showerror("Error", f"No se pudo abrir archivo: {ex}")

            tree.bind("<Double-1>", lambda e: ver_examen())

            btn_row = tk.Frame(parent, bg="#e8edf2")
            btn_row.pack(fill="x", pady=(2, 0))
            tk.Button(btn_row, text="Ver archivo", bg="#3498db", fg="white",
                      font=("Segoe UI", 9), relief="flat", padx=10, pady=1,
                      command=ver_examen, cursor="hand2").pack(side="left", padx=2)

            for r in examenes:
                archivo = r["archivo"] or ""
                tree.insert("", "end", values=(r["fecha"], r["tipo"], r["nombre"],
                                               (r["resultados"] or "")[:60], archivo))
        else:
            tk.Label(parent, text="  Sin examenes registrados", font=("Segoe UI", 9),
                     bg="#e8edf2", fg="#999", anchor="w").pack(fill="x")

    def _seccion_signos_vitales(self, parent, animal_id):
        conn = self.get_db()
        registros = conn.execute(
            "SELECT rm.id, rm.fecha, sv.temperatura, sv.frecuencia_cardiaca, "
            "sv.frecuencia_respiratoria, sv.presion_sistolica, sv.presion_diastolica "
            "FROM registros_medicos rm "
            "LEFT JOIN signos_vitales sv ON sv.id_registro = rm.id "
            "WHERE rm.id_animal=? AND (sv.temperatura IS NOT NULL OR sv.frecuencia_cardiaca IS NOT NULL) "
            "ORDER BY rm.fecha DESC LIMIT 10", (animal_id,)).fetchall()
        conn.close()
        if not registros:
            return
        sec, lbl = self._make_section(parent, "Signos Vitales", "#e74c3c")
        f = self._make_table_frame(parent)
        cols = ("sv_f", "sv_temp", "sv_fc", "sv_fr", "sv_pa")
        tree = ttk.Treeview(f, columns=cols, show="headings", height=4)
        tree.heading("sv_f", text="Fecha")
        tree.heading("sv_temp", text="Temp (C)")
        tree.heading("sv_fc", text="FC (lpm)")
        tree.heading("sv_fr", text="FR (rpm)")
        tree.heading("sv_pa", text="PA (mmHg)")
        tree.column("sv_f", width=90)
        tree.column("sv_temp", width=80)
        tree.column("sv_fc", width=80)
        tree.column("sv_fr", width=80)
        tree.column("sv_pa", width=100)
        tree.pack(fill="x")
        for r in registros:
            pa = f"{r['presion_sistolica']}/{r['presion_diastolica']}" if r["presion_sistolica"] else ""
            tree.insert("", "end", values=(r["fecha"],
                        f"{r['temperatura']:.1f}" if r["temperatura"] else "-",
                        r["frecuencia_cardiaca"] or "-",
                        r["frecuencia_respiratoria"] or "-", pa or "-"))

    def _grafico_peso(self, parent, historial):
        pesos = [(r["fecha"], r["peso"]) for r in historial if r["peso"]]
        if len(pesos) < 2:
            return
        pesos.sort(key=lambda x: x[0])
        sec, lbl = self._make_section(parent, "Evolucion del Peso", "#2ecc71")
        try:
            from matplotlib.figure import Figure
            from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
            fig = Figure(figsize=(4, 1.8), dpi=80, facecolor="white")
            ax = fig.add_subplot(111)
            fechas = [p[0] for p in pesos]
            valores = [p[1] for p in pesos]
            ax.plot(range(len(valores)), valores, marker="o", color="#2ecc71", linewidth=2, markersize=6)
            ax.fill_between(range(len(valores)), valores, alpha=0.15, color="#2ecc71")
            ax.set_xticks(range(len(fechas)))
            ax.set_xticklabels(fechas, rotation=30, fontsize=7)
            ax.set_ylabel("Kg", fontsize=8)
            fig.tight_layout()
            canvas_chart = FigureCanvasTkAgg(fig, master=parent)
            canvas_chart.draw()
            canvas_chart.get_tk_widget().pack(fill="x", padx=5, pady=5)
        except Exception:
            f = self._make_table_frame(parent)
            cols = ("pf", "pp", "pc")
            tree = ttk.Treeview(f, columns=cols, show="headings", height=4)
            tree.heading("pf", text="Fecha")
            tree.heading("pp", text="Peso (kg)")
            tree.heading("pc", text="Cambio")
            tree.column("pf", width=100)
            tree.column("pp", width=80)
            tree.column("pc", width=80)
            tree.pack(fill="x")
            prev = None
            for fecha, p in pesos:
                cambio = ""
                if prev is not None:
                    diff = p - prev
                    if diff > 0: cambio = f"+{diff:.1f} kg"
                    elif diff < 0: cambio = f"{diff:.1f} kg"
                    else: cambio = "="
                tree.insert("", "end", values=(fecha, f"{p:.1f}", cambio))
                prev = p

    def _diagnosticos_recurrentes(self, parent, animal_id):
        conn = self.get_db()
        rows = conn.execute(
            "SELECT diagnostico, COUNT(*) as total FROM registros_medicos "
            "WHERE id_animal=? AND diagnostico != '' GROUP BY diagnostico HAVING total > 1 ORDER BY total DESC",
            (animal_id,)).fetchall()
        conn.close()
        if not rows:
            return
        sec, lbl = self._make_section(parent, "Diagnosticos Recurrentes", "#e74c3c")
        f = self._make_table_frame(parent)
        for r in rows:
            row_f = tk.Frame(f, bg="white", padx=8, pady=2)
            row_f.pack(fill="x")
            tk.Label(row_f, text=r["diagnostico"], font=("Segoe UI", 10),
                     bg="white", fg="#c0392b").pack(side="left")
            tk.Label(row_f, text=f"({r['total']} veces)", font=("Segoe UI", 9, "bold"),
                     bg="white", fg="#e74c3c").pack(side="left", padx=10)

    # ---------- DUENNO ----------
    def add_owner_dialog(self):
        dialog = tk.Toplevel(self.root)
        dialog.title("Nuevo Dueño")
        dialog.geometry("600x550")
        dialog.resizable(False, False)
        dialog.configure(bg="#f0f4f8")
        dialog.transient(self.root)
        dialog.grab_set()
        center_dialog(dialog, self.root)

        main_frame = tk.Frame(dialog, bg="#f0f4f8")
        main_frame.pack(fill='both', expand=True, padx=15, pady=10)

        # Owner fields
        tk.Label(main_frame, text="DATOS DEL DUEÑO", font=("Segoe UI", 11, "bold"),
                 bg="#f0f4f8", fg="#2c3e50").grid(row=0, column=0, columnspan=2, pady=(0,5), sticky='w')

        fields = [("DNI", "dni"), ("Nombre", "nombre"), ("Telefono", "telefono"),
                  ("Email", "email"), ("Direccion", "direccion")]
        entries = {}
        for i, (label, key) in enumerate(fields):
            tk.Label(main_frame, text=label, bg="#f0f4f8", font=("Segoe UI", 10)).grid(
                row=i+1, column=0, sticky="e", padx=5, pady=4)
            e = tk.Entry(main_frame, width=35, font=("Segoe UI", 10))
            e.grid(row=i+1, column=1, padx=5, pady=4, sticky='w')
            entries[key] = e

        def buscar_dni():
            num = entries["dni"].get().strip()
            if not num:
                return
            try:
                resp = requests.get('https://api.apis.net.pe/v1/dni?numero=' + num, timeout=10)
                if resp.status_code == 200:
                    data = resp.json()
                    if 'nombre' in data:
                        entries["nombre"].delete(0, 'end')
                        entries["nombre"].insert(0, data['nombre'])
                    if 'direccion' in data:
                        entries["direccion"].delete(0, 'end')
                        entries["direccion"].insert(0, data['direccion'])
            except Exception as ex:
                messagebox.showerror("Error", f"No se pudo consultar DNI: {ex}")
        tk.Button(main_frame, text="🔍", bg="#3498db", fg="white", relief="flat",
                  font=("Segoe UI", 8), cursor="hand2", command=buscar_dni).grid(
                  row=1, column=2, padx=(0, 5), pady=4)

        # Pet section
        tk.Label(main_frame, text="REGISTRAR MASCOTAS", font=("Segoe UI", 11, "bold"),
                 bg="#f0f4f8", fg="#2c3e50").grid(row=7, column=0, columnspan=3, pady=(10,5), sticky='w')

        razas_perro = [
            "Labrador Retriever", "Golden Retriever", "Pastor Aleman", "Bulldog Frances",
            "Bulldog Ingles", "Beagle", "Caniche", "Chihuahua", "Yorkshire Terrier",
            "Boxer", "Rottweiler", "Doberman", "Husky Siberiano", "Shih Tzu",
            "Pug", "Border Collie", "Dalmata", "Cocker Spaniel", "Pomerania",
            "Bichon Frisk", "Maltes", "Schnauzer", "Pastor Belga", "Samoyedo",
            "Gran Danes", "San Bernardo", "Akita Inu", "Shiba Inu", "Pit Bull",
            "American Staffordshire", "Weimaraner", "Galgos", "Whippet",
            "Terrier Escoces", "West Highland White Terrier", "Corgi",
            "Carlino", "Boston Terrier", "Jack Russell Terrier", "Fox Terrier",
            "Basenji", "Chow Chow", "Shar Pei", "Lhasa Apso", "Bulldog Americano",
            "Mastin Napolitano", "Boyero de Berna", "Papillon", "Habanero",
            "Cavalier King Charles Spaniel", "Otra"
        ]
        razas_gato = [
            "Mestizo", "Persa", "Siames", "Maine Coon", "Bengala",
            "Sphynx", "Ragdoll", "British Shorthair", "Scottish Fold",
            "Abisinio", "Bosque de Noruega", "Ruso Azul", "Oriental",
            "Cornish Rex", "Devon Rex", "Burmes", "Tonquines",
            "Siberiano", "Savannah", "Birmano", "Angora Turco",
            "La Perm", "Mau Egipcio", "Sokoke", "American Shorthair",
            "Exotico", "Himalayo", "Ocicat", "Selkirk Rex",
            "Toyger", "Otra"
        ]

        def actualizar_razas(*args):
            esp = e_pet_especie.get()
            if esp == "Perro":
                e_pet_raza.set_full_values(razas_perro)
                e_pet_raza.set("")
            elif esp == "Gato":
                e_pet_raza.set_full_values(razas_gato)
                e_pet_raza.set("")
            else:
                e_pet_raza.set_full_values([])
                e_pet_raza.set("")

        pet_fields_frame = tk.Frame(main_frame, bg="#f0f4f8")
        pet_fields_frame.grid(row=8, column=0, columnspan=3, sticky='ew')

        pet_labels = ["Nombre", "Especie", "Raza", "Edad", "Sexo"]
        for j, lbl in enumerate(pet_labels):
            tk.Label(pet_fields_frame, text=lbl, bg="#f0f4f8", font=("Segoe UI", 9)).grid(row=0, column=j, padx=3)
        e_pet_nombre = tk.Entry(pet_fields_frame, width=18, font=("Segoe UI", 9))
        e_pet_nombre.grid(row=1, column=0, padx=3)
        e_pet_especie = ttk.Combobox(pet_fields_frame, values=["Perro", "Gato", "Otro"], width=15, font=("Segoe UI", 9))
        e_pet_especie.grid(row=1, column=1, padx=3)
        e_pet_especie.set("Perro")
        e_pet_especie.bind("<<ComboboxSelected>>", actualizar_razas)
        e_pet_raza = AutocompleteEntry(pet_fields_frame, full_values=razas_perro, width=16, font=("Segoe UI", 9))
        e_pet_raza.grid(row=1, column=2, padx=3)
        e_pet_edad = tk.Spinbox(pet_fields_frame, from_=0, to=50, width=6, font=("Segoe UI", 9))
        e_pet_edad.grid(row=1, column=3, padx=3)
        e_pet_sexo = ttk.Combobox(pet_fields_frame, values=["Macho", "Hembra"], width=10, font=("Segoe UI", 9))
        e_pet_sexo.grid(row=1, column=4, padx=3)
        e_pet_sexo.set("Macho")

        pet_list_frame = tk.Frame(main_frame, bg="white", highlightbackground="#ccc", highlightthickness=1)
        pet_list_frame.grid(row=9, column=0, columnspan=3, sticky='ew', pady=5, ipady=5)
        pet_listbox = tk.Listbox(pet_list_frame, height=4, font=("Segoe UI", 9))
        pet_listbox.pack(fill='x', padx=3, pady=3)
        pets_to_add = []

        def add_pet():
            nom = e_pet_nombre.get().strip()
            if not nom:
                messagebox.showwarning("Error", "Ingrese nombre de la mascota")
                return
            pets_to_add.append({
                "nombre": nom,
                "especie": e_pet_especie.get(),
                "raza": e_pet_raza.get().strip(),
                "edad": int(e_pet_edad.get()) if e_pet_edad.get().isdigit() else 0,
                "sexo": e_pet_sexo.get(),
            })
            pet_listbox.insert('end', f"{nom} - {e_pet_especie.get()} ({e_pet_sexo.get()})")
            e_pet_nombre.delete(0, 'end')
            e_pet_raza.set("")
            e_pet_especie.set("Perro")
            e_pet_sexo.set("Macho")
            e_pet_edad.delete(0, 'end')
            e_pet_edad.insert(0, "0")
            e_pet_nombre.focus_set()

        def remove_pet():
            sel = pet_listbox.curselection()
            if sel:
                idx = sel[0]
                pet_listbox.delete(idx)
                pets_to_add.pop(idx)

        btn_pet_frame = tk.Frame(main_frame, bg="#f0f4f8")
        btn_pet_frame.grid(row=10, column=0, columnspan=3, pady=3)
        tk.Button(btn_pet_frame, text="+ Agregar Mascota", bg="#3498db", fg="white",
                  font=("Segoe UI", 9), relief="flat", padx=10, pady=2,
                  command=add_pet, cursor="hand2").pack(side='left', padx=3)
        tk.Button(btn_pet_frame, text="- Quitar", bg="#e74c3c", fg="white",
                  font=("Segoe UI", 9), relief="flat", padx=10, pady=2,
                  command=remove_pet, cursor="hand2").pack(side='left', padx=3)

        def save():
            if not validate_required([entries["nombre"]], ["Nombre"]):
                return
            conn = self.get_db()
            cursor = conn.execute(
                "INSERT INTO duenos (dni, nombre, telefono, email, direccion) VALUES (?, ?, ?, ?, ?)",
                (entries["dni"].get(), entries["nombre"].get(), entries["telefono"].get(),
                 entries["email"].get(), entries["direccion"].get()))
            dueno_id = cursor.lastrowid
            for pet in pets_to_add:
                conn.execute(
                    "INSERT INTO animales (nombre, especie, raza, edad, sexo, id_dueno) VALUES (?,?,?,?,?,?)",
                    (pet["nombre"], pet["especie"], pet["raza"], pet["edad"], pet["sexo"], dueno_id))
            conn.commit()
            conn.close()
            dialog.destroy()
            self.show_medical_history()
            msg = f"Dueño registrado"
            if pets_to_add:
                msg += f" con {len(pets_to_add)} mascota(s)"
            messagebox.showinfo("Exito", msg)

        tk.Button(main_frame, text="Guardar Todo", bg="#2ecc71", fg="white",
                  font=("Segoe UI", 10, "bold"), relief="flat", padx=30, pady=6,
                  command=save, cursor="hand2").grid(row=11, column=0, columnspan=3, pady=10)

    def show_owner_detail(self, dueno_id, detail_frame=None, reload_callback=None):
        conn = self.get_db()
        d = conn.execute("SELECT * FROM duenos WHERE id=?", (dueno_id,)).fetchone()
        if not d:
            conn.close(); return
        otras = conn.execute(
            "SELECT id, nombre, especie, raza FROM animales WHERE id_dueno=? ORDER BY nombre",
            (dueno_id,)).fetchall()
        conn.close()
        dia = tk.Toplevel(self.root)
        dia.title("Ficha del Dueño")
        w = 520 if otras else 400
        dia.geometry(f"{w}x360")
        dia.resizable(False, False)
        dia.configure(bg="#e8edf2")
        dia.transient(self.root)
        dia.grab_set()
        center_dialog(dia, self.root)
        data = [
            ("DNI", d["dni"] or "N/A"),
            ("Nombre", d["nombre"]),
            ("Telefono", d["telefono"] or "N/A"),
            ("Email", d["email"] or "N/A"),
            ("Direccion", d["direccion"] or "N/A"),
        ]
        for i, (lbl, val) in enumerate(data):
            tk.Label(dia, text=lbl + ":", font=("Segoe UI", 10, "bold"),
                     bg="#e8edf2", fg="#2c3e50").grid(row=i, column=0, sticky="e", padx=8, pady=3)
            tk.Label(dia, text=val, font=("Segoe UI", 10),
                     bg="#e8edf2", fg="#555").grid(row=i, column=1, sticky="w", padx=8, pady=3)
        if otras:
            row_offset = len(data)
            tk.Frame(dia, bg="#ddd", height=1).grid(row=row_offset, column=0, columnspan=2, sticky="ew", pady=6, padx=8)
            tk.Label(dia, text="Otras mascotas del dueño:", font=("Segoe UI", 10, "bold"),
                     bg="#e8edf2", fg="#2c3e50").grid(row=row_offset+1, column=0, columnspan=2, sticky="w", padx=8)
            for j, m in enumerate(otras):
                fr = tk.Frame(dia, bg="#e8edf2")
                fr.grid(row=row_offset+2+j, column=0, columnspan=2, sticky="ew", padx=20, pady=1)
                tk.Label(fr, text="🐾", bg="#e8edf2").pack(side="left")
                lbl_m = tk.Label(fr, text=f"{m['nombre']} ({m['especie'] or '?'} - {m['raza'] or '?'})",
                                 font=("Segoe UI", 10), bg="#e8edf2", fg="#2980b9", cursor="hand2")
                lbl_m.pack(side="left", padx=4)
                if detail_frame and reload_callback:
                    lbl_m.bind("<Button-1>", lambda e, mid=m["id"]: (
                        dia.destroy(),
                        self.show_detail(detail_frame, mid, reload_callback)
                    ))
            btn_row = row_offset + 2 + len(otras)
        else:
            btn_row = len(data)
        tk.Button(dia, text="Cerrar", bg="#e74c3c", fg="white", relief="flat",
                  command=dia.destroy, cursor="hand2").grid(row=btn_row, column=1, pady=15, sticky="e")

    def edit_owner_dialog(self, dueno_id, animal_id, detail_frame, reload_callback):
        conn = self.get_db()
        d = conn.execute("SELECT * FROM duenos WHERE id=?", (dueno_id,)).fetchone()
        conn.close()
        if not d:
            return
        dialog = tk.Toplevel(self.root)
        dialog.title("Editar Dueño")
        dialog.geometry("450x280")
        dialog.resizable(False, False)
        dialog.configure(bg="#f0f4f8")
        dialog.transient(self.root)
        dialog.grab_set()
        center_dialog(dialog, self.root)
        fields = [("DNI", "dni"), ("Nombre", "nombre"), ("Telefono", "telefono"),
                  ("Email", "email"), ("Direccion", "direccion")]
        entries = {}
        for i, (label, key) in enumerate(fields):
            tk.Label(dialog, text=label, bg="#f0f4f8", font=("Segoe UI", 10)).grid(
                row=i, column=0, sticky="e", padx=8, pady=3)
            e = tk.Entry(dialog, width=40, font=("Segoe UI", 10))
            e.insert(0, d[key] or "")
            e.grid(row=i, column=1, padx=8, pady=3)
            entries[key] = e
        def save():
            if not validate_required([entries["nombre"]], ["Nombre"]):
                return
            conn = self.get_db()
            conn.execute("UPDATE duenos SET dni=?, nombre=?, telefono=?, email=?, direccion=? WHERE id=?",
                         (entries["dni"].get(), entries["nombre"].get(), entries["telefono"].get(),
                          entries["email"].get(), entries["direccion"].get(), dueno_id))
            conn.commit()
            conn.close()
            dialog.destroy()
            self.show_detail(detail_frame, animal_id, reload_callback)
            messagebox.showinfo("Exito", "Dueño actualizado")
        add_keyboard_shortcuts(dialog, save)
        tk.Button(dialog, text="Guardar", bg="#2ecc71", fg="white",
                  font=("Segoe UI", 10), relief="flat", padx=20, pady=5,
                  command=save, cursor="hand2").grid(row=len(fields), column=1, pady=15, sticky="e")

    def export_pdf(self, animal_id):
        try:
            from reportlab.lib.pagesizes import A4
            from reportlab.pdfgen import canvas
            from reportlab.lib.units import cm
            conn = self.get_db()
            a = conn.execute(
                "SELECT a.*, d.nombre as dname, d.telefono as dtel, d.direccion as ddir "
                "FROM animales a JOIN duenos d ON a.id_dueno = d.id WHERE a.id=?",
                (animal_id,)).fetchone()
            regs = conn.execute(
                "SELECT * FROM registros_medicos WHERE id_animal=? ORDER BY fecha DESC",
                (animal_id,)).fetchall()
            vacs = conn.execute(
                "SELECT * FROM vacunas WHERE id_animal=? ORDER BY fecha DESC", (animal_id,)).fetchall()
            alergias = conn.execute(
                "SELECT * FROM alergias WHERE id_animal=? ORDER BY alergeno", (animal_id,)).fetchall()
            medicacion = conn.execute(
                "SELECT * FROM medicacion WHERE id_animal=? AND activo=1 ORDER BY fecha_inicio DESC", (animal_id,)).fetchall()
            examenes = conn.execute(
                "SELECT * FROM examenes_auxiliares WHERE id_animal=? ORDER BY fecha DESC", (animal_id,)).fetchall()
            grooming = conn.execute(
                "SELECT hg.*, sg.nombre as snombre FROM historial_grooming hg "
                "LEFT JOIN servicios_grooming sg ON hg.id_servicio=sg.id "
                "WHERE hg.id_animal=? ORDER BY hg.fecha DESC", (animal_id,)).fetchall()
            conn.close()
            if not a: return
            from tkinter import filedialog
            path = filedialog.asksaveasfilename(defaultextension=".pdf",
                filetypes=[("PDF", "*.pdf")], title="Guardar Historial como PDF")
            if not path: return
            c = canvas.Canvas(path, pagesize=A4)
            w, h = A4
            c.setFont("Helvetica-Bold", 18)
            c.drawString(2*cm, h-2*cm, f"Historial Clinico - {a['nombre']}")
            c.setFont("Helvetica", 11)
            y = h - 3.5*cm
            c.drawString(2*cm, y, f"Especie: {a['especie']}   Raza: {a['raza'] or 'N/A'}")
            y -= 0.8*cm
            c.drawString(2*cm, y, f"Edad: {a['edad']} años   Peso: {a['peso']} kg   Sexo: {a['sexo'] or 'N/A'}")
            y -= 0.8*cm
            c.drawString(2*cm, y, f"Dueño: {a['dname']}   Tel: {a['dtel'] or 'N/A'}")
            y -= 0.8*cm
            c.drawString(2*cm, y, f"Direccion: {a['ddir'] or 'N/A'}")
            y -= 1.5*cm

            # --- REGISTROS MEDICOS ---
            c.setFont("Helvetica-Bold", 14)
            c.drawString(2*cm, y, "Registros Medicos")
            y -= 0.8*cm
            c.setFont("Helvetica", 10)
            for r in regs:
                if y < 3*cm:
                    c.showPage(); c.setFont("Helvetica", 10); y = h - 2*cm
                c.drawString(2*cm, y, f"{r['fecha']} {r['hora'] or ''} - {r['doctor'] or ''}")
                y -= 0.5*cm
                items_pdf = []
                if r['diagnostico']: items_pdf.append(f"Dx: {r['diagnostico']}")
                if r['tratamiento']: items_pdf.append(f"Tx: {r['tratamiento']}")
                if r['observaciones']: items_pdf.append(f"Obs: {r['observaciones'][:80]}")
                for line in items_pdf:
                    c.drawString(2.5*cm, y, line); y -= 0.5*cm
                y -= 0.3*cm

            # --- VACUNAS ---
            if vacs:
                y -= 0.5*cm
                c.setFont("Helvetica-Bold", 14)
                c.drawString(2*cm, y, "Vacunas y Desparasitaciones")
                y -= 0.8*cm; c.setFont("Helvetica", 10)
                for v in vacs:
                    if y < 3*cm:
                        c.showPage(); c.setFont("Helvetica", 10); y = h - 2*cm
                    c.drawString(2*cm, y, f"{v['fecha']} - {v['nombre']} ({v['tipo'].capitalize()})")
                    y -= 0.5*cm
                    if v['proxima_dosis']:
                        c.drawString(2.5*cm, y, f"Proxima dosis: {v['proxima_dosis']}")
                        y -= 0.5*cm

            # --- ALERGIAS ---
            if alergias:
                y -= 0.5*cm
                c.setFont("Helvetica-Bold", 14)
                c.drawString(2*cm, y, "Alergias")
                y -= 0.8*cm; c.setFont("Helvetica", 10)
                for al in alergias:
                    if y < 3*cm:
                        c.showPage(); c.setFont("Helvetica", 10); y = h - 2*cm
                    c.drawString(2*cm, y, f"{al['alergeno']} ({al['tipo'] or ''}) - Severidad: {al['severidad']}")
                    y -= 0.5*cm

            # --- MEDICACION ACTIVA ---
            if medicacion:
                y -= 0.5*cm
                c.setFont("Helvetica-Bold", 14)
                c.drawString(2*cm, y, "Medicacion Activa")
                y -= 0.8*cm; c.setFont("Helvetica", 10)
                for m in medicacion:
                    if y < 3*cm:
                        c.showPage(); c.setFont("Helvetica", 10); y = h - 2*cm
                    c.drawString(2*cm, y, f"{m['medicamento']} {m['dosis']} - {m['frecuencia']} ({m['via'] or 'N/A'})")
                    y -= 0.5*cm
                    c.drawString(2.5*cm, y, f"Inicio: {m['fecha_inicio']}   Fin: {m['fecha_fin'] or 'Indefinido'}")
                    y -= 0.5*cm

            # --- EXAMENES AUXILIARES ---
            if examenes:
                y -= 0.5*cm
                c.setFont("Helvetica-Bold", 14)
                c.drawString(2*cm, y, "Examenes Auxiliares")
                y -= 0.8*cm; c.setFont("Helvetica", 10)
                for ex in examenes:
                    if y < 3*cm:
                        c.showPage(); c.setFont("Helvetica", 10); y = h - 2*cm
                    c.drawString(2*cm, y, f"{ex['fecha']} - {ex['nombre']} ({ex['tipo']})")
                    y -= 0.5*cm
                    if ex['resultados']:
                        c.drawString(2.5*cm, y, f"Resultados: {ex['resultados'][:80]}")
                        y -= 0.5*cm

            # --- GROOMING ---
            if grooming:
                y -= 0.5*cm
                c.setFont("Helvetica-Bold", 14)
                c.drawString(2*cm, y, "Historial Grooming")
                y -= 0.8*cm; c.setFont("Helvetica", 10)
                for g in grooming:
                    if y < 3*cm:
                        c.showPage(); c.setFont("Helvetica", 10); y = h - 2*cm
                    c.drawString(2*cm, y, f"{g['fecha']} - {g['snombre'] or 'N/A'} - S/.{g['precio']:.2f}")
                    y -= 0.5*cm

            c.save()
            messagebox.showinfo("Exito", f"PDF guardado en {path}")
        except Exception as ex:
            messagebox.showerror("Error", f"No se pudo exportar PDF: {ex}")

    def generar_receta(self, animal_id, tree):
        sel = tree.selection()
        if not sel:
            messagebox.showinfo("Receta", "Selecciona una consulta primero")
            return
        try:
            from reportlab.lib.pagesizes import A5
            from reportlab.pdfgen import canvas
            from reportlab.lib.units import cm
            conn = self.get_db()
            a = conn.execute(
                "SELECT a.*, d.nombre as dname, d.telefono as dtel, d.direccion as ddir "
                "FROM animales a JOIN duenos d ON a.id_dueno = d.id WHERE a.id=?",
                (animal_id,)).fetchone()
            r = conn.execute("SELECT * FROM registros_medicos WHERE id=?", (int(sel[0]),)).fetchone()
            conn.close()
            if not a or not r:
                messagebox.showerror("Error", "No se encontraron datos"); return
            from tkinter import filedialog
            path = filedialog.asksaveasfilename(defaultextension=".pdf",
                filetypes=[("PDF", "*.pdf")], title="Guardar Receta")
            if not path: return
            c = canvas.Canvas(path, pagesize=A5)
            w, h = A5
            # Header
            c.setFont("Helvetica-Bold", 16)
            c.drawCentredString(w/2, h-1.5*cm, "RECETA MEDICA VETERINARIA")
            c.setFont("Helvetica", 9)
            c.drawCentredString(w/2, h-2*cm, "SCRIPTYFY Veterinaria")
            c.line(1.5*cm, h-2.3*cm, w-1.5*cm, h-2.3*cm)
            y = h-3*cm
            # Data
            c.setFont("Helvetica-Bold", 10)
            c.drawString(1.5*cm, y, f"Paciente: {a['nombre']}")
            c.setFont("Helvetica", 10)
            c.drawString(8*cm, y, f"Fecha: {r['fecha']}")
            y -= 0.6*cm
            c.setFont("Helvetica", 10)
            c.drawString(1.5*cm, y, f"Especie: {a['especie']}  |  Raza: {a['raza'] or 'N/A'}  |  Peso: {a['peso']} kg")
            y -= 0.6*cm
            c.drawString(1.5*cm, y, f"Dueño: {a['dname']}  |  Tel: {a['dtel'] or 'N/A'}")
            y -= 1*cm
            # Doctor & Diagnosis
            if r['doctor']:
                c.setFont("Helvetica-Bold", 10)
                c.drawString(1.5*cm, y, f"Medico: {r['doctor']}")
                y -= 0.6*cm
            c.setFont("Helvetica-Bold", 10)
            c.drawString(1.5*cm, y, "Diagnostico:")
            y -= 0.5*cm
            c.setFont("Helvetica", 10)
            diag = r['diagnostico'] or r['diagnostico_definitivo'] or r['diagnostico_presuntivo'] or ''
            c.drawString(2*cm, y, diag)
            y -= 1*cm
            # Treatment / Prescription
            c.setFont("Helvetica-Bold", 11)
            c.drawString(1.5*cm, y, "PRESCRIPCION:")
            y -= 0.7*cm
            c.setFont("Helvetica", 10)
            tx = r['tratamiento'] or ''
            lines = []
            for line in tx.split('\n'):
                while len(line) > 70:
                    lines.append(line[:70]); line = line[70:]
                lines.append(line)
            for line in lines:
                c.drawString(2*cm, y, line); y -= 0.5*cm
                if y < 2*cm:
                    c.showPage(); c.setFont("Helvetica", 10); y = h-2*cm
            # Observations
            if r['observaciones']:
                y -= 0.5*cm
                c.setFont("Helvetica-Bold", 10)
                c.drawString(1.5*cm, y, "Observaciones:")
                y -= 0.5*cm
                c.setFont("Helvetica", 9)
                c.drawString(2*cm, y, r['observaciones'][:200])
                y -= 1*cm
            # Footer
            c.line(1.5*cm, y, w-1.5*cm, y)
            y -= 0.5*cm
            c.setFont("Helvetica", 8)
            c.drawString(1.5*cm, y, "Firma del Medico: _________________________")
            c.drawRightString(w-1.5*cm, y, f"Emision: {r['fecha']}")
            y -= 0.5*cm
            c.drawCentredString(w/2, y, "Esta receta es valida solo para el paciente indicado")
            c.save()
            # Open PDF
            os.startfile(path)
            # WhatsApp option
            if a['dtel']:
                resp = messagebox.askyesno("WhatsApp", "Enviar receta por WhatsApp?")
                if resp:
                    import webbrowser, urllib.parse
                    tel = a['dtel'].replace(' ', '').replace('-', '').replace('+', '')
                    msg = f"Hola {a['dname']}, adjunto la receta de {a['nombre']} - {r['fecha']}"
                    url = f"https://wa.me/{tel}?text={urllib.parse.quote(msg)}"
                    webbrowser.open(url)
        except Exception as ex:
            messagebox.showerror("Error", f"No se pudo generar receta: {ex}")

    def print_card(self, animal_id):
        try:
            from reportlab.lib.pagesizes import A6
            from reportlab.pdfgen import canvas
            from reportlab.lib.units import cm
            conn = self.get_db()
            a = conn.execute(
                "SELECT a.*, d.nombre as dname, d.telefono as dtel "
                "FROM animales a JOIN duenos d ON a.id_dueno = d.id WHERE a.id=?",
                (animal_id,)).fetchone()
            conn.close()
            if not a: return
            from tkinter import filedialog
            path = filedialog.asksaveasfilename(defaultextension=".pdf",
                filetypes=[("PDF", "*.pdf")], title="Guardar Ficha del Paciente")
            if not path: return
            c = canvas.Canvas(path, pagesize=A6)
            c.setFont("Helvetica-Bold", 14)
            c.drawString(1*cm, 10*cm, a['nombre'])
            c.setFont("Helvetica", 10)
            y = 9*cm
            for lbl, val in [("Especie", a['especie']), ("Raza", a['raza'] or 'N/A'),
                             ("Edad", f"{a['edad']} años"), ("Peso", f"{a['peso']} kg"),
                             ("Dueño", a['dname']), ("Tel", a['dtel'] or 'N/A')]:
                c.drawString(1*cm, y, f"{lbl}: {val}")
                y -= 0.7*cm
            c.showPage()
            c.save()
            messagebox.showinfo("Exito", f"Ficha guardada en {path}")
        except Exception as ex:
            messagebox.showerror("Error", f"No se pudo guardar: {ex}")

    def export_excel(self, animal_id):
        try:
            import csv
            conn = self.get_db()
            a = conn.execute(
                "SELECT a.*, d.nombre as dname FROM animales a JOIN duenos d ON a.id_dueno = d.id WHERE a.id=?",
                (animal_id,)).fetchone()
            regs = conn.execute(
                "SELECT * FROM registros_medicos WHERE id_animal=? ORDER BY fecha DESC",
                (animal_id,)).fetchall()
            conn.close()
            if not a:
                return
            from tkinter import filedialog
            path = filedialog.asksaveasfilename(defaultextension=".csv",
                filetypes=[("CSV", "*.csv")], title="Guardar Historial como CSV")
            if not path:
                return
            with open(path, "w", newline="", encoding="utf-8-sig") as f:
                w = csv.writer(f)
                w.writerow(["Historial Clinico - " + a["nombre"]])
                w.writerow(["Especie", a["especie"], "Raza", a["raza"] or "N/A",
                            "Edad", str(a["edad"]), "Peso", str(a["peso"]) + " kg"])
                w.writerow(["Dueño", a["dname"]])
                w.writerow([])
                w.writerow(["Fecha", "Hora", "Peso", "Doctor", "Diagnostico", "Tratamiento", "Observaciones"])
                for r in regs:
                    w.writerow([r["fecha"], r["hora"] or "", str(r["peso"]) if r["peso"] else "",
                                r["doctor"] or "", r["diagnostico"], r["tratamiento"],
                                r["observaciones"] or ""])
            messagebox.showinfo("Exito", f"Historial exportado a {path}")
        except Exception as ex:
            messagebox.showerror("Error", f"No se pudo exportar: {ex}")

    # ---------- ANIMAL ----------
    def add_animal_dialog(self, callback=None):
        dialog = tk.Toplevel(self.root)
        dialog.title("NUEVO")
        dialog.geometry("680x560")
        dialog.resizable(True, True)
        dialog.configure(bg="#eef2f7")
        dialog.transient(self.root)
        dialog.grab_set()
        center_dialog(dialog, self.root)

        selected_photo = tk.StringVar(value="")

        header = tk.Frame(dialog, bg="#2c3e50", height=40)
        header.pack(fill="x", side="top")
        header.pack_propagate(False)
        tk.Label(header, text="REGISTRAR NUEVO PACIENTE", bg="#2c3e50",
                 fg="white", font=("Segoe UI", 14, "bold")).pack(expand=True)

        main = tk.Frame(dialog, bg="#eef2f7", padx=15, pady=8)
        main.pack(fill="both", expand=True)
        main.columnconfigure(1, weight=1)

        razas_perro = [
            "Labrador Retriever", "Golden Retriever", "Pastor Aleman", "Bulldog Frances",
            "Bulldog Ingles", "Beagle", "Caniche", "Chihuahua", "Yorkshire Terrier",
            "Boxer", "Rottweiler", "Doberman", "Husky Siberiano", "Shih Tzu",
            "Pug", "Border Collie", "Dalmata", "Cocker Spaniel", "Pomerania",
            "Bichon Frisé", "Maltes", "Schnauzer", "Pastor Belga", "Samoyedo",
            "Gran Danes", "San Bernardo", "Akita Inu", "Shiba Inu", "Pit Bull",
            "American Staffordshire", "Weimaraner", "Galgos", "Whippet",
            "Terrier Escoces", "West Highland White Terrier", "Corgi",
            "Carlino", "Boston Terrier", "Jack Russell Terrier", "Fox Terrier",
            "Basenji", "Chow Chow", "Shar Pei", "Lhasa Apso", "Bulldog Americano",
            "Mastin Napolitano", "Boyero de Berna", "Papillon", "Habanero",
            "Cavalier King Charles Spaniel", "Otra"
        ]
        razas_gato = [
            "Mestizo", "Persa", "Siamés", "Maine Coon", "Bengala",
            "Sphynx", "Ragdoll", "British Shorthair", "Scottish Fold",
            "Abisinio", "Bosque de Noruega", "Ruso Azul", "Oriental",
            "Cornish Rex", "Devon Rex", "Burmes", "Tonquines",
            "Siberiano", "Savannah", "Birmano", "Angora Turco",
            "La Perm", "Mau Egipcio", "Sokoke", "American Shorthair",
            "Exotico", "Himalayo", "Ocicat", "Selkirk Rex",
            "Toyger", "Otra"
        ]

        def actualizar_razas(*args):
            esp = e_especie.get()
            if esp == "Perro":
                e_raza.set_full_values(razas_perro)
                e_raza.set("")
            elif esp == "Gato":
                e_raza.set_full_values(razas_gato)
                e_raza.set("")
            else:
                e_raza.set_full_values([])
                e_raza.set("")

        pf = {"bg": "#eef2f7", "font": ("Segoe UI", 10)}
        ef = {"width": 22, "font": ("Segoe UI", 10)}

        r = 0
        tk.Label(main, text="Nombre", **pf).grid(row=r, column=0, sticky="e", padx=8, pady=3)
        e_nombre = tk.Entry(main, width=30, font=("Segoe UI", 10))
        e_nombre.grid(row=r, column=1, padx=8, pady=3, sticky="ew")
        r += 1

        tk.Label(main, text="Especie", **pf).grid(row=r, column=0, sticky="e", padx=8, pady=3)
        e_especie = ttk.Combobox(main, values=["Perro", "Gato", "Otros"], width=22, font=("Segoe UI", 10))
        e_especie.grid(row=r, column=1, padx=8, pady=3, sticky="w")
        e_especie.bind("<<ComboboxSelected>>", actualizar_razas)
        r += 1

        tk.Label(main, text="Raza", **pf).grid(row=r, column=0, sticky="e", padx=8, pady=3)
        e_raza = AutocompleteEntry(main, full_values=[], width=22, font=("Segoe UI", 10))
        e_raza.grid(row=r, column=1, padx=10, pady=5, sticky="w")
        r += 1

        tk.Label(main, text="Edad", **pf).grid(row=r, column=0, sticky="e", padx=8, pady=3)
        edad_frame = tk.Frame(main, bg="#eef2f7")
        edad_frame.grid(row=r, column=1, padx=10, pady=5, sticky="w")
        e_edad = tk.Entry(edad_frame, width=10, font=("Segoe UI", 10))
        e_edad.pack(side="left")
        tk.Label(edad_frame, text="años", bg="#eef2f7", font=("Segoe UI", 9)).pack(side="left", padx=3)
        from tkcalendar import DateEntry
        tk.Label(edad_frame, text="  Fec.Nac:", bg="#eef2f7", font=("Segoe UI", 9)).pack(side="left", padx=(10, 2))
        e_fecnac = DateEntry(edad_frame, width=12, font=("Segoe UI", 9),
                             background="#2c3e50", foreground="white",
                             borderwidth=2, date_pattern="yyyy-mm-dd")
        e_fecnac.pack(side="left")
        def calc_edad(*args):
            try:
                nac = e_fecnac.get_date()
                hoy = date.today()
                edad_calc = hoy.year - nac.year - ((hoy.month, hoy.day) < (nac.month, nac.day))
                e_edad.delete(0, "end")
                e_edad.insert(0, str(edad_calc))
            except:
                pass
        e_fecnac.bind("<<DateEntrySelected>>", calc_edad)
        r += 1

        tk.Label(main, text="Peso (kg)", **pf).grid(row=r, column=0, sticky="e", padx=8, pady=3)
        e_peso = tk.Entry(main, **ef)
        e_peso.grid(row=r, column=1, padx=8, pady=3, sticky="w")
        r += 1

        tk.Label(main, text="Sexo", **pf).grid(row=r, column=0, sticky="e", padx=8, pady=3)
        e_sexo = ttk.Combobox(main, values=["Macho", "Hembra"], width=22, font=("Segoe UI", 10))
        e_sexo.grid(row=r, column=1, padx=10, pady=5, sticky="w")
        r += 1

        tk.Label(main, text="Esterilizado", **pf).grid(row=r, column=0, sticky="e", padx=8, pady=3)
        e_esterilizado = ttk.Combobox(main, values=["No", "Si"], width=22, font=("Segoe UI", 10))
        e_esterilizado.set("No")
        e_esterilizado.grid(row=r, column=1, padx=10, pady=5, sticky="w")
        r += 1

        tk.Label(main, text="Color", **pf).grid(row=r, column=0, sticky="e", padx=8, pady=3)
        e_color = tk.Entry(main, **ef)
        e_color.grid(row=r, column=1, padx=8, pady=3, sticky="w")
        r += 1

        lf_dueno = tk.LabelFrame(main, text="Datos del Dueño", bg="#eef2f7",
                                 font=("Segoe UI", 10, "bold"), relief="groove")
        lf_dueno.grid(row=r, column=0, columnspan=2, padx=8, pady=6, sticky="ew")
        lf_dueno.columnconfigure(1, weight=1)
        lf_dueno.columnconfigure(3, weight=1)

        ld = {"bg": "#eef2f7", "font": ("Segoe UI", 9)}
        ed = {"width": 35, "font": ("Segoe UI", 10)}

        tk.Label(lf_dueno, text="DNI/RUC:", **ld).grid(row=0, column=0, sticky="e", padx=5, pady=2)
        dni_frame = tk.Frame(lf_dueno, bg="#eef2f7")
        dni_frame.grid(row=0, column=1, padx=5, pady=2, sticky="w")
        e_dueno_dni = tk.Entry(dni_frame, width=14, font=("Segoe UI", 10))
        e_dueno_dni.pack(side="left")

        def lookup_document():
            doc = e_dueno_dni.get().strip()
            if not doc:
                return
            import urllib.request, urllib.error, json
            try:
                if len(doc) == 8:
                    url = f"https://api.apis.net.pe/v1/dni?numero={doc}"
                elif len(doc) == 11:
                    url = f"https://api.apis.net.pe/v1/ruc?numero={doc}"
                else:
                    messagebox.showerror("Error", "DNI debe tener 8 dígitos o RUC 11 dígitos")
                    return
                req = urllib.request.Request(url, headers={"Accept": "application/json", "User-Agent": "Mozilla/5.0", "Referer": "https://apis.net.pe/"})
                with urllib.request.urlopen(req, timeout=5) as resp:
                    data = json.loads(resp.read().decode())
                if "error" in data:
                    messagebox.showerror("Error", f"DNI/RUC no válido: {data['error']}. Verifique el número e intente de nuevo.")
                    return
                if len(doc) == 8:
                    if "nombre" in data and data["nombre"]:
                        e_dueno_nombre.delete("1.0", "end")
                        e_dueno_nombre.insert("1.0", data["nombre"].strip())
                elif len(doc) == 11:
                    if "nombre" in data and data["nombre"]:
                        e_dueno_nombre.delete("1.0", "end")
                        e_dueno_nombre.insert("1.0", data["nombre"])
                    if "direccion" in data and data["direccion"]:
                        e_dueno_dir.delete(0, "end")
                        e_dueno_dir.insert(0, data["direccion"])
            except urllib.error.HTTPError as e:
                messagebox.showerror("Error", f"No se encontraron datos para el {('DNI' if len(doc)==8 else 'RUC')} ingresado. Verifique el número (HTTP {e.code})")
            except Exception as ex:
                messagebox.showerror("Error", f"No se pudo consultar: {ex}")

        tk.Button(dni_frame, text="🔍", bg="#3498db", fg="white", relief="flat",
                  width=2, font=("Segoe UI", 9), command=lookup_document,
                  cursor="hand2").pack(side="left", padx=(3, 0))

        tk.Label(lf_dueno, text="Nombre:", **ld).grid(row=0, column=2, sticky="ne", padx=(15, 5), pady=3)
        e_dueno_nombre = tk.Text(lf_dueno, height=2, width=30, font=("Segoe UI", 10), wrap="word", relief="solid", bd=1)
        e_dueno_nombre.grid(row=0, column=3, padx=5, pady=2, sticky="ew")

        tk.Label(lf_dueno, text="Celular:", **ld).grid(row=1, column=0, sticky="e", padx=5, pady=2)
        e_dueno_tel = tk.Entry(lf_dueno, **ed)
        e_dueno_tel.grid(row=1, column=1, padx=5, pady=2, sticky="ew")

        tk.Label(lf_dueno, text="Dirección:", **ld).grid(row=1, column=2, sticky="e", padx=(15, 5), pady=2)
        e_dueno_dir = tk.Entry(lf_dueno, **ed)
        e_dueno_dir.grid(row=1, column=3, padx=5, pady=2, sticky="ew")

        tk.Label(lf_dueno, text="Email:", **ld).grid(row=2, column=0, sticky="e", padx=5, pady=2)
        e_dueno_email = tk.Entry(lf_dueno, **ed)
        e_dueno_email.grid(row=2, column=1, padx=5, pady=2, sticky="ew")
        r += 1

        def choose_photo():
            path = filedialog.askopenfilename(
                filetypes=[("Imagenes", "*.jpg *.jpeg *.png *.bmp *.gif")])
            if path:
                selected_photo.set(path)
                photo_label.config(text="Foto: " + os.path.basename(path))

        photo_row = r
        tk.Button(main, text="📷 Seleccionar Foto", bg="#3498db", fg="white",
                  relief="flat", command=choose_photo, cursor="hand2",
                  font=("Segoe UI", 10)).grid(row=photo_row, column=0, padx=8, pady=4)
        photo_label = tk.Label(main, text="Ninguna foto seleccionada",
                               bg="#eef2f7", fg="#999", font=("Segoe UI", 9))
        photo_label.grid(row=photo_row, column=1, padx=8, pady=4, sticky="w")

        def save():
            if not e_dueno_nombre.get("1.0", "end-1c").strip():
                messagebox.showerror("Error", "El nombre del Dueño es obligatorio")
                return
            if not validate_required([e_nombre, e_edad, e_peso], ["Nombre", "Edad", "Peso"]):
                return
            try:
                int(e_edad.get())
                float(e_peso.get())
            except ValueError:
                messagebox.showerror("Error", "Edad debe ser numero entero y Peso debe ser numero")
                return
            try:
                conn = self.get_db()
                cur = conn.execute(
                    "INSERT INTO duenos (nombre, dni, telefono, direccion, email) VALUES (?, ?, ?, ?, ?)",
                    (e_dueno_nombre.get("1.0", "end-1c").strip(), e_dueno_dni.get().strip(),
                     e_dueno_tel.get().strip(), e_dueno_dir.get().strip(),
                     e_dueno_email.get().strip()))
                id_dueno = cur.lastrowid
                fecnac = ""
                try:
                    fecnac = e_fecnac.get()
                except:
                    pass
                cur = conn.execute(
                    "INSERT INTO animales (nombre, especie, raza, edad, peso, sexo, color, fecha_nacimiento, esterilizado, id_dueno, foto) "
                    "VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
                    (e_nombre.get(), e_especie.get(), e_raza.get(),
                     int(e_edad.get()), float(e_peso.get()),
                     e_sexo.get(), e_color.get(), fecnac,
                     1 if e_esterilizado.get() == "Si" else 0,
                     id_dueno, ""))
                animal_id = cur.lastrowid
                if selected_photo.get():
                    ext = os.path.splitext(selected_photo.get())[1]
                    dest = os.path.join(PHOTOS_DIR, str(animal_id) + ext)
                    shutil.copy2(selected_photo.get(), dest)
                    conn.execute("UPDATE animales SET foto=? WHERE id=?",
                                 (str(animal_id) + ext, animal_id))
                conn.commit()
                conn.close()
                dialog.destroy()
                if callback:
                    callback()
                self.show_medical_history()
                messagebox.showinfo("Exito", "Animal registrado")
            except Exception as ex:
                messagebox.showerror("Error", str(ex))

        btn_frame = tk.Frame(dialog, bg="#eef2f7")
        btn_frame.pack(fill="x", pady=(0, 15))
        tk.Button(btn_frame, text="Guardar", bg="#27ae60", fg="white",
                  font=("Segoe UI", 11, "bold"), relief="flat", padx=30, pady=8,
                  command=save, cursor="hand2").pack(pady=5)
        dialog.bind("<Escape>", lambda e: dialog.destroy())
        dialog.bind("<Control-Return>", lambda e: save())

    def edit_animal_dialog(self, animal_id, reload_callback):
        conn = self.get_db()
        a = conn.execute("SELECT * FROM animales WHERE id=?", (animal_id,)).fetchone()
        duenos = conn.execute("SELECT id, nombre FROM duenos ORDER BY nombre").fetchall()
        conn.close()
        if not a:
            return
        dialog = tk.Toplevel(self.root)
        dialog.title("Editar Animal")
        dialog.geometry("500x520")
        dialog.resizable(False, False)
        dialog.configure(bg="#e8edf2")
        dialog.transient(self.root)
        dialog.grab_set()
        center_dialog(dialog, self.root)
        tk.Label(dialog, text="Nombre", bg="#e8edf2", font=("Segoe UI", 10)).grid(row=0, column=0, sticky="e", padx=8, pady=3)
        e_nombre = tk.Entry(dialog, width=40, font=("Segoe UI", 10))
        e_nombre.insert(0, a["nombre"])
        e_nombre.grid(row=0, column=1, padx=8, pady=3)
        tk.Label(dialog, text="Especie", bg="#e8edf2", font=("Segoe UI", 10)).grid(row=1, column=0, sticky="e", padx=8, pady=3)
        e_especie = tk.Entry(dialog, width=40, font=("Segoe UI", 10))
        e_especie.insert(0, a["especie"] or "")
        e_especie.grid(row=1, column=1, padx=8, pady=3)
        tk.Label(dialog, text="Raza", bg="#e8edf2", font=("Segoe UI", 10)).grid(row=2, column=0, sticky="e", padx=8, pady=3)
        e_raza = tk.Entry(dialog, width=40, font=("Segoe UI", 10))
        e_raza.insert(0, a["raza"] or "")
        e_raza.grid(row=2, column=1, padx=8, pady=3)
        tk.Label(dialog, text="Edad", bg="#e8edf2", font=("Segoe UI", 10)).grid(row=3, column=0, sticky="e", padx=8, pady=3)
        e_edad = tk.Entry(dialog, width=40, font=("Segoe UI", 10))
        e_edad.insert(0, str(a["edad"]))
        e_edad.grid(row=3, column=1, padx=8, pady=3)
        tk.Label(dialog, text="Peso (kg)", bg="#e8edf2", font=("Segoe UI", 10)).grid(row=4, column=0, sticky="e", padx=8, pady=3)
        e_peso = tk.Entry(dialog, width=40, font=("Segoe UI", 10))
        e_peso.insert(0, str(a["peso"]))
        e_peso.grid(row=4, column=1, padx=8, pady=3)
        tk.Label(dialog, text="Sexo", bg="#e8edf2", font=("Segoe UI", 10)).grid(row=5, column=0, sticky="e", padx=8, pady=3)
        e_sexo = ttk.Combobox(dialog, values=["Macho", "Hembra"], width=37, font=("Segoe UI", 10))
        e_sexo.set(a["sexo"] or "")
        e_sexo.grid(row=5, column=1, padx=8, pady=3)

        tk.Label(dialog, text="Color", bg="#e8edf2", font=("Segoe UI", 10)).grid(row=6, column=0, sticky="e", padx=8, pady=3)
        e_color = tk.Entry(dialog, width=40, font=("Segoe UI", 10))
        e_color.insert(0, a["color"] or "")
        e_color.grid(row=6, column=1, padx=8, pady=3)

        dueno_map = {d["nombre"]: d["id"] for d in duenos}
        tk.Label(dialog, text="Dueño", bg="#e8edf2", font=("Segoe UI", 10)).grid(row=7, column=0, sticky="e", padx=8, pady=3)
        e_dueno = ttk.Combobox(dialog, values=list(dueno_map.keys()), width=37, font=("Segoe UI", 10))
        dueno_orig = conn.execute("SELECT nombre FROM duenos WHERE id=?", (a["id_dueno"],)).fetchone()
        if dueno_orig:
            e_dueno.set(dueno_orig["nombre"])
        e_dueno.grid(row=7, column=1, padx=8, pady=3)
        def save():
            conn = self.get_db()
            conn.execute("UPDATE animales SET nombre=?, especie=?, raza=?, edad=?, peso=?, sexo=?, color=?, id_dueno=? WHERE id=?",
                         (e_nombre.get(), e_especie.get(), e_raza.get(), int(e_edad.get()),
                          float(e_peso.get()), e_sexo.get(), e_color.get(),
                          dueno_map[e_dueno.get()], animal_id))
            conn.commit()
            conn.close()
            dialog.destroy()
            reload_callback()
            messagebox.showinfo("Exito", "Animal actualizado")
        tk.Button(dialog, text="Guardar", bg="#2ecc71", fg="white",
                  font=("Segoe UI", 10), relief="flat", padx=20, pady=5,
                  command=save, cursor="hand2").grid(row=8, column=1, pady=15, sticky="e")

    def edit_registro_dialog(self, reg_id, animal_id, callback):
        conn = self.get_db()
        r = conn.execute("SELECT * FROM registros_medicos WHERE id=?", (reg_id,)).fetchone()
        conn.close()
        if not r:
            return
        dialog = tk.Toplevel(self.root)
        dialog.title("Editar Consulta")
        dialog.geometry("720x680")
        dialog.resizable(True, True)
        dialog.configure(bg="#e8edf2")
        dialog.transient(self.root)
        dialog.grab_set()
        center_dialog(dialog, self.root)

        # Scrollable canvas
        canvas = tk.Canvas(dialog, bg="#e8edf2", highlightthickness=0)
        v_scroll = ttk.Scrollbar(dialog, orient="vertical", command=canvas.yview)
        scrollable = tk.Frame(canvas, bg="#e8edf2")
        scrollable.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.create_window((0, 0), window=scrollable, anchor="nw")
        canvas.configure(yscrollcommand=v_scroll.set)
        canvas.pack(side="left", fill="both", expand=True)
        v_scroll.pack(side="right", fill="y")
        canvas.bind("<MouseWheel>", lambda e: canvas.yview_scroll(int(-1*(e.delta/120)), "units"))

        main = scrollable

        pf = {"bg": "#e8edf2", "font": ("Segoe UI", 10)}
        ef = {"width": 44, "font": ("Segoe UI", 10)}

        def parse_exam(val):
            if not val:
                return "", ""
            if val.startswith("FILE:"):
                return "", val[5:]
            if "|||FILE:" in val:
                parts = val.split("|||FILE:", 1)
                return parts[0], parts[1]
            return val, ""

        row_i = 0
        tk.Label(main, text="Peso (kg)", **pf).grid(row=row_i, column=0, sticky="e", padx=12, pady=6)
        e_peso = tk.Entry(main, **ef)
        e_peso.insert(0, str(r["peso"]) if r["peso"] else "")
        e_peso.grid(row=row_i, column=1, padx=12, pady=6, sticky="w")
        row_i += 1

        e_doctor = self._make_doctor_frame(main, row_i, r["doctor"] or "")
        row_i += 1

        tk.Label(main, text="Anamnesis", **pf).grid(row=row_i, column=0, sticky="ne", padx=12, pady=6)
        e_anamnesis = tk.Text(main, width=46, height=3, font=("Segoe UI", 10))
        e_anamnesis.insert("1.0", r["anamnesis"] or "")
        e_anamnesis.grid(row=row_i, column=1, padx=12, pady=6, sticky="ew")
        row_i += 1

        tk.Label(main, text="Diagnóstico Presuntivo", **pf).grid(row=row_i, column=0, sticky="e", padx=12, pady=6)
        e_dx_pres = tk.Entry(main, **ef)
        e_dx_pres.insert(0, r["diagnostico_presuntivo"] or "")
        e_dx_pres.grid(row=row_i, column=1, padx=12, pady=6, sticky="ew")
        row_i += 1

        tk.Label(main, text="Diagnóstico Definitivo", **pf).grid(row=row_i, column=0, sticky="e", padx=12, pady=6)
        e_dx_def = tk.Entry(main, **ef)
        e_dx_def.insert(0, r["diagnostico_definitivo"] or "")
        e_dx_def.grid(row=row_i, column=1, padx=12, pady=6, sticky="ew")
        row_i += 1

        tk.Label(main, text="Diagnóstico", **pf).grid(row=row_i, column=0, sticky="e", padx=12, pady=6)
        e_diag = tk.Entry(main, **ef)
        e_diag.insert(0, r["diagnostico"] or "")
        e_diag.grid(row=row_i, column=1, padx=12, pady=6, sticky="ew")
        row_i += 1

        tk.Label(main, text="Tratamiento", **pf).grid(row=row_i, column=0, sticky="e", padx=12, pady=6)
        e_trat = tk.Entry(main, **ef)
        e_trat.insert(0, r["tratamiento"] or "")
        e_trat.grid(row=row_i, column=1, padx=12, pady=6, sticky="ew")
        row_i += 1

        tk.Label(main, text="Observaciones", **pf).grid(row=row_i, column=0, sticky="ne", padx=12, pady=6)
        e_obs = tk.Text(main, width=46, height=2, font=("Segoe UI", 10))
        e_obs.insert("1.0", r["observaciones"] or "")
        e_obs.grid(row=row_i, column=1, padx=12, pady=6, sticky="ew")
        row_i += 1

        # Exámenes Complementarios
        lf_exam = tk.LabelFrame(main, text="Exámenes Complementarios", bg="#e8edf2",
                                 font=("Segoe UI", 10, "bold"), relief="groove")
        lf_exam.grid(row=row_i, column=0, columnspan=2, padx=12, pady=12, sticky="ew")
        lf_exam.columnconfigure(1, weight=1)
        lf_exam.columnconfigure(3, weight=1)

        def make_exam_row(lf, label, row, init_val):
            txt, img = parse_exam(init_val)
            tk.Label(lf, text=label+":", bg="#e8edf2", font=("Segoe UI", 9)).grid(row=row, column=0, sticky="e", padx=5, pady=3)
            entry = tk.Entry(lf, width=25, font=("Segoe UI", 10))
            entry.insert(0, txt)
            entry.grid(row=row, column=1, padx=5, pady=3, sticky="ew")
            img_path = tk.StringVar(value=img)
            fn = os.path.basename(img) if img else ""
            img_label = tk.Label(lf, text=fn[:20] if fn else "", bg="#e8edf2",
                                 fg="#2980b9" if fn else "#999", font=("Segoe UI", 8),
                                 cursor="hand2" if fn else "")
            if fn:
                img_label.config(cursor="hand2")
                img_label.bind("<Button-1>", lambda e, p=img: self._open_file(p))
            img_label.grid(row=row, column=2, padx=(5, 0), pady=3, sticky="w")
            def pick_image(ip=img_path, il=img_label):
                path = filedialog.askopenfilename(filetypes=[("Imagenes", "*.jpg *.jpeg *.png *.bmp *.gif")])
                if path:
                    ip.set(path)
                    il.config(text=os.path.basename(path)[:20], fg="#2980b9", cursor="hand2")
                    il.bind("<Button-1>", lambda e, p=path: self._open_file(p))
            tk.Button(lf, text="📷", bg="#3498db", fg="white", relief="flat",
                      width=2, font=("Segoe UI", 9), command=pick_image,
                      cursor="hand2").grid(row=row, column=3, padx=(4, 0), pady=3, sticky="w")
            return entry, img_path

        e_eco, ip_eco = make_exam_row(lf_exam, "Ecografía", 0, r["exam_ecografia"])
        e_radio, ip_radio = make_exam_row(lf_exam, "Radiografía", 1, r["exam_radiografia"])
        e_hemo, ip_hemo = make_exam_row(lf_exam, "Hemograma", 2, r["exam_hemograma"])
        e_bioq, ip_bioq = make_exam_row(lf_exam, "Bioquímico", 3, r["exam_bioquimico"])
        e_orina, ip_orina = make_exam_row(lf_exam, "Orina", 4, r["exam_orina"])
        e_otros, ip_otros = make_exam_row(lf_exam, "Otros", 5, r["exam_otros"])
        row_i += 1

        def save():
            conn = self.get_db()
            peso = e_peso.get().strip()
            peso_val = float(peso) if peso else None

            def get_exam_val(entry, ip):
                txt = entry.get().strip()
                f = ip.get().strip()
                if f:
                    return f"{txt}|||FILE:{f}" if txt else f"FILE:{f}"
                return txt

            conn.execute(
                "UPDATE registros_medicos SET peso=?, doctor=?, diagnostico=?, tratamiento=?, observaciones=?, "
                "anamnesis=?, diagnostico_presuntivo=?, diagnostico_definitivo=?, "
                "exam_ecografia=?, exam_radiografia=?, exam_hemograma=?, exam_bioquimico=?, exam_orina=?, exam_otros=? "
                "WHERE id=?",
                (peso_val, e_doctor.get(), e_diag.get(), e_trat.get(), e_obs.get("1.0", "end-1c"),
                 e_anamnesis.get("1.0", "end-1c"), e_dx_pres.get(), e_dx_def.get(),
                 get_exam_val(e_eco, ip_eco), get_exam_val(e_radio, ip_radio),
                 get_exam_val(e_hemo, ip_hemo), get_exam_val(e_bioq, ip_bioq),
                 get_exam_val(e_orina, ip_orina), get_exam_val(e_otros, ip_otros), reg_id))
            conn.commit()
            conn.close()
            dialog.destroy()
            callback()
            messagebox.showinfo("Exito", "Registro actualizado")

        btn_frame = tk.Frame(main, bg="#e8edf2")
        btn_frame.grid(row=row_i, column=0, columnspan=2, pady=10)
        tk.Button(btn_frame, text="Guardar", bg="#2ecc71", fg="white",
                  font=("Segoe UI", 11, "bold"), relief="flat", padx=30, pady=8,
                  command=save, cursor="hand2").pack()

    def _add_signos_vitales_dialog(self, animal_id, detail_frame, reload_callback):
        dialog = tk.Toplevel(self.root)
        dialog.title("Agregar Signos Vitales")
        dialog.geometry("400x300")
        dialog.resizable(False, False)
        dialog.configure(bg="#e8edf2")
        dialog.transient(self.root)
        dialog.grab_set()
        center_dialog(dialog, self.root)
        conn = self.get_db()
        registros = conn.execute(
            "SELECT id, fecha, diagnostico FROM registros_medicos WHERE id_animal=? ORDER BY fecha DESC",
            (animal_id,)).fetchall()
        conn.close()
        reg_map = {f"{r['fecha']} - {r['diagnostico'][:30]}": r["id"] for r in registros}
        tk.Label(dialog, text="Consulta:", bg="#e8edf2", font=("Segoe UI", 10)).grid(row=0, column=0, sticky="e", padx=8, pady=3)
        e_reg = ttk.Combobox(dialog, values=list(reg_map.keys()), width=35, font=("Segoe UI", 10))
        e_reg.grid(row=0, column=1, padx=8, pady=3)
        if reg_map:
            e_reg.set(list(reg_map.keys())[0])
        fields_data = [(1, "Temperatura (C)", "temp"), (2, "Frec. Cardiaca (lpm)", "fc"),
                       (3, "Frec. Respiratoria (rpm)", "fr"), (4, "Presion Sistolica", "sis"),
                       (5, "Presion Diastolica", "dis")]
        entries = {}
        for row, label, key in fields_data:
            tk.Label(dialog, text=label, bg="#e8edf2", font=("Segoe UI", 10)).grid(row=row, column=0, sticky="e", padx=10, pady=4)
            e = tk.Entry(dialog, width=20, font=("Segoe UI", 10))
            e.grid(row=row, column=1, padx=10, pady=4, sticky="w")
            entries[key] = e

        def save():
            if not e_reg.get():
                messagebox.showerror("Error", "Seleccione una consulta"); return
            conn = self.get_db()
            reg_id = reg_map[e_reg.get()]
            conn.execute(
                "INSERT OR REPLACE INTO signos_vitales (id_registro, temperatura, frecuencia_cardiaca, frecuencia_respiratoria, presion_sistolica, presion_diastolica) VALUES (?,?,?,?,?,?)",
                (reg_id,
                 float(entries["temp"].get()) if entries["temp"].get() else None,
                 int(entries["fc"].get()) if entries["fc"].get() else None,
                 int(entries["fr"].get()) if entries["fr"].get() else None,
                 int(entries["sis"].get()) if entries["sis"].get() else None,
                 int(entries["dis"].get()) if entries["dis"].get() else None))
            conn.commit()
            conn.close()
            dialog.destroy()
            self.show_detail(detail_frame, animal_id, reload_callback)

        tk.Button(dialog, text="Guardar", bg="#2ecc71", fg="white",
                  font=("Segoe UI", 10), relief="flat", padx=20, pady=5,
                  command=save, cursor="hand2").grid(row=6, column=1, pady=15, sticky="e")

    def _make_doctor_frame(self, parent, row, default=""):
        frame = tk.Frame(parent, bg="#e8edf2")
        frame.grid(row=row, column=0, columnspan=2, padx=12, pady=6, sticky="w")
        tk.Label(frame, text="Doctor", bg="#e8edf2", font=("Segoe UI", 10)).pack(side="left", padx=(0, 8))
        valores = self.get_doctores()
        cb = ttk.Combobox(frame, values=valores, width=32, font=("Segoe UI", 10), state="normal")
        if default:
            cb.set(default)
        cb.pack(side="left")
        tk.Button(frame, text="+", bg="#2ecc71", fg="white", font=("Segoe UI", 9, "bold"),
                  relief="flat", width=2, cursor="hand2",
                  command=lambda: (self.manage_doctores_dialog(
                      lambda: cb.configure(values=self.get_doctores())))).pack(side="left", padx=(4, 0))
        return cb
    def _open_file(self, path):
        if path and os.path.exists(path):
            os.startfile(path)
    def add_appointment_dialog(self, animal_id=None):
        from tkcalendar import DateEntry
        dialog = tk.Toplevel(self.root)
        dialog.title("Nueva Cita")
        dialog.geometry("450x300")
        dialog.resizable(False, False)
        dialog.configure(bg="#f0f4f8")
        dialog.transient(self.root)
        dialog.grab_set()
        center_dialog(dialog, self.root)

        conn = self.get_db()
        animales = conn.execute(
            "SELECT a.id, a.nombre, d.nombre as dname "
            "FROM animales a JOIN duenos d ON a.id_dueno = d.id "
            "ORDER BY a.nombre").fetchall()
        conn.close()
        animal_map = {a["nombre"] + " (" + a["dname"] + ")": a["id"] for a in animales}

        tk.Label(dialog, text="Animal", bg="#f0f4f8", font=("Segoe UI", 10)).grid(
            row=0, column=0, sticky="e", padx=10, pady=8)
        e_animal = AutocompleteEntry(dialog, full_values=list(animal_map.keys()),
                                        width=37, font=("Segoe UI", 10))
        e_animal.grid(row=0, column=1, padx=10, pady=8)

        tk.Label(dialog, text="Tipo", bg="#f0f4f8", font=("Segoe UI", 10)).grid(
            row=1, column=0, sticky="e", padx=10, pady=8)
        e_tipo = ttk.Combobox(dialog, values=["veterinaria", "grooming"], width=37, font=("Segoe UI", 10))
        e_tipo.set("veterinaria")
        e_tipo.grid(row=1, column=1, padx=10, pady=8)

        if animal_id:
            for name, aid in animal_map.items():
                if aid == animal_id:
                    e_animal.set(name)
                    break

        tk.Label(dialog, text="Fecha", bg="#f0f4f8", font=("Segoe UI", 10)).grid(
            row=2, column=0, sticky="e", padx=10, pady=8)
        e_fecha = DateEntry(dialog, width=37, font=("Segoe UI", 10),
                            background="#2c3e50", foreground="white",
                            borderwidth=2, date_pattern="yyyy-mm-dd")
        e_fecha.grid(row=2, column=1, padx=10, pady=8)

        def cargar_servicios():
            tipo = e_tipo.get()
            conn = self.get_db()
            if tipo == "grooming":
                rows = conn.execute("SELECT nombre, precio FROM servicios_grooming").fetchall()
            else:
                rows = conn.execute("SELECT nombre, precio FROM servicios_medicos").fetchall()
            conn.close()
            servicios = [(r['nombre'], r['precio']) for r in rows]
            servicios.append(("Otros", 0))
            e_servicio['values'] = [s[0] for s in servicios]
            e_servicio.servicio_precios = {s[0]: s[1] for s in servicios}
            e_servicio.set("")
            e_precio.delete(0, 'end')
            e_precio.configure(state='normal')

        tk.Label(dialog, text="Servicio", bg="#f0f4f8", font=("Segoe UI", 10)).grid(
            row=3, column=0, sticky="e", padx=10, pady=8)
        e_servicio = ttk.Combobox(dialog, values=[], width=37, font=("Segoe UI", 10))
        e_servicio.grid(row=3, column=1, padx=10, pady=8)
        e_servicio.servicio_precios = {}

        def servicio_selected(*args):
            nom = e_servicio.get()
            precio = e_servicio.servicio_precios.get(nom)
            if precio is not None:
                e_precio.configure(state='normal')
                e_precio.delete(0, 'end')
                e_precio.insert(0, str(precio))
                if nom != "Otros":
                    e_precio.configure(state='readonly')

        e_servicio.bind("<<ComboboxSelected>>", servicio_selected)
        e_tipo.bind("<<ComboboxSelected>>", lambda e: cargar_servicios())

        tk.Label(dialog, text="Precio", bg="#f0f4f8", font=("Segoe UI", 10)).grid(
            row=4, column=0, sticky="e", padx=10, pady=8)
        e_precio = tk.Entry(dialog, width=40, font=("Segoe UI", 10))
        e_precio.grid(row=4, column=1, padx=10, pady=8)

        cargar_servicios()

        def save():
            animal_id = animal_map[e_animal.get()]
            conn = self.get_db()
            id_dueno = conn.execute(
                "SELECT id_dueno FROM animales WHERE id=?", (animal_id,)).fetchone()[0]
            animal_nombre = conn.execute(
                "SELECT nombre FROM animales WHERE id=?", (animal_id,)).fetchone()[0]
            dueno_info = conn.execute(
                "SELECT nombre, telefono FROM duenos WHERE id=?", (id_dueno,)).fetchone()
            conn.execute(
                "INSERT INTO citas (id_animal, id_dueno, fecha, motivo, tipo, precio) VALUES (?, ?, ?, ?, ?, ?)",
                (animal_id, id_dueno, e_fecha.get(), e_servicio.get(), e_tipo.get(), float(e_precio.get() or 0)))
            conn.commit()
            conn.close()
            dialog.destroy()
            receipt = (
                "--- COMPROBANTE DE CITA ---\n\n"
                f"Animal: {animal_nombre}\n"
                f"Due\u00f1o: {dueno_info['nombre']}\n"
                f"Tel: {dueno_info['telefono'] or 'N/A'}\n"
                f"Fecha: {e_fecha.get()}\n"
                f"Servicio: {e_servicio.get()}\n\n"
                "-----------------------------"
            )
            messagebox.showinfo("Cita Agendada", receipt)
            if self._current_view == 'dashboard':
                self.show_dashboard()

        tk.Button(dialog, text="Guardar", bg="#2ecc71", fg="white",
                  font=("Segoe UI", 10), relief="flat", padx=20, pady=5,
                  command=save, cursor="hand2").grid(row=5, column=1, pady=15, sticky="e")

    # ---------- REGISTRO MEDICO ----------
    def add_medical_record_dialog(self):
        dialog = tk.Toplevel(self.root)
        dialog.title("Nueva Consulta")
        dialog.geometry("720x680")
        dialog.resizable(True, True)
        dialog.configure(bg="#e8edf2")
        dialog.transient(self.root)
        dialog.grab_set()
        center_dialog(dialog, self.root)

        conn = self.get_db()
        animales = conn.execute(
            "SELECT a.id, a.nombre, d.nombre as dname "
            "FROM animales a JOIN duenos d ON a.id_dueno = d.id "
            "ORDER BY a.nombre").fetchall()
        conn.close()
        animal_map = {a["nombre"] + " (" + a["dname"] + ")": a["id"] for a in animales}

        # Scrollable canvas
        canvas = tk.Canvas(dialog, bg="#e8edf2", highlightthickness=0)
        v_scroll = ttk.Scrollbar(dialog, orient="vertical", command=canvas.yview)
        scrollable = tk.Frame(canvas, bg="#e8edf2")
        scrollable.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.create_window((0, 0), window=scrollable, anchor="nw")
        canvas.configure(yscrollcommand=v_scroll.set)
        canvas.pack(side="left", fill="both", expand=True)
        v_scroll.pack(side="right", fill="y")
        canvas.bind("<MouseWheel>", lambda e: canvas.yview_scroll(int(-1*(e.delta/120)), "units"))

        main = scrollable

        pf = {"bg": "#e8edf2", "font": ("Segoe UI", 10)}
        ef = {"width": 44, "font": ("Segoe UI", 10)}

        r = 0
        tk.Label(main, text="Animal", **pf).grid(row=r, column=0, sticky="e", padx=12, pady=6)
        e_animal = AutocompleteEntry(main, full_values=list(animal_map.keys()),
                                        width=42, font=("Segoe UI", 10))
        e_animal.grid(row=r, column=1, padx=12, pady=6, sticky="w")
        r += 1

        tk.Label(main, text="Peso (kg)", **pf).grid(row=r, column=0, sticky="e", padx=12, pady=6)
        e_peso = tk.Entry(main, **ef)
        e_peso.grid(row=r, column=1, padx=12, pady=6, sticky="w")
        r += 1

        # --- SIGNOS VITALES ---
        sv_frame = tk.LabelFrame(main, text="Signos Vitales", bg="#e8edf2",
                                  font=("Segoe UI", 10, "bold"), relief="groove")
        sv_frame.grid(row=r, column=0, columnspan=2, padx=12, pady=6, sticky="ew")
        tk.Label(sv_frame, text="Temperatura (\u00b0C)", bg="#e8edf2", font=("Segoe UI", 9)).grid(row=0, column=0, padx=5, pady=2)
        e_temp = tk.Entry(sv_frame, width=10, font=("Segoe UI", 9))
        e_temp.grid(row=1, column=0, padx=5, pady=2)
        tk.Label(sv_frame, text="Frec. Cardiaca (lpm)", bg="#e8edf2", font=("Segoe UI", 9)).grid(row=0, column=1, padx=5, pady=2)
        e_fc = tk.Entry(sv_frame, width=10, font=("Segoe UI", 9))
        e_fc.grid(row=1, column=1, padx=5, pady=2)
        tk.Label(sv_frame, text="Frec. Respiratoria (rpm)", bg="#e8edf2", font=("Segoe UI", 9)).grid(row=0, column=2, padx=5, pady=2)
        e_fr = tk.Entry(sv_frame, width=10, font=("Segoe UI", 9))
        e_fr.grid(row=1, column=2, padx=5, pady=2)
        tk.Label(sv_frame, text="Presion Arterial", bg="#e8edf2", font=("Segoe UI", 9)).grid(row=0, column=3, padx=5, pady=2)
        pres_frame = tk.Frame(sv_frame, bg="#e8edf2")
        pres_frame.grid(row=1, column=3, padx=5, pady=2)
        e_pres_sis = tk.Entry(pres_frame, width=5, font=("Segoe UI", 9))
        e_pres_sis.pack(side="left")
        tk.Label(pres_frame, text="/", bg="#e8edf2", font=("Segoe UI", 9)).pack(side="left")
        e_pres_dias = tk.Entry(pres_frame, width=5, font=("Segoe UI", 9))
        e_pres_dias.pack(side="left")
        r += 1

        e_doctor = self._make_doctor_frame(main, r)
        r += 1

        tk.Label(main, text="Anamnesis", **pf).grid(row=r, column=0, sticky="ne", padx=12, pady=6)
        e_anamnesis = tk.Text(main, width=46, height=3, font=("Segoe UI", 10))
        e_anamnesis.grid(row=r, column=1, padx=12, pady=6, sticky="ew")
        r += 1

        tk.Label(main, text="Diagnóstico Presuntivo", **pf).grid(row=r, column=0, sticky="e", padx=12, pady=6)
        e_dx_pres = tk.Entry(main, **ef)
        e_dx_pres.grid(row=r, column=1, padx=12, pady=6, sticky="ew")
        r += 1

        tk.Label(main, text="Diagnóstico Definitivo", **pf).grid(row=r, column=0, sticky="e", padx=12, pady=6)
        e_dx_def = tk.Entry(main, **ef)
        e_dx_def.grid(row=r, column=1, padx=12, pady=6, sticky="ew")
        r += 1

        tk.Label(main, text="Diagnóstico", **pf).grid(row=r, column=0, sticky="e", padx=12, pady=6)
        e_diag = tk.Entry(main, **ef)
        e_diag.grid(row=r, column=1, padx=12, pady=6, sticky="ew")
        r += 1

        tk.Label(main, text="Tratamiento", **pf).grid(row=r, column=0, sticky="e", padx=12, pady=6)
        e_trat = tk.Entry(main, **ef)
        e_trat.grid(row=r, column=1, padx=12, pady=6, sticky="ew")
        r += 1

        tk.Label(main, text="Observaciones", **pf).grid(row=r, column=0, sticky="ne", padx=12, pady=6)
        e_obs = tk.Text(main, width=46, height=2, font=("Segoe UI", 10))
        e_obs.grid(row=r, column=1, padx=12, pady=6, sticky="ew")
        r += 1

        # Exámenes Complementarios
        lf_exam = tk.LabelFrame(main, text="Exámenes Complementarios", bg="#e8edf2",
                                 font=("Segoe UI", 10, "bold"), relief="groove")
        lf_exam.grid(row=r, column=0, columnspan=2, padx=12, pady=12, sticky="ew")
        lf_exam.columnconfigure(1, weight=1)
        lf_exam.columnconfigure(3, weight=1)

        def make_exam_row(lf, label, row):
            tk.Label(lf, text=label+":", bg="#e8edf2", font=("Segoe UI", 9)).grid(row=row, column=0, sticky="e", padx=5, pady=3)
            entry = tk.Entry(lf, width=25, font=("Segoe UI", 10))
            entry.grid(row=row, column=1, padx=5, pady=3, sticky="ew")
            img_path = tk.StringVar()
            img_label = tk.Label(lf, text="", bg="#e8edf2", fg="#999", font=("Segoe UI", 8))
            img_label.grid(row=row, column=2, padx=(5, 0), pady=3, sticky="w")
            def pick_image(ip=img_path, il=img_label):
                path = filedialog.askopenfilename(filetypes=[("Imagenes", "*.jpg *.jpeg *.png *.bmp *.gif")])
                if path:
                    ip.set(path)
                    il.config(text=os.path.basename(path)[:20], fg="#2980b9", cursor="hand2")
                    il.bind("<Button-1>", lambda e, p=path: self._open_file(p))
            tk.Button(lf, text="📷", bg="#3498db", fg="white", relief="flat",
                      width=2, font=("Segoe UI", 9), command=pick_image,
                      cursor="hand2").grid(row=row, column=3, padx=(4, 0), pady=3, sticky="w")
            return entry, img_path

        e_eco, ip_eco = make_exam_row(lf_exam, "Ecografía", 0)
        e_radio, ip_radio = make_exam_row(lf_exam, "Radiografía", 1)
        e_hemo, ip_hemo = make_exam_row(lf_exam, "Hemograma", 2)
        e_bioq, ip_bioq = make_exam_row(lf_exam, "Bioquímico", 3)
        e_orina, ip_orina = make_exam_row(lf_exam, "Orina", 4)
        e_otros, ip_otros = make_exam_row(lf_exam, "Otros", 5)
        r += 1

        # --- INSUMOS UTILIZADOS ---
        insumos_frame = tk.LabelFrame(main, text="Insumos Utilizados (descuentan stock)",
                                      bg="#e8edf2", font=("Segoe UI", 10, "bold"), relief="groove")
        insumos_frame.grid(row=r, column=0, columnspan=2, padx=12, pady=6, sticky="ew")
        conn_ins = self.get_db()
        prod_rows = conn_ins.execute("SELECT id, nombre, stock FROM productos WHERE activo=1 AND stock > 0 ORDER BY nombre").fetchall()
        conn_ins.close()
        prod_map = {p['nombre'] + ' (stock: ' + str(p['stock']) + ')': p['id'] for p in prod_rows}
        tk.Label(insumos_frame, text="Producto", bg="#e8edf2", font=("Segoe UI", 9)).grid(row=0, column=0, padx=5, pady=2)
        e_insumo = AutocompleteEntry(insumos_frame, full_values=list(prod_map.keys()), width=30, font=("Segoe UI", 9))
        e_insumo.grid(row=1, column=0, padx=5, pady=2)
        tk.Label(insumos_frame, text="Cant.", bg="#e8edf2", font=("Segoe UI", 9)).grid(row=0, column=1, padx=5, pady=2)
        e_ins_cant = tk.Spinbox(insumos_frame, from_=1, to=999, width=5, font=("Segoe UI", 9))
        e_ins_cant.grid(row=1, column=1, padx=5, pady=2)
        insumos_usados = []
        ins_listbox = tk.Listbox(insumos_frame, height=3, font=("Segoe UI", 8))
        ins_listbox.grid(row=2, column=0, columnspan=3, sticky='ew', padx=5, pady=2)
        def agregar_insumo():
            nom = e_insumo.get().strip()
            if not nom or nom not in prod_map:
                messagebox.showwarning("Error", "Seleccione un producto valido")
                return
            cant = int(e_ins_cant.get() or 1)
            prod_id = prod_map[nom]
            insumos_usados.append({'nombre': nom.split(' (stock:')[0], 'cantidad': cant, 'id': prod_id})
            ins_listbox.insert('end', '{} x {}'.format(nom.split(' (stock:')[0], cant))
            e_insumo.set("")
            e_ins_cant.delete(0, 'end')
            e_ins_cant.insert(0, '1')
        def quitar_insumo():
            sel = ins_listbox.curselection()
            if sel:
                ins_listbox.delete(sel[0])
                insumos_usados.pop(sel[0])
        tk.Button(insumos_frame, text="+", bg="#2ecc71", fg="white", relief="flat",
                  width=3, font=("Segoe UI", 9, "bold"), command=agregar_insumo, cursor="hand2").grid(row=1, column=2, padx=2)
        tk.Button(insumos_frame, text="-", bg="#e74c3c", fg="white", relief="flat",
                  width=3, font=("Segoe UI", 9, "bold"), command=quitar_insumo, cursor="hand2").grid(row=2, column=2, padx=2, sticky='n')
        r += 1

        # --- PROXIMO CONTROL ---
        tk.Label(main, text="Proximo Control", **pf).grid(row=r, column=0, sticky="e", padx=12, pady=6)
        from tkcalendar import DateEntry
        e_prox_control = DateEntry(main, width=20, font=("Segoe UI", 10),
                                   background="#2c3e50", foreground="white",
                                   borderwidth=2, date_pattern="yyyy-mm-dd")
        e_prox_control.grid(row=r, column=1, padx=12, pady=6, sticky="w")
        r += 1

        # --- PENDIENTE DE COBRO ---
        var_pendiente = tk.IntVar(value=0)
        tk.Checkbutton(main, text="Pendiente de cobro", variable=var_pendiente,
                       bg="#e8edf2", font=("Segoe UI", 10)).grid(row=r, column=0, columnspan=2, padx=12, pady=4, sticky="w")
        r += 1

        def save():
            conn = self.get_db()
            peso = e_peso.get().strip()
            peso_val = float(peso) if peso else None

            def get_exam_val(entry, ip):
                txt = entry.get().strip()
                f = ip.get().strip()
                if f:
                    return f"{txt}|||FILE:{f}" if txt else f"FILE:{f}"
                return txt

            cursor = conn.execute(
                "INSERT INTO registros_medicos (id_animal, fecha, hora, peso, doctor, diagnostico, tratamiento, observaciones, "
                "anamnesis, diagnostico_presuntivo, diagnostico_definitivo, "
                "exam_ecografia, exam_radiografia, exam_hemograma, exam_bioquimico, exam_orina, exam_otros, "
                "proximo_control, pendiente_cobro) "
                "VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
                (animal_map[e_animal.get()], str(date.today()),
                 datetime.now().strftime("%H:%M"), peso_val, e_doctor.get(),
                 e_diag.get(), e_trat.get(), e_obs.get("1.0", "end-1c"),
                 e_anamnesis.get("1.0", "end-1c"), e_dx_pres.get(), e_dx_def.get(),
                 get_exam_val(e_eco, ip_eco), get_exam_val(e_radio, ip_radio),
                 get_exam_val(e_hemo, ip_hemo), get_exam_val(e_bioq, ip_bioq),
                 get_exam_val(e_orina, ip_orina), get_exam_val(e_otros, ip_otros),
                 e_prox_control.get(), var_pendiente.get()))
            registro_id = cursor.lastrowid
            if e_temp.get().strip() or e_fc.get().strip() or e_fr.get().strip() or e_pres_sis.get().strip():
                conn.execute(
                    "INSERT INTO signos_vitales (id_registro, temperatura, frecuencia_cardiaca, "
                    "frecuencia_respiratoria, presion_sistolica, presion_diastolica) "
                    "VALUES (?, ?, ?, ?, ?, ?)",
                    (registro_id,
                     float(e_temp.get()) if e_temp.get().strip() else None,
                     int(e_fc.get()) if e_fc.get().strip() else None,
                     int(e_fr.get()) if e_fr.get().strip() else None,
                     int(e_pres_sis.get()) if e_pres_sis.get().strip() else None,
                     int(e_pres_dias.get()) if e_pres_dias.get().strip() else None))
            for ins in insumos_usados:
                conn.execute("UPDATE productos SET stock = stock - ? WHERE id = ? AND stock >= ?",
                            (ins['cantidad'], ins['id'], ins['cantidad']))
            conn.commit()
            conn.close()
            dialog.destroy()
            self.show_medical_history()
            msg = "Registro agregado"
            if insumos_usados:
                msg += " con descuento de stock"
            messagebox.showinfo("Exito", msg)

        btn_frame = tk.Frame(main, bg="#e8edf2")
        btn_frame.grid(row=r, column=0, columnspan=2, pady=10)
        tk.Button(btn_frame, text="Guardar", bg="#2ecc71", fg="white",
                  font=("Segoe UI", 11, "bold"), relief="flat", padx=30, pady=8,
                  command=save, cursor="hand2").pack()

    def add_medical_record_for(self, animal_id, callback=None):
        dialog = tk.Toplevel(self.root)
        dialog.title("Nueva Consulta")
        dialog.geometry("720x680")
        dialog.resizable(True, True)
        dialog.configure(bg="#e8edf2")
        dialog.transient(self.root)
        dialog.grab_set()
        center_dialog(dialog, self.root)

        # Scrollable canvas
        canvas = tk.Canvas(dialog, bg="#e8edf2", highlightthickness=0)
        v_scroll = ttk.Scrollbar(dialog, orient="vertical", command=canvas.yview)
        scrollable = tk.Frame(canvas, bg="#e8edf2")
        scrollable.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.create_window((0, 0), window=scrollable, anchor="nw")
        canvas.configure(yscrollcommand=v_scroll.set)
        canvas.pack(side="left", fill="both", expand=True)
        v_scroll.pack(side="right", fill="y")
        canvas.bind("<MouseWheel>", lambda e: canvas.yview_scroll(int(-1*(e.delta/120)), "units"))

        main = scrollable

        pf = {"bg": "#e8edf2", "font": ("Segoe UI", 10)}
        ef = {"width": 44, "font": ("Segoe UI", 10)}

        r = 0
        tk.Label(main, text="Peso (kg)", **pf).grid(row=r, column=0, sticky="e", padx=12, pady=6)
        e_peso = tk.Entry(main, **ef)
        e_peso.grid(row=r, column=1, padx=12, pady=6, sticky="w")
        r += 1

        e_doctor = self._make_doctor_frame(main, r)
        r += 1

        tk.Label(main, text="Anamnesis", **pf).grid(row=r, column=0, sticky="ne", padx=12, pady=6)
        e_anamnesis = tk.Text(main, width=46, height=3, font=("Segoe UI", 10))
        e_anamnesis.grid(row=r, column=1, padx=12, pady=6, sticky="ew")
        r += 1

        tk.Label(main, text="Diagnóstico Presuntivo", **pf).grid(row=r, column=0, sticky="e", padx=12, pady=6)
        e_dx_pres = tk.Entry(main, **ef)
        e_dx_pres.grid(row=r, column=1, padx=12, pady=6, sticky="ew")
        r += 1

        tk.Label(main, text="Diagnóstico Definitivo", **pf).grid(row=r, column=0, sticky="e", padx=12, pady=6)
        e_dx_def = tk.Entry(main, **ef)
        e_dx_def.grid(row=r, column=1, padx=12, pady=6, sticky="ew")
        r += 1

        tk.Label(main, text="Diagnóstico", **pf).grid(row=r, column=0, sticky="e", padx=12, pady=6)
        e_diag = tk.Entry(main, **ef)
        e_diag.grid(row=r, column=1, padx=12, pady=6, sticky="ew")
        r += 1

        tk.Label(main, text="Tratamiento", **pf).grid(row=r, column=0, sticky="e", padx=12, pady=6)
        e_trat = tk.Entry(main, **ef)
        e_trat.grid(row=r, column=1, padx=12, pady=6, sticky="ew")
        r += 1

        tk.Label(main, text="Observaciones", **pf).grid(row=r, column=0, sticky="ne", padx=12, pady=6)
        e_obs = tk.Text(main, width=46, height=2, font=("Segoe UI", 10))
        e_obs.grid(row=r, column=1, padx=12, pady=6, sticky="ew")
        r += 1

        # Exámenes Complementarios
        lf_exam = tk.LabelFrame(main, text="Exámenes Complementarios", bg="#e8edf2",
                                 font=("Segoe UI", 10, "bold"), relief="groove")
        lf_exam.grid(row=r, column=0, columnspan=2, padx=12, pady=12, sticky="ew")
        lf_exam.columnconfigure(1, weight=1)
        lf_exam.columnconfigure(3, weight=1)

        def make_exam_row(lf, label, row):
            tk.Label(lf, text=label+":", bg="#e8edf2", font=("Segoe UI", 9)).grid(row=row, column=0, sticky="e", padx=5, pady=3)
            entry = tk.Entry(lf, width=25, font=("Segoe UI", 10))
            entry.grid(row=row, column=1, padx=5, pady=3, sticky="ew")
            img_path = tk.StringVar()
            img_label = tk.Label(lf, text="", bg="#e8edf2", fg="#999", font=("Segoe UI", 8))
            img_label.grid(row=row, column=2, padx=(5, 0), pady=3, sticky="w")
            def pick_image():
                path = filedialog.askopenfilename(filetypes=[("Imagenes", "*.jpg *.jpeg *.png *.bmp *.gif")])
                if path:
                    img_path.set(path)
                    img_label.config(text=os.path.basename(path)[:20], fg="#2980b9", cursor="hand2")
                    img_label.bind("<Button-1>", lambda e, p=path: self._open_file(p))
            tk.Button(lf, text="📷", bg="#3498db", fg="white", relief="flat",
                      width=2, font=("Segoe UI", 9), command=pick_image,
                      cursor="hand2").grid(row=row, column=3, padx=(4, 0), pady=3, sticky="w")
            return entry, img_path

        e_eco, ip_eco = make_exam_row(lf_exam, "Ecografía", 0)
        e_radio, ip_radio = make_exam_row(lf_exam, "Radiografía", 1)
        e_hemo, ip_hemo = make_exam_row(lf_exam, "Hemograma", 2)
        e_bioq, ip_bioq = make_exam_row(lf_exam, "Bioquímico", 3)
        e_orina, ip_orina = make_exam_row(lf_exam, "Orina", 4)
        e_otros, ip_otros = make_exam_row(lf_exam, "Otros", 5)
        r += 1

        def save():
            conn = self.get_db()
            peso = e_peso.get().strip()
            peso_val = float(peso) if peso else None

            def get_exam_val(entry, ip):
                txt = entry.get().strip()
                f = ip.get().strip()
                if f:
                    return f"{txt}|||FILE:{f}" if txt else f"FILE:{f}"
                return txt

            conn.execute(
                "INSERT INTO registros_medicos (id_animal, fecha, hora, peso, doctor, diagnostico, tratamiento, observaciones, "
                "anamnesis, diagnostico_presuntivo, diagnostico_definitivo, "
                "exam_ecografia, exam_radiografia, exam_hemograma, exam_bioquimico, exam_orina, exam_otros) "
                "VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
                (animal_id, str(date.today()),
                 datetime.now().strftime("%H:%M"), peso_val, e_doctor.get(),
                 e_diag.get(), e_trat.get(), e_obs.get("1.0", "end-1c"),
                 e_anamnesis.get("1.0", "end-1c"), e_dx_pres.get(), e_dx_def.get(),
                 get_exam_val(e_eco, ip_eco), get_exam_val(e_radio, ip_radio),
                 get_exam_val(e_hemo, ip_hemo), get_exam_val(e_bioq, ip_bioq),
                 get_exam_val(e_orina, ip_orina), get_exam_val(e_otros, ip_otros)))
            conn.commit()
            conn.close()
            dialog.destroy()
            if callback:
                callback()
            messagebox.showinfo("Exito", "Registro agregado")

        btn_frame = tk.Frame(main, bg="#e8edf2")
        btn_frame.grid(row=r, column=0, columnspan=2, pady=10)
        tk.Button(btn_frame, text="Guardar", bg="#2ecc71", fg="white",
                  font=("Segoe UI", 11, "bold"), relief="flat", padx=30, pady=8,
                  command=save, cursor="hand2").pack()



    # ---------- SERVICIOS Y PRODUCTOS (COMBINADO) ----------
    def show_servicios_productos(self):
        self._current_view = 'servicios_productos'
        self.clear_frame()
        main = tk.Frame(self.content_area, bg='#f0f4f8')
        main.pack(fill='both', expand=True, padx=20, pady=15)

        tk.Label(main, text='Servicios y Productos', font=('Segoe UI', 16, 'bold'),
                 bg='#f0f4f8', fg='#2c3e50').pack(anchor='w')

        notebook = ttk.Notebook(main)
        notebook.pack(fill='both', expand=True, pady=(10, 0))

        # Tab 1: Productos
        tab_prod = tk.Frame(notebook, bg='#f0f4f8')
        notebook.add(tab_prod, text='Productos')

        top_prod = tk.Frame(tab_prod, bg='#f0f4f8')
        top_prod.pack(fill='x')
        tk.Label(top_prod, text='Inventario', font=('Segoe UI', 14, 'bold'),
                 bg='#f0f4f8', fg='#2c3e50').pack(side='left')
        tk.Button(top_prod, text='+ Nuevo Producto', bg='#2ecc71', fg='white',
                  font=('Segoe UI', 9), relief='flat', padx=10, pady=3,
                  command=lambda: self.add_producto_dialog(self.show_servicios_productos), cursor='hand2').pack(side='right')

        cols_p = ('id', 'nombre', 'categoria', 'precio_compra', 'precio_venta', 'stock', 'fecha_venc')
        tree_p = ttk.Treeview(tab_prod, columns=cols_p, show='headings', height=12)
        tree_p.heading('id', text='ID')
        tree_p.heading('nombre', text='Nombre')
        tree_p.heading('categoria', text='Categoria')
        tree_p.heading('precio_compra', text='Costo')
        tree_p.heading('precio_venta', text='Venta')
        tree_p.heading('stock', text='Stock')
        tree_p.heading('fecha_venc', text='Vencimiento')
        tree_p.column('id', width=40)
        tree_p.column('nombre', width=200)
        tree_p.column('categoria', width=120)
        tree_p.column('precio_compra', width=70)
        tree_p.column('precio_venta', width=70)
        tree_p.column('stock', width=60)
        tree_p.column('fecha_venc', width=90)
        tree_p.pack(fill='both', expand=True, pady=5)

        def load_productos():
            for item in tree_p.get_children():
                tree_p.delete(item)
            conn = self.get_db()
            rows = conn.execute('SELECT * FROM productos WHERE activo=1 ORDER BY nombre').fetchall()
            conn.close()
            for r in rows:
                item = tree_p.insert('', 'end', values=(
                    r['id'], r['nombre'], r['categoria'],
                    f'S/.{r["precio_compra"]:.2f}', f'S/.{r["precio_venta"]:.2f}',
                    r['stock'], r['fecha_vencimiento'] or ''))
                stock = r['stock']
                if stock <= 2:
                    tree_p.tag_configure('critical', background='#fadbd8')
                    tree_p.item(item, tags=('critical',))
                elif stock <= 5:
                    tree_p.tag_configure('low', background='#fef9e7')
                    tree_p.item(item, tags=('low',))

        load_productos()

        tree_p.bind('<Double-1>', lambda e: self.add_producto_dialog(self.show_servicios_productos, tree_p))

        btn_prod = tk.Frame(tab_prod, bg='#f0f4f8')
        btn_prod.pack(fill='x')
        tk.Button(btn_prod, text='Eliminar', bg='#e74c3c', fg='white',
                  font=('Segoe UI', 9), relief='flat', padx=10, pady=2,
                  command=lambda: self._delete_producto(tree_p, load_productos), cursor='hand2').pack(side='left', padx=2)
        tk.Button(btn_prod, text='Exportar CSV', bg='#3498db', fg='white',
                  font=('Segoe UI', 9), relief='flat', padx=10, pady=2,
                  command=lambda: self._export_to_csv(
                      [[tree_p.item(i, 'values')[0], tree_p.item(i, 'values')[1], tree_p.item(i, 'values')[2], tree_p.item(i, 'values')[3], tree_p.item(i, 'values')[4], tree_p.item(i, 'values')[5], tree_p.item(i, 'values')[6]] for i in tree_p.get_children()],
                      ['ID', 'Nombre', 'Categoria', 'Costo', 'Venta', 'Stock', 'Vencimiento'],
                      'productos.csv'
                  ), cursor='hand2').pack(side='left', padx=2)
        tk.Button(btn_prod, text='Exportar Excel', bg='#27ae60', fg='white',
                  font=('Segoe UI', 9), relief='flat', padx=10, pady=2,
                  command=lambda: self._export_to_excel(
                      [[tree_p.item(i, 'values')[0], tree_p.item(i, 'values')[1], tree_p.item(i, 'values')[2], tree_p.item(i, 'values')[3], tree_p.item(i, 'values')[4], tree_p.item(i, 'values')[5], tree_p.item(i, 'values')[6]] for i in tree_p.get_children()],
                      ['ID', 'Nombre', 'Categoria', 'Costo', 'Venta', 'Stock', 'Vencimiento'],
                      'productos.xlsx'
                  ), cursor='hand2').pack(side='left', padx=2)

        # Tab 2: Servicios Grooming
        tab_serv = tk.Frame(notebook, bg='#f0f4f8')
        notebook.add(tab_serv, text='Servicios Grooming')

        top_serv = tk.Frame(tab_serv, bg='#f0f4f8')
        top_serv.pack(fill='x')
        tk.Button(top_serv, text='+ Nuevo Servicio', bg='#2ecc71', fg='white',
                  font=('Segoe UI', 9), relief='flat', padx=10, pady=3,
                  command=lambda: [self.add_grooming_service_dialog(), self.show_servicios_productos()], cursor='hand2').pack(side='right')

        cols_s = ('id', 'nombre', 'descripcion', 'precio', 'tipo')
        tree_s = ttk.Treeview(tab_serv, columns=cols_s, show='headings', height=12)
        tree_s.heading('id', text='ID')
        tree_s.heading('nombre', text='Nombre')
        tree_s.heading('descripcion', text='Descripcion')
        tree_s.heading('precio', text='Precio')
        tree_s.heading('tipo', text='Tipo')
        tree_s.column('id', width=40)
        tree_s.column('nombre', width=200)
        tree_s.column('descripcion', width=300)
        tree_s.column('precio', width=70)
        tree_s.column('tipo', width=100)
        tree_s.pack(fill='both', expand=True, pady=5)

        def load_servicios():
            for item in tree_s.get_children():
                tree_s.delete(item)
            conn = self.get_db()
            rows = conn.execute('SELECT * FROM servicios_grooming ORDER BY nombre').fetchall()
            conn.close()
            for r in rows:
                tree_s.insert('', 'end', values=(r['id'], r['nombre'], r['descripcion'],
                                               f'S/.{r["precio"]:.2f}', r['tipo']))

        load_servicios()

        def delete_servicio():
            sel = tree_s.selection()
            if not sel: return
            vals = tree_s.item(sel[0], 'values')
            if messagebox.askyesno('Confirmar', 'Eliminar servicio?'):
                conn = self.get_db()
                conn.execute('DELETE FROM servicios_grooming WHERE id=?', (int(vals[0]),))
                conn.commit()
                conn.close()
                load_servicios()

        tree_s.bind('<Double-1>', lambda e: self.edit_grooming_service_dialog(tree_s))

        btn_serv = tk.Frame(tab_serv, bg='#f0f4f8')
        btn_serv.pack(fill='x')
        tk.Button(btn_serv, text='Eliminar', bg='#e74c3c', fg='white',
                  font=('Segoe UI', 9), relief='flat', padx=10, pady=2,
                  command=delete_servicio, cursor='hand2').pack(side='left', padx=2)

        # Tab 3: Servicios Medicos
        tab_med = tk.Frame(notebook, bg='#f0f4f8')
        notebook.add(tab_med, text='Servicios Medicos')

        top_med = tk.Frame(tab_med, bg='#f0f4f8')
        top_med.pack(fill='x')
        tk.Button(top_med, text='+ Nuevo Servicio Medico', bg='#2ecc71', fg='white',
                  font=('Segoe UI', 9), relief='flat', padx=10, pady=3,
                  command=lambda: [self.add_servicio_medico_dialog(), self.show_servicios_productos()], cursor='hand2').pack(side='right')

        cols_m = ('id', 'nombre', 'descripcion', 'precio', 'tipo')
        tree_m = ttk.Treeview(tab_med, columns=cols_m, show='headings', height=12)
        tree_m.heading('id', text='ID')
        tree_m.heading('nombre', text='Nombre')
        tree_m.heading('descripcion', text='Descripcion')
        tree_m.heading('precio', text='Precio')
        tree_m.heading('tipo', text='Tipo')
        tree_m.column('id', width=40)
        tree_m.column('nombre', width=200)
        tree_m.column('descripcion', width=300)
        tree_m.column('precio', width=70)
        tree_m.column('tipo', width=100)
        tree_m.pack(fill='both', expand=True, pady=5)

        def load_servicios_med():
            for item in tree_m.get_children():
                tree_m.delete(item)
            conn = self.get_db()
            rows = conn.execute('SELECT * FROM servicios_medicos ORDER BY nombre').fetchall()
            conn.close()
            for r in rows:
                tree_m.insert('', 'end', values=(r['id'], r['nombre'], r['descripcion'],
                                               f'S/.{r["precio"]:.2f}', r['tipo']))

        load_servicios_med()

        def delete_servicio_med():
            sel = tree_m.selection()
            if not sel: return
            vals = tree_m.item(sel[0], 'values')
            if messagebox.askyesno('Confirmar', 'Eliminar servicio medico?'):
                conn = self.get_db()
                conn.execute('DELETE FROM servicios_medicos WHERE id=?', (int(vals[0]),))
                conn.commit()
                conn.close()
                load_servicios_med()

        tree_m.bind('<Double-1>', lambda e: self.edit_servicio_medico_dialog(tree_m))

        btn_med = tk.Frame(tab_med, bg='#f0f4f8')
        btn_med.pack(fill='x')
        tk.Button(btn_med, text='Eliminar', bg='#e74c3c', fg='white',
                  font=('Segoe UI', 9), relief='flat', padx=10, pady=2,
                  command=delete_servicio_med, cursor='hand2').pack(side='left', padx=2)



    # ---------- GROOMING / PELUQUERIA ----------
    def show_grooming(self):
        self._current_view = 'grooming'
        self.clear_frame()
        main = tk.Frame(self.content_area, bg='#f0f4f8')
        main.pack(fill='both', expand=True, padx=20, pady=15)

        tk.Label(main, text='Peluqueria / Grooming', font=('Segoe UI', 16, 'bold'),
                 bg='#f0f4f8', fg='#2c3e50').pack(anchor='w')

        notebook = ttk.Notebook(main)
        notebook.pack(fill='both', expand=True, pady=(10, 0))

        # Tab 1: Servicios
        tab1 = tk.Frame(notebook, bg='#f0f4f8')
        notebook.add(tab1, text='Servicios')

        top1 = tk.Frame(tab1, bg='#f0f4f8')
        top1.pack(fill='x')
        tk.Button(top1, text='+ Nuevo Servicio', bg='#2ecc71', fg='white',
                  font=('Segoe UI', 9), relief='flat', padx=10, pady=3,
                  command=self.add_grooming_service_dialog, cursor='hand2').pack(side='right')

        cols_s = ('id', 'nombre', 'descripcion', 'precio', 'tipo')
        tree_s = ttk.Treeview(tab1, columns=cols_s, show='headings', height=12)
        tree_s.heading('id', text='ID')
        tree_s.heading('nombre', text='Nombre')
        tree_s.heading('descripcion', text='Descripcion')
        tree_s.heading('precio', text='Precio')
        tree_s.heading('tipo', text='Tipo')
        tree_s.column('id', width=40)
        tree_s.column('nombre', width=200)
        tree_s.column('descripcion', width=300)
        tree_s.column('precio', width=70)
        tree_s.column('tipo', width=100)
        tree_s.pack(fill='both', expand=True, pady=5)

        def load_servicios():
            for item in tree_s.get_children():
                tree_s.delete(item)
            conn = self.get_db()
            rows = conn.execute('SELECT * FROM servicios_grooming ORDER BY nombre').fetchall()
            conn.close()
            for r in rows:
                tree_s.insert('', 'end', values=(r['id'], r['nombre'], r['descripcion'],
                                               f'S/.{r["precio"]:.2f}', r['tipo']))

        load_servicios()

        def delete_servicio():
            sel = tree_s.selection()
            if not sel: return
            vals = tree_s.item(sel[0], 'values')
            if messagebox.askyesno('Confirmar', 'Eliminar servicio?'):
                conn = self.get_db()
                conn.execute('DELETE FROM servicios_grooming WHERE id=?', (int(vals[0]),))
                conn.commit()
                conn.close()
                load_servicios()

        btn_s = tk.Frame(tab1, bg='#f0f4f8')
        btn_s.pack(fill='x')
        tk.Button(btn_s, text='Eliminar', bg='#e74c3c', fg='white',
                  font=('Segoe UI', 9), relief='flat', padx=10, pady=2,
                  command=delete_servicio, cursor='hand2').pack(side='left', padx=2)

        # Tab 2: Historial Grooming
        tab2 = tk.Frame(notebook, bg='#f0f4f8')
        notebook.add(tab2, text='Historial Grooming')

        tabla_frame = tk.Frame(tab2, bg='#f0f4f8')
        tabla_frame.pack(fill='both', expand=True)

        # Top: search by DNI or name
        search_frame = tk.Frame(tabla_frame, bg='#f0f4f8')
        search_frame.pack(fill='x', pady=(0, 5))
        tk.Label(search_frame, text='Buscar por DNI o Dueño:', bg='#f0f4f8',
                 font=('Segoe UI', 9)).pack(side='left', padx=(0, 5))
        e_buscar = tk.Entry(search_frame, width=30, font=('Segoe UI', 10))
        e_buscar.pack(side='left', padx=2)

        cols_h = ('fecha', 'animal', 'dueno', 'dueno_dni', 'servicio', 'precio', 'visitas')
        tree_h = ttk.Treeview(tabla_frame, columns=cols_h, show='headings', height=12)
        tree_h.heading('fecha', text='Fecha')
        tree_h.heading('animal', text='Animal')
        tree_h.heading('dueno', text='Dueño')
        tree_h.heading('dueno_dni', text='DNI')
        tree_h.heading('servicio', text='Servicio')
        tree_h.heading('precio', text='Precio')
        tree_h.heading('visitas', text='Visitas')
        tree_h.column('fecha', width=80)
        tree_h.column('animal', width=130)
        tree_h.column('dueno', width=130)
        tree_h.column('dueno_dni', width=80)
        tree_h.column('servicio', width=130)
        tree_h.column('precio', width=60)
        tree_h.column('visitas', width=50)
        tree_h.pack(fill='both', expand=True, pady=5)

        hist_btn_frame = tk.Frame(tab2, bg='white')
        hist_btn_frame.pack(fill='x', padx=10, pady=3)
        tk.Button(hist_btn_frame, text='Exportar CSV', bg='#3498db', fg='white',
                  font=('Segoe UI', 9), relief='flat', padx=10, pady=2,
                  command=lambda: self._export_to_csv(
                      [[tree_h.item(i, 'values')[j] for j in range(len(tree_h['columns']))] for i in tree_h.get_children()],
                      [tree_h.heading(c)['text'] for c in tree_h['columns']],
                      'historial_grooming.csv'
                  ), cursor='hand2').pack(side='left', padx=2)
        tk.Button(hist_btn_frame, text='Excel', bg='#27ae60', fg='white',
                  font=('Segoe UI', 9), relief='flat', padx=10, pady=2,
                  command=lambda: self._export_to_excel(
                      [[tree_h.item(i, 'values')[j] for j in range(len(tree_h['columns']))] for i in tree_h.get_children()],
                      [tree_h.heading(c)['text'] for c in tree_h['columns']],
                      'historial_grooming.xlsx'
                  ), cursor='hand2').pack(side='left', padx=2)

        def load_historial(filter_text=''):
            for item in tree_h.get_children():
                tree_h.delete(item)
            conn = self.get_db()
            rows = conn.execute(
                'SELECT hg.fecha, a.nombre as animal, d.nombre as dueno, d.dni as dueno_dni, '
                'sg.nombre as servicio, hg.precio, hg.id_animal '
                'FROM historial_grooming hg '
                'JOIN animales a ON hg.id_animal = a.id '
                'JOIN duenos d ON a.id_dueno = d.id '
                'LEFT JOIN servicios_grooming sg ON hg.id_servicio = sg.id '
                'ORDER BY hg.fecha DESC').fetchall()
            visitas_count = conn.execute(
                'SELECT id_animal, COUNT(*) as total FROM historial_grooming GROUP BY id_animal'
            ).fetchall()
            conn.close()
            visitas_map = {r['id_animal']: r['total'] for r in visitas_count}
            for r in rows:
                if filter_text:
                    fl = filter_text.lower()
                    if fl not in r['dueno'].lower() and fl not in (r['dueno_dni'] or '').lower() and fl not in r['animal'].lower():
                        continue
                tree_h.insert('', 'end', values=(
                    r['fecha'], r['animal'], r['dueno'], r['dueno_dni'] or '',
                    r['servicio'] or 'N/A', f'S/.{r["precio"]:.2f}',
                    visitas_map.get(r['id_animal'], 1)))

        load_historial()
        e_buscar.bind('<KeyRelease>', lambda e: load_historial(e_buscar.get()))

        # Tab 3: Citas Grooming
        tab3 = tk.Frame(notebook, bg='#f0f4f8')
        notebook.add(tab3, text='Citas Grooming')

        cols_g = ('fecha', 'animal', 'dueno', 'motivo')
        tree_g = ttk.Treeview(tab3, columns=cols_g, show='headings', height=12)
        tree_g.heading('fecha', text='Fecha')
        tree_g.heading('animal', text='Animal')
        tree_g.heading('dueno', text='Dueño')
        tree_g.heading('motivo', text='Motivo')
        tree_g.column('fecha', width=100)
        tree_g.column('animal', width=180)
        tree_g.column('dueno', width=180)
        tree_g.column('motivo', width=250)
        tree_g.pack(fill='both', expand=True, pady=5)

        def load_citas_grooming():
            for item in tree_g.get_children():
                tree_g.delete(item)
            conn = self.get_db()
            rows = conn.execute(
                'SELECT c.fecha, a.nombre as animal, d.nombre as dueno, c.motivo '
                'FROM citas c JOIN animales a ON c.id_animal = a.id '
                'JOIN duenos d ON c.id_dueno = d.id '
                "WHERE c.tipo = 'grooming' AND c.estado = 'pendiente' ORDER BY c.fecha").fetchall()
            conn.close()
            for r in rows:
                tree_g.insert('', 'end', values=(r['fecha'], r['animal'], r['dueno'], r['motivo']))

        load_citas_grooming()

    def add_grooming_service_dialog(self):
        conn = self.get_db()
        tipos_g = [r['tipo'] for r in conn.execute('SELECT DISTINCT tipo FROM servicios_grooming WHERE tipo!="" ORDER BY tipo').fetchall()]
        conn.close()
        if not tipos_g: tipos_g = ['Baño', 'Corte', 'Combo', 'Higiene', 'Otros']

        dialog = tk.Toplevel(self.root)
        dialog.title('Nuevo Servicio de Grooming')
        dialog.geometry('450x250')
        dialog.resizable(False, False)
        dialog.configure(bg='#e8edf2')
        dialog.transient(self.root)
        dialog.grab_set()
        center_dialog(dialog, self.root)

        tk.Label(dialog, text='Nombre', bg='#e8edf2', font=('Segoe UI', 10)).grid(row=0, column=0, sticky='e', padx=8, pady=3)
        e_nombre = tk.Entry(dialog, width=35, font=('Segoe UI', 10))
        e_nombre.grid(row=0, column=1, padx=8, pady=3)
        tk.Label(dialog, text='Descripcion', bg='#e8edf2', font=('Segoe UI', 10)).grid(row=1, column=0, sticky='e', padx=8, pady=3)
        e_desc = tk.Entry(dialog, width=35, font=('Segoe UI', 10))
        e_desc.grid(row=1, column=1, padx=8, pady=3)
        tk.Label(dialog, text='Precio', bg='#e8edf2', font=('Segoe UI', 10)).grid(row=2, column=0, sticky='e', padx=8, pady=3)
        tk.Label(dialog, text='S/.', bg='#e8edf2', font=('Segoe UI', 10, 'bold'), fg='#27ae60').grid(row=2, column=1, sticky='w')
        e_precio = tk.Entry(dialog, width=30, font=('Segoe UI', 10))
        e_precio.grid(row=2, column=1, padx=(25, 10), pady=6, sticky='w')
        tk.Label(dialog, text='Tipo', bg='#e8edf2', font=('Segoe UI', 10)).grid(row=3, column=0, sticky='e', padx=8, pady=3)
        tipo_frame = tk.Frame(dialog, bg='#e8edf2')
        tipo_frame.grid(row=3, column=1, sticky='w', padx=8, pady=3)
        e_tipo = ttk.Combobox(tipo_frame, values=tipos_g, width=28, font=('Segoe UI', 10))
        e_tipo.set('Baño')
        e_tipo.pack(side='left', padx=2)
        tk.Button(tipo_frame, text='+', font=('Segoe UI', 10, 'bold'), bg='#3498db', fg='white',
                  width=2, relief='flat', cursor='hand2',
                  command=lambda: self._add_categoria_dialog(tipos_g, e_tipo)
                  ).pack(side='left', padx=1)

        def save():
            conn = self.get_db()
            conn.execute(
                'INSERT INTO servicios_grooming (nombre, descripcion, precio, duracion_minutos, tipo) VALUES (?,?,?,?,?)',
                (e_nombre.get(), e_desc.get(), float(e_precio.get() or 0), 0, e_tipo.get()))
            conn.commit()
            conn.close()
            dialog.destroy()
            self.show_servicios_productos()

        tk.Button(dialog, text='Guardar', bg='#2ecc71', fg='white',
                  font=('Segoe UI', 10), relief='flat', padx=20, pady=5,
                  command=save, cursor='hand2').grid(row=4, column=1, pady=15, sticky='e')

    def add_servicio_medico_dialog(self):
        conn = self.get_db()
        tipos_m = [r['tipo'] for r in conn.execute('SELECT DISTINCT tipo FROM servicios_medicos WHERE tipo!="" ORDER BY tipo').fetchall()]
        conn.close()
        if not tipos_m: tipos_m = ['consulta', 'cirugia', 'hospitalizacion', 'medicacion', 'procedimiento', 'laboratorio', 'otros']

        dialog = tk.Toplevel(self.root)
        dialog.title('Nuevo Servicio Medico')
        dialog.geometry('450x250')
        dialog.resizable(False, False)
        dialog.configure(bg='#e8edf2')
        dialog.transient(self.root)
        dialog.grab_set()
        center_dialog(dialog, self.root)

        tk.Label(dialog, text='Nombre', bg='#e8edf2', font=('Segoe UI', 10)).grid(row=0, column=0, sticky='e', padx=8, pady=3)
        e_nombre = tk.Entry(dialog, width=35, font=('Segoe UI', 10))
        e_nombre.grid(row=0, column=1, padx=8, pady=3)
        tk.Label(dialog, text='Descripcion', bg='#e8edf2', font=('Segoe UI', 10)).grid(row=1, column=0, sticky='e', padx=8, pady=3)
        e_desc = tk.Entry(dialog, width=35, font=('Segoe UI', 10))
        e_desc.grid(row=1, column=1, padx=8, pady=3)
        tk.Label(dialog, text='Precio', bg='#e8edf2', font=('Segoe UI', 10)).grid(row=2, column=0, sticky='e', padx=8, pady=3)
        tk.Label(dialog, text='S/.', bg='#e8edf2', font=('Segoe UI', 10, 'bold'), fg='#27ae60').grid(row=2, column=1, sticky='w')
        e_precio = tk.Entry(dialog, width=30, font=('Segoe UI', 10))
        e_precio.grid(row=2, column=1, padx=(25, 10), pady=6, sticky='w')
        tk.Label(dialog, text='Tipo', bg='#e8edf2', font=('Segoe UI', 10)).grid(row=3, column=0, sticky='e', padx=8, pady=3)
        tipo_frame = tk.Frame(dialog, bg='#e8edf2')
        tipo_frame.grid(row=3, column=1, sticky='w', padx=8, pady=3)
        e_tipo = ttk.Combobox(tipo_frame, values=tipos_m, width=28, font=('Segoe UI', 10))
        e_tipo.set('consulta')
        e_tipo.pack(side='left', padx=2)
        tk.Button(tipo_frame, text='+', font=('Segoe UI', 10, 'bold'), bg='#3498db', fg='white',
                  width=2, relief='flat', cursor='hand2',
                  command=lambda: self._add_categoria_dialog(tipos_m, e_tipo)
                  ).pack(side='left', padx=1)

        def save():
            conn = self.get_db()
            conn.execute(
                'INSERT INTO servicios_medicos (nombre, descripcion, precio, tipo) VALUES (?,?,?,?)',
                (e_nombre.get(), e_desc.get(), float(e_precio.get() or 0), e_tipo.get()))
            conn.commit()
            conn.close()
            dialog.destroy()
            self.show_servicios_productos()

        tk.Button(dialog, text='Guardar', bg='#2ecc71', fg='white',
                  font=('Segoe UI', 10), relief='flat', padx=20, pady=5,
                  command=save, cursor='hand2').grid(row=4, column=1, pady=15, sticky='e')

    def edit_grooming_service_dialog(self, tree):
        sel = tree.selection()
        if not sel: return
        vals = tree.item(sel[0], 'values')
        edit_id = int(vals[0])

        conn = self.get_db()
        data = conn.execute('SELECT * FROM servicios_grooming WHERE id=?', (edit_id,)).fetchone()
        tipos_g = [r['tipo'] for r in conn.execute('SELECT DISTINCT tipo FROM servicios_grooming WHERE tipo!="" ORDER BY tipo').fetchall()]
        conn.close()
        if not data: return
        if not tipos_g: tipos_g = ['Baño', 'Corte', 'Combo', 'Higiene', 'Otros']

        dialog = tk.Toplevel(self.root)
        dialog.title('Editar Servicio Grooming')
        dialog.geometry('450x250')
        dialog.resizable(False, False)
        dialog.configure(bg='#e8edf2')
        dialog.transient(self.root)
        dialog.grab_set()
        center_dialog(dialog, self.root)

        tk.Label(dialog, text='Nombre', bg='#e8edf2', font=('Segoe UI', 10)).grid(row=0, column=0, sticky='e', padx=8, pady=3)
        e_nombre = tk.Entry(dialog, width=35, font=('Segoe UI', 10))
        e_nombre.insert(0, data['nombre'])
        e_nombre.grid(row=0, column=1, padx=8, pady=3)
        tk.Label(dialog, text='Descripcion', bg='#e8edf2', font=('Segoe UI', 10)).grid(row=1, column=0, sticky='e', padx=8, pady=3)
        e_desc = tk.Entry(dialog, width=35, font=('Segoe UI', 10))
        e_desc.insert(0, data['descripcion'] or '')
        e_desc.grid(row=1, column=1, padx=8, pady=3)
        tk.Label(dialog, text='Precio', bg='#e8edf2', font=('Segoe UI', 10)).grid(row=2, column=0, sticky='e', padx=8, pady=3)
        tk.Label(dialog, text='S/.', bg='#e8edf2', font=('Segoe UI', 10, 'bold'), fg='#27ae60').grid(row=2, column=1, sticky='w')
        e_precio = tk.Entry(dialog, width=30, font=('Segoe UI', 10))
        e_precio.insert(0, str(data['precio']))
        e_precio.grid(row=2, column=1, padx=(25, 10), pady=6, sticky='w')
        tk.Label(dialog, text='Tipo', bg='#e8edf2', font=('Segoe UI', 10)).grid(row=3, column=0, sticky='e', padx=8, pady=3)
        tipo_frame = tk.Frame(dialog, bg='#e8edf2')
        tipo_frame.grid(row=3, column=1, sticky='w', padx=8, pady=3)
        e_tipo = ttk.Combobox(tipo_frame, values=tipos_g, width=28, font=('Segoe UI', 10))
        e_tipo.set(data['tipo'])
        e_tipo.pack(side='left', padx=2)
        tk.Button(tipo_frame, text='+', font=('Segoe UI', 10, 'bold'), bg='#3498db', fg='white',
                  width=2, relief='flat', cursor='hand2',
                  command=lambda: self._add_categoria_dialog(tipos_g, e_tipo)
                  ).pack(side='left', padx=1)

        def save():
            conn = self.get_db()
            conn.execute(
                'UPDATE servicios_grooming SET nombre=?, descripcion=?, precio=?, tipo=? WHERE id=?',
                (e_nombre.get(), e_desc.get(), float(e_precio.get() or 0), e_tipo.get(), edit_id))
            conn.commit()
            conn.close()
            dialog.destroy()
            self.show_servicios_productos()

        btn_frame = tk.Frame(dialog, bg='#e8edf2')
        btn_frame.grid(row=4, column=0, columnspan=2, pady=10)
        tk.Button(btn_frame, text='Eliminar', bg='#e74c3c', fg='white',
                  font=('Segoe UI', 10), relief='flat', padx=15, pady=5,
                  command=lambda: self._delete_grooming_service(edit_id, dialog), cursor='hand2').pack(side='left', padx=5)
        tk.Button(btn_frame, text='Guardar', bg='#2ecc71', fg='white',
                  font=('Segoe UI', 10), relief='flat', padx=20, pady=5,
                  command=save, cursor='hand2').pack(side='left', padx=5)

    def _delete_grooming_service(self, sid, dialog):
        if messagebox.askyesno('Confirmar', 'Eliminar servicio de grooming?'):
            conn = self.get_db()
            conn.execute('DELETE FROM servicios_grooming WHERE id=?', (sid,))
            conn.commit()
            conn.close()
            dialog.destroy()
            self.show_servicios_productos()

    def edit_servicio_medico_dialog(self, tree):
        sel = tree.selection()
        if not sel: return
        vals = tree.item(sel[0], 'values')
        edit_id = int(vals[0])

        conn = self.get_db()
        data = conn.execute('SELECT * FROM servicios_medicos WHERE id=?', (edit_id,)).fetchone()
        tipos_m = [r['tipo'] for r in conn.execute('SELECT DISTINCT tipo FROM servicios_medicos WHERE tipo!="" ORDER BY tipo').fetchall()]
        conn.close()
        if not data: return
        if not tipos_m: tipos_m = ['consulta', 'cirugia', 'hospitalizacion', 'medicacion', 'procedimiento', 'laboratorio', 'otros']

        dialog = tk.Toplevel(self.root)
        dialog.title('Editar Servicio Medico')
        dialog.geometry('450x250')
        dialog.resizable(False, False)
        dialog.configure(bg='#e8edf2')
        dialog.transient(self.root)
        dialog.grab_set()
        center_dialog(dialog, self.root)

        tk.Label(dialog, text='Nombre', bg='#e8edf2', font=('Segoe UI', 10)).grid(row=0, column=0, sticky='e', padx=8, pady=3)
        e_nombre = tk.Entry(dialog, width=35, font=('Segoe UI', 10))
        e_nombre.insert(0, data['nombre'])
        e_nombre.grid(row=0, column=1, padx=8, pady=3)
        tk.Label(dialog, text='Descripcion', bg='#e8edf2', font=('Segoe UI', 10)).grid(row=1, column=0, sticky='e', padx=8, pady=3)
        e_desc = tk.Entry(dialog, width=35, font=('Segoe UI', 10))
        e_desc.insert(0, data['descripcion'] or '')
        e_desc.grid(row=1, column=1, padx=8, pady=3)
        tk.Label(dialog, text='Precio', bg='#e8edf2', font=('Segoe UI', 10)).grid(row=2, column=0, sticky='e', padx=8, pady=3)
        tk.Label(dialog, text='S/.', bg='#e8edf2', font=('Segoe UI', 10, 'bold'), fg='#27ae60').grid(row=2, column=1, sticky='w')
        e_precio = tk.Entry(dialog, width=30, font=('Segoe UI', 10))
        e_precio.insert(0, str(data['precio']))
        e_precio.grid(row=2, column=1, padx=(25, 10), pady=6, sticky='w')
        tk.Label(dialog, text='Tipo', bg='#e8edf2', font=('Segoe UI', 10)).grid(row=3, column=0, sticky='e', padx=8, pady=3)
        tipo_frame = tk.Frame(dialog, bg='#e8edf2')
        tipo_frame.grid(row=3, column=1, sticky='w', padx=8, pady=3)
        e_tipo = ttk.Combobox(tipo_frame, values=tipos_m, width=28, font=('Segoe UI', 10))
        e_tipo.set(data['tipo'])
        e_tipo.pack(side='left', padx=2)
        tk.Button(tipo_frame, text='+', font=('Segoe UI', 10, 'bold'), bg='#3498db', fg='white',
                  width=2, relief='flat', cursor='hand2',
                  command=lambda: self._add_categoria_dialog(tipos_m, e_tipo)
                  ).pack(side='left', padx=1)

        def save():
            conn = self.get_db()
            conn.execute(
                'UPDATE servicios_medicos SET nombre=?, descripcion=?, precio=?, tipo=? WHERE id=?',
                (e_nombre.get(), e_desc.get(), float(e_precio.get() or 0), e_tipo.get(), edit_id))
            conn.commit()
            conn.close()
            dialog.destroy()
            self.show_servicios_productos()

        btn_frame = tk.Frame(dialog, bg='#e8edf2')
        btn_frame.grid(row=4, column=0, columnspan=2, pady=10)
        tk.Button(btn_frame, text='Eliminar', bg='#e74c3c', fg='white',
                  font=('Segoe UI', 10), relief='flat', padx=15, pady=5,
                  command=lambda: self._delete_medico_service(edit_id, dialog), cursor='hand2').pack(side='left', padx=5)
        tk.Button(btn_frame, text='Guardar', bg='#2ecc71', fg='white',
                  font=('Segoe UI', 10), relief='flat', padx=20, pady=5,
                  command=save, cursor='hand2').pack(side='left', padx=5)

    def _delete_medico_service(self, sid, dialog):
        if messagebox.askyesno('Confirmar', 'Eliminar servicio medico?'):
            conn = self.get_db()
            conn.execute('DELETE FROM servicios_medicos WHERE id=?', (sid,))
            conn.commit()
            conn.close()
            dialog.destroy()
            self.show_servicios_productos()

    def add_grooming_from_detail(self, animal_id, detail_frame, reload_callback):
        conn = self.get_db()
        servicios = conn.execute('SELECT * FROM servicios_grooming ORDER BY nombre').fetchall()
        animal = conn.execute('SELECT a.id, a.nombre, d.nombre as dueno, d.dni FROM animales a JOIN duenos d ON a.id_dueno=d.id WHERE a.id=?', (animal_id,)).fetchone()
        visitas = conn.execute('SELECT COUNT(*) FROM historial_grooming WHERE id_animal=?', (animal_id,)).fetchone()[0]
        conn.close()
        if not servicios:
            messagebox.showinfo('Info', 'No hay servicios de grooming. Cree uno primero.')
            return

        dialog = tk.Toplevel(self.root)
        dialog.title(f'Registrar Grooming - {animal["nombre"]}')
        dialog.geometry('420x300')
        dialog.resizable(False, False)
        dialog.configure(bg='#e8edf2')
        dialog.transient(self.root)
        dialog.grab_set()
        center_dialog(dialog, self.root)

        info_f = tk.Frame(dialog, bg='#e8edf2')
        info_f.grid(row=0, column=0, columnspan=2, padx=10, pady=6, sticky='w')
        tk.Label(info_f, text=f'Dueño: {animal["dueno"]} | DNI: {animal["dni"] or "N/A"} | Visitas: {visitas}',
                 font=('Segoe UI', 9, 'bold'), bg='#e8edf2', fg='#9b59b6').pack()

        svc_map = {s['nombre']: s for s in servicios}
        svc_names = list(svc_map.keys()) + ['Otros']
        tk.Label(dialog, text='Servicio:', bg='#e8edf2', font=('Segoe UI', 10)).grid(row=1, column=0, sticky='e', padx=8, pady=3)
        e_servicio = ttk.Combobox(dialog, values=svc_names, width=35, font=('Segoe UI', 10))
        e_servicio.set(list(svc_map.keys())[0])
        e_servicio.grid(row=1, column=1, padx=8, pady=3)
        tk.Label(dialog, text='Precio:', bg='#e8edf2', font=('Segoe UI', 10)).grid(row=2, column=0, sticky='e', padx=8, pady=3)
        e_precio = tk.Entry(dialog, width=35, font=('Segoe UI', 10))
        e_precio.grid(row=2, column=1, padx=8, pady=3)
        tk.Label(dialog, text='Estilista:', bg='#e8edf2', font=('Segoe UI', 10)).grid(row=3, column=0, sticky='e', padx=8, pady=3)
        e_estilista = tk.Entry(dialog, width=35, font=('Segoe UI', 10))
        e_estilista.grid(row=3, column=1, padx=8, pady=3)
        tk.Label(dialog, text='Observaciones:', bg='#e8edf2', font=('Segoe UI', 10)).grid(row=4, column=0, sticky='e', padx=8, pady=3)
        e_obs = tk.Entry(dialog, width=35, font=('Segoe UI', 10))
        e_obs.grid(row=4, column=1, padx=8, pady=3)

        def on_servicio_select(*args):
            s = svc_map.get(e_servicio.get())
            if s:
                e_precio.delete(0, 'end')
                e_precio.insert(0, str(s['precio']))
            else:
                e_precio.delete(0, 'end')
        e_servicio.bind('<<ComboboxSelected>>', on_servicio_select)
        on_servicio_select()

        def save():
            conn = self.get_db()
            svc_name = e_servicio.get()
            s = svc_map.get(svc_name)
            svc_id = s['id'] if s else None
            conn.execute(
                'INSERT INTO historial_grooming (id_animal, id_servicio, fecha, observaciones, precio, estilista) VALUES (?,?,?,?,?,?)',
                (animal_id, svc_id, str(date.today()), e_obs.get(), float(e_precio.get() or 0), e_estilista.get()))
            conn.commit()
            conn.close()
            dialog.destroy()
            self.show_detail(detail_frame, animal_id, reload_callback)
            messagebox.showinfo('Exito', 'Grooming registrado')

        tk.Button(dialog, text='Guardar', bg='#2ecc71', fg='white',
                  font=('Segoe UI', 10), relief='flat', padx=20, pady=5,
                  command=save, cursor='hand2').grid(row=5, column=1, pady=15, sticky='e')

    def _seccion_historial_grooming(self, parent, animal_id):
        conn = self.get_db()
        rows = conn.execute(
            'SELECT hg.fecha, sg.nombre as servicio, hg.precio, hg.estilista, hg.observaciones '
            'FROM historial_grooming hg '
            'LEFT JOIN servicios_grooming sg ON hg.id_servicio = sg.id '
            'WHERE hg.id_animal=? ORDER BY hg.fecha DESC', (animal_id,)).fetchall()
        total_visitas = conn.execute(
            'SELECT COUNT(*) as total FROM historial_grooming WHERE id_animal=?', (animal_id,)).fetchone()[0]
        conn.close()
        sec, lbl = self._make_section(parent, 'Historial Grooming', '#9b59b6')
        lbl_visitas = tk.Label(sec, text=f'({total_visitas} visita(s))', font=('Segoe UI', 9),
                               bg='#e8edf2', fg='#9b59b6')
        lbl_visitas.pack(side='left', padx=(8, 0))
        if not rows:
            tk.Label(parent, text='  Sin historial de grooming', font=('Segoe UI', 9),
                     bg='#e8edf2', fg='#999', anchor='w').pack(fill='x')
            return
        f = self._make_table_frame(parent)
        cols = ('gf', 'gs', 'gp', 'ge', 'go')
        tree = ttk.Treeview(f, columns=cols, show='headings', height=4)
        tree.heading('gf', text='Fecha')
        tree.heading('gs', text='Servicio')
        tree.heading('gp', text='Precio')
        tree.heading('ge', text='Estilista')
        tree.heading('go', text='Observaciones')
        tree.column('gf', width=90)
        tree.column('gs', width=150)
        tree.column('gp', width=70)
        tree.column('ge', width=100)
        tree.column('go', width=180)
        tree.pack(fill='x')
        for r in rows:
            tree.insert("", "end", values=(r["fecha"], r["servicio"] or "N/A",
                        f"S/.{r['precio']:.2f}", r["estilista"] or "", r["observaciones"] or ""))

    # ---------- DOCUMENTOS ADJUNTOS ----------
    def _seccion_documentos(self, parent, animal_id, detail_frame, reload_callback):
        conn = self.get_db()
        docs = conn.execute(
            "SELECT * FROM documentos_adjuntos WHERE id_animal=? ORDER BY fecha DESC",
            (animal_id,)).fetchall()
        conn.close()

        sec, lbl = self._make_section(parent, "Documentos Adjuntos", "#8e44ad")
        tk.Button(sec, text="+ Subir Documento", bg="#8e44ad", fg="white",
                  font=("Segoe UI", 9), relief="flat", padx=10, pady=2, cursor="hand2",
                  command=lambda: self._add_documento(animal_id, detail_frame, reload_callback)
                  ).pack(side="right")

        if not docs:
            tk.Label(parent, text="  Sin documentos adjuntos", font=("Segoe UI", 9),
                     bg="#e8edf2", fg="#999", anchor="w").pack(fill="x")
            return

        f = self._make_table_frame(parent)
        cols = ("df", "dn", "dt", "do")
        tree = ttk.Treeview(f, columns=cols, show="headings", height=4)
        tree.heading("df", text="Fecha")
        tree.heading("dn", text="Nombre")
        tree.heading("dt", text="Tipo")
        tree.heading("do", text="Observaciones")
        tree.column("df", width=90)
        tree.column("dn", width=200)
        tree.column("dt", width=100)
        tree.column("do", width=250)
        tree.pack(fill="x")

        tipos = {"receta": "📋 Receta", "resultado": "🧪 Resultado",
                 "radiografia": "📷 Radiografía", "otro": "📄 Otro"}
        for d in docs:
            tipo_str = tipos.get(d["tipo"], d["tipo"])
            tree.insert("", "end", iid=str(d["id"]), values=(
                d["fecha"], d["nombre"], tipo_str, d["observaciones"] or ""))

        def open_doc():
            sel = tree.selection()
            if not sel: return
            d = next((x for x in docs if str(x["id"]) == sel[0]), None)
            if d and d["archivo"]:
                ruta = os.path.join(DOCS_DIR, d["archivo"])
                if os.path.exists(ruta):
                    os.startfile(ruta)
        def delete_doc():
            sel = tree.selection()
            if not sel: return
            d = next((x for x in docs if str(x["id"]) == sel[0]), None)
            if not d or not messagebox.askyesno("Confirmar", "Eliminar documento?"): return
            conn = self.get_db()
            conn.execute("DELETE FROM documentos_adjuntos WHERE id=?", (d["id"],))
            conn.commit()
            conn.close()
            if d["archivo"]:
                ruta = os.path.join(DOCS_DIR, d["archivo"])
                if os.path.exists(ruta): os.remove(ruta)
            self.show_detail(detail_frame, animal_id, reload_callback)

        tree.bind("<Double-1>", lambda e: open_doc())
        br = tk.Frame(parent, bg="#e8edf2")
        br.pack(fill="x", pady=(4, 0))
        tk.Button(br, text="📂 Abrir", bg="#3498db", fg="white",
                  font=("Segoe UI", 9), relief="flat", padx=10, pady=2, cursor="hand2",
                  command=open_doc).pack(side="left", padx=2)
        tk.Button(br, text="🗑 Eliminar", bg="#e74c3c", fg="white",
                  font=("Segoe UI", 9), relief="flat", padx=10, pady=2, cursor="hand2",
                  command=delete_doc).pack(side="left", padx=2)

    def _add_documento(self, animal_id, detail_frame, reload_callback):
        from tkinter import filedialog
        path = filedialog.askopenfilename(
            title="Seleccionar archivo",
            filetypes=[("Todos los archivos", "*.*"),
                       ("PDF", "*.pdf"), ("Imagen", "*.png;*.jpg;*.jpeg"),
                       ("Word", "*.doc;*.docx"), ("Excel", "*.xls;*.xlsx")])
        if not path: return

        import shutil
        from datetime import datetime
        ext = os.path.splitext(path)[1]
        fname = f"{animal_id}_{datetime.now().strftime('%Y%m%d%H%M%S')}{ext}"
        dest = os.path.join(DOCS_DIR, fname)
        shutil.copy2(path, dest)

        dia = tk.Toplevel(self.root)
        dia.title("Detalles del Documento")
        dia.geometry("450x300")
        dia.configure(bg="#f0f4f8")
        dia.transient(self.root)
        dia.grab_set()
        center_dialog(dia, self.root)

        tk.Label(dia, text="Nombre:", bg="#f0f4f8", font=("Segoe UI", 10)).grid(row=0, column=0, sticky="e", padx=8, pady=6)
        e_nombre = tk.Entry(dia, width=40, font=("Segoe UI", 10))
        e_nombre.insert(0, os.path.splitext(os.path.basename(path))[0])
        e_nombre.grid(row=0, column=1, padx=8, pady=6)

        tk.Label(dia, text="Tipo:", bg="#f0f4f8", font=("Segoe UI", 10)).grid(row=1, column=0, sticky="e", padx=8, pady=6)
        cb_tipo = ttk.Combobox(dia, values=["receta", "resultado", "radiografia", "otro"], state="readonly", width=37)
        cb_tipo.set("otro")
        cb_tipo.grid(row=1, column=1, padx=8, pady=6)

        tk.Label(dia, text="Observaciones:", bg="#f0f4f8", font=("Segoe UI", 10)).grid(row=2, column=0, sticky="ne", padx=8, pady=6)
        t_obs = tk.Text(dia, width=38, height=4, font=("Segoe UI", 10))
        t_obs.grid(row=2, column=1, padx=8, pady=6)

        def save():
            conn = self.get_db()
            conn.execute(
                "INSERT INTO documentos_adjuntos (id_animal, nombre, tipo, fecha, archivo, observaciones) VALUES (?,?,?,?,?,?)",
                (animal_id, e_nombre.get(), cb_tipo.get(), str(date.today()), fname, t_obs.get("1.0", "end-1c")))
            conn.commit()
            conn.close()
            dia.destroy()
            self.show_detail(detail_frame, animal_id, reload_callback)
            messagebox.showinfo("Exito", "Documento guardado")
        add_keyboard_shortcuts(dia, save)
        tk.Button(dia, text="Guardar", bg="#2ecc71", fg="white",
                  font=("Segoe UI", 10), relief="flat", padx=20, pady=5,
                  command=save, cursor="hand2").grid(row=3, column=1, pady=15, sticky="e")

    # ---------- LINEA DE TIEMPO ----------
    def _linea_tiempo(self, parent, animal_id):
        conn = self.get_db()
        eventos = []
        for r in conn.execute("SELECT fecha, diagnostico FROM registros_medicos WHERE id_animal=? ORDER BY fecha", (animal_id,)).fetchall():
            desc = "Consulta" + (": " + r["diagnostico"][:60] if r["diagnostico"] else "")
            eventos.append(("📋", r["fecha"], desc))
        for v in conn.execute("SELECT fecha, tipo, nombre FROM vacunas WHERE id_animal=? ORDER BY fecha", (animal_id,)).fetchall():
            eventos.append(("💉", v["fecha"], f"{v['tipo'].capitalize()}: {v['nombre']}"))
        for g in conn.execute("SELECT hg.fecha, COALESCE(sg.nombre,'Grooming') as desc FROM historial_grooming hg LEFT JOIN servicios_grooming sg ON hg.id_servicio=sg.id WHERE hg.id_animal=? ORDER BY hg.fecha", (animal_id,)).fetchall():
            eventos.append(("✂", g["fecha"], "Grooming: " + g["desc"]))
        for c in conn.execute("SELECT fecha, motivo FROM citas WHERE id_animal=? ORDER BY fecha", (animal_id,)).fetchall():
            eventos.append(("📅", c["fecha"], "Cita: " + (c["motivo"] or "")))
        conn.close()

        eventos.sort(key=lambda x: x[1])

        sec, lbl = self._make_section(parent, "Linea de Tiempo", "#2c3e50")
        if not eventos:
            tk.Label(parent, text="  Sin eventos registrados", font=("Segoe UI", 9),
                     bg="#e8edf2", fg="#999", anchor="w").pack(fill="x")
            return

        canvas_tl = tk.Canvas(parent, bg="#e8edf2", highlightthickness=0, height=min(300, len(eventos)*40))
        canvas_tl.pack(fill="x", pady=(0, 8))
        canvas_tl.bind("<MouseWheel>", lambda e: canvas_tl.yview_scroll(int(-1*(e.delta/120)), "units"))

        inner_tl = tk.Frame(canvas_tl, bg="#e8edf2")
        canvas_tl.create_window((0, 0), window=inner_tl, anchor="nw", tags="inner")

        fechas_unicas = sorted(set(e[1] for e in eventos))
        color_idx = 0
        colores = ["#3498db", "#2ecc71", "#e67e22", "#9b59b6", "#e74c3c", "#1abc9c"]

        prev_date = None
        for ev in eventos:
            icono, fecha, desc = ev
            fr = tk.Frame(inner_tl, bg="#e8edf2")
            fr.pack(fill="x", pady=2)

            if fecha != prev_date:
                tk.Label(fr, text=fecha, font=("Segoe UI", 8, "bold"),
                         bg="#e8edf2", fg="#7f8c8d", width=12, anchor="e").pack(side="left")
                prev_date = fecha
            else:
                tk.Label(fr, text="", width=12, bg="#e8edf2").pack(side="left")

            color = colores[color_idx % len(colores)]
            color_idx += 1
            dot = tk.Label(fr, text="●", fg=color, bg="#e8edf2", font=("Segoe UI", 10))
            dot.pack(side="left", padx=4)

            tk.Label(fr, text=f"{icono} {desc}", font=("Segoe UI", 9),
                     bg="#e8edf2", fg="#333", anchor="w").pack(side="left", padx=4)

        inner_tl.update_idletasks()
        canvas_tl.configure(scrollregion=canvas_tl.bbox("all"), height=min(300, inner_tl.winfo_reqheight()))

    # ---------- PRODUCTOS ----------
    def show_productos(self):
        self._current_view = 'productos'
        self.clear_frame()
        main = tk.Frame(self.content_area, bg='#f0f4f8')
        main.pack(fill='both', expand=True, padx=20, pady=15)

        top = tk.Frame(main, bg='#f0f4f8')
        top.pack(fill='x')
        tk.Label(top, text='Productos / Inventario', font=('Segoe UI', 16, 'bold'),
                 bg='#f0f4f8', fg='#2c3e50').pack(side='left')
        tk.Button(top, text='+ Nuevo Producto', bg='#2ecc71', fg='white',
                  font=('Segoe UI', 9), relief='flat', padx=10, pady=3,
                  command=lambda: self.add_producto_dialog(self.show_productos), cursor='hand2').pack(side='right')

        cols = ('id', 'nombre', 'desc', 'precio_compra', 'precio_venta', 'stock', 'categoria')
        tree = ttk.Treeview(main, columns=cols, show='headings', height=18)
        tree.heading('id', text='ID')
        tree.heading('nombre', text='Nombre')
        tree.heading('desc', text='Descripcion')
        tree.heading('precio_compra', text='Costo')
        tree.heading('precio_venta', text='Venta')
        tree.heading('stock', text='Stock')
        tree.heading('categoria', text='Categoria')
        tree.column('id', width=40)
        tree.column('nombre', width=200)
        tree.column('desc', width=250)
        tree.column('precio_compra', width=70)
        tree.column('precio_venta', width=70)
        tree.column('stock', width=60)
        tree.column('categoria', width=100)
        tree.pack(fill='both', expand=True, pady=10)

        def load_productos():
            for item in tree.get_children():
                tree.delete(item)
            conn = self.get_db()
            rows = conn.execute('SELECT * FROM productos WHERE activo=1 ORDER BY nombre').fetchall()
            conn.close()
            for r in rows:
                item = tree.insert('', 'end', values=(
                    r['id'], r['nombre'], r['descripcion'],
                    f'S/.{r["precio_compra"]:.2f}', f'S/.{r["precio_venta"]:.2f}',
                    r['stock'], r['categoria']))
                stock = r['stock']
                if stock <= 2:
                    tree.tag_configure('critical', background='#fadbd8')
                    tree.item(item, tags=('critical',))
                elif stock <= 5:
                    tree.tag_configure('low', background='#fef9e7')
                    tree.item(item, tags=('low',))

        load_productos()

        btn_frame = tk.Frame(main, bg='#f0f4f8')
        btn_frame.pack(fill='x')
        tk.Button(btn_frame, text='Editar', bg='#f39c12', fg='white',
                  font=('Segoe UI', 9), relief='flat', padx=10, pady=2,
                  command=lambda: self.add_producto_dialog(self.show_productos, tree), cursor='hand2').pack(side='left', padx=2)
        tk.Button(btn_frame, text='Eliminar', bg='#e74c3c', fg='white',
                  font=('Segoe UI', 9), relief='flat', padx=10, pady=2,
                  command=lambda: self._delete_producto(tree, load_productos), cursor='hand2').pack(side='left', padx=2)

        exp_btn_frame = tk.Frame(main, bg='#f0f4f8')
        exp_btn_frame.pack(fill='x', padx=8, pady=3)
        tk.Button(exp_btn_frame, text='Exportar CSV', bg='#3498db', fg='white',
                  font=('Segoe UI', 9), relief='flat', padx=10, pady=2,
                  command=lambda: self._export_to_csv(
                      [[tree.item(i, 'values')[0], tree.item(i, 'values')[1], tree.item(i, 'values')[2], tree.item(i, 'values')[3], tree.item(i, 'values')[4]] for i in tree.get_children()],
                      ['ID', 'Nombre', 'Precio', 'Stock', 'Activo'],
                      'productos.csv'
                  ), cursor='hand2').pack(side='left', padx=2)
        tk.Button(exp_btn_frame, text='Exportar Excel', bg='#27ae60', fg='white',
                  font=('Segoe UI', 9), relief='flat', padx=10, pady=2,
                  command=lambda: self._export_to_excel(
                      [[tree.item(i, 'values')[0], tree.item(i, 'values')[1], tree.item(i, 'values')[2], tree.item(i, 'values')[3], tree.item(i, 'values')[4]] for i in tree.get_children()],
                      ['ID', 'Nombre', 'Precio', 'Stock', 'Activo'],
                      'productos.xlsx'
                  ), cursor='hand2').pack(side='left', padx=2)

    def _delete_producto(self, tree, reload_cb):
        sel = tree.selection()
        if not sel: return
        vals = tree.item(sel[0], 'values')
        if messagebox.askyesno('Confirmar', 'Desactivar producto?'):
            conn = self.get_db()
            conn.execute('UPDATE productos SET activo=0 WHERE id=?', (int(vals[0]),))
            conn.commit()
            conn.close()
            reload_cb()

    def _add_categoria_dialog(self, existing, combo):
        dialog = tk.Toplevel(self.root)
        dialog.title('Nueva Categoria')
        dialog.geometry('300x120')
        dialog.resizable(False, False)
        dialog.configure(bg='#e8edf2')
        dialog.transient(self.root)
        dialog.grab_set()
        center_dialog(dialog, self.root)

        tk.Label(dialog, text='Nombre de categoria:', bg='#e8edf2', font=('Segoe UI', 10)).pack(pady=(10, 5))
        e_nueva = tk.Entry(dialog, width=30, font=('Segoe UI', 10))
        e_nueva.pack(pady=5)
        e_nueva.focus_set()

        def add():
            val = e_nueva.get().strip()
            if val and val not in existing:
                existing.append(val)
                combo['values'] = existing
            combo.set(val)
            dialog.destroy()

        tk.Button(dialog, text='Agregar', bg='#2ecc71', fg='white',
                  font=('Segoe UI', 10), relief='flat', padx=15,
                  command=add, cursor='hand2').pack(pady=5)

    def add_producto_dialog(self, callback=None, tree=None):
        edit_id = None
        if tree:
            sel = tree.selection()
            if sel:
                vals = tree.item(sel[0], 'values')
                edit_id = int(vals[0])

        dialog = tk.Toplevel(self.root)
        title = 'Editar Producto' if edit_id else 'Nuevo Producto'
        dialog.title(title)
        dialog.geometry('450x400')
        dialog.resizable(False, False)
        dialog.configure(bg='#e8edf2')
        dialog.transient(self.root)
        dialog.grab_set()
        center_dialog(dialog, self.root)

        data = {}
        if edit_id:
            conn = self.get_db()
            d = conn.execute('SELECT * FROM productos WHERE id=?', (edit_id,)).fetchone()
            conn.close()
            if d:
                for k in d.keys(): data[k] = d[k]

        conn = self.get_db()
        existing_categorias = [r['categoria'] for r in conn.execute('SELECT DISTINCT categoria FROM productos WHERE categoria!="" ORDER BY categoria').fetchall()]
        conn.close()

        row = 0
        tk.Label(dialog, text='Nombre', bg='#e8edf2', font=('Segoe UI', 10)).grid(row=row, column=0, sticky='e', padx=8, pady=3)
        e_nombre = tk.Entry(dialog, width=35, font=('Segoe UI', 10))
        if 'nombre' in data: e_nombre.insert(0, str(data['nombre']))
        e_nombre.grid(row=row, column=1, padx=10, pady=5, columnspan=2, sticky='w')
        row += 1

        tk.Label(dialog, text='Descripcion', bg='#e8edf2', font=('Segoe UI', 10)).grid(row=row, column=0, sticky='e', padx=8, pady=3)
        e_desc = tk.Entry(dialog, width=35, font=('Segoe UI', 10))
        if 'descripcion' in data: e_desc.insert(0, str(data['descripcion'] or ''))
        e_desc.grid(row=row, column=1, padx=10, pady=5, columnspan=2, sticky='w')
        row += 1

        tk.Label(dialog, text='Precio Compra', bg='#e8edf2', font=('Segoe UI', 10)).grid(row=row, column=0, sticky='e', padx=8, pady=3)
        tk.Label(dialog, text='S/.', bg='#e8edf2', font=('Segoe UI', 10, 'bold'), fg='#27ae60').grid(row=row, column=1, sticky='w')
        e_precio_compra = tk.Entry(dialog, width=30, font=('Segoe UI', 10))
        if 'precio_compra' in data: e_precio_compra.insert(0, str(data['precio_compra']))
        e_precio_compra.grid(row=row, column=1, padx=(25, 10), pady=5, sticky='w')
        row += 1

        tk.Label(dialog, text='Precio Venta', bg='#e8edf2', font=('Segoe UI', 10)).grid(row=row, column=0, sticky='e', padx=8, pady=3)
        tk.Label(dialog, text='S/.', bg='#e8edf2', font=('Segoe UI', 10, 'bold'), fg='#27ae60').grid(row=row, column=1, sticky='w')
        e_precio_venta = tk.Entry(dialog, width=30, font=('Segoe UI', 10))
        if 'precio_venta' in data: e_precio_venta.insert(0, str(data['precio_venta']))
        e_precio_venta.grid(row=row, column=1, padx=(25, 10), pady=5, sticky='w')
        row += 1

        tk.Label(dialog, text='Stock', bg='#e8edf2', font=('Segoe UI', 10)).grid(row=row, column=0, sticky='e', padx=8, pady=3)
        stock_frame = tk.Frame(dialog, bg='#e8edf2')
        stock_frame.grid(row=row, column=1, sticky='w', padx=8, pady=3)
        e_stock = tk.Spinbox(stock_frame, from_=0, to=99999, width=8, font=('Segoe UI', 12), justify='center')
        if 'stock' in data: e_stock.delete(0, 'end'); e_stock.insert(0, str(data['stock']))
        else: e_stock.delete(0, 'end'); e_stock.insert(0, '0')
        e_stock.pack(side='left', padx=2)
        row += 1

        tk.Label(dialog, text='Fecha Vencimiento', bg='#e8edf2', font=('Segoe UI', 10)).grid(row=row, column=0, sticky='e', padx=8, pady=3)
        from tkcalendar import DateEntry
        e_fecha_venc = DateEntry(dialog, width=32, font=('Segoe UI', 10),
                                 background='#2c3e50', foreground='white',
                                 borderwidth=2, date_pattern='yyyy-mm-dd')
        if 'fecha_vencimiento' in data and data['fecha_vencimiento']:
            from datetime import datetime
            try:
                e_fecha_venc.set_date(datetime.strptime(data['fecha_vencimiento'], '%Y-%m-%d').date())
            except:
                pass
        e_fecha_venc.grid(row=row, column=1, padx=10, pady=5, sticky='w')
        row += 1

        tk.Label(dialog, text='Categoria', bg='#e8edf2', font=('Segoe UI', 10)).grid(row=row, column=0, sticky='e', padx=8, pady=3)
        cat_frame = tk.Frame(dialog, bg='#e8edf2')
        cat_frame.grid(row=row, column=1, sticky='w', padx=8, pady=3)
        e_categoria = ttk.Combobox(cat_frame, values=existing_categorias, width=28, font=('Segoe UI', 10))
        if 'categoria' in data: e_categoria.set(data['categoria'])
        e_categoria.pack(side='left', padx=2)
        tk.Button(cat_frame, text='+', font=('Segoe UI', 10, 'bold'), bg='#3498db', fg='white',
                  width=2, relief='flat', cursor='hand2',
                  command=lambda: self._add_categoria_dialog(existing_categorias, e_categoria)
                  ).pack(side='left', padx=1)
        row += 1

        def save():
            conn = self.get_db()
            if edit_id:
                conn.execute(
                    'UPDATE productos SET nombre=?, descripcion=?, precio_compra=?, precio_venta=?, stock=?, categoria=?, fecha_vencimiento=? WHERE id=?',
                    (e_nombre.get(), e_desc.get(),
                     float(e_precio_compra.get() or 0), float(e_precio_venta.get() or 0),
                     int(e_stock.get() or 0), e_categoria.get(), e_fecha_venc.get(), edit_id))
            else:
                conn.execute(
                    'INSERT INTO productos (nombre, descripcion, precio_compra, precio_venta, stock, categoria, fecha_vencimiento) VALUES (?,?,?,?,?,?,?)',
                    (e_nombre.get(), e_desc.get(),
                     float(e_precio_compra.get() or 0), float(e_precio_venta.get() or 0),
                     int(e_stock.get() or 0), e_categoria.get(), e_fecha_venc.get()))
            conn.commit()
            conn.close()
            dialog.destroy()
            if callback: callback()

        btn_frame = tk.Frame(dialog, bg='#e8edf2')
        btn_frame.grid(row=row, column=0, columnspan=2, pady=10)
        if edit_id:
            tk.Button(btn_frame, text='Eliminar', bg='#e74c3c', fg='white',
                      font=('Segoe UI', 10), relief='flat', padx=15, pady=5,
                      command=lambda: self._delete_producto_from_dialog(edit_id, dialog, callback), cursor='hand2').pack(side='left', padx=5)
        tk.Button(btn_frame, text='Guardar', bg='#2ecc71', fg='white',
                  font=('Segoe UI', 10), relief='flat', padx=20, pady=5,
                  command=save, cursor='hand2').pack(side='left', padx=5)

    def _delete_producto_from_dialog(self, pid, dialog, callback):
        if messagebox.askyesno('Confirmar', 'Desactivar producto?'):
            conn = self.get_db()
            conn.execute('UPDATE productos SET activo=0 WHERE id=?', (pid,))
            conn.commit()
            conn.close()
            dialog.destroy()
            if callback: callback()

    def aperturar_caja_dialog(self, callback=None):
        if self._caja_abierta:
            messagebox.showinfo("Info", f"Caja ya esta abierta (Saldo inicial: S/.{self._saldo_inicial:.2f})")
            if callback: callback()
            return
        dialog = tk.Toplevel(self.root)
        dialog.title("Aperturar Caja")
        dialog.geometry("350x150")
        dialog.resizable(False, False)
        dialog.configure(bg="#e8edf2")
        dialog.transient(self.root)
        dialog.grab_set()
        dialog.update_idletasks()
        x = self.root.winfo_x() + (self.root.winfo_width() - dialog.winfo_width()) // 2
        y = self.root.winfo_y() + (self.root.winfo_height() - dialog.winfo_height()) // 2
        dialog.geometry("+%d+%d" % (x, y))

        tk.Label(dialog, text="Aperturar Caja del Dia", font=("Segoe UI", 12, "bold"),
                 bg="#e8edf2", fg="#2c3e50").pack(pady=(10, 5))
        tk.Label(dialog, text="Monto inicial (S/):", bg="#e8edf2", font=("Segoe UI", 10)).pack()
        e_monto = tk.Entry(dialog, width=25, font=("Segoe UI", 12), justify="center")
        e_monto.pack(pady=5)
        e_monto.insert(0, "0.00")
        e_monto.focus_set()

        def abrir():
            try:
                monto = float(e_monto.get())
            except:
                messagebox.showerror("Error", "Ingrese un monto valido")
                return
            self._caja_abierta = True
            self._saldo_inicial = monto
            conn = self.get_db()
            conn.execute("INSERT INTO caja (fecha, tipo, concepto, monto) VALUES (?, 'apertura', 'Apertura de caja', ?)",
                         (str(date.today()), monto))
            conn.commit()
            conn.close()
            dialog.destroy()
            messagebox.showinfo("Exito", f"Caja aperturada con S/.{monto:.2f}")
            if callback: callback()

        tk.Button(dialog, text="Abrir Caja", bg="#2ecc71", fg="white",
                  font=("Segoe UI", 10), relief="flat", padx=20, pady=5,
                  command=abrir, cursor="hand2").pack(pady=10)

    def show_ventas(self):
        self._current_view = 'ventas'
        self.clear_frame()
        main = tk.Frame(self.content_area, bg='#f0f4f8')
        main.pack(fill='both', expand=True, padx=20, pady=15)

        top_bar = tk.Frame(main, bg='#f0f4f8')
        top_bar.pack(fill='x')
        tk.Label(top_bar, text='Ventas / Caja', font=('Segoe UI', 16, 'bold'),
                 bg='#f0f4f8', fg='#2c3e50').pack(side='left')

        # Caja status
        if self._caja_abierta:
            tk.Label(top_bar, text=f'  Caja: S/.{self._saldo_inicial:.2f} (Abierta)',
                     font=('Segoe UI', 11, 'bold'), bg='#f0f4f8', fg='#2ecc71').pack(side='left', padx=10)
        else:
            tk.Button(top_bar, text='🔓 Aperturar Caja', bg='#e67e22', fg='white',
                      font=('Segoe UI', 9), relief='flat', padx=10, pady=3,
                      command=lambda: self.aperturar_caja_dialog(self.show_ventas),
                      cursor='hand2').pack(side='left', padx=10)

        if self._user_role in ('admin', 'recepcion'):
            try:
                c = self.get_db()
                rows = c.execute("SELECT COUNT(*) as cnt, COALESCE(SUM(total),0) as tot FROM cobros_pendientes").fetchone()
                c.close()
                if rows and rows['cnt'] > 0:
                    pend_frame = tk.Frame(main, bg='#fff3cd', highlightbackground='#ffc107', highlightthickness=1)
                    pend_frame.pack(fill='x', pady=(8, 0))
                    tk.Label(pend_frame,
                             text=f"💰 {rows['cnt']} cobro(s) pendiente(s) - Total: S/.{rows['tot']:.2f}",
                             font=('Segoe UI', 11, 'bold'), bg='#fff3cd', fg='#856404').pack(side='left', padx=10, pady=6)
                    tk.Button(pend_frame, text='Ir a Cobros Pendientes', bg='#ffc107', fg='#333',
                              font=('Segoe UI', 9, 'bold'), relief='flat', padx=10, pady=2,
                              command=self.show_cobros_pendientes, cursor='hand2').pack(side='right', padx=10, pady=4)
            except:
                pass

        panes = tk.Frame(main, bg='#f0f4f8')
        panes.pack(fill='both', expand=True, pady=(10, 0))

        # LEFT: POS / Cart
        left = tk.Frame(panes, bg='white', highlightbackground='#ddd', highlightthickness=1)
        left.pack(side='left', fill='both', expand=True, padx=(0, 8))

        tk.Label(left, text='Punto de Venta', font=('Segoe UI', 12, 'bold'),
                 bg='white', fg='#2c3e50', padx=10, pady=8).pack(anchor='w')

        # Client selection by DNI/RUC with API consultation
        client_frame = tk.Frame(left, bg='white', padx=10, pady=3)
        client_frame.pack(fill='x')
        tk.Label(client_frame, text='DNI / RUC:', bg='white', font=('Segoe UI', 10)).pack(side='left')
        e_dni = tk.Entry(client_frame, width=15, font=('Segoe UI', 10))
        e_dni.pack(side='left', padx=3)

        def consultar_api():
            num = e_dni.get().strip()
            if not num or not num.isdigit():
                return
            try:
                nombre = None
                direccion = None
                if len(num) == 8:
                    # DNI - api.apis.net.pe (no necesita key)
                    resp = requests.get('https://api.apis.net.pe/v1/dni?numero=' + num, timeout=10)
                    if resp.status_code == 200:
                        data = resp.json()
                        nombre = data.get('nombre', '') or ''
                        apellidos = []
                        for k in ['apellidoPaterno', 'apellidoMaterno', 'apellidos']:
                            if data.get(k):
                                apellidos.append(data[k])
                        if apellidos:
                            nombre = nombre + ' ' + ' '.join(apellidos)
                        nombre = nombre.strip()
                        direccion = data.get('direccion', '') or ''
                elif len(num) == 11:
                    # RUC - api.apis.net.pe (no necesita key)
                    resp = requests.get('https://api.apis.net.pe/v1/ruc?numero=' + num, timeout=10)
                    if resp.status_code == 200:
                        data = resp.json()
                        nombre = data.get('nombre', '') or data.get('razonSocial', '') or ''
                        direccion = data.get('direccion', '') or ''
                        if not nombre:
                            # Try SUNAT scraping as fallback
                            try:
                            # Skip scraping, just use what we have
                                pass
                            except:
                                pass
                else:
                    messagebox.showwarning('Formato', 'DNI: 8 digitos, RUC: 11 digitos')
                    return

                if nombre:
                    e_cliente.config(state='normal')
                    e_cliente.delete(0, 'end')
                    e_cliente.insert(0, nombre.strip())
                    e_cliente.config(state='readonly')
                    conn = self.get_db()
                    row = conn.execute("SELECT id, nombre FROM duenos WHERE dni=?", (num,)).fetchone()
                    if not row:
                        if direccion:
                            conn.execute("INSERT INTO duenos (dni, nombre, direccion) VALUES (?,?,?)",
                                        (num, nombre.strip(), direccion.strip()))
                        else:
                            conn.execute("INSERT INTO duenos (dni, nombre) VALUES (?,?)", (num, nombre.strip()))
                        conn.commit()
                        row = conn.execute("SELECT id, nombre FROM duenos WHERE dni=?", (num,)).fetchone()
                    conn.close()
                    if row:
                        self._venta_cliente_id = row['id']
                        self._venta_cliente_nombre = row['nombre']
                else:
                    messagebox.showinfo('Consulta', 'No se encontraron datos en API\nBuscando en BD local...')
                    buscar_cliente_local()
            except Exception as ex:
                messagebox.showinfo('API', 'Error de conexion: ' + str(ex) + '\nBuscando en BD local...')
                buscar_cliente_local()

        tk.Button(client_frame, text='Consultar', bg='#3498db', fg='white',
                  font=('Segoe UI', 9), relief='flat', padx=6, pady=1,
                  command=consultar_api, cursor='hand2').pack(side='left', padx=(0, 5))
        tk.Label(client_frame, text='Cliente:', bg='white', font=('Segoe UI', 10)).pack(side='left', padx=(5, 3))
        e_cliente = tk.Entry(client_frame, width=25, font=('Segoe UI', 10), state='readonly')
        e_cliente.pack(side='left')
        self._venta_cliente_id = None
        self._venta_cliente_nombre = ''
        self._venta_animal_id = None

        # Animal selection
        animal_frame = tk.Frame(left, bg='white', padx=10, pady=2)
        animal_frame.pack(fill='x')
        tk.Label(animal_frame, text='Mascota:', bg='white', font=('Segoe UI', 10)).pack(side='left')
        self._venta_animal_entry = AutocompleteEntry(animal_frame, width=25, font=('Segoe UI', 10))
        self._venta_animal_entry.pack(side='left', padx=3)

        def load_animales():
            conn = self.get_db()
            rows = conn.execute("SELECT a.id, a.nombre, d.nombre as dueno FROM animales a JOIN duenos d ON a.id_dueno=d.id ORDER BY a.nombre").fetchall()
            conn.close()
            return [f"{r['nombre']} ({r['dueno']})" for r in rows]

        self._venta_animal_entry.set_full_values(load_animales())

        def on_animal_select(*args):
            display = self._venta_animal_entry.get()
            if not display:
                self._venta_animal_id = None
                return
            conn = self.get_db()
            row = conn.execute("SELECT id FROM animales WHERE nombre=?", (display.split(' (')[0],)).fetchone()
            if row:
                self._venta_animal_id = row['id']
            conn.close()

        if hasattr(self._venta_animal_entry, 'entry'):
            self._venta_animal_entry.entry.bind('<FocusOut>', lambda e: on_animal_select())
            self._venta_animal_entry.entry.bind('<KeyRelease>', lambda e: on_animal_select())

        def buscar_cliente_local(*args):
            dni = e_dni.get().strip()
            if not dni:
                return
            conn = self.get_db()
            row = conn.execute("SELECT id, nombre, telefono, direccion FROM duenos WHERE dni=?", (dni,)).fetchone()
            conn.close()
            if row:
                e_cliente.config(state='normal')
                e_cliente.delete(0, 'end')
                e_cliente.insert(0, row['nombre'])
                e_cliente.config(state='readonly')
                self._venta_cliente_id = row['id']
                self._venta_cliente_nombre = row['nombre']
            else:
                if messagebox.askyesno("Cliente no encontrado", f"DNI {dni} no registrado. Desea registrarlo?"):
                    self.add_owner_dialog()
                    conn = self.get_db()
                    row = conn.execute("SELECT id, nombre FROM duenos WHERE dni=?", (dni,)).fetchone()
                    conn.close()
                    if row:
                        e_cliente.config(state='normal')
                        e_cliente.delete(0, 'end')
                        e_cliente.insert(0, row['nombre'])
                        e_cliente.config(state='readonly')
                        self._venta_cliente_id = row['id']
                        self._venta_cliente_nombre = row['nombre']

        e_dni.bind('<FocusOut>', buscar_cliente_local)
        e_dni.bind('<Return>', buscar_cliente_local)

        # Document type selector
        doc_frame = tk.Frame(left, bg='white', padx=10, pady=2)
        doc_frame.pack(fill='x')
        tk.Label(doc_frame, text='Comprobante:', bg='white', font=('Segoe UI', 10)).pack(side='left')
        tipo_comprobante_var = tk.StringVar(value='Boleta')
        for val in ['Boleta', 'Factura', 'Nota Venta']:
            tk.Radiobutton(doc_frame, text=val, variable=tipo_comprobante_var, value=val,
                          bg='white', font=('Segoe UI', 9)).pack(side='left', padx=3)

        # Buscador unico: Productos / Grooming / Servicios Medicos
        item_frame = tk.Frame(left, bg='white', padx=10, pady=2)
        item_frame.pack(fill='x')

        def get_all_items():
            conn = self.get_db()
            items = {}
            rows = conn.execute('SELECT id, nombre, precio_venta FROM productos WHERE activo=1 AND stock>0 ORDER BY nombre').fetchall()
            for r in rows:
                key = f"[Producto] {r['nombre']} - S/.{r['precio_venta']:.2f}"
                items[key] = {"tipo": "producto", "id": r["id"], "nombre": r["nombre"], "precio": r["precio_venta"]}
            rows = conn.execute('SELECT id, nombre, precio FROM servicios_grooming ORDER BY nombre').fetchall()
            for r in rows:
                key = f"[Grooming] {r['nombre']} - S/.{r['precio']:.2f}"
                items[key] = {"tipo": "grooming", "id": r["id"], "nombre": r["nombre"], "precio": r["precio"]}
            rows = conn.execute('SELECT id, nombre, precio FROM servicios_medicos ORDER BY nombre').fetchall()
            for r in rows:
                key = f"[Serv.Med] {r['nombre']} - S/.{r['precio']:.2f}"
                items[key] = {"tipo": "servicio_medico", "id": r["id"], "nombre": r["nombre"], "precio": r["precio"]}
            conn.close()
            return items

        all_items = get_all_items()
        tk.Label(item_frame, text='Buscar:', bg='white', font=('Segoe UI', 10)).pack(side='left')
        e_buscar = AutocompleteEntry(item_frame, full_values=list(all_items.keys()), width=55, font=('Segoe UI', 10))
        e_buscar.pack(side='left', padx=3)
        tk.Label(item_frame, text='Cant:', bg='white', font=('Segoe UI', 10)).pack(side='left', padx=(3, 2))
        e_cant = tk.Spinbox(item_frame, from_=1, to=99, width=5, font=('Segoe UI', 10))
        e_cant.pack(side='left')

        def add_item_to_cart():
            key = e_buscar.get() if hasattr(e_buscar, 'entry') else e_buscar.get()
            if not key:
                messagebox.showwarning('Error', 'Escriba o seleccione un item')
                return
            try:
                qty = int(e_cant.get())
            except:
                qty = 1
            if key in all_items:
                item = all_items[key]
                if item["precio"] == 0 and item["tipo"] == "servicio_medico":
                    precio_dialog = tk.Toplevel(self.root)
                    precio_dialog.title('Precio')
                    precio_dialog.geometry('280x130')
                    precio_dialog.configure(bg='#f0f4f8')
                    precio_dialog.transient(self.root)
                    precio_dialog.grab_set()
                    center_dialog(precio_dialog, self.root)
                    tk.Label(precio_dialog, text='Precio para ' + item['nombre'] + ':', bg='#f0f4f8', font=('Segoe UI', 10)).pack(pady=(10, 5))
                    e_precio = tk.Entry(precio_dialog, width=20, font=('Segoe UI', 12), justify='center')
                    e_precio.pack(pady=5)
                    e_precio.insert(0, '0')
                    e_precio.focus_set()
                    def confirm_precio(precio_item=item):
                        try:
                            p = float(e_precio.get())
                        except:
                            p = 0
                        if p <= 0:
                            messagebox.showwarning('Error', 'Ingrese un precio valido')
                            return
                        precio_dialog.destroy()
                        self._add_to_cart(precio_item["tipo"], precio_item["id"], precio_item["nombre"], qty, p)
                        e_buscar.set('')
                        update_cart_display()
                    tk.Button(precio_dialog, text='Agregar al Carrito', bg='#3498db', fg='white',
                            font=('Segoe UI', 10), relief='flat', padx=15, pady=3,
                            command=confirm_precio, cursor='hand2').pack(pady=5)
                else:
                    self._add_to_cart(item["tipo"], item["id"], item["nombre"], qty, item["precio"])
                    e_buscar.set('')
            else:
                # Custom item: ask for price
                precio_dialog = tk.Toplevel(self.root)
                precio_dialog.title('Item Personalizado')
                precio_dialog.geometry('300x150')
                precio_dialog.configure(bg='#f0f4f8')
                precio_dialog.transient(self.root)
                precio_dialog.grab_set()
                center_dialog(precio_dialog, self.root)
                tk.Label(precio_dialog, text='Item: ' + key, bg='#f0f4f8', font=('Segoe UI', 10)).pack(pady=(5, 2))
                tk.Label(precio_dialog, text='Precio unitario (S/):', bg='#f0f4f8', font=('Segoe UI', 10)).pack()
                e_precio = tk.Entry(precio_dialog, width=20, font=('Segoe UI', 12), justify='center')
                e_precio.pack(pady=5)
                e_precio.insert(0, '0')
                e_precio.focus_set()
                def confirm_precio():
                    try:
                        p = float(e_precio.get())
                    except:
                        p = 0
                    if p <= 0:
                        messagebox.showwarning('Error', 'Ingrese un precio valido')
                        return
                    precio_dialog.destroy()
                    self._add_to_cart('producto', 0, key, qty, p)
                    e_buscar.set('')
                    update_cart_display()
                tk.Button(precio_dialog, text='Agregar al Carrito', bg='#3498db', fg='white',
                        font=('Segoe UI', 10), relief='flat', padx=15, pady=3,
                        command=confirm_precio, cursor='hand2').pack(pady=5)
            update_cart_display()

        tk.Button(item_frame, text='Agregar', bg='#2ecc71', fg='white',
                  font=('Segoe UI', 9, 'bold'), relief='flat', padx=12, pady=2,
                  command=add_item_to_cart, cursor='hand2').pack(side='left', padx=(3, 0))

        # Cart display
        cart_frame = tk.Frame(left, bg='white', padx=10, pady=3)
        cart_frame.pack(fill='both', expand=True)
        tk.Label(cart_frame, text='Carrito:', font=('Segoe UI', 11, 'bold'),
                 bg='white', fg='#2c3e50').pack(anchor='w')

        cart_cols = ('item', 'qty', 'price', 'subtotal')
        cart_tree = ttk.Treeview(cart_frame, columns=cart_cols, show='headings', height=6)
        cart_tree.heading('item', text='Item')
        cart_tree.heading('qty', text='Cant')
        cart_tree.heading('price', text='P.Unit')
        cart_tree.heading('subtotal', text='Subtotal')
        cart_tree.column('item', width=300)
        cart_tree.column('qty', width=50)
        cart_tree.column('price', width=80)
        cart_tree.column('subtotal', width=80)
        cart_tree.pack(fill='both', expand=True)
        cart_tree.tag_configure('producto', background='#e8f5e9')
        cart_tree.tag_configure('grooming', background='#e3f2fd')
        cart_tree.tag_configure('servicio_medico', background='#fff3e0')

        def edit_cart_cell(event):
            col = cart_tree.identify_column(event.x)
            row_id = cart_tree.identify_row(event.y)
            if not row_id:
                return
            vals = list(cart_tree.item(row_id, 'values'))
            # col '#1'=item, '#2'=qty, '#3'=price, '#4'=subtotal
            if col == '#2':
                dialog = tk.Toplevel(self.root)
                dialog.title('Editar Cantidad')
                dialog.geometry('260x120')
                dialog.configure(bg='#f0f4f8')
                dialog.transient(self.root); dialog.grab_set()
                center_dialog(dialog, self.root)
                tk.Label(dialog, text='Nueva cantidad para:', bg='#f0f4f8', font=('Segoe UI', 10)).pack(pady=(8, 0))
                tk.Label(dialog, text=vals[0], bg='#f0f4f8', font=('Segoe UI', 10, 'bold')).pack()
                e_new = tk.Entry(dialog, width=10, font=('Segoe UI', 14), justify='center')
                e_new.pack(pady=6)
                e_new.insert(0, vals[1])
                e_new.focus_set(); e_new.select_range(0, 'end')
                def save_qty():
                    try:
                        new_qty = int(e_new.get())
                        if new_qty <= 0: raise ValueError
                        price = float(vals[2].replace('S/.', ''))
                        vals[1] = str(new_qty)
                        vals[3] = f'S/.{new_qty * price:.2f}'
                        cart_tree.item(row_id, values=vals)
                        dialog.destroy()
                        update_cart_display()
                    except:
                        messagebox.showwarning('Error', 'Ingrese un numero valido')
                e_new.bind('<Return>', lambda e: save_qty())
                tk.Button(dialog, text='Aceptar', bg='#2ecc71', fg='white',
                          font=('Segoe UI', 10), relief='flat', padx=20, pady=3,
                          command=save_qty, cursor='hand2').pack(pady=4)
            elif col == '#3':
                dialog = tk.Toplevel(self.root)
                dialog.title('Editar Precio')
                dialog.geometry('260x120')
                dialog.configure(bg='#f0f4f8')
                dialog.transient(self.root); dialog.grab_set()
                center_dialog(dialog, self.root)
                tk.Label(dialog, text='Nuevo precio para:', bg='#f0f4f8', font=('Segoe UI', 10)).pack(pady=(8, 0))
                tk.Label(dialog, text=vals[0], bg='#f0f4f8', font=('Segoe UI', 10, 'bold')).pack()
                e_new = tk.Entry(dialog, width=10, font=('Segoe UI', 14), justify='center')
                e_new.pack(pady=6)
                e_new.insert(0, vals[2].replace('S/.', ''))
                e_new.focus_set(); e_new.select_range(0, 'end')
                def save_price():
                    try:
                        new_price = float(e_new.get())
                        if new_price <= 0: raise ValueError
                        qty = int(vals[1])
                        vals[2] = f'S/.{new_price:.2f}'
                        vals[3] = f'S/.{qty * new_price:.2f}'
                        cart_tree.item(row_id, values=vals)
                        dialog.destroy()
                        update_cart_display()
                    except:
                        messagebox.showwarning('Error', 'Ingrese un precio valido')
                e_new.bind('<Return>', lambda e: save_price())
                tk.Button(dialog, text='Aceptar', bg='#2ecc71', fg='white',
                          font=('Segoe UI', 10), relief='flat', padx=20, pady=3,
                          command=save_price, cursor='hand2').pack(pady=4)

        cart_tree.bind('<Double-1>', edit_cart_cell)

        total_var = tk.StringVar(value='Total: S/.0.00')
        tk.Label(cart_frame, textvariable=total_var, font=('Segoe UI', 14, 'bold'),
                 bg='white', fg='#e74c3c').pack(anchor='e', pady=(3, 0))

        def remove_from_cart():
            sel = cart_tree.selection()
            if not sel: return
            for item in sel:
                cart_tree.delete(item)
            update_cart_display()

        btn_cart = tk.Frame(cart_frame, bg='white')
        btn_cart.pack(fill='x', pady=(3, 0))
        tk.Button(btn_cart, text='Quitar', bg='#e74c3c', fg='white',
                  font=('Segoe UI', 9), relief='flat', padx=10, pady=2,
                  command=remove_from_cart, cursor='hand2').pack(side='left', padx=2)
        tk.Button(btn_cart, text='Limpiar', bg='#95a5a6', fg='white',
                  font=('Segoe UI', 9), relief='flat', padx=10, pady=2,
                  command=lambda: ([cart_tree.delete(i) for i in cart_tree.get_children()], update_cart_display()), cursor='hand2').pack(side='left', padx=2)
        tk.Button(btn_cart, text='Cobrar', bg='#2ecc71', fg='white',
                  font=('Segoe UI', 12, 'bold'), relief='flat', padx=30, pady=5,
                  command=self._checkout, cursor='hand2').pack(side='right', padx=2)

        # RIGHT: Today sales / history
        right = tk.Frame(panes, bg='white', highlightbackground='#ddd', highlightthickness=1)
        right.pack(side='left', fill='both', expand=True)

        tk.Label(right, text='Ventas de Hoy', font=('Segoe UI', 12, 'bold'),
                 bg='white', fg='#2c3e50', padx=10, pady=8).pack(anchor='w')

        # Cash summary
        cash_summary = tk.Frame(right, bg='white', padx=10)
        cash_summary.pack(fill='x')
        daily_total_var = tk.StringVar(value='Total Hoy: S/.0.00')
        tk.Label(cash_summary, textvariable=daily_total_var, font=('Segoe UI', 12, 'bold'),
                 bg='white', fg='#1abc9c').pack(side='left')

        def show_all_sales():
            for item in venta_tree.get_children():
                venta_tree.delete(item)
            conn = self.get_db()
            rows = conn.execute(
                'SELECT v.id, v.fecha, v.total, v.tipo, COALESCE(d.nombre, "-") as cliente '
                'FROM ventas v LEFT JOIN duenos d ON v.id_cliente = d.id '
                'ORDER BY v.id DESC LIMIT 200').fetchall()
            all_total = conn.execute('SELECT COALESCE(SUM(total), 0) FROM ventas').fetchone()[0]
            conn.close()
            for r in rows:
                venta_tree.insert("", "end", values=(r["id"], r["fecha"], r["cliente"],
                    f"S/.{r['total']:.2f}", r["tipo"]))
            daily_total_var.set(f"Total General: S/.{all_total:.2f}")
            btn_historial.config(text='Ventas Hoy', command=load_daily_sales)

        btn_historial = tk.Button(cash_summary, text='Historial Ventas', bg='#8e44ad', fg='white',
                  font=('Segoe UI', 9), relief='flat', padx=8, pady=2,
                  command=show_all_sales, cursor='hand2')
        btn_historial.pack(side='left', padx=5)

        tk.Button(cash_summary, text='Exportar CSV', bg='#3498db', fg='white',
                  font=('Segoe UI', 9), relief='flat', padx=8, pady=2,
                  command=lambda: self._export_to_csv(
                      [[venta_tree.item(i, 'values')[j] for j in range(5)] for i in venta_tree.get_children()],
                      ['ID', 'Fecha', 'Cliente', 'Total', 'Tipo'],
                      f'ventas_{str(date.today())}.csv'
                  ), cursor='hand2').pack(side='left', padx=5)
        tk.Button(cash_summary, text='Excel', bg='#27ae60', fg='white',
                  font=('Segoe UI', 9), relief='flat', padx=8, pady=2,
                  command=lambda: self._export_to_excel(
                      [[venta_tree.item(i, 'values')[j] for j in range(5)] for i in venta_tree.get_children()],
                      ['ID', 'Fecha', 'Cliente', 'Total', 'Tipo'],
                      f'ventas_{str(date.today())}.xlsx'
                  ), cursor='hand2').pack(side='left', padx=2)

        venta_cols = ('id', 'fecha', 'cliente', 'total', 'tipo')
        venta_tree = ttk.Treeview(right, columns=venta_cols, show='headings', height=15)
        venta_tree.heading('id', text='ID')
        venta_tree.heading('fecha', text='Fecha')
        venta_tree.heading('cliente', text='Cliente')
        venta_tree.heading('total', text='Total')
        venta_tree.heading('tipo', text='Tipo')
        venta_tree.column('id', width=40)
        venta_tree.column('fecha', width=90)
        venta_tree.column('cliente', width=180)
        venta_tree.column('total', width=80)
        venta_tree.column('tipo', width=80)
        venta_tree.pack(fill='both', expand=True, padx=8, pady=3)
        venta_tree.bind('<Double-1>', lambda e: self._editar_venta(venta_tree))

        def load_daily_sales():
            for item in venta_tree.get_children():
                venta_tree.delete(item)
            conn = self.get_db()
            rows = conn.execute(
                'SELECT v.id, v.fecha, v.total, v.tipo, COALESCE(d.nombre, "-") as cliente '
                'FROM ventas v LEFT JOIN duenos d ON v.id_cliente = d.id '
                'WHERE v.fecha = ? ORDER BY v.id DESC', (str(date.today()),)).fetchall()
            daily_total = conn.execute(
                'SELECT COALESCE(SUM(total), 0) FROM ventas WHERE fecha = ?',
                (str(date.today()),)).fetchone()[0]
            conn.close()
            for r in rows:
                venta_tree.insert("", "end", values=(r["id"], r["fecha"], r["cliente"],
                    f"S/.{r['total']:.2f}", r["tipo"]))
            daily_total_var.set(f"Total Hoy: S/.{daily_total:.2f}")

        load_daily_sales()

        def update_cart_display():
            total = 0.0
            for item in cart_tree.get_children():
                vals = cart_tree.item(item, 'values')
                try:
                    subtotal_str = vals[3].replace('S/.', '')
                    total += float(subtotal_str)
                except:
                    pass
            total_var.set(f"Total: S/.{total:.2f}")

        self._cart_items = []
        self._cart_tree = cart_tree
        self._cart_total_var = total_var
        self._cart_update = update_cart_display
        self._load_daily_sales = load_daily_sales
        self._daily_total_var = daily_total_var

    def _add_to_cart(self, item_type, ref_id, name, qty, price):
        # Check stock for products
        if item_type == 'producto':
            conn = self.get_db()
            stock = conn.execute('SELECT stock FROM productos WHERE id=?', (ref_id,)).fetchone()
            conn.close()
            if stock and stock[0] < qty:
                messagebox.showerror('Stock insuficiente', f'Stock disponible: {stock[0]}')
                return
        # Check if already in cart and update qty
        for item in self._cart_tree.get_children():
            vals = self._cart_tree.item(item, 'values')
            if vals[0] == name:
                new_qty = int(vals[1]) + qty
                subtotal = new_qty * price
                self._cart_tree.item(item, values=(name, str(new_qty), f'S/.{price:.2f}', f'S/.{subtotal:.2f}'))
                return
        subtotal = qty * price
        self._cart_tree.insert('', 'end', values=(name, str(qty), f'S/.{price:.2f}', f'S/.{subtotal:.2f}'),
                               tags=(item_type,))

    def _checkout(self):
        try:
            if not self._caja_abierta:
                messagebox.showwarning("Caja Cerrada", "Debe aperturar la caja primero")
                return
            cart_items = []
            total = 0.0
            for item in self._cart_tree.get_children():
                vals = self._cart_tree.item(item, 'values')
                cart_items.append(vals)
                try:
                    subtotal_str = vals[3].replace('S/.', '')
                    total += float(subtotal_str)
                except:
                    pass
            if not cart_items:
                messagebox.showwarning('Carrito vacio', 'Agregue items al carrito')
                return
            if not self._venta_cliente_id:
                conn = self.get_db()
                row = conn.execute("SELECT id, nombre FROM duenos WHERE nombre='VARIADO'").fetchone()
                if not row:
                    conn.execute("INSERT INTO duenos (dni, nombre, telefono) VALUES ('00000000', 'VARIADO', '000000000')")
                    conn.commit()
                    row = conn.execute("SELECT id, nombre FROM duenos WHERE nombre='VARIADO'").fetchone()
                conn.close()
                if row:
                    self._venta_cliente_id = row['id']
                    self._venta_cliente_nombre = row['nombre']
            # Payment dialog
            pay_dialog = tk.Toplevel(self.root)
            pay_dialog.title('Forma de Pago')
            pay_dialog.geometry('400x420')
            pay_dialog.configure(bg='#f0f4f8')
            pay_dialog.transient(self.root)
            pay_dialog.grab_set()
            pay_dialog.update_idletasks()
            x = self.root.winfo_x() + (self.root.winfo_width() - pay_dialog.winfo_width()) // 2
            y = self.root.winfo_y() + (self.root.winfo_height() - pay_dialog.winfo_height()) // 2
            pay_dialog.geometry(f'+{x}+{y}')

            bg_card = '#ffffff'
            bg_main = '#f0f4f8'
            fg_dark = '#2c3e50'
            fg_label = '#7f8c8d'
            fg_accent = '#2ecc71'
            font_bold = ('Segoe UI', 10, 'bold')
            font_normal = ('Segoe UI', 9)

            # Header
            header = tk.Frame(pay_dialog, bg='#2c3e50', height=60)
            header.pack(fill='x')
            header.pack_propagate(False)
            tk.Label(header, text='COBRAR', font=('Segoe UI', 14, 'bold'),
                     bg='#2c3e50', fg='#ffffff').pack(side='left', padx=16, pady=16)
            tk.Label(header, text=f'S/.{total:.2f}', font=('Segoe UI', 18, 'bold'),
                     bg='#2c3e50', fg='#2ecc71').pack(side='right', padx=16, pady=16)

            # Payment method (multi-select with amounts)
            met_card = tk.Frame(pay_dialog, bg=bg_card, padx=14, pady=10,
                                highlightbackground='#dcdde1', highlightthickness=1)
            met_card.pack(fill='x', padx=14, pady=(14, 0))
            tk.Label(met_card, text='METODO DE PAGO (marque uno o varios)', font=font_bold,
                     bg=bg_card, fg=fg_label).pack(anchor='w')
            met_inner = tk.Frame(met_card, bg=bg_card)
            met_inner.pack(fill='x', pady=(6, 0))

            pagos = {}
            pago_vars = {}
            pago_entries = {}
            for val, txt in [('efectivo', 'Efectivo'), ('yape', 'Yape'), ('transferencia', 'Transferencia')]:
                row = tk.Frame(met_inner, bg=bg_card)
                row.pack(fill='x', pady=2)
                var = tk.BooleanVar(value=False)
                pago_vars[val] = var
                cb = tk.Checkbutton(row, text=txt, variable=var, bg=bg_card, fg=fg_dark,
                                    font=font_normal, selectcolor='#ecf0f1', activebackground=bg_card)
                cb.pack(side='left', padx=(0, 6))
                ent = tk.Entry(row, width=12, font=('Segoe UI', 11, 'bold'), justify='center',
                               bg='#ffffff', fg=fg_dark, relief='solid', bd=1)
                ent.pack(side='left', padx=(0, 4))
                ent.config(state='disabled')
                pago_entries[val] = ent
                def toggle_entry(*args, e=ent, v=var):
                    e.config(state='normal' if v.get() else 'disabled')
                    if not v.get():
                        e.delete(0, 'end')
                    checked = [kv for kv, kvv in pago_vars.items() if kvv.get()]
                    if len(checked) == 1:
                        for kv in checked:
                            pago_entries[kv].delete(0, 'end')
                            pago_entries[kv].insert(0, f'{total:.2f}')
                var.trace_add('write', toggle_entry)
                pagos[val] = txt

            # Print option
            print_card = tk.Frame(pay_dialog, bg=bg_card, padx=14, pady=10,
                                  highlightbackground='#dcdde1', highlightthickness=1)
            print_card.pack(fill='x', padx=14, pady=(8, 0))
            tk.Label(print_card, text='IMPRESION', font=font_bold,
                     bg=bg_card, fg=fg_label).pack(anchor='w')
            print_var = tk.StringVar(value='')
            print_inner = tk.Frame(print_card, bg=bg_card)
            print_inner.pack(fill='x', pady=(6, 0))
            for val, txt in [('ticket', 'Ticket'), ('pdf', 'PDF'), ('none', 'No imprimir')]:
                rb = tk.Radiobutton(print_inner, text=f'  {txt}', variable=print_var, value=val,
                                    bg=bg_card, fg=fg_dark, selectcolor='#ecf0f1',
                                    font=font_normal, activebackground=bg_card, activeforeground='#2ecc71')
                rb.pack(side='left', padx=4)

            def pay_confirm():
                selected = []
                suma = 0.0
                for val, txt in pagos.items():
                    if pago_vars[val].get():
                        try:
                            amt = float(pago_entries[val].get() or '0')
                        except:
                            amt = 0
                        if amt > 0:
                            selected.append((val, txt, amt))
                            suma += amt
                if not selected:
                    messagebox.showwarning('Error', 'Seleccione al menos un metodo de pago')
                    return
                if abs(suma - total) > 0.01:
                    messagebox.showwarning('Error', f'La suma de pagos (S/.{suma:.2f}) no coincide con el total (S/.{total:.2f})')
                    return
                print_op = print_var.get()
                if not print_op:
                    messagebox.showwarning('Error', 'Seleccione opcion de impresion')
                    return
                pay_dialog.destroy()
                try:
                    self._finalizar_venta(total, selected, '', print_op)
                except Exception as ex:
                    import traceback
                    messagebox.showerror('Error al finalizar venta',
                        f'Datos: total={total}, selected={selected}\n\n{ex}\n\n{traceback.format_exc()}')

            btn_frame = tk.Frame(pay_dialog, bg=bg_main)
            btn_frame.pack(fill='x', padx=14, pady=(14, 14))
            tk.Button(btn_frame, text='CONFIRMAR PAGO', bg='#2ecc71', fg='white',
                      font=('Segoe UI', 11, 'bold'), relief='flat', padx=25, pady=10,
                      command=pay_confirm, cursor='hand2', activebackground='#27ae60',
                      activeforeground='white').pack(fill='x')
        except Exception as e:
            import traceback
            messagebox.showerror('Error Checkout', f'{e}\n\n{traceback.format_exc()}')

    def _editar_venta(self, tree):
        try:
            sel = tree.selection()
            if not sel:
                return
            vals = tree.item(sel[0], 'values')
            venta_id = int(vals[0])
        except:
            messagebox.showwarning('Error', 'Seleccione una venta valida')
            return

        conn = self.get_db()
        venta_info = conn.execute(
            'SELECT v.id, v.fecha, v.total, v.tipo_comprobante, v.metodo_pago, v.id_cliente, COALESCE(d.nombre, "-") as cliente '
            'FROM ventas v LEFT JOIN duenos d ON v.id_cliente = d.id WHERE v.id=?', (venta_id,)).fetchone()
        items = conn.execute('SELECT id, nombre, cantidad, precio_unitario, subtotal FROM venta_items WHERE id_venta=?', (venta_id,)).fetchall()
        conn.close()
        if not venta_info:
            messagebox.showerror('Error', 'Venta no encontrada')
            return

        dlg = tk.Toplevel(self.root)
        dlg.title(f'Editar Venta #{venta_id}')
        dlg.geometry('650x560')
        dlg.configure(bg='#f0f4f8')
        dlg.transient(self.root)
        dlg.grab_set()
        dlg.update_idletasks()
        x = self.root.winfo_x() + (self.root.winfo_width() - dlg.winfo_width()) // 2
        y = self.root.winfo_y() + (self.root.winfo_height() - dlg.winfo_height()) // 2
        dlg.geometry(f'+{x}+{y}')

        info_frame = tk.Frame(dlg, bg='#ffffff', padx=14, pady=8,
                              highlightbackground='#dcdde1', highlightthickness=1)
        info_frame.pack(fill='x', padx=14, pady=(12, 4))
        tk.Label(info_frame, text=f'Venta #{venta_id}  |  {venta_info["cliente"]}  |  {venta_info["fecha"]}',
                 font=('Segoe UI', 11, 'bold'), bg='#ffffff', fg='#2c3e50').pack(anchor='w')
        tk.Label(info_frame,
                 text=f'Pago: {venta_info["metodo_pago"] or "N/A"}  |  Comprobante: {venta_info["tipo_comprobante"] or "N/A"}',
                 font=('Segoe UI', 9), bg='#ffffff', fg='#7f8c8d').pack(anchor='w', pady=(2, 0))

        cols = ('item_id', 'nombre', 'cantidad', 'precio', 'subtotal')
        tree_frame = tk.Frame(dlg, bg='#f0f4f8')
        tree_frame.pack(fill='both', expand=True, padx=14, pady=4)
        t = ttk.Treeview(tree_frame, columns=cols, show='headings', height=9)
        vsb = ttk.Scrollbar(tree_frame, orient='vertical', command=t.yview)
        t.configure(yscrollcommand=vsb.set)
        t.pack(side='left', fill='both', expand=True)
        vsb.pack(side='right', fill='y')
        for c, h in zip(cols, ['ID', 'Producto', 'Cant', 'Precio', 'Subtotal']):
            t.heading(c, text=h)
            t.column(c, width=40 if c == 'item_id' else 70 if c == 'cantidad' else 130)

        for row in items:
            t.insert('', 'end', values=(row['id'], row['nombre'], row['cantidad'],
                     f"S/.{row['precio_unitario']:.2f}", f"S/.{row['subtotal']:.2f}"))

        def recalc_total():
            tot = 0.0
            for child in t.get_children():
                vals = t.item(child, 'values')
                try:
                    tot += float(vals[4].replace('S/.', ''))
                except:
                    pass
            lbl_total.config(text=f'Total: S/.{tot:.2f}')

        def edit_cell(event):
            col = t.identify_column(event.x)
            row_id = t.identify_row(event.y)
            if not row_id:
                return
            col_idx = int(col.replace('#', '')) - 1
            if col_idx not in (2, 3):
                return
            x, y, w, h = t.bbox(row_id, col)
            val = t.item(row_id, 'values')[col_idx]
            e = tk.Entry(t, font=('Segoe UI', 10))
            e.place(x=x, y=y, width=w, height=h)
            e.insert(0, val.replace('S/.', ''))
            e.focus()
            def save_edit(event=None):
                new_val = e.get().strip()
                e.destroy()
                if not new_val:
                    return
                try:
                    new_num = float(new_val)
                    vals = list(t.item(row_id, 'values'))
                    vals[col_idx] = str(int(new_num)) if col_idx == 2 else f'S/.{new_num:.2f}'
                    qty = int(vals[2])
                    price = float(vals[3].replace('S/.', ''))
                    vals[4] = f'S/.{qty * price:.2f}'
                    t.item(row_id, values=tuple(vals))
                    recalc_total()
                except:
                    messagebox.showwarning('Error', 'Valor invalido')
            e.bind('<Return>', save_edit)
            e.bind('<FocusOut>', save_edit)
        t.bind('<Double-1>', edit_cell)

        lbl_total = tk.Label(dlg, text='Total: S/.0.00', font=('Segoe UI', 12, 'bold'),
                             bg='#f0f4f8', fg='#e74c3c')
        lbl_total.pack(pady=2)
        recalc_total()

        def eliminar_item():
            sel = t.selection()
            if not sel:
                messagebox.showwarning('Seleccion', 'Seleccione un item')
                return
            if messagebox.askyesno('Confirmar', 'Eliminar este item de la venta?'):
                t.delete(sel[0])
                recalc_total()

        def agregar_producto():
            try:
                conn = self.get_db()
                items_map = {}
                rows = conn.execute('SELECT id, nombre, precio_venta FROM productos WHERE activo=1 ORDER BY nombre').fetchall()
                for r in rows:
                    key = f"[Producto] {r['nombre']} - S/.{r['precio_venta']:.2f}"
                    items_map[key] = {'tipo': 'producto', 'id': r['id'], 'nombre': r['nombre'], 'precio': r['precio_venta']}
                rows = conn.execute('SELECT id, nombre, precio FROM servicios_grooming ORDER BY nombre').fetchall()
                for r in rows:
                    key = f"[Grooming] {r['nombre']} - S/.{r['precio']:.2f}"
                    items_map[key] = {'tipo': 'grooming', 'id': r['id'], 'nombre': r['nombre'], 'precio': r['precio']}
                rows = conn.execute('SELECT id, nombre, precio FROM servicios_medicos ORDER BY nombre').fetchall()
                for r in rows:
                    key = f"[Serv.Med] {r['nombre']} - S/.{r['precio']:.2f}"
                    items_map[key] = {'tipo': 'servicio_medico', 'id': r['id'], 'nombre': r['nombre'], 'precio': r['precio']}
                conn.close()

                adlg = tk.Toplevel(dlg)
                adlg.title('Agregar a Venta')
                adlg.geometry('500x200')
                adlg.configure(bg='#f0f4f8')
                adlg.transient(dlg)
                adlg.grab_set()
                adlg.update_idletasks()
                ax = dlg.winfo_x() + (dlg.winfo_width() - adlg.winfo_width()) // 2
                ay = dlg.winfo_y() + (dlg.winfo_height() - adlg.winfo_height()) // 2
                adlg.geometry(f'+{ax}+{ay}')

                tk.Label(adlg, text='Buscar producto, grooming o servicio:', bg='#f0f4f8',
                         font=('Segoe UI', 10, 'bold'), fg='#2c3e50').pack(pady=(12, 4))
                search_var = tk.StringVar()
                e_search = AutocompleteEntry(adlg, full_values=list(items_map.keys()), width=55,
                                             font=('Segoe UI', 10))
                e_search.pack(pady=2)

                row2 = tk.Frame(adlg, bg='#f0f4f8')
                row2.pack(pady=6)
                tk.Label(row2, text='Cantidad:', bg='#f0f4f8', font=('Segoe UI', 10)).pack(side='left')
                qty_e = tk.Entry(row2, width=8, font=('Segoe UI', 12, 'bold'), justify='center')
                qty_e.pack(side='left', padx=6)
                qty_e.insert(0, '1')
                tk.Label(row2, text='', bg='#f0f4f8').pack(side='left', fill='x', expand=True)

                def confirm_add():
                    try:
                        key = e_search.get() if hasattr(e_search, 'entry') else e_search.get()
                        if key not in items_map:
                            messagebox.showwarning('Seleccion', 'Seleccione un item de la lista')
                            return
                        item = items_map[key]
                        try:
                            qty = int(qty_e.get())
                        except:
                            qty = 1
                        if qty < 1:
                            qty = 1
                        next_id = 0
                        for child in t.get_children():
                            vid = int(t.item(child, 'values')[0])
                            if vid < next_id:
                                next_id = vid
                        next_id -= 1
                        price = item['precio']
                        subtotal = qty * price
                        t.insert('', 'end', values=(next_id, item['nombre'], qty,
                                 f'S/.{price:.2f}', f'S/.{subtotal:.2f}'))
                        recalc_total()
                        adlg.destroy()
                    except Exception as ex:
                        messagebox.showerror('Error', f'No se pudo agregar: {ex}')

                btn_frame = tk.Frame(adlg, bg='#f0f4f8')
                btn_frame.pack(fill='x', padx=14, pady=(4, 10))
                tk.Button(btn_frame, text='Agregar a Venta', bg='#2ecc71', fg='white',
                          font=('Segoe UI', 10, 'bold'), relief='flat', padx=18, pady=6,
                          command=confirm_add, cursor='hand2').pack(side='right')
            except Exception as ex:
                messagebox.showerror('Error', f'Error al abrir dialogo: {ex}')

        def save_changes():
            new_total = 0.0
            conn = self.get_db()
            existing_ids = set(row['id'] for row in items)
            seen_ids = set()
            for child in t.get_children():
                vals = t.item(child, 'values')
                item_id = int(vals[0])
                seen_ids.add(item_id)
                new_qty = int(vals[2])
                new_price = float(vals[3].replace('S/.', ''))
                new_subtotal = new_qty * new_price
                new_total += new_subtotal
                if item_id > 0:
                    conn.execute('UPDATE venta_items SET cantidad=?, precio_unitario=?, subtotal=? WHERE id=?',
                                (new_qty, new_price, new_subtotal, item_id))
                else:
                    conn.execute('INSERT INTO venta_items (id_venta, tipo_item, nombre, cantidad, precio_unitario, subtotal) VALUES (?,?,?,?,?,?)',
                                (venta_id, 'producto', vals[1], new_qty, new_price, new_subtotal))
            for old_id in existing_ids:
                if old_id not in seen_ids:
                    conn.execute('DELETE FROM venta_items WHERE id=?', (old_id,))
            conn.execute('UPDATE ventas SET total=? WHERE id=?', (new_total, venta_id))
            conn.commit()
            conn.close()
            dlg.destroy()
            messagebox.showinfo('Exito', f'Venta #{venta_id} actualizada. Nuevo total: S/.{new_total:.2f}')

        btn_frame = tk.Frame(dlg, bg='#f0f4f8')
        btn_frame.pack(fill='x', padx=14, pady=(6, 12))
        tk.Button(btn_frame, text='Guardar Cambios', bg='#2ecc71', fg='white',
                  font=('Segoe UI', 11, 'bold'), relief='flat', padx=20, pady=6,
                  command=save_changes, cursor='hand2').pack(side='left', padx=5)
        tk.Button(btn_frame, text='Agregar Producto', bg='#3498db', fg='white',
                  font=('Segoe UI', 10, 'bold'), relief='flat', padx=12, pady=6,
                  command=agregar_producto, cursor='hand2').pack(side='left', padx=5)
        tk.Button(btn_frame, text='Eliminar Item', bg='#e74c3c', fg='white',
                  font=('Segoe UI', 10, 'bold'), relief='flat', padx=12, pady=6,
                  command=eliminar_item, cursor='hand2').pack(side='left', padx=5)
        tk.Button(btn_frame, text='Cancelar', bg='#95a5a6', fg='white',
                  font=('Segoe UI', 10, 'bold'), relief='flat', padx=12, pady=6,
                  command=dlg.destroy, cursor='hand2').pack(side='right', padx=5)

    def _finalizar_venta(self, total, modo_pago, tipo_comprobante='', print_op='pdf'):
        cart_items = []
        cart_tags = {}
        for item in self._cart_tree.get_children():
            vals = self._cart_tree.item(item, 'values')
            tags = self._cart_tree.item(item, 'tags')
            cart_items.append(vals)
            # Store the tag (item type) for each item, referenced by name
            cart_tags[vals[0]] = tags[0] if tags else 'producto'

        conn = self.get_db()
        metodo_str = ', '.join(f'{t}:{m:.2f}' for _, t, m in modo_pago)
        id_animal = getattr(self, '_venta_animal_id', None)
        cursor = conn.execute(
            'INSERT INTO ventas (id_cliente, id_animal, fecha, total, tipo, tipo_comprobante, tipo_pago, metodo_pago) VALUES (?,?,?,?,?,?,?,?)',
            (self._venta_cliente_id or None, id_animal, str(date.today()), total, metodo_str,
             tipo_comprobante or 'N/A', metodo_str, metodo_str))
        venta_id = cursor.lastrowid

        for vals in cart_items:
            tipo_item = cart_tags.get(vals[0], 'producto')
            # Determine referencia_id from item type and name
            if tipo_item == 'producto':
                conn.execute('UPDATE productos SET stock = stock - ? WHERE nombre = ? AND activo=1',
                            (int(vals[1]), vals[0]))
            conn.execute(
                'INSERT INTO venta_items (id_venta, tipo_item, referencia_id, nombre, cantidad, precio_unitario, subtotal) VALUES (?,?,?,?,?,?,?)',
                (venta_id, tipo_item, None, vals[0], int(vals[1]),
                 float(vals[2].replace('S/.', '')), float(vals[3].replace('S/.', ''))))

        for val, txt, monto in modo_pago:
            conn.execute("INSERT INTO caja (fecha, tipo, concepto, monto, referencia_tipo, referencia_id) VALUES (?, 'ingreso', ?, ?, 'venta', ?)",
                        (str(date.today()), f'Venta #{venta_id} ({txt})', monto, venta_id))

        conn.commit()
        conn.close()

        for i in self._cart_tree.get_children():
            self._cart_tree.delete(i)
        self._cart_update()
        self._load_daily_sales()

        if print_op != 'none':
            try:
                ticket_items_for_pdf = []
                for vals in cart_items:
                    ticket_items_for_pdf.append([
                        vals[0],
                        vals[1],
                        f"S/.{float(vals[2].replace('S/.','')):.2f}",
                        f"S/.{float(vals[3].replace('S/.','')):.2f}"
                    ])
                pdf_path = self._generar_ticket_pdf(venta_id, ticket_items_for_pdf, total, metodo_str)
                if pdf_path:
                    if print_op == 'ticket':
                        try:
                            import subprocess
                            subprocess.run(['print', pdf_path], shell=True, timeout=10)
                        except:
                            import os
                            os.startfile(pdf_path)
                    else:
                        if messagebox.askyesno('PDF', 'Abrir PDF?'):
                            import os
                            os.startfile(pdf_path)
            except Exception as e:
                messagebox.showwarning('Documento', f'No se pudo generar: {e}')

        msg = f'Venta registrada. Total: S/.{total:.2f}'
        msg += '\nPago: ' + metodo_str
        messagebox.showinfo('Exito', msg)

    def show_creditos(self):
        self._current_view = 'creditos'
        self.clear_frame()
        main = tk.Frame(self.content_area, bg='#f0f4f8')
        main.pack(fill='both', expand=True, padx=20, pady=15)

        top = tk.Frame(main, bg='#f0f4f8')
        top.pack(fill='x')
        tk.Label(top, text='Cuentas por Cobrar', font=('Segoe UI', 16, 'bold'),
                 bg='#f0f4f8', fg='#2c3e50').pack(side='left')

        conn = self.get_db()
        total_pend = conn.execute(
            'SELECT COALESCE(SUM(saldo), 0) FROM creditos WHERE estado="pendiente"').fetchone()[0]
        conn.close()
        tk.Label(top, text=f'  Total Pendiente: S/.{total_pend:.2f}',
                 font=('Segoe UI', 12, 'bold'), bg='#f0f4f8', fg='#e74c3c').pack(side='left', padx=15)

        cols = ('id', 'cliente', 'total', 'saldo', 'fecha', 'vencimiento', 'estado')
        tree = ttk.Treeview(main, columns=cols, show='headings', height=20)
        for c, h in zip(cols, ['ID', 'Cliente', 'Total', 'Saldo', 'Fecha', 'Venc.', 'Estado']):
            tree.heading(c, text=h)
            tree.column(c, width=100 if c not in ('cliente',) else 200)
        tree.pack(fill='both', expand=True, pady=10)

        def load_creditos():
            for i in tree.get_children():
                tree.delete(i)
            conn = self.get_db()
            rows = conn.execute(
                'SELECT c.id, c.id_venta, c.total, c.saldo, c.fecha_venta, c.fecha_vencimiento, c.estado, '
                'COALESCE(d.nombre, "Directo") as cliente '
                'FROM creditos c LEFT JOIN duenos d ON c.id_cliente=d.id '
                'ORDER BY c.estado, c.fecha_venta DESC').fetchall()
            conn.close()
            for r in rows:
                tree.insert('', 'end', values=(
                    r['id'], r['cliente'], f"S/.{r['total']:.2f}", f"S/.{r['saldo']:.2f}",
                    r['fecha_venta'], r['fecha_vencimiento'] or '-', r['estado']))

        load_creditos()

        def pagar_credito():
            sel = tree.selection()
            if not sel:
                messagebox.showwarning('Seleccion', 'Seleccione un credito')
                return
            vals = tree.item(sel[0], 'values')
            credito_id = int(vals[0])
            saldo_str = vals[3].replace('S/.', '')
            saldo = float(saldo_str)
            if saldo <= 0:
                messagebox.showinfo('Info', 'Credito ya pagado')
                return

            pago_dialog = tk.Toplevel(self.root)
            pago_dialog.title(f'Pagar Credito #{credito_id}')
            pago_dialog.geometry('300x180')
            pago_dialog.configure(bg='#f0f4f8')
            pago_dialog.transient(self.root)
            pago_dialog.grab_set()
            pago_dialog.update_idletasks()
            x = self.root.winfo_x() + (self.root.winfo_width() - pago_dialog.winfo_width()) // 2
            y = self.root.winfo_y() + (self.root.winfo_height() - pago_dialog.winfo_height()) // 2
            pago_dialog.geometry(f'+{x}+{y}')

            tk.Label(pago_dialog, text=f'Saldo pendiente: S/.{saldo:.2f}',
                     font=('Segoe UI', 11, 'bold'), bg='#f0f4f8', fg='#e74c3c').pack(pady=(10, 5))
            tk.Label(pago_dialog, text='Monto a pagar:', bg='#f0f4f8', font=('Segoe UI', 10)).pack()
            e_monto = tk.Entry(pago_dialog, width=25, font=('Segoe UI', 12), justify='center')
            e_monto.pack(pady=5)
            e_monto.insert(0, f'{saldo:.2f}')

            def do_pago():
                try:
                    monto = float(e_monto.get())
                except:
                    messagebox.showerror('Error', 'Monto invalido')
                    return
                if monto <= 0 or monto > saldo + 0.01:
                    messagebox.showerror('Error', 'Monto debe ser > 0 y <= saldo')
                    return
                conn = self.get_db()
                conn.execute('INSERT INTO pagos_credito (id_credito, monto, fecha) VALUES (?,?,?)',
                            (credito_id, monto, str(date.today())))
                nuevo_saldo = saldo - monto
                if nuevo_saldo <= 0.01:
                    conn.execute('UPDATE creditos SET saldo=0, estado="pagado" WHERE id=?', (credito_id,))
                else:
                    conn.execute('UPDATE creditos SET saldo=? WHERE id=?', (round(nuevo_saldo, 2), credito_id))
                conn.execute("INSERT INTO caja (fecha, tipo, concepto, monto, referencia_tipo, referencia_id) VALUES (?, 'ingreso', ?, ?, 'pago_credito', ?)",
                            (str(date.today()), f'Pago Credito #{credito_id}', monto, credito_id))
                conn.commit()
                conn.close()
                pago_dialog.destroy()
                load_creditos()
                messagebox.showinfo('Exito', f'Pago registrado: S/.{monto:.2f}')

            tk.Button(pago_dialog, text='Registrar Pago', bg='#2ecc71', fg='white',
                      font=('Segoe UI', 10, 'bold'), relief='flat', padx=20, pady=5,
                      command=do_pago, cursor='hand2').pack(pady=10)

        btn_frame = tk.Frame(main, bg='#f0f4f8')
        btn_frame.pack(fill='x')
        tk.Button(btn_frame, text='+ Registrar Pago', bg='#2ecc71', fg='white',
                  font=('Segoe UI', 10), relief='flat', padx=15, pady=4,
                  command=pagar_credito, cursor='hand2').pack(side='left', padx=2)
        tk.Button(btn_frame, text='Exportar CSVs', bg='#3498db', fg='white',
                  font=('Segoe UI', 9), relief='flat', padx=10, pady=4,
                  command=lambda: self._export_to_csv(
                      [[tree.item(i, 'values')[j] for j in range(7)] for i in tree.get_children()],
                      ['ID', 'Cliente', 'Total', 'Saldo', 'Fecha', 'Venc.', 'Estado'],
                      f'creditos_{str(date.today())}.csv'
                  ), cursor='hand2').pack(side='left', padx=2)

    def _export_to_csv(self, data, headers, filename):
        import csv, os
        path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'reportes')
        os.makedirs(path, exist_ok=True)
        filepath = os.path.join(path, filename)
        with open(filepath, 'w', newline='', encoding='utf-8-sig') as f:
            w = csv.writer(f)
            w.writerow(headers)
            w.writerows(data)
        messagebox.showinfo('Exportado', f'Reporte guardado:\n{filepath}')
        return filepath

    def _export_to_excel(self, data, headers, filename):
        import openpyxl, os
        from openpyxl.styles import Font, PatternFill
        path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'reportes')
        os.makedirs(path, exist_ok=True)
        filepath = os.path.join(path, filename)
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Reporte'
        hdr_font = Font(bold=True, color='FFFFFF')
        hdr_fill = PatternFill('solid', fgColor='2c3e50')
        for c, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=c, value=h)
            cell.font = hdr_font
            cell.fill = hdr_fill
        for r, row in enumerate(data, 2):
            for c, val in enumerate(row, 1):
                ws.cell(row=r, column=c, value=val)
        for col in ws.columns:
            max_len = max(len(str(c.value or '')) for c in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)
        wb.save(filepath)
        messagebox.showinfo('Exportado', f'Reporte guardado:\n{filepath}')
        return filepath

    def _generar_ticket_pdf(self, venta_id, items, total, metodo_str=''):
        try:
            import os
            from reportlab.lib.pagesizes import A5
            from reportlab.platypus import SimpleDocTemplate, Paragraph, Table, TableStyle, Spacer, Image
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib import colors
            from reportlab.lib.units import mm

            ticket_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'tickets')
            os.makedirs(ticket_dir, exist_ok=True)
            path = os.path.join(ticket_dir, f'ticket_{venta_id}.pdf')

            doc = SimpleDocTemplate(path, pagesize=A5,
                                    rightMargin=10*mm, leftMargin=10*mm,
                                    topMargin=10*mm, bottomMargin=10*mm)

            styles = getSampleStyleSheet()
            normal = styles['Normal']

            story = []

            # Header with logo
            logo_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'photos', 'logo.png')
            if os.path.exists(logo_path):
                logo = Image(logo_path, width=22*mm, height=22*mm)
                hdr_cells = [[logo, Paragraph(
                    'SCRIPTYFY V.02<br/><font size=9>Veterinaria y Servicios</font><br/>'
                    '<font size=8>RUC: 12345678901</font>',
                    ParagraphStyle('HdrRight', parent=normal, fontSize=12, alignment=2, leading=14))]]
            else:
                hdr_cells = [[Paragraph('SCRIPTYFY V.02', ParagraphStyle('H1', parent=normal, fontSize=14, alignment=0)),
                              Paragraph('Veterinaria<br/>RUC: 12345678901', ParagraphStyle('H2', parent=normal, fontSize=10, alignment=2, leading=13))]]
            hdr_table = Table(hdr_cells, colWidths=[28*mm, 100*mm])
            hdr_table.setStyle(TableStyle([
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('LEFTPADDING', (0, 0), (-1, -1), 2),
                ('RIGHTPADDING', (0, 0), (-1, -1), 2),
                ('TOPPADDING', (0, 0), (-1, -1), 2),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
            ]))
            story.append(hdr_table)
            story.append(Spacer(1, 2*mm))

            # Separator line
            sep = Table([['']], colWidths=[128*mm])
            sep.setStyle(TableStyle([('LINEBELOW', (0, 0), (-1, 0), 0.5, colors.HexColor('#2c3e50'))]))
            story.append(sep)
            story.append(Spacer(1, 2*mm))

            # Document info
            from datetime import date
            story.append(Paragraph(
                f'<b>BOLETA DE VENTA</b><br/>'
                f'Serie: S001-{venta_id:06d}<br/>'
                f'Fecha: {date.today()}<br/>'
                f'Pago: {metodo_str or "N/A"}',
                ParagraphStyle('DocInfo', parent=normal, fontSize=9, leading=13)))
            story.append(Spacer(1, 3*mm))

            # Items table
            avail_w = 128*mm
            table_data = [['Item', 'Cant', 'P.Unit', 'Subtotal']]
            for vals in items:
                table_data.append([vals[0], str(vals[1]), f"S/.{float(vals[2].replace('S/.','')):.2f}",
                                  f"S/.{float(vals[3].replace('S/.','')):.2f}"])
            table_data.append(['', '', 'TOTAL:', f'S/.{total:.2f}'])

            t = Table(table_data, colWidths=[avail_w*0.45, avail_w*0.13, avail_w*0.21, avail_w*0.21])
            t.setStyle(TableStyle([
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 8),
                ('ALIGN', (1, 0), (-1, -1), 'CENTER'),
                ('ALIGN', (0, 0), (0, -1), 'LEFT'),
                ('GRID', (0, 0), (-2, -2), 0.5, colors.grey),
                ('LINEBELOW', (0, -1), (-1, -1), 0.5, colors.HexColor('#2c3e50')),
                ('LINEABOVE', (0, -1), (-1, -1), 1, colors.black),
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2c3e50')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('TOPPADDING', (0, 0), (-1, -1), 4),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
                ('LEFTPADDING', (0, 0), (-1, -1), 4),
                ('RIGHTPADDING', (0, 0), (-1, -1), 4),
            ]))
            story.append(t)
            story.append(Spacer(1, 4*mm))
            story.append(Paragraph('Gracias por su preferencia!', ParagraphStyle('Footer', parent=normal, fontSize=9, alignment=1)))

            doc.build(story)
            return path
        except Exception as e:
            messagebox.showerror('Error ticket', f'No se pudo generar el ticket: {e}')
            return None

    def _marcar_para_cobro(self, animal_id, tree, detail_frame, reload_callback):
        sel = tree.selection()
        if not sel:
            messagebox.showerror("Error", "Selecciona una consulta primero")
            return
        reg_id = int(sel[0])
        conn = self.get_db()
        reg = conn.execute("SELECT * FROM registros_medicos WHERE id=?", (reg_id,)).fetchone()
        servicios = conn.execute("SELECT id, nombre, precio FROM servicios_medicos ORDER BY nombre").fetchall()
        conn.close()
        if not reg:
            messagebox.showerror("Error", "Registro no encontrado")
            return

        dialog = tk.Toplevel(self.root)
        dialog.title("Servicios para cobro")
        dialog.geometry("500x450")
        dialog.configure(bg='#f0f4f8')
        dialog.transient(self.root)
        dialog.grab_set()
        dialog.update_idletasks()
        x = self.root.winfo_x() + (self.root.winfo_width() - dialog.winfo_width()) // 2
        y = self.root.winfo_y() + (self.root.winfo_height() - dialog.winfo_height()) // 2
        dialog.geometry(f'+{x}+{y}')

        tk.Label(dialog, text="Seleccione los servicios a cobrar", font=("Segoe UI", 12, "bold"),
                 bg='#f0f4f8', fg='#2c3e50').pack(pady=(10, 5))

        container = tk.Frame(dialog, bg='white')
        container.pack(fill='both', expand=True, padx=10, pady=5)

        canvas = tk.Canvas(container, bg='white', highlightthickness=0)
        scroll = ttk.Scrollbar(container, orient='vertical', command=canvas.yview)
        scrollable = tk.Frame(canvas, bg='white')
        scrollable.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.create_window((0, 0), window=scrollable, anchor='nw')
        canvas.configure(yscrollcommand=scroll.set)
        canvas.pack(side='left', fill='both', expand=True)
        scroll.pack(side='right', fill='y')

        vars = []
        price_vars = []
        entries = []
        for s in servicios:
            row = tk.Frame(scrollable, bg='white')
            row.pack(fill='x', pady=2, padx=5)
            var = tk.BooleanVar(value=(s['nombre'] == 'Consulta General'))
            vars.append(var)
            cb = tk.Checkbutton(row, text=s['nombre'], variable=var, bg='white',
                                font=("Segoe UI", 10), anchor='w')
            cb.pack(side='left', fill='x', expand=True)
            tk.Label(row, text="S/.", bg='white', font=("Segoe UI", 10)).pack(side='left')
            pv = tk.StringVar(value=f"{s['precio']:.2f}")
            price_vars.append(pv)
            ent = tk.Entry(row, width=8, font=("Segoe UI", 11, "bold"), justify='center',
                           textvariable=pv, relief='solid', bd=1)
            ent.pack(side='left', padx=2)
            entries.append(ent)

        total_frame = tk.Frame(dialog, bg='#f0f4f8')
        total_frame.pack(fill='x', padx=10, pady=5)
        tk.Label(total_frame, text="Total:", font=("Segoe UI", 14, "bold"),
                 bg='#f0f4f8', fg='#2c3e50').pack(side='left')
        total_var = tk.StringVar(value="S/.0.00")
        tk.Label(total_frame, textvariable=total_var, font=("Segoe UI", 14, "bold"),
                 bg='#f0f4f8', fg='#e74c3c').pack(side='right')

        def update_total(*args):
            tot = 0.0
            for i in range(len(servicios)):
                if vars[i].get():
                    try:
                        tot += float(price_vars[i].get() or '0')
                    except:
                        pass
            total_var.set(f"S/.{tot:.2f}")

        for pv in price_vars:
            pv.trace_add('write', update_total)
        for v in vars:
            v.trace_add('write', update_total)
        update_total()

        btn_frame = tk.Frame(dialog, bg='#f0f4f8')
        btn_frame.pack(fill='x', padx=10, pady=10)

        def save():
            selected_ids = []
            total = 0.0
            for i, s in enumerate(servicios):
                if vars[i].get():
                    try:
                        p = float(price_vars[i].get() or '0')
                    except:
                        p = 0
                    if p > 0:
                        selected_ids.append({"id": s["id"], "precio": p, "nombre": s["nombre"]})
                        total += p
            if not selected_ids:
                messagebox.showwarning("Error", "Seleccione al menos un servicio")
                return

            conn = self.get_db()
            conn.execute("UPDATE registros_medicos SET pendiente_cobro=1, facturado=0 WHERE id=?", (reg_id,))
            conn.execute(
                "INSERT INTO cobros_pendientes (id_registro, id_animal, servicios, total, created_by, fecha) VALUES (?,?,?,?,?,?)",
                (reg_id, animal_id, json.dumps(selected_ids), total, self._user_id, str(date.today())))
            conn.commit()
            conn.close()
            dialog.destroy()
            if reload_callback:
                reload_callback()
            messagebox.showinfo("Exito", "Consulta marcada para cobro")

        tk.Button(btn_frame, text="Guardar", bg='#2ecc71', fg='white',
                  font=("Segoe UI", 11, "bold"), relief='flat', padx=30, pady=5,
                  command=save, cursor='hand2').pack()

    def show_cobros_pendientes(self):
        self._current_view = 'cobros'
        self.clear_frame()
        main = tk.Frame(self.content_area, bg='#f0f4f8')
        main.pack(fill='both', expand=True, padx=20, pady=15)

        conn = self.get_db()
        rows = conn.execute("""
            SELECT cp.*, a.nombre as animal_nombre, d.nombre as dueno_nombre,
                   r.fecha as consulta_fecha, r.doctor, r.diagnostico,
                   u.nombre as doctor_nombre
            FROM cobros_pendientes cp
            JOIN registros_medicos r ON cp.id_registro = r.id
            JOIN animales a ON cp.id_animal = a.id
            JOIN duenos d ON a.id_dueno = d.id
            LEFT JOIN usuarios u ON cp.created_by = u.id
            ORDER BY cp.fecha DESC
        """).fetchall()
        conn.close()

        top_bar = tk.Frame(main, bg='#f0f4f8')
        top_bar.pack(fill='x')
        tk.Label(top_bar, text=f"💰 Cobros Pendientes ({len(rows)})",
                 font=("Segoe UI", 16, "bold"), bg='#f0f4f8', fg='#2c3e50').pack(side='left')

        if not rows:
            tk.Label(main, text="No hay cobros pendientes", font=("Segoe UI", 14),
                     bg='#f0f4f8', fg='#999').pack(expand=True, pady=40)
            return

        cols = ('fecha', 'paciente', 'dueno', 'doctor', 'diagnostico', 'total')
        tree = ttk.Treeview(main, columns=cols, show='headings', height=20)
        for c, h, w in [('fecha', 'Fecha', 100), ('paciente', 'Paciente', 140),
                        ('dueno', 'Dueño', 160), ('doctor', 'Doctor', 140),
                        ('diagnostico', 'Diagnóstico', 250), ('total', 'Total', 80)]:
            tree.heading(c, text=h)
            tree.column(c, width=w)
        tree.pack(fill='both', expand=True, pady=10)

        for r in rows:
            diag = r['diagnostico'] or ''
            tree.insert('', 'end', iid=str(r['id']), values=(
                r['consulta_fecha'], r['animal_nombre'], r['dueno_nombre'],
                r['doctor_nombre'] or r['doctor'] or '', diag[:60],
                f"S/.{r['total']:.2f}"))

        btn_frame = tk.Frame(main, bg='#f0f4f8')
        btn_frame.pack(fill='x')

        tk.Button(btn_frame, text="💵 Cobrar", bg='#2ecc71', fg='white',
                  font=("Segoe UI", 11, "bold"), relief='flat', padx=25, pady=5,
                  command=lambda: self._cobrar_dialog(tree, rows),
                  cursor='hand2').pack(side='left', padx=2)

        tk.Button(btn_frame, text="🔄 Actualizar", bg='#3498db', fg='white',
                  font=("Segoe UI", 9), relief='flat', padx=15, pady=4,
                  command=self.show_cobros_pendientes,
                  cursor='hand2').pack(side='left', padx=2)

    def _cobrar_dialog(self, tree, rows):
        sel = tree.selection()
        if not sel:
            messagebox.showwarning("Seleccion", "Seleccione un cobro pendiente")
            return

        cp_id = int(sel[0])
        cp = None
        for r in rows:
            if r['id'] == cp_id:
                cp = r
                break
        if not cp:
            messagebox.showerror("Error", "Registro no encontrado")
            return

        servicios = json.loads(cp['servicios']) if isinstance(cp['servicios'], str) else cp['servicios']

        dialog = tk.Toplevel(self.root)
        dialog.title("Cobrar - Agregar productos/servicios extra")
        dialog.geometry("720x680")
        dialog.configure(bg='#f0f4f8')
        dialog.transient(self.root)
        dialog.grab_set()
        dialog.update_idletasks()
        x = self.root.winfo_x() + (self.root.winfo_width() - dialog.winfo_width()) // 2
        y = self.root.winfo_y() + (self.root.winfo_height() - dialog.winfo_height()) // 2
        dialog.geometry(f'+{x}+{y}')

        # Scrollable main area
        canvas = tk.Canvas(dialog, bg='#f0f4f8', highlightthickness=0)
        scroll = ttk.Scrollbar(dialog, orient='vertical', command=canvas.yview)
        scroll.pack(side='right', fill='y')
        canvas.pack(side='left', fill='both', expand=True)
        scrollable = tk.Frame(canvas, bg='#f0f4f8')
        scrollable.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.create_window((0, 0), window=scrollable, anchor='nw')
        canvas.configure(yscrollcommand=scroll.set)

        conn = self.get_db()

        # === INFO ===
        info_frame = tk.Frame(scrollable, bg='white', padx=14, pady=8,
                              highlightbackground='#ddd', highlightthickness=1)
        info_frame.pack(fill='x', padx=14, pady=(12, 4))
        dueno_info = conn.execute("SELECT id_dueno FROM animales WHERE id=?", (cp['id_animal'],)).fetchone()
        id_dueno = dueno_info['id_dueno'] if dueno_info else None
        dueno_row = conn.execute("SELECT nombre, dni FROM duenos WHERE id=?", (id_dueno,)).fetchone() if id_dueno else None
        dueno_nombre = dueno_row['nombre'] if dueno_row else cp['dueno_nombre']
        info_text = (
            f"Paciente: {cp['animal_nombre']}    Dueño: {dueno_nombre}\n"
            f"Doctor: {cp['doctor_nombre'] or cp['doctor'] or 'N/A'}\n"
            f"Fecha: {cp['consulta_fecha']}    Diagnóstico: {cp['diagnostico'] or 'N/A'}"
        )
        tk.Label(info_frame, text=info_text, font=("Segoe UI", 10),
                 bg='white', fg='#2c3e50', justify='left').pack(anchor='w')

        # Other pets of same owner
        otros = []
        if id_dueno:
            otros = conn.execute("SELECT id, nombre FROM animales WHERE id_dueno=? AND id!=?",
                                  (id_dueno, cp['id_animal'])).fetchall()
        conn.close()

        # === ITEMS LIST (all items to be charged) ===
        all_items = []  # list of {type, name, qty, price, ref_id, removable}
        for s in servicios:
            all_items.append({"type": "consulta", "name": s["nombre"], "qty": 1,
                              "price": s["precio"], "ref_id": s.get("id"), "removable": False})

        extra_items = []  # items added in this dialog

        items_frame = tk.LabelFrame(scrollable, text=" Items a cobrar ", font=("Segoe UI", 10, "bold"),
                                    bg='white', fg='#2c3e50', padx=8, pady=4)
        items_frame.pack(fill='x', padx=14, pady=(4, 0))

        def rebuild_items_list():
            for w in items_frame.winfo_children():
                w.destroy()
            all_items.clear()
            for s in servicios:
                all_items.append({"type": "consulta", "name": s["nombre"], "qty": 1,
                                  "price": s["precio"], "ref_id": s.get("id"), "removable": False})
            all_items.extend(extra_items)
            tot = 0.0
            for it in all_items:
                row = tk.Frame(items_frame, bg='white')
                row.pack(fill='x', pady=1)
                subt = it["qty"] * it["price"]
                txt = f"  {it['name']}"
                if it["qty"] > 1:
                    txt += f" x{it['qty']}"
                txt += f"   S/.{subt:.2f}"
                tk.Label(row, text=txt, font=("Segoe UI", 10), bg='white',
                         fg='#333' if it["type"] != "consulta" else '#2c3e50').pack(side='left')
                if it.get("removable"):
                    tk.Button(row, text="✕", font=("Segoe UI", 8), bg='#e74c3c', fg='white',
                              relief='flat', padx=4, pady=0, cursor='hand2',
                              command=lambda item=it: remove_item(item)).pack(side='right')
                tot += subt
            tk.Frame(items_frame, bg='#ddd', height=1).pack(fill='x', pady=4)
            total_label = tk.Label(items_frame, text=f"Total: S/.{tot:.2f}",
                                   font=("Segoe UI", 12, "bold"), bg='white', fg='#e74c3c')
            total_label.pack(anchor='e')
            # Update payment entries
            for val, kv in pago_vars.items():
                if kv.get():
                    checked = [k for k, v in pago_vars.items() if v.get()]
                    if len(checked) == 1:
                        pago_entries[val].delete(0, 'end')
                        pago_entries[val].insert(0, f'{tot:.2f}')

        def remove_item(item):
            if item in extra_items:
                extra_items.remove(item)
            rebuild_items_list()

        rebuild_items_list()

        # === ADD PRODUCTS ===
        add_frame = tk.LabelFrame(scrollable, text=" Agregar productos / servicios ", font=("Segoe UI", 10, "bold"),
                                  bg='white', fg='#2c3e50', padx=8, pady=4)
        add_frame.pack(fill='x', padx=14, pady=(4, 0))

        # Productos
        conn2 = self.get_db()
        prod_rows = conn2.execute("SELECT id, nombre, precio_venta, stock FROM productos WHERE stock>0 ORDER BY nombre").fetchall()
        conn2.close()
        tk.Label(add_frame, text="Producto:", font=("Segoe UI", 9), bg='white').pack(anchor='w')
        prod_row = tk.Frame(add_frame, bg='white')
        prod_row.pack(fill='x')
        prod_var = tk.StringVar()
        prod_combo = ttk.Combobox(prod_row, textvariable=prod_var, font=("Segoe UI", 10),
                                  values=[f"{p['id']}: {p['nombre']} S/.{p['precio_venta']:.2f} (stock:{p['stock']})" for p in prod_rows],
                                  width=50)
        prod_combo.pack(side='left', padx=(0, 4))
        tk.Label(prod_row, text="Cant:", bg='white', font=("Segoe UI", 9)).pack(side='left')
        qty_var = tk.StringVar(value="1")
        qty_spin = tk.Spinbox(prod_row, from_=1, to=99, width=4, font=("Segoe UI", 10),
                              textvariable=qty_var)
        qty_spin.pack(side='left', padx=2)

        def add_producto():
            txt = prod_var.get().strip()
            if not txt or ':' not in txt:
                messagebox.showwarning("Seleccion", "Seleccione un producto")
                return
            try:
                pid = int(txt.split(':')[0].strip())
            except:
                messagebox.showwarning("Error", "Producto invalido")
                return
            qty = int(qty_var.get() or '1')
            conn3 = self.get_db()
            p = conn3.execute("SELECT nombre, precio_venta, stock FROM productos WHERE id=?", (pid,)).fetchone()
            conn3.close()
            if not p:
                return
            if p['stock'] < qty:
                messagebox.showwarning("Stock", f"Stock insuficiente: {p['stock']}")
                return
            extra_items.append({"type": "producto", "name": p['nombre'], "qty": qty,
                                "price": p['precio_venta'], "ref_id": pid, "removable": True, "stock_item": True})
            prod_var.set('')
            rebuild_items_list()

        tk.Button(prod_row, text="+Agregar", bg='#27ae60', fg='white',
                  font=("Segoe UI", 9), relief='flat', padx=8, pady=2,
                  command=add_producto, cursor='hand2').pack(side='left', padx=4)

        # Servicios Médicos
        conn3 = self.get_db()
        sv_rows = conn3.execute("SELECT id, nombre, precio FROM servicios_medicos ORDER BY nombre").fetchall()
        conn3.close()
        tk.Label(add_frame, text="Servicio veterinario:", font=("Segoe UI", 9), bg='white').pack(anchor='w')
        sv_row = tk.Frame(add_frame, bg='white')
        sv_row.pack(fill='x')
        sv_var = tk.StringVar()
        sv_combo = ttk.Combobox(sv_row, textvariable=sv_var, font=("Segoe UI", 10),
                                values=[f"{s['id']}: {s['nombre']} S/.{s['precio']:.2f}" for s in sv_rows],
                                width=50)
        sv_combo.pack(side='left', padx=(0, 4))

        def add_servicio():
            txt = sv_var.get().strip()
            if not txt or ':' not in txt:
                messagebox.showwarning("Seleccion", "Seleccione un servicio")
                return
            try:
                sid = int(txt.split(':')[0].strip())
            except:
                messagebox.showwarning("Error", "Servicio invalido")
                return
            conn4 = self.get_db()
            s = conn4.execute("SELECT nombre, precio FROM servicios_medicos WHERE id=?", (sid,)).fetchone()
            conn4.close()
            if not s:
                return
            extra_items.append({"type": "servicio", "name": s['nombre'] + " (extra)", "qty": 1,
                                "price": s['precio'], "ref_id": sid, "removable": True})
            sv_var.set('')
            rebuild_items_list()

        tk.Button(sv_row, text="+Agregar", bg='#2980b9', fg='white',
                  font=("Segoe UI", 9), relief='flat', padx=8, pady=2,
                  command=add_servicio, cursor='hand2').pack(side='left', padx=4)

        # Other pets of same owner (for grooming, etc)
        if otros:
            tk.Label(add_frame, text="Servicio para otra mascota del mismo dueño:", font=("Segoe UI", 9), bg='white').pack(anchor='w')
            otro_row = tk.Frame(add_frame, bg='white')
            otro_row.pack(fill='x')
            tk.Label(otro_row, text="Mascota:", bg='white', font=("Segoe UI", 9)).pack(side='left')
            otro_var = tk.StringVar()
            otro_combo = ttk.Combobox(otro_row, textvariable=otro_var, font=("Segoe UI", 10),
                                      values=[f"{o['id']}: {o['nombre']}" for o in otros], width=20)
            otro_combo.pack(side='left', padx=2)
            tk.Label(otro_row, text="Servicio:", bg='white', font=("Segoe UI", 9)).pack(side='left')
            otro_sv_var = tk.StringVar()
            otro_sv_combo = ttk.Combobox(otro_row, textvariable=otro_sv_var, font=("Segoe UI", 10),
                                         values=[f"{s['id']}: {s['nombre']} S/.{s['precio']:.2f}" for s in sv_rows],
                                         width=25)
            otro_sv_combo.pack(side='left', padx=2)

            def add_otro():
                mascota_txt = otro_var.get().strip()
                svc_txt = otro_sv_var.get().strip()
                if not mascota_txt or ':' not in mascota_txt:
                    messagebox.showwarning("Error", "Seleccione mascota")
                    return
                if not svc_txt or ':' not in svc_txt:
                    messagebox.showwarning("Error", "Seleccione servicio")
                    return
                mascota_nombre = mascota_txt.split(':', 1)[1].strip()
                svc_id = int(svc_txt.split(':')[0].strip())
                conn5 = self.get_db()
                s = conn5.execute("SELECT nombre, precio FROM servicios_medicos WHERE id=?", (svc_id,)).fetchone()
                conn5.close()
                if not s:
                    return
                extra_items.append({"type": "otro", "name": f"{mascota_nombre}: {s['nombre']}", "qty": 1,
                                    "price": s['precio'], "ref_id": svc_id, "removable": True})
                otro_var.set('')
                otro_sv_var.set('')
                rebuild_items_list()

            tk.Button(otro_row, text="+Agregar", bg='#8e44ad', fg='white',
                      font=("Segoe UI", 9), relief='flat', padx=8, pady=2,
                      command=add_otro, cursor='hand2').pack(side='left', padx=4)

        # Custom item
        tk.Label(add_frame, text="Item personalizado:", font=("Segoe UI", 9), bg='white').pack(anchor='w')
        cust_row = tk.Frame(add_frame, bg='white')
        cust_row.pack(fill='x')
        tk.Label(cust_row, text="Nombre:", bg='white', font=("Segoe UI", 9)).pack(side='left')
        cust_name = tk.Entry(cust_row, width=25, font=("Segoe UI", 10))
        cust_name.pack(side='left', padx=2)
        tk.Label(cust_row, text="Precio:", bg='white', font=("Segoe UI", 9)).pack(side='left')
        cust_price = tk.Entry(cust_row, width=8, font=("Segoe UI", 10))
        cust_price.pack(side='left', padx=2)

        def add_custom():
            name = cust_name.get().strip()
            price_txt = cust_price.get().strip()
            if not name or not price_txt:
                messagebox.showwarning("Error", "Ingrese nombre y precio")
                return
            try:
                price = float(price_txt)
            except:
                messagebox.showwarning("Error", "Precio invalido")
                return
            extra_items.append({"type": "custom", "name": name, "qty": 1,
                                "price": price, "ref_id": None, "removable": True})
            cust_name.delete(0, 'end')
            cust_price.delete(0, 'end')
            rebuild_items_list()

        tk.Button(cust_row, text="+Agregar", bg='#f39c12', fg='white',
                  font=("Segoe UI", 9), relief='flat', padx=8, pady=2,
                  command=add_custom, cursor='hand2').pack(side='left', padx=4)

        # === PAYMENT METHOD ===
        met_frame = tk.LabelFrame(scrollable, text=" Metodo de Pago ", font=("Segoe UI", 10, "bold"),
                                  bg='white', fg='#2c3e50', padx=8, pady=4)
        met_frame.pack(fill='x', padx=14, pady=(4, 0))

        pago_vars = {}
        pago_entries = {}
        for val, txt in [('efectivo', 'Efectivo'), ('yape', 'Yape'), ('transferencia', 'Transferencia')]:
            row = tk.Frame(met_frame, bg='white')
            row.pack(fill='x', pady=2)
            var = tk.BooleanVar(value=False)
            pago_vars[val] = var
            cb = tk.Checkbutton(row, text=txt, variable=var, bg='white', font=("Segoe UI", 10))
            cb.pack(side='left', padx=(0, 6))
            ent = tk.Entry(row, width=14, font=("Segoe UI", 11, "bold"), justify='center', relief='solid', bd=1)
            ent.pack(side='left')
            ent.config(state='disabled')
            pago_entries[val] = ent
            def toggle_entry(*args, e=ent, v=var):
                e.config(state='normal' if v.get() else 'disabled')
                if not v.get():
                    e.delete(0, 'end')
                checked = [kv for kv, kvv in pago_vars.items() if kvv.get()]
                if len(checked) == 1:
                    tot = sum(it["qty"] * it["price"] for it in all_items)
                    for kv in checked:
                        pago_entries[kv].delete(0, 'end')
                        pago_entries[kv].insert(0, f'{tot:.2f}')
            var.trace_add('write', toggle_entry)

        def confirm():
            selected = []
            suma = 0.0
            for val, txt in [('efectivo', 'Efectivo'), ('yape', 'Yape'), ('transferencia', 'Transferencia')]:
                if pago_vars[val].get():
                    try:
                        amt = float(pago_entries[val].get() or '0')
                    except:
                        amt = 0
                    if amt > 0:
                        selected.append((val, txt, amt))
                        suma += amt
            if not selected:
                messagebox.showwarning("Error", "Seleccione al menos un metodo de pago")
                return
            total_final = sum(it["qty"] * it["price"] for it in all_items)
            if abs(suma - total_final) > 0.01:
                messagebox.showwarning("Error", f"La suma de pagos (S/.{suma:.2f}) no coincide con el total (S/.{total_final:.2f})")
                return

            conn = self.get_db()
            id_cliente = id_dueno
            metodo_str = ', '.join(f'{t}:{m:.2f}' for _, t, m in selected)
            cursor = conn.execute(
                'INSERT INTO ventas (id_cliente, id_animal, fecha, total, tipo, tipo_comprobante, tipo_pago, metodo_pago) VALUES (?,?,?,?,?,?,?,?)',
                (id_cliente, cp['id_animal'], str(date.today()), total_final, metodo_str,
                 'Boleta', metodo_str, metodo_str))
            venta_id = cursor.lastrowid

            for it in all_items:
                tipo_item = it["type"]
                ref_id = it.get("ref_id")
                if tipo_item == "consulta":
                    tipo_item = "servicio_medico"
                elif tipo_item == "producto":
                    tipo_item = "producto"
                elif tipo_item in ("servicio", "otro"):
                    tipo_item = "servicio_medico"
                else:
                    tipo_item = "servicio_medico"
                conn.execute(
                    'INSERT INTO venta_items (id_venta, tipo_item, referencia_id, nombre, cantidad, precio_unitario, subtotal) VALUES (?,?,?,?,?,?,?)',
                    (venta_id, tipo_item, ref_id, it["name"], it["qty"], it["price"], it["qty"] * it["price"]))
                # Update stock for products
                if tipo_item == "producto" and ref_id:
                    conn.execute("UPDATE productos SET stock = stock - ? WHERE id=?", (it["qty"], ref_id))

            for val, txt, monto in selected:
                conn.execute("INSERT INTO caja (fecha, tipo, concepto, monto, referencia_tipo, referencia_id) VALUES (?, 'ingreso', ?, ?, 'venta', ?)",
                            (str(date.today()), f'Venta #{venta_id} ({txt})', monto, venta_id))

            conn.execute("UPDATE registros_medicos SET facturado=1, cobrado_por=?, id_venta=? WHERE id=?",
                        (self._user_id, venta_id, cp['id_registro']))
            conn.execute("DELETE FROM cobros_pendientes WHERE id=?", (cp_id,))
            conn.commit()
            conn.close()

            dialog.destroy()

            ticket_items = [[it["name"], str(it["qty"]), f"S/.{it['price']:.2f}", f"S/.{it['qty'] * it['price']:.2f}"] for it in all_items]
            pdf_path = self._generar_ticket_pdf(venta_id, ticket_items, total_final, metodo_str)
            if pdf_path:
                if messagebox.askyesno('PDF', 'Abrir ticket?'):
                    os.startfile(pdf_path)

            messagebox.showinfo('Exito', f'Cobro registrado. Venta #{venta_id}\nTotal: S/.{total_final:.2f}')
            self.show_cobros_pendientes()

        btn_frame = tk.Frame(scrollable, bg='#f0f4f8')
        btn_frame.pack(fill='x', padx=14, pady=(10, 14))
        tk.Button(btn_frame, text="CONFIRMAR COBRO", bg='#2ecc71', fg='white',
                  font=("Segoe UI", 11, "bold"), relief='flat', padx=25, pady=8,
                  command=confirm, cursor='hand2').pack(fill='x')

    def _global_search(self, query):
        query = query.strip()
        if not query:
            messagebox.showinfo('Buscar', 'Ingrese un termino de busqueda')
            return
        conn = self.get_db()
        duenos = conn.execute(
            'SELECT id, nombre, dni FROM duenos WHERE nombre LIKE ? OR dni LIKE ? LIMIT 10',
            (f'%{query}%', f'%{query}%')).fetchall()
        animales = conn.execute(
            'SELECT a.id, a.nombre, a.especie, d.nombre FROM animales a '
            'JOIN duenos d ON a.id_dueno=d.id '
            'WHERE a.nombre LIKE ? OR a.raza LIKE ? LIMIT 10',
            (f'%{query}%', f'%{query}%')).fetchall()
        productos = conn.execute(
            'SELECT id, nombre, precio_venta, stock FROM productos WHERE nombre LIKE ? LIMIT 10',
            (f'%{query}%',)).fetchall()
        conn.close()

        if not duenos and not animales and not productos:
            messagebox.showinfo('Buscar', f'Sin resultados para: {query}')
            return

        win = tk.Toplevel(self.root)
        win.title(f'Resultados: {query}')
        win.geometry('600x400')
        win.configure(bg='#f0f4f8')
        win.transient(self.root)
        win.grab_set()
        win.update_idletasks()
        x = self.root.winfo_x() + (self.root.winfo_width() - win.winfo_width()) // 2
        y = self.root.winfo_y() + (self.root.winfo_height() - win.winfo_height()) // 2
        win.geometry(f'+{x}+{y}')

        row = 0
        if duenos:
            tk.Label(win, text='Duennos', font=('Segoe UI', 11, 'bold'),
                     bg='#f0f4f8', fg='#2c3e50', padx=8, pady=3).grid(row=row, column=0, columnspan=3, sticky='w')
            row += 1
            for d in duenos:
                tk.Button(win, text=f"{d['nombre']} (DNI: {d['dni'] or 'N/A'})",
                          font=('Segoe UI', 9), bg='white', relief='flat', anchor='w', padx=10,
                          cursor='hand2',
                          command=lambda did=d['id']: (win.destroy(), self.show_owner_detail(did))
                         ).grid(row=row, column=0, columnspan=3, sticky='ew', padx=10, pady=1)
                row += 1

        if animales:
            tk.Label(win, text='Mascotas', font=('Segoe UI', 11, 'bold'),
                     bg='#f0f4f8', fg='#2c3e50', padx=8, pady=3).grid(row=row, column=0, columnspan=3, sticky='w')
            row += 1
            for a in animales:
                tk.Button(win, text=f"{a['nombre']} ({a['especie']}) - Duenio: {a['nombre']}",
                          font=('Segoe UI', 9), bg='white', relief='flat', anchor='w', padx=10,
                          cursor='hand2',
                          command=lambda aid=a['id']: (win.destroy(), self.show_animal_detail(aid))
                         ).grid(row=row, column=0, columnspan=3, sticky='ew', padx=10, pady=1)
                row += 1

        if productos:
            tk.Label(win, text='Productos', font=('Segoe UI', 11, 'bold'),
                     bg='#f0f4f8', fg='#2c3e50', padx=8, pady=3).grid(row=row, column=0, columnspan=3, sticky='w')
            row += 1
            for p in productos:
                tk.Label(win, text=f"{p['nombre']} - S/.{p['precio_venta']:.2f} (Stock: {p['stock']})",
                         font=('Segoe UI', 9), bg='white', fg='#555', anchor='w', padx=10
                        ).grid(row=row, column=0, columnspan=3, sticky='ew', padx=10, pady=1)
                row += 1

        tk.Button(win, text='Cerrar', bg='#e74c3c', fg='white', relief='flat', padx=20,
                  command=win.destroy, cursor='hand2').grid(row=row, column=0, columnspan=3, pady=10)

    def _get_all_products(self):
        conn = self.get_db()
        rows = conn.execute('SELECT nombre FROM productos WHERE activo=1').fetchall()
        conn.close()
        return [dict(r) for r in rows]

    def manage_usuarios_dialog(self):
        if self._user_role != 'admin':
            messagebox.showerror('Acceso denegado', 'Solo administradores pueden gestionar usuarios')
            return
        dialog = tk.Toplevel(self.root)
        dialog.title('Gestion de Usuarios')
        dialog.geometry('600x400')
        dialog.configure(bg='#f0f4f8')
        dialog.transient(self.root)
        dialog.grab_set()
        dialog.update_idletasks()
        x = self.root.winfo_x() + (self.root.winfo_width() - dialog.winfo_width()) // 2
        y = self.root.winfo_y() + (self.root.winfo_height() - dialog.winfo_height()) // 2
        dialog.geometry(f'+{x}+{y}')

        cols = ('id', 'user', 'nombre', 'rol', 'activo')
        tree = ttk.Treeview(dialog, columns=cols, show='headings', height=15)
        for c, h in zip(cols, ['ID', 'Usuario', 'Nombre', 'Rol', 'Activo']):
            tree.heading(c, text=h)
            tree.column(c, width=100 if c != 'nombre' else 200)
        tree.pack(fill='both', expand=True, padx=10, pady=10)

        def load_usuarios():
            for i in tree.get_children():
                tree.delete(i)
            conn = self.get_db()
            rows = conn.execute('SELECT id, username, nombre, rol, activo FROM usuarios ORDER BY id').fetchall()
            conn.close()
            for r in rows:
                tree.insert('', 'end', values=(r['id'], r['username'], r['nombre'], r['rol'],
                    'Si' if r['activo'] else 'No'))

        load_usuarios()

        def add_usuario():
            if self._user_role != 'admin':
                return
            sub = tk.Toplevel(dialog)
            sub.title('Nuevo Usuario')
            sub.geometry('350x250')
            sub.configure(bg='#f0f4f8')
            sub.transient(dialog)
            sub.grab_set()
            fields = {}
            for i, (lbl, key) in enumerate([('Usuario', 'username'), ('Contraseña', 'password'),
                                            ('Nombre', 'nombre'), ('Rol', 'rol')]):
                tk.Label(sub, text=lbl, bg='#f0f4f8', font=('Segoe UI', 10)).grid(row=i, column=0, sticky='e', padx=8, pady=3)
                if key == 'rol':
                    e = ttk.Combobox(sub, values=['admin', 'doctor', 'recepcion', 'empleado'], width=27)
                else:
                    e = tk.Entry(sub, width=30, font=('Segoe UI', 10))
                    if key == 'password':
                        e.config(show='*')
                e.grid(row=i, column=1, padx=8, pady=3)
                fields[key] = e
            def save():
                conn = self.get_db()
                try:
                    conn.execute('INSERT INTO usuarios (username, password, nombre, rol) VALUES (?,?,?,?)',
                                (fields['username'].get(), fields['password'].get(),
                                 fields['nombre'].get(), fields['rol'].get()))
                    conn.commit()
                    sub.destroy()
                    load_usuarios()
                except Exception as ex:
                    messagebox.showerror('Error', str(ex))
                finally:
                    conn.close()
            tk.Button(sub, text='Guardar', bg='#2ecc71', fg='white',
                      font=('Segoe UI', 10), relief='flat', padx=15,
                      command=save, cursor='hand2').grid(row=4, column=1, pady=15, sticky='e')

        btn_frame = tk.Frame(dialog, bg='#f0f4f8')
        btn_frame.pack(fill='x', padx=10, pady=(0, 10))
        tk.Button(btn_frame, text='+ Nuevo Usuario', bg='#3498db', fg='white',
                  font=('Segoe UI', 10), relief='flat', padx=15, pady=3,
                  command=add_usuario, cursor='hand2').pack(side='left', padx=2)

    def _is_admin(self):
        return self._user_role == 'admin'

if __name__ == "__main__":
    root = tk.Tk()
    app = App(root)
    root.mainloop()
