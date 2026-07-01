import os, sys, json, uuid, logging, csv, io, urllib.parse, hashlib, time, math, re
from datetime import date, datetime, timedelta
from functools import wraps
import requests
import pdfkit, fitz
from flask import Flask, render_template, request, redirect, session, url_for, send_from_directory, Response, abort, jsonify, flash

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
import database
import storage
import sunat_api

database.init_db()

app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY', os.urandom(32).hex())
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024
app.config['PERMANENT_SESSION_LIFETIME'] = timedelta(hours=8)
app.config['SESSION_PERMANENT'] = True
app.jinja_env.auto_reload = True
app.config['TEMPLATES_AUTO_RELOAD'] = True

# Load AI config
AI_CONFIG = {}
_cfg_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "config.json")
if os.path.exists(_cfg_path):
    try:
        with open(_cfg_path) as _f:
            AI_CONFIG = json.load(_f)
    except: pass

# In-memory rate limiting
_login_attempts = {}

def check_login_rate_limit(key):
    return True

def hash_pwd(pwd):
    return hashlib.sha256(pwd.encode()).hexdigest()

def verify_pwd(stored, pwd):
    hashed = hash_pwd(pwd)
    if len(stored) == 64:
        return stored == hashed
    return stored == pwd

@app.errorhandler(404)
def not_found(e):
    return render_template("admin_error.html", code=404, message="P\u00e1gina no encontrada"), 404

@app.errorhandler(500)
def server_error(e):
    logger.exception("Error interno")
    return render_template("admin_error.html", code=500, message="Error interno del servidor"), 500

@app.after_request
def no_cache(resp):
    resp.headers["Cache-Control"] = "no-cache, no-store, must-revalidate"
    resp.headers["Pragma"] = "no-cache"
    resp.headers["Expires"] = "0"
    return resp

def login_required(f):
    from functools import wraps
    @wraps(f)
    def wrapper(*args, **kwargs):
        if "doctor" not in session:
            return redirect(url_for("index"))
        return f(*args, **kwargs)
    return wrapper

def role_required(*roles):
    def decorator(f):
        @wraps(f)
        def wrapper(*args, **kwargs):
            if "doctor" not in session:
                return redirect(url_for("index"))
            if session.get("rol") not in roles:
                flash("No tienes permiso para acceder a esta sección", "danger")
                return redirect(url_for("dashboard"))
            return f(*args, **kwargs)
        return wrapper
    return decorator

def _ctx(**kw):
    return dict(doctor=session.get("doctor",""), rol=session.get("rol",""),
                hoy=str(date.today()), **kw)

def _excel_file(title, columns, data, keys, subtitles=None, col_widths=None, tab_color="1abc9c"):
    import openpyxl, io
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, Color
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = title[:31]
    ws.sheet_properties.tabColor = tab_color
    hf = Font(bold=True, color="FFFFFF", size=11)
    hfill = PatternFill(start_color=tab_color, end_color=tab_color, fill_type="solid")
    ha = Alignment(horizontal="center", vertical="center")
    tbc = Color(rgb="CCCCCC")
    tb = Border(left=Side(style="thin",color=tbc),right=Side(style="thin",color=tbc),top=Side(style="thin",color=tbc),bottom=Side(style="thin",color=tbc))
    ncol = len(columns)
    row = 1
    if title:
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncol)
        c = ws.cell(row=1, column=1, value=title)
        c.font = Font(bold=True, size=14, color=tab_color); c.alignment = ha
        ws.row_dimensions[1].height = 30
        row = 2
        if subtitles:
            for end_col, text in subtitles:
                ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=end_col)
                ws.cell(row=2, column=1, value=text).font = Font(bold=False, color="555555", size=11)
            row = 3
    hr = row + 1
    for ci, cn in enumerate(columns, 1):
        c = ws.cell(row=hr, column=ci, value=cn); c.font = hf; c.fill = hfill; c.alignment = ha; c.border = tb
    ws.row_dimensions[hr].height = 22
    for ri, r in enumerate(data, hr + 1):
        rd = dict(r) if not isinstance(r, dict) else r
        for ci, key in enumerate(keys, 1):
            v = rd.get(key, "")
            c = ws.cell(row=ri, column=ci, value=v); c.border = tb; c.font = Font(size=10)
            if isinstance(v, float):
                c.number_format = "#,##0.00"
    if col_widths:
        for letter, w in col_widths.items():
            ws.column_dimensions[letter].width = w
    else:
        for ci in range(1, ncol + 1):
            letter = chr(64 + ci)
            ws.column_dimensions[letter].width = max(len(str(columns[ci-1])) + 4, 14)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()

@app.route('/photos/<filename>')
def serve_photo(filename):
    if storage.S3_ENABLED:
        return redirect(storage.get_url(filename, 'photos'))
    return send_from_directory(storage.LOCAL_DIRS['photos'], filename)

@app.route('/examenes/<filename>')
def serve_examen(filename):
    if storage.S3_ENABLED:
        return redirect(storage.get_url(filename, 'examenes'))
    return send_from_directory(storage.LOCAL_DIRS['examenes'], filename)

@app.route('/productos_photos/<filename>')
def serve_producto_photo(filename):
    if storage.S3_ENABLED:
        return redirect(storage.get_url(filename, 'productos'))
    return send_from_directory(storage.LOCAL_DIRS['productos'], filename)

@app.route('/comprobantes/<filename>')
def serve_comprobante(filename):
    if storage.S3_ENABLED:
        return redirect(storage.get_url(filename, 'comprobantes'))
    return send_from_directory(storage.LOCAL_DIRS['comprobantes'], filename)

@app.route("/", methods=["GET"])
def index():
    if "doctor" in session:
        if session.get("rol") == "admin":
            return redirect(url_for("dashboard"))
        return redirect(url_for("pacientes"))
    conn = database.get_db()
    doctores = [r["nombre"] for r in database.fetchall(conn,
        "SELECT DISTINCT nombre FROM doctores ORDER BY nombre")]
    conn.close()
    return render_template("web_login.html", doctores=doctores, error=request.args.get("error"))

@app.route("/login", methods=["POST"])
def login():
    doctor = request.form.get("doctor", "").strip()
    if not doctor:
        return redirect(url_for("index", error="Selecciona un doctor"))
    if not check_login_rate_limit(request.remote_addr or "local"):
        return redirect(url_for("index", error="Demasiados intentos. Espera 5 minutos."))
    conn = database.get_db()
    exists = database.fetchone(conn, "SELECT id FROM doctores WHERE nombre=?", (doctor,))
    conn.close()
    if exists:
        session.permanent = True
        session["doctor"] = doctor
        session["rol"] = "doctor"
        return redirect(url_for("pacientes"))
    return redirect(url_for("index", error="Doctor no encontrado"))

@app.route("/admin_login", methods=["POST"])
def admin_login():
    user = request.form.get("username", "").strip()
    pwd = request.form.get("password", "").strip()
    if not check_login_rate_limit("admin_"+user):
        return redirect(url_for("index", error="Demasiados intentos. Espera 5 minutos."))
    conn = database.get_db()
    u = database.fetchone(conn,
        "SELECT * FROM usuarios WHERE username=? AND activo=1", (user,))
    conn.close()
    if u and verify_pwd(u["password"], pwd):
        session.permanent = True
        session["doctor"] = u["nombre"]
        session["rol"] = u["rol"]
        session["user_id"] = u["id"]
        return redirect(url_for("dashboard"))
    return redirect(url_for("index", error="Credenciales inv\u00e1lidas"))

@app.route("/logout")
def logout():
    session.clear()
    response = redirect(url_for("index"))
    response.set_cookie('session', '', expires=0)
    return response

_ultimo_backup = ""

@app.route("/dashboard")
@login_required
def dashboard():
    global _ultimo_backup
    hoy = str(date.today())
    if _ultimo_backup != hoy:
        database.backup_database()
        _ultimo_backup = hoy
    conn = database.get_db()
    ayer = str(date.today() - timedelta(days=1))
    hace_7 = str(date.today() - timedelta(days=6))

    td = database.fetchone(conn, "SELECT COUNT(*) as cnt FROM duenos")["cnt"]
    ta = database.fetchone(conn, "SELECT COUNT(*) as cnt FROM animales")["cnt"]
    ch = database.fetchone(conn, "SELECT COUNT(*) as cnt FROM citas WHERE fecha=? AND estado='pendiente'", (hoy,))["cnt"]
    co_hoy = database.fetchone(conn, "SELECT COUNT(*) as cnt FROM registros_medicos WHERE fecha=?", (hoy,))["cnt"]
    co_ayer = database.fetchone(conn, "SELECT COUNT(*) as cnt FROM registros_medicos WHERE fecha=?", (ayer,))["cnt"]
    ventas_hoy = database.fetchone(conn, "SELECT COALESCE(SUM(total),0) as tot FROM ventas WHERE fecha=?", (hoy,))["tot"]
    prod_bajos = database.fetchall(conn, "SELECT nombre, stock FROM productos WHERE activo=1 AND stock<=5 ORDER BY stock LIMIT 10")
    citas_list = database.fetchall(conn,
        "SELECT c.*, a.nombre as animal_nombre, d.nombre as dueno_nombre FROM citas c "
        "JOIN animales a ON c.id_animal=a.id JOIN duenos d ON c.id_dueno=d.id "
        "WHERE c.fecha=? AND c.estado='pendiente' ORDER BY c.id LIMIT 10", (hoy,))

    cp_count = database.fetchone(conn, "SELECT COUNT(*) as cnt FROM cobros_pendientes WHERE cobrado=0")["cnt"]
    cp_total = database.fetchone(conn, "SELECT COALESCE(SUM(total),0) as tot FROM cobros_pendientes WHERE cobrado=0")["tot"]

    cred = database.fetchone(conn,
        "SELECT COUNT(*) as cnt, COALESCE(SUM(saldo),0) as tot FROM creditos WHERE estado='pendiente'")
    cred_count = cred["cnt"] if cred else 0
    cred_total = cred["tot"] if cred else 0

    # Ventas ultimos 7 dias en una sola query
    rows = database.fetchall(conn,
        "SELECT fecha, COALESCE(SUM(total),0) as t FROM ventas "
        "WHERE fecha>=? AND fecha<=? GROUP BY fecha ORDER BY fecha",
        (hace_7, hoy))
    ventas_dict = {r["fecha"]: float(r["t"]) for r in rows}
    dias = []
    ventas_semana = []
    for i in range(6, -1, -1):
        d = str(date.today() - timedelta(days=i))
        dias.append(d[-5:])
        ventas_semana.append(ventas_dict.get(d, 0))

    # Top 5 productos mas vendidos
    top_productos = [dict(r) for r in database.fetchall(conn, """
        SELECT vi.nombre, SUM(vi.cantidad) as total_vendido
        FROM venta_items vi JOIN ventas v ON vi.id_venta=v.id
        WHERE v.estado='completada'
        GROUP BY vi.nombre ORDER BY total_vendido DESC LIMIT 5
    """)]

    # Distribucion de especies
    especies = [dict(r) for r in database.fetchall(conn,
        "SELECT COALESCE(NULLIF(especie,''),'Sin especificar') as especie, COUNT(*) as cnt "
        "FROM animales GROUP BY especie ORDER BY cnt DESC")]

    conn.close()
    return render_template("admin_dashboard.html", td=td, ta=ta, ch=ch,
                           co_hoy=co_hoy, co_ayer=co_ayer,
                           ventas_hoy=ventas_hoy, prod_bajos=prod_bajos,
                           citas_list=citas_list, ventas_semana=ventas_semana,
                           dias=dias, top_productos=json.dumps(top_productos, ensure_ascii=False),
                           cp_count=cp_count, cp_total=cp_total,
                           cred_count=cred_count, cred_total=cred_total,
                           especies=json.dumps(especies, ensure_ascii=False),
                           active="dashboard", title="Dashboard", **_ctx())

@app.route("/nuevo_paciente_consulta", methods=["GET","POST"])
def web_nuevo_paciente_consulta():
    if "doctor" not in session:
        return redirect(url_for("index"))
    conn = database.get_db()
    if request.method == "GET":
        search_dni = request.args.get("dni","").strip()
        if search_dni:
            prefill = None
            encontrado = database.fetchone(conn, "SELECT id,nombre,telefono,direccion FROM duenos WHERE dni=? AND dni!=''", (search_dni,))
            if encontrado:
                prefill = dict(encontrado)
                prefill["dueno_dni"] = search_dni
            else:
                try:
                    resp = requests.get(f"https://graphperu.daustinn.com/api/query/{search_dni}", timeout=10)
                    if resp.status_code == 200:
                        data = resp.json()
                        if data.get("fullName"):
                            prefill = {
                                "nombre": data["names"] or "", "telefono": "", "direccion": "",
                                "dueno_dni": search_dni,
                                "apellido_paterno": data.get("paternalLastName","") or "",
                                "apellido_materno": data.get("maternalLastName","") or ""
                            }
                except Exception:
                    pass
            if prefill:
                conn.close()
                return render_template("web_nuevo_paciente_consulta.html", doctor=session["doctor"], prefill_dueno=prefill)
            conn.close()
            return render_template("web_nuevo_paciente_consulta.html", doctor=session["doctor"], dni_error="No se encontr\u00f3 DNI. Escribe manualmente.")
        conn.close()
        return render_template("web_nuevo_paciente_consulta.html", doctor=session["doctor"])
        dni = request.form.get("dni","").strip()
        dueno_nombre = request.form.get("dueno_nombre","").strip()
        telefono = request.form.get("telefono","").strip()
        direccion = request.form.get("direccion","").strip()
        dueno_id = None
        if dni:
            exist = database.fetchone(conn, "SELECT id FROM duenos WHERE dni=? AND dni!=''", (dni,))
            if exist: dueno_id = exist["id"]
        if not dueno_id and dueno_nombre:
            database.execute(conn,"INSERT INTO duenos (nombre,dni,telefono,direccion) VALUES (?,?,?,?)",
                (dueno_nombre, dni, telefono, direccion))
            conn.commit()
            dueno_id = database.fetchone(conn, "SELECT MAX(id) as id FROM duenos")["id"]
        if not dueno_id:
            conn.close()
            return Response(json.dumps({"error":"Debes ingresar nombre del due\u00f1o"}), mimetype="application/json", status=400)
        nombre = request.form.get("nombre","").strip()
        if not nombre:
            conn.close()
            return Response(json.dumps({"error":"Nombre del paciente requerido"}), mimetype="application/json", status=400)
        esterilizado = 1 if request.form.get("esterilizado") == "1" else 0
        database.execute(conn,"INSERT INTO animales (nombre,especie,raza,sexo,esterilizado,id_dueno) VALUES (?,?,?,?,?,?)",
            (nombre, request.form.get("especie","Perro"), request.form.get("raza",""),
             request.form.get("sexo",""), esterilizado, dueno_id))
        conn.commit()
        nuevo_id = database.fetchone(conn, "SELECT MAX(id) as id FROM animales")["id"]
        if 'foto' in request.files and request.files['foto'] and request.files['foto'].filename:
            try:
                fname, _ = storage.save_file(request.files['foto'], 'photos')
                database.execute(conn, "UPDATE animales SET foto=? WHERE id=?", (fname, nuevo_id))
                conn.commit()
            except Exception as e:
                logger.warning("Error al guardar foto: %s", e)
        conn.close()
        return redirect(url_for("consulta", animal_id=nuevo_id))
    conn.close()
    return render_template("web_nuevo_paciente_consulta.html", doctor=session["doctor"])

@app.route("/editar_paciente/<int:animal_id>", methods=["GET","POST"])
def web_editar_paciente(animal_id):
    if "doctor" not in session:
        return redirect(url_for("index"))
    conn = database.get_db()
    p = database.fetchone(conn,
        "SELECT a.*, d.nombre as dueno_nombre, d.telefono as dueno_telefono, d.direccion as dueno_direccion, d.email as dueno_email, d.dni as dueno_dni "
        "FROM animales a JOIN duenos d ON a.id_dueno = d.id WHERE a.id=?", (animal_id,))
    if not p:
        conn.close()
        return "Paciente no encontrado", 404
    if request.method == "POST":
        database.execute(conn,
            "UPDATE animales SET nombre=?, especie=?, raza=?, edad=?, peso=?, sexo=?, color=?, esterilizado=? WHERE id=?",
            (request.form["nombre"], request.form.get("especie",""), request.form.get("raza",""),
             int(request.form.get("edad",0)), float(request.form.get("peso",0)),
             request.form.get("sexo",""), request.form.get("color",""),
             1 if request.form.get("esterilizado") == "1" else 0, animal_id))
        dueno_id = p["id_dueno"]
        database.execute(conn,
            "UPDATE duenos SET nombre=?, dni=?, telefono=?, direccion=?, email=? WHERE id=?",
            (request.form.get("dueno_nombre",""), request.form.get("dueno_dni",""),
             request.form.get("dueno_telefono",""), request.form.get("dueno_direccion",""),
             request.form.get("dueno_email",""), dueno_id))
        conn.commit()
        conn.close()
        return redirect(url_for("consulta", animal_id=animal_id))
    conn.close()
    return render_template("web_editar_paciente.html", p=p, doctor=session["doctor"], dueno=p)

@app.route("/pacientes")
@login_required
def pacientes():
    conn = database.get_db()
    rows = database.fetchall(conn,
        "SELECT a.id, a.nombre, a.especie, a.raza, a.edad, d.nombre as dueno_nombre "
        "FROM animales a JOIN duenos d ON a.id_dueno = d.id ORDER BY a.nombre")
    doctorname = session["doctor"]
    pendientes = database.fetchone(conn,
        "SELECT COUNT(*) as cnt FROM cobros_pendientes cp "
        "JOIN registros_medicos r ON cp.id_registro = r.id WHERE cp.cobrado=0 AND r.doctor=?", (doctorname,))
    conn.close()
    return render_template("web_pacientes.html", pacientes=rows, doctor=doctorname,
                           pendientes=pendientes["cnt"] if pendientes else 0)

# ===== DUE\u00d1OS =====
@app.route("/duenos")
@login_required
def listar_duenos():
    conn = database.get_db()
    q = request.args.get("q","").strip()
    page = request.args.get("page", 1, type=int)
    per_page = 25
    deuda_f = request.args.get("deuda","")
    pacientes_f = request.args.get("pacientes","")

    where = " WHERE 1=1"
    params = []
    if q:
        where += " AND (d.dni LIKE ? OR d.nombre LIKE ? OR d.telefono LIKE ? OR d.email LIKE ?)"
        params.extend([f"%{q}%", f"%{q}%", f"%{q}%", f"%{q}%"])

    # Quick filters
    filtro = None
    filtro_param = ""
    filtro_val = ""
    if deuda_f == "si":
        where += " AND (SELECT COALESCE(SUM(saldo_pendiente),0) FROM ventas WHERE id_cliente=d.id AND (estado_pago='pendiente' OR COALESCE(saldo_pendiente,0)>0)) > 0"
        filtro = "deuda"; filtro_param = "deuda"; filtro_val = "si"
    elif deuda_f == "no":
        where += " AND (SELECT COALESCE(SUM(saldo_pendiente),0) FROM ventas WHERE id_cliente=d.id AND (estado_pago='pendiente' OR COALESCE(saldo_pendiente,0)>0)) = 0"
        filtro = "nodeuda"; filtro_param = "deuda"; filtro_val = "no"
    elif pacientes_f == "si":
        where += " AND (SELECT COUNT(*) FROM animales WHERE id_dueno=d.id) > 0"
        filtro = "pacientes"; filtro_param = "pacientes"; filtro_val = "si"

    # Total count
    cnt = database.fetchone(conn, "SELECT COUNT(*) as c FROM duenos d"+where, params)["c"]
    total_pages = max(1, (cnt + per_page - 1) // per_page)
    page = min(page, total_pages)
    offset = (page - 1) * per_page

    sql = "SELECT d.*, (SELECT COUNT(*) FROM animales WHERE id_dueno=d.id) as num_pacientes FROM duenos d"+where+" ORDER BY d.id DESC LIMIT ? OFFSET ?"
    rows = database.fetchall(conn, sql, params+[per_page, offset])

    duenos = []
    for d in rows:
        dd = dict(d)
        row = database.fetchone(conn, "SELECT COALESCE(SUM(saldo_pendiente),0) as deuda FROM ventas WHERE id_cliente=? AND (estado_pago='pendiente' OR COALESCE(saldo_pendiente,0)>0)", (dd["id"],))
        dd["deuda"] = row["deuda"] if row else 0
        duenos.append(dd)
    conn.close()
    return render_template("admin_duenos.html", duenos=duenos, q=q, page=page, total_pages=total_pages, total=cnt,
        filtro=filtro, filtro_param=filtro_param, filtro_val=filtro_val,
        title="Due\u00f1os", active="duenos", **_ctx())

@app.route("/duenos/exportar")
@login_required
def exportar_duenos():
    conn = database.get_db()
    q = request.args.get("q","").strip()
    where = " WHERE 1=1"
    params = []
    if q:
        where += " AND (dni LIKE ? OR nombre LIKE ? OR telefono LIKE ? OR email LIKE ?)"
        params.extend([f"%{q}%", f"%{q}%", f"%{q}%", f"%{q}%"])
    rows = database.fetchall(conn, "SELECT * FROM duenos d"+where+" ORDER BY nombre", params)
    conn.close()
    title = "Due\u00f1os" + (f" - Busqueda: {q}" if q else "")
    cols = ["ID","Nombre","DNI/RUC","Tel\u00e9fono","Email","Direcci\u00f3n"]
    keys = ["id","nombre","dni","telefono","email","direccion"]
    data = _excel_file(title, cols, rows, keys, col_widths={"A":8,"B":30,"C":16,"D":16,"E":30,"F":35})
    return Response(data, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    headers={"Content-Disposition": "attachment;filename=duenos.xlsx"})

@app.route("/duenos/nuevo", methods=["GET", "POST"])
@login_required
def crear_dueno():
    conn = database.get_db()
    if request.method == "POST":
        database.execute(conn,
            "INSERT INTO duenos (nombre, telefono, email, direccion, dni) VALUES (?,?,?,?,?)",
            (request.form["nombre"], request.form.get("telefono",""),
             request.form.get("email",""), request.form.get("direccion",""),
             request.form.get("dni","")))
        conn.commit(); conn.close()
        return redirect(url_for("listar_duenos"))
    conn.close()
    return redirect(url_for("listar_duenos"))

@app.route("/duenos/<int:id>")
@login_required
def ver_dueno(id):
    conn = database.get_db()
    d = database.fetchone(conn, "SELECT * FROM duenos WHERE id=?", (id,))
    if not d: conn.close(); return redirect(url_for("listar_duenos"))
    animales = database.fetchall(conn, "SELECT * FROM animales WHERE id_dueno=? ORDER BY id DESC", (id,))
    ventas = database.fetchall(conn, "SELECT * FROM ventas WHERE id_cliente=? ORDER BY fecha DESC, id DESC LIMIT 20", (id,))
    total_ventas = database.fetchone(conn, "SELECT COUNT(*) as c FROM ventas WHERE id_cliente=?", (id,))["c"]
    deuda = database.fetchone(conn, "SELECT COALESCE(SUM(saldo_pendiente),0) as total FROM ventas WHERE id_cliente=? AND (estado_pago='pendiente' OR COALESCE(saldo_pendiente,0)>0)", (id,))["total"]
    total_gastado = database.fetchone(conn, "SELECT COALESCE(SUM(total),0) as t FROM ventas WHERE id_cliente=?", (id,))["t"]
    ultima_cita = database.fetchone(conn, "SELECT fecha FROM citas WHERE id_dueno=? ORDER BY fecha DESC LIMIT 1", (id,))
    ultima_venta = database.fetchone(conn, "SELECT fecha FROM ventas WHERE id_cliente=? ORDER BY fecha DESC LIMIT 1", (id,))
    ultimo_registro = database.fetchone(conn,
        "SELECT r.fecha FROM registros_medicos r JOIN animales a ON r.id_animal=a.id WHERE a.id_dueno=? ORDER BY r.fecha DESC, r.hora DESC LIMIT 1", (id,))
    ultima_visita = None
    fechas = []
    if ultima_cita and ultima_cita["fecha"]: fechas.append(ultima_cita["fecha"])
    if ultima_venta and ultima_venta["fecha"]: fechas.append(ultima_venta["fecha"])
    if ultimo_registro and ultimo_registro["fecha"]: fechas.append(ultimo_registro["fecha"])
    if fechas: ultima_visita = max(fechas)
    conn.close()
    return render_template("admin_dueno_detail.html", d=d, animales=animales, ventas=ventas, total_ventas=total_ventas,
        deuda=deuda, total_gastado=total_gastado, ultima_visita=ultima_visita,
        title=d["nombre"], active="duenos", **_ctx())

@app.route("/duenos/<int:id>/editar", methods=["GET", "POST"])
@login_required
def editar_dueno(id):
    conn = database.get_db()
    if request.method == "POST":
        database.execute(conn,
            "UPDATE duenos SET nombre=?, telefono=?, email=?, direccion=?, dni=? WHERE id=?",
            (request.form["nombre"], request.form.get("telefono",""), request.form.get("email",""),
             request.form.get("direccion",""), request.form.get("dni",""), id))
        conn.commit(); conn.close()
        return redirect(url_for("listar_duenos", edited=id))
    item = database.fetchone(conn, "SELECT * FROM duenos WHERE id=?", (id,))
    conn.close()
    if not item:
        return redirect(url_for("listar_duenos"))
    return render_template("admin_dueno_form.html", item=item, title="Editar Due\u00f1o", active="duenos", **_ctx())

@app.route("/duenos/<int:id>/eliminar")
@login_required
def eliminar_dueno(id):
    conn = database.get_db()
    database.execute(conn, "DELETE FROM citas WHERE id_dueno=?", (id,))
    database.execute(conn, "DELETE FROM animales WHERE id_dueno=?", (id,))
    database.execute(conn, "DELETE FROM duenos WHERE id=?", (id,))
    conn.commit();     conn.close()
    return redirect(url_for("listar_duenos"))

@app.route("/api/duenos/<int:id>/relaciones")
@login_required
def api_duenos_relaciones(id):
    conn = database.get_db()
    animales = database.fetchone(conn, "SELECT COUNT(*) as c FROM animales WHERE id_dueno=?", (id,))["c"]
    citas = database.fetchone(conn, "SELECT COUNT(*) as c FROM citas WHERE id_dueno=?", (id,))["c"]
    ventas = database.fetchone(conn, "SELECT COUNT(*) as c FROM ventas WHERE id_cliente=?", (id,))["c"]
    creditos = database.fetchone(conn, "SELECT COUNT(*) as c FROM creditos WHERE id_cliente=?", (id,))["c"]
    conn.close()
    return {"animales": animales, "citas": citas, "ventas": ventas, "creditos": creditos}

# ===== ANIMALES =====
@app.route("/animales")
@login_required
def listar_animales():
    conn = database.get_db()
    animales = database.fetchall(conn,
        "SELECT a.*, d.nombre as dueno_nombre FROM animales a "
        "JOIN duenos d ON a.id_dueno = d.id ORDER BY a.id DESC")
    conn.close()
    return render_template("admin_list.html", items=animales,
        cols=["id","nombre","especie","raza","edad","dueno_nombre"],
        labels=["ID","Nombre","Especie","Raza","Edad","Due\u00f1o"],
        edit="animales", create="animales", detail="animales",
        title="Pacientes", active="animales", **_ctx())

@app.route("/animales/nuevo", methods=["GET", "POST"])
@login_required
def crear_animal():
    conn = database.get_db()
    duenos = database.fetchall(conn, "SELECT id, nombre FROM duenos ORDER BY nombre")
    if request.method == "POST":
        database.execute(conn,
            "INSERT INTO animales (nombre, especie, raza, edad, peso, sexo, color, id_dueno) VALUES (?,?,?,?,?,?,?,?)",
            (request.form["nombre"], request.form.get("especie",""), request.form.get("raza",""),
             int(request.form.get("edad",0)), float(request.form.get("peso",0)),
             request.form.get("sexo",""), request.form.get("color",""), int(request.form["id_dueno"])))
        conn.commit(); conn.close()
        return redirect(url_for("listar_animales"))
    conn.close()
    return render_template("admin_animal_form.html", duenos=duenos, item=None, title="Nuevo Paciente", active="animales", **_ctx())

@app.route("/animales/<int:id>/editar", methods=["GET", "POST"])
@login_required
def editar_animal(id):
    conn = database.get_db()
    duenos = database.fetchall(conn, "SELECT id, nombre FROM duenos ORDER BY nombre")
    if request.method == "POST":
        database.execute(conn,
            "UPDATE animales SET nombre=?, especie=?, raza=?, edad=?, peso=?, sexo=?, color=?, id_dueno=? WHERE id=?",
            (request.form["nombre"], request.form.get("especie",""), request.form.get("raza",""),
             int(request.form.get("edad",0)), float(request.form.get("peso",0)),
             request.form.get("sexo",""), request.form.get("color",""), int(request.form["id_dueno"]), id))
        conn.commit(); conn.close()
        return redirect(url_for("historial_global", edited_animal=id))
    item = database.fetchone(conn, "SELECT * FROM animales WHERE id=?", (id,))
    conn.close()
    return render_template("admin_animal_form.html", duenos=duenos, item=item, title="Editar Paciente", active="animales", **_ctx())

@app.route("/animales/<int:id>/eliminar")
@login_required
def eliminar_animal(id):
    conn = database.get_db()
    database.execute(conn, "DELETE FROM animales WHERE id=?", (id,))
    conn.commit(); conn.close()
    return redirect(url_for("listar_animales"))

@app.route("/animales/<int:id>")
@login_required
def ver_animal(id):
    conn = database.get_db()
    item = database.fetchone(conn,
        "SELECT a.*, d.nombre as dueno_nombre, d.telefono as dueno_telefono, d.direccion as dueno_direccion "
        "FROM animales a JOIN duenos d ON a.id_dueno = d.id WHERE a.id=?", (id,))
    historial = database.fetchall(conn,
        "SELECT * FROM registros_medicos WHERE id_animal=? ORDER BY fecha DESC LIMIT 20", (id,))
    hermanos = database.fetchall(conn,
        "SELECT id, nombre, especie, raza, foto FROM animales WHERE id_dueno=? AND id!=? ORDER BY nombre", (item["id_dueno"], id))
    duenos = database.fetchall(conn, "SELECT id, nombre FROM duenos ORDER BY nombre")
    ultimo_peso = database.fetchone(conn,
        "SELECT fecha, peso FROM registros_medicos WHERE id_animal=? AND peso IS NOT NULL ORDER BY fecha DESC LIMIT 1", (id,))
    conn.close()
    return render_template("admin_animal_detail.html", a=item, historial=historial, hermanos=hermanos, duenos=duenos, ultimo_peso=ultimo_peso, title="Paciente: "+item["nombre"], active="animales", **_ctx())

# ===== CITAS =====
@app.route("/citas")
@login_required
def listar_citas():
    conn = database.get_db()
    q = request.args.get("q","").strip()
    estado_f = request.args.get("estado","")
    hoy_f = request.args.get("hoy","")
    desde = request.args.get("desde","").strip()
    hasta = request.args.get("hasta","").strip()
    page = int(request.args.get("page","1"))
    per_page = 50

    where = " WHERE 1=1"
    params = []
    if q:
        where += " AND (a.nombre LIKE ? OR d.nombre LIKE ? OR c.motivo LIKE ?)"
        params.extend([f"%{q}%", f"%{q}%", f"%{q}%"])
    if estado_f:
        where += " AND c.estado=?"
        params.append(estado_f)
    if hoy_f:
        where += " AND c.fecha=?"
        params.append(str(date.today()))
    if desde:
        where += " AND c.fecha >= ?"; params.append(desde)
    if hasta:
        where += " AND c.fecha <= ?"; params.append(hasta)

    total = database.fetchone(conn,
        "SELECT COUNT(*) as c FROM citas c JOIN animales a ON c.id_animal=a.id JOIN duenos d ON c.id_dueno=d.id"+where,
        params)["c"]
    total_pages = max(1, (total + per_page - 1) // per_page)
    page = min(page, total_pages)
    offset = (page - 1) * per_page

    citas = database.fetchall(conn,
        "SELECT c.*, a.nombre as animal_nombre, d.nombre as dueno_nombre "
        "FROM citas c JOIN animales a ON c.id_animal=a.id JOIN duenos d ON c.id_dueno=d.id"+where+
        " ORDER BY CASE WHEN c.fecha >= date('now') THEN 0 ELSE 1 END, CASE WHEN c.fecha >= date('now') THEN c.fecha END ASC, c.fecha DESC, c.id DESC LIMIT ? OFFSET ?", params+[per_page, offset])

    animales = database.fetchall(conn,
        "SELECT a.*, d.nombre as dueno_nombre FROM animales a JOIN duenos d ON a.id_dueno=d.id ORDER BY a.nombre")
    servicios = database.fetchall(conn,
        "SELECT id, nombre, precio FROM servicios_medicos ORDER BY nombre")
    servicios_grooming = database.fetchall(conn,
        "SELECT id, nombre, precio FROM servicios_grooming ORDER BY nombre")
    conn.close()
    precios_cita = {}
    for s in servicios:
        precios_cita[s["nombre"].lower()] = s["precio"]
    for s in servicios_grooming:
        precios_cita[s["nombre"].lower()] = s["precio"]
    hoy_str = str(date.today())
    return render_template("admin_citas.html", citas=citas, animales=animales,
        precios_cita=json.dumps(precios_cita), q=q, estado_f=estado_f, desde=desde, hasta=hasta,
        page=page, total_pages=total_pages, total=total, hoy_str=hoy_str,
        title="Citas", active="citas", **_ctx())

@app.route("/citas/nuevo", methods=["GET", "POST"])
@login_required
def crear_cita():
    conn = database.get_db()
    if request.method == "POST":
        database.execute(conn,"INSERT INTO citas (id_animal,id_dueno,fecha,hora,motivo,tipo,precio) VALUES (?,?,?,?,?,?,?)",
            (request.form["id_animal"],request.form["id_dueno"],request.form["fecha"],
             request.form.get("hora",""),request.form["motivo"],
             request.form.get("tipo","veterinaria"),float(request.form.get("precio",0))))
        conn.commit(); conn.close()
        return redirect(url_for("listar_citas"))
    animales = database.fetchall(conn,"SELECT a.*,d.nombre as dueno_nombre FROM animales a JOIN duenos d ON a.id_dueno=d.id")
    conn.close()
    return redirect(url_for("listar_citas"))

@app.route("/citas/<int:id>/completar")
@login_required
def completar_cita(id):
    conn = database.get_db()
    database.execute(conn, "UPDATE citas SET estado='completada' WHERE id=?", (id,))
    conn.commit(); conn.close()
    return redirect(url_for("listar_citas"))

@app.route("/citas/<int:id>/cancelar")
@login_required
def cancelar_cita(id):
    conn = database.get_db()
    database.execute(conn, "UPDATE citas SET estado='cancelada' WHERE id=?", (id,))
    conn.commit(); conn.close()
    return redirect(url_for("listar_citas"))

@app.route("/citas/<int:id>/editar", methods=["POST"])
@login_required
def editar_cita(id):
    conn = database.get_db()
    database.execute(conn,
        "UPDATE citas SET fecha=?, hora=?, tipo=?, motivo=?, precio=? WHERE id=?",
        (request.form["fecha"], request.form.get("hora",""), request.form.get("tipo","veterinaria"),
         request.form["motivo"], float(request.form.get("precio",0)), id))
    conn.commit(); conn.close()
    return redirect(url_for("listar_citas"))

# ===== PRODUCTOS / INVENTARIO =====
def registrar_movimiento(conn, id_producto, tipo, cantidad, referencia_tipo='', referencia_id=None, descripcion=''):
    prod = database.fetchone(conn, "SELECT stock FROM productos WHERE id=?", (id_producto,))
    stock_anterior = prod["stock"] - cantidad if prod else 0
    database.execute(conn, """
        INSERT INTO stock_movimientos (id_producto, tipo, cantidad, stock_anterior, stock_nuevo, referencia_tipo, referencia_id, descripcion, fecha)
        VALUES (?,?,?,?,?,?,?,?,datetime('now','localtime'))
    """, (id_producto, tipo, cantidad, stock_anterior, prod["stock"] if prod else 0, referencia_tipo, referencia_id, descripcion))

@app.route("/productos")
@login_required
def listar_productos():
    conn = database.get_db()
    q = request.args.get("q","").strip()
    cat_f = request.args.get("categoria","")
    stock_bajo = request.args.get("stock_bajo","")
    tab = request.args.get("tab","productos")
    page = int(request.args.get("page","1"))
    per_page = 50

    where = " WHERE activo=1"
    params = []
    if q:
        where += " AND (nombre LIKE ? OR descripcion LIKE ?)"
        params.extend([f"%{q}%", f"%{q}%"])
    if cat_f:
        where += " AND categoria=?"
        params.append(cat_f)
    if stock_bajo:
        where += " AND ((stock_minimo IS NOT NULL AND stock<=stock_minimo) OR (stock_minimo IS NULL AND stock<=5))"

    total = database.fetchone(conn, "SELECT COUNT(*) as c FROM productos"+where, params)["c"]
    total_pages = max(1, (total + per_page - 1) // per_page)
    page = min(page, total_pages)
    offset = (page - 1) * per_page

    items = database.fetchall(conn,
        "SELECT *, (precio_venta - precio_compra) as ganancia FROM productos"+where+
        " ORDER BY nombre LIMIT ? OFFSET ?", params+[per_page, offset])

    categorias = database.get_categorias(conn)
    stock_critico = database.fetchone(conn,
        "SELECT COUNT(*) as c FROM productos WHERE activo=1 AND ((stock_minimo IS NOT NULL AND stock<=stock_minimo) OR (stock_minimo IS NULL AND stock<=5))", [])["c"]

    grooming_servicios = database.fetchall(conn, "SELECT * FROM servicios_grooming ORDER BY nombre")
    servicios_medicos = database.fetchall(conn, "SELECT * FROM servicios_medicos ORDER BY nombre")
    conn.close()
    return render_template("admin_productos.html", items=items, q=q, cat_f=cat_f, stock_bajo=stock_bajo,
        page=page, total_pages=total_pages, total=total, categorias=categorias, stock_critico=stock_critico,
        tab=tab, grooming_servicios=grooming_servicios, servicios_medicos=servicios_medicos,
        title="Inventario", active="productos", **_ctx())

@app.route("/productos/nuevo", methods=["GET", "POST"])
@login_required
@role_required('admin')
def crear_producto():
    conn = database.get_db()
    if request.method == "POST":
        stock = int(request.form.get("stock",0))
        foto = ''
        if 'foto' in request.files and request.files['foto'] and request.files['foto'].filename:
            try:
                fname, _ = storage.save_file(request.files['foto'], 'productos')
                foto = fname
            except Exception as e:
                logger.warning("Error al guardar foto producto: %s", e)
        else:
            existing = database.fetchone(conn,
                "SELECT foto FROM productos WHERE nombre=? AND foto!='' AND foto IS NOT NULL LIMIT 1",
                (request.form["nombre"],))
            if existing and existing["foto"]:
                foto = existing["foto"]
        por_mayor = 1 if request.form.get("por_mayor") == '1' else 0
        descuento_mayorista = float(request.form.get("descuento_mayorista", 0) or 0)
        colores = request.form.get("colores","")
        en_catalogo = 1 if request.form.get("en_catalogo") == '1' else 0
        vende_por_kilo = 1 if request.form.get("vende_por_kilo") == '1' else 0
        precio_kilo = float(request.form.get("precio_kilo", 0) or 0)
        database.execute(conn,"INSERT INTO productos (nombre,descripcion,precio_compra,precio_venta,stock,categoria,stock_minimo,codigo_barras,foto,por_mayor,descuento_mayorista,colores,en_catalogo,vende_por_kilo,precio_kilo) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
            (request.form["nombre"],request.form.get("descripcion",""),float(request.form.get("precio_compra",0) or 0),
             float(request.form.get("precio_venta",0) or 0),stock,request.form.get("categoria","General"),
             int(request.form.get("stock_minimo",5)),request.form.get("codigo_barras",""),foto,por_mayor,descuento_mayorista,colores,en_catalogo,vende_por_kilo,precio_kilo))
        pid = database.fetchone(conn, "SELECT MAX(id) as id FROM productos")["id"]
        if stock > 0:
            registrar_movimiento(conn, pid, 'entrada', stock, 'inicial', None, 'Stock inicial')
        conn.commit(); conn.close()
        return redirect(url_for("ver_producto", id=pid))
    conn.close()
    categorias = database.get_categorias() + ["+ Nueva Categor\u00eda"]
    return render_template("admin_form.html", fields=[
        ("nombre","Nombre",True),("descripcion","Descripci\u00f3n","textarea"),
        ("precio_compra","Precio Compra","number"),("precio_venta","Precio Venta","number"),
        ("stock","Stock","number"),("stock_minimo","Stock M\u00ednimo","number"),
        ("categoria","Categor\u00eda","select",categorias),
        ("codigo_barras","C\u00f3digo de Barras (SKU)"),
        ("foto","Foto del Producto","file"),
        ("por_mayor","Vendible por mayor","checkbox"),
        ("descuento_mayorista","Descuento Mayorista (%)","number"),
        ("colores","Colores Disponibles","textarea"),
        ("vende_por_kilo","Vender por kilo (Ej: 21 soles el kilo)","checkbox"),
        ("precio_kilo","Precio por Kilo (S/)","number"),
        ("en_catalogo","Mostrar en cat\u00e1logo","checkbox"),
    ], item=None, title="Nuevo Producto", active="productos", form_enctype="multipart/form-data", **_ctx())

@app.route("/productos/<int:id>/editar", methods=["GET", "POST"])
@login_required
@role_required('admin')
def editar_producto(id):
    conn = database.get_db()
    if request.method == "POST":
        nuevo_stock = int(request.form.get("stock",0))
        old = database.fetchone(conn, "SELECT stock, foto FROM productos WHERE id=?", (id,))
        diff = nuevo_stock - old["stock"] if old else 0
        foto = old["foto"] if old else ''
        if 'foto' in request.files and request.files['foto'] and request.files['foto'].filename:
            try:
                fname, _ = storage.save_file(request.files['foto'], 'productos')
                foto = fname
            except Exception as e:
                logger.warning("Error al guardar foto producto: %s", e)
        por_mayor = 1 if request.form.get("por_mayor") == '1' else 0
        descuento_mayorista = float(request.form.get("descuento_mayorista", 0) or 0)
        colores = request.form.get("colores","")
        en_catalogo = 1 if request.form.get("en_catalogo") == '1' else 0
        vende_por_kilo = 1 if request.form.get("vende_por_kilo") == '1' else 0
        precio_kilo = float(request.form.get("precio_kilo", 0) or 0)
        database.execute(conn,"UPDATE productos SET nombre=?,descripcion=?,precio_compra=?,precio_venta=?,stock=?,categoria=?,stock_minimo=?,codigo_barras=?,foto=?,por_mayor=?,descuento_mayorista=?,colores=?,en_catalogo=?,vende_por_kilo=?,precio_kilo=? WHERE id=?",
            (request.form["nombre"],request.form.get("descripcion",""),float(request.form.get("precio_compra",0) or 0),
             float(request.form.get("precio_venta",0) or 0),nuevo_stock,request.form.get("categoria","General"),
             int(request.form.get("stock_minimo",5)),request.form.get("codigo_barras",""),foto,por_mayor,descuento_mayorista,colores,en_catalogo,vende_por_kilo,precio_kilo,id))
        if diff > 0:
            registrar_movimiento(conn, id, 'entrada', diff, 'ajuste', None, 'Ajuste manual (+{})'.format(diff))
        elif diff < 0:
            registrar_movimiento(conn, id, 'salida', abs(diff), 'ajuste', None, 'Ajuste manual ({})'.format(diff))
        conn.commit(); conn.close()
        return redirect(url_for("ver_producto", id=id))
    item = database.fetchone(conn,"SELECT * FROM productos WHERE id=?",(id,))
    conn.close()
    categorias = database.get_categorias() + ["+ Nueva Categor\u00eda"]
    if item and item["categoria"] and item["categoria"] not in categorias:
        categorias.insert(-1, item["categoria"])
    return render_template("admin_form.html", fields=[
        ("nombre","Nombre",True),("descripcion","Descripci\u00f3n","textarea"),
        ("precio_compra","Precio Compra","number"),("precio_venta","Precio Venta","number"),
        ("stock","Stock","number"),("stock_minimo","Stock M\u00ednimo","number"),
        ("categoria","Categor\u00eda","select",categorias),
        ("codigo_barras","C\u00f3digo de Barras (SKU)"),
        ("foto","Foto del Producto","file"),
        ("por_mayor","Vendible por mayor","checkbox"),
        ("descuento_mayorista","Descuento Mayorista (%)","number"),
        ("colores","Colores Disponibles","textarea"),
        ("vende_por_kilo","Vender por kilo (Ej: 21 soles el kilo)","checkbox"),
        ("precio_kilo","Precio por Kilo (S/)","number"),
        ("en_catalogo","Mostrar en cat\u00e1logo","checkbox"),
    ], item=item, title="Editar Producto", active="productos", form_enctype="multipart/form-data", **_ctx())

@app.route("/catalogo/<tipo>")
@login_required
def catalogo(tipo):
    if tipo not in ('menor', 'mayor'):
        return redirect(url_for("listar_productos"))
    conn = database.get_db()
    items = database.fetchall(conn,
        "SELECT * FROM productos WHERE activo=1 AND stock>0 AND en_catalogo=1 ORDER BY categoria, nombre")
    conn.close()
    by_cat = {}
    for it in items:
        cat = it["categoria"] or "General"
        by_cat.setdefault(cat, []).append(it)
    return render_template("catalogo.html", items=items, by_cat=by_cat, tipo=tipo,
        title=f"Cat\u00e1logo por {tipo.capitalize()}", active="productos", **_ctx())

@app.route("/catalogo/<tipo>/pdf")
@login_required
def catalogo_pdf(tipo):
    if tipo not in ('menor', 'mayor'):
        return redirect(url_for("listar_productos"))
    conn = database.get_db()
    items = database.fetchall(conn,
        "SELECT * FROM productos WHERE activo=1 AND stock>0 AND en_catalogo=1 ORDER BY categoria, nombre")
    conn.close()
    by_cat = {}
    for it in items:
        cat = it["categoria"] or "General"
        by_cat.setdefault(cat, []).append(it)
    html = render_template("catalogo.html", items=items, by_cat=by_cat, tipo=tipo,
        title=f"Cat\u00e1logo por {tipo.capitalize()}", active="productos", **_ctx())
    base_url = request.host_url.rstrip('/')
    html = html.replace(' src="/', f' src="{base_url}/')
    html = html.replace(' href="/', f' href="{base_url}/')
    import re
    html = re.sub(r'<div class="btn-group">.*?</div>', '', html, flags=re.DOTALL)
    html = html.replace('mix-blend-mode:multiply', '')
    wk_path = r'C:\Program Files\wkhtmltopdf\bin\wkhtmltopdf.exe'
    cfg = pdfkit.configuration(wkhtmltopdf=wk_path)
    opts = {
        'page-size': 'A4',
        'margin-top': '15mm',
        'margin-right': '10mm',
        'margin-bottom': '15mm',
        'margin-left': '10mm',
        'encoding': 'UTF-8',
        'no-outline': None,
        'enable-local-file-access': None,
    }
    pdf = pdfkit.from_string(html, False, configuration=cfg, options=opts)
    return Response(pdf, mimetype='application/pdf',
        headers={'Content-Disposition': f'attachment; filename=catalogo_{tipo}.pdf'})

@app.route("/productos/<int:id>/eliminar")
@login_required
@role_required('admin')
def eliminar_producto(id):
    conn = database.get_db()
    database.execute(conn,"UPDATE productos SET activo=0 WHERE id=?",(id,))
    conn.commit(); conn.close()
    return redirect(url_for("listar_productos"))

@app.route("/productos/eliminar-masivo", methods=["POST"])
@login_required
@role_required('admin')
def eliminar_productos_masivo():
    ids = request.form.getlist("ids")
    if ids:
        conn = database.get_db()
        placeholders = ",".join("?" for _ in ids)
        database.execute(conn, f"UPDATE productos SET activo=0 WHERE id IN ({placeholders})", ids)
        conn.commit(); conn.close()
    return redirect(url_for("listar_productos"))

@app.route("/productos/<int:id>")
@login_required
def ver_producto(id):
    conn = database.get_db()
    item = database.fetchone(conn, "SELECT * FROM productos WHERE id=?", (id,))
    if not item: conn.close(); return redirect(url_for("listar_productos"))
    ventas = database.fetchall(conn,
        "SELECT vi.*, v.fecha, v.id as venta_id FROM venta_items vi JOIN ventas v ON vi.id_venta=v.id WHERE vi.tipo_item='producto' AND vi.referencia_id=? ORDER BY v.fecha DESC LIMIT 20", (id,))
    total_vendido = database.fetchone(conn, "SELECT COALESCE(SUM(cantidad),0) as c FROM venta_items WHERE tipo_item='producto' AND referencia_id=?", (id,))["c"]
    total_ingresos = database.fetchone(conn, "SELECT COALESCE(SUM(cantidad*precio_unitario),0) as t FROM venta_items WHERE tipo_item='producto' AND referencia_id=?", (id,))["t"]
    movimientos = database.fetchall(conn,
        "SELECT * FROM stock_movimientos WHERE id_producto=? ORDER BY fecha DESC LIMIT 50", (id,))
    conn.close()
    return render_template("admin_producto_detail.html", item=item, ventas=ventas,
        total_vendido=total_vendido, total_ingresos=total_ingresos, movimientos=movimientos,
        title=item["nombre"], active="productos", **_ctx())

@app.route("/api/productos/crear", methods=["POST"])
@login_required
@role_required('admin')
def api_productos_crear():
    nombre = request.form.get("nombre","").strip()
    precio = request.form.get("precio","0")
    if not nombre: return Response(json.dumps({"error":"Nombre requerido"}), mimetype="application/json", status=400)
    conn = database.get_db()
    database.execute(conn, "INSERT INTO productos (nombre,precio_venta,stock,activo) VALUES (?,?,?,?)", (nombre, float(precio), 0, 1))
    conn.commit()
    nuevo_id = database.fetchone(conn, "SELECT MAX(id) as id FROM productos")["id"]
    conn.close()
    return Response(json.dumps({"id":nuevo_id,"nombre":nombre,"precio_venta":float(precio)}), mimetype="application/json")

# ===== VENTAS =====
@app.route("/productos/<int:id>/stock", methods=["POST"])
@login_required
@role_required('admin')
def producto_actualizar_stock(id):
    conn = database.get_db()
    old = database.fetchone(conn, "SELECT stock FROM productos WHERE id=?", (id,))
    nuevo = int(request.form.get("stock", 0))
    diff = nuevo - old["stock"] if old else 0
    database.execute(conn, "UPDATE productos SET stock=? WHERE id=?", (nuevo, id))
    if diff != 0:
        tipo = "entrada" if diff > 0 else "salida"
        registrar_movimiento(conn, id, tipo, abs(diff), "ajuste", id, "Ajuste inline desde inventario")
    conn.commit(); conn.close()
    return "", 200

@app.route("/ventas")
@login_required
@role_required('admin', 'recepcion')
def listar_ventas():
    conn = database.get_db()
    q = request.args.get("q","").strip()
    desde = request.args.get("desde","").strip()
    hasta = request.args.get("hasta","").strip()
    page = int(request.args.get("page","1"))
    per_page = 50
    ok = request.args.get("ok","")
    error = request.args.get("error","")
    where = "WHERE 1=1"
    params = []
    if q:
        # Search also by product name in venta_items
        item_ids = database.fetchall(conn,
            "SELECT DISTINCT id_venta FROM venta_items WHERE nombre LIKE ?", [f"%{q}%"])
        item_ids = [r["id_venta"] for r in item_ids]
        if item_ids:
            placeholders = ",".join("?" for _ in item_ids)
            where += f" AND (cliente_dni LIKE ? OR cliente_nombre LIKE ? OR CAST(id AS TEXT) LIKE ? OR id IN ({placeholders}))"
            params.extend([f"%{q}%", f"%{q}%", f"%{q}%"] + item_ids)
        else:
            where += " AND (cliente_dni LIKE ? OR cliente_nombre LIKE ? OR CAST(id AS TEXT) LIKE ?)"
            params.extend([f"%{q}%", f"%{q}%", f"%{q}%"])
    if desde:
        where += " AND fecha >= ?"; params.append(desde)
    if hasta:
        where += " AND fecha <= ?"; params.append(hasta)
    total = database.fetchone(conn, f"SELECT COUNT(*) as cnt FROM ventas {where}", params[:])["cnt"]
    offset = (page - 1) * per_page
    sql = f"SELECT * FROM ventas {where} ORDER BY fecha DESC, id DESC LIMIT ? OFFSET ?"
    params.extend([per_page, offset])
    ventas = database.fetchall(conn, sql, params)
    total_pages = max(1, (total + per_page - 1) // per_page)
    conn.close()
    return render_template("admin_ventas.html", ventas=ventas, title="Ventas", active="ventas",
        q=q, desde=desde, hasta=hasta, page=page, total_pages=total_pages,
        ok=ok, error=error, **_ctx())

@app.route("/ventas/exportar")
@login_required
@role_required('admin', 'recepcion')
def exportar_ventas():
    conn = database.get_db()
    q = request.args.get("q","").strip()
    desde = request.args.get("desde","").strip()
    hasta = request.args.get("hasta","").strip()
    where = "WHERE 1=1"
    params = []
    if q:
        item_ids = database.fetchall(conn,
            "SELECT DISTINCT id_venta FROM venta_items WHERE nombre LIKE ?", [f"%{q}%"])
        item_ids = [r["id_venta"] for r in item_ids]
        if item_ids:
            placeholders = ",".join("?" for _ in item_ids)
            where += f" AND (cliente_dni LIKE ? OR cliente_nombre LIKE ? OR CAST(id AS TEXT) LIKE ? OR id IN ({placeholders}))"
            params.extend([f"%{q}%", f"%{q}%", f"%{q}%"] + item_ids)
        else:
            where += " AND (cliente_dni LIKE ? OR cliente_nombre LIKE ? OR CAST(id AS TEXT) LIKE ?)"
            params.extend([f"%{q}%", f"%{q}%", f"%{q}%"])
    if desde:
        where += " AND fecha >= ?"; params.append(desde)
    if hasta:
        where += " AND fecha <= ?"; params.append(hasta)
    rows = database.fetchall(conn, "SELECT * FROM ventas "+where+" ORDER BY fecha DESC, id DESC", params)
    conn.close()
    data = []
    for r in rows:
        rd = dict(r)
        s = rd.get("serie",""); n = rd.get("numero")
        rd["_comprobante"] = f"{s}-{n:05d}" if s and n else ""
        data.append(rd)
    title = "Ventas" + (f" ({desde} a {hasta})" if desde or hasta else "")
    cols = ["ID","Fecha","Cliente","DNI/RUC","Subtotal","Descuento","IGV","Total","Tipo","Estado","Comprobante","Serie","M\u00e9todo Pago"]
    keys = ["id","fecha","cliente_nombre","cliente_dni","subtotal","descuento","igv","total","tipo","estado","_comprobante","tipo_comprobante","metodo_pago"]
    xldata = _excel_file(title, cols, data, keys, col_widths={"A":8,"B":14,"C":28,"D":16,"E":12,"F":12,"G":12,"H":12,"I":10,"J":14,"K":16,"L":16,"M":16})
    return Response(xldata, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    headers={"Content-Disposition": f"attachment;filename=ventas{f'_{desde}_{hasta}' if desde or hasta else ''}.xlsx"})

@app.route("/ventas/nueva", methods=["GET","POST"])
@login_required
@role_required('admin', 'recepcion')
def nueva_venta():
    conn = database.get_db()
    productos = database.fetchall(conn,"SELECT id,nombre,precio_venta,stock,codigo_barras,por_mayor,descuento_mayorista,vende_por_kilo,precio_kilo FROM productos WHERE activo=1 ORDER BY nombre")

    servicios_med = database.fetchall(conn,"SELECT id,nombre,precio,tipo FROM servicios_medicos ORDER BY nombre")
    servicios_groom = database.fetchall(conn,"SELECT id,nombre,precio,tipo FROM servicios_grooming ORDER BY nombre")
    combos = database.fetchall(conn,"SELECT id,nombre,precio_total,descripcion FROM combos WHERE activo=1 ORDER BY nombre")
    clientes = database.fetchall(conn,"SELECT id,nombre FROM duenos ORDER BY nombre")
    por_mayor_count = database.fetchone(conn, "SELECT COUNT(*) as c FROM productos WHERE activo=1 AND por_mayor=1", [])["c"]
    conn.close()
    return render_template("admin_venta_form.html", productos=productos, servicios_med=servicios_med,
        servicios_groom=servicios_groom, combos=combos, clientes=clientes, title="Nueva Venta", active="ventas", por_mayor_count=por_mayor_count, **_ctx())

def _generar_comprobante(conn, tipo):
    if tipo not in ('Boleta', 'Factura', 'Nota de Venta'):
        tipo = 'Boleta'
    row = database.fetchone(conn, "SELECT serie, ultimo_numero FROM series WHERE tipo=?", (tipo,))
    if not row:
        database.execute(conn, "INSERT INTO series (tipo, serie, ultimo_numero) VALUES (?,?,0)", (tipo, 'B001'))
        serie, num = 'B001', 0
    else:
        serie, num = row["serie"], row["ultimo_numero"]
    num += 1
    database.execute(conn, "UPDATE series SET ultimo_numero=? WHERE tipo=?", (num, tipo))
    return serie, num

@app.route("/ventas/guardar", methods=["POST"])
@login_required
@role_required('admin', 'recepcion')
def guardar_venta():
    conn = database.get_db()
    items_json = request.form.get("items","[]")
    try: items = json.loads(items_json)
    except: items = []
    if not items:
        conn.close()
        return redirect(url_for("nueva_venta"))
    total = sum(float(it.get("subtotal",0)) for it in items)
    try:
        # Validate stock before proceeding (only for products, not services)
        for it in items:
            pid = it.get("id")
            tipo = it.get("tipo","producto")
            kilo_mode = it.get("kiloMode", False)
            if pid and tipo == "producto" and not kilo_mode:
                cant = int(it.get("cant",1))
                prod = database.fetchone(conn, "SELECT stock FROM productos WHERE id=? AND activo=1", (pid,))
                if not prod or prod["stock"] < cant:
                    conn.close()
                    return f"Stock insuficiente para {it.get('nombre','producto')}. Disponible: {prod['stock'] if prod else 0}, requerido: {cant}", 400
        # Auto-create dueño if DNI+name provided and not exists
        dni_cliente = request.form.get("cliente_dni","").strip()
        nombre_cliente = request.form.get("cliente_nombre","").strip() or "Cliente Variado"
        if dni_cliente and nombre_cliente:
            exist = database.fetchone(conn, "SELECT id FROM duenos WHERE dni=? AND dni!=''", (dni_cliente,))
            if not exist:
                database.execute(conn,"INSERT INTO duenos (nombre,dni) VALUES (?,?)", (nombre_cliente, dni_cliente))
        # Descuento + IGV
        subtotal_sin_desc = total
        descuento_global = float(request.form.get("descuento_global","0"))
        subtotal_con_desc = subtotal_sin_desc - descuento_global
        tipo_comp = request.form.get("tipo_comprobante","Boleta")
        if tipo_comp not in ('Boleta', 'Factura', 'Nota de Venta'):
            tipo_comp = 'Boleta'
        serie, numero = _generar_comprobante(conn, tipo_comp)
        igv = 0
        if tipo_comp == "Factura":
            igv = round(subtotal_con_desc * 0.18, 2)
        total_final = subtotal_con_desc + igv
        metodo_pago_raw = request.form.get("metodo_pago","efectivo")
        try:
            mp_parsed = json.loads(metodo_pago_raw)
            if isinstance(mp_parsed, list):
                metodo_pago = json.dumps(mp_parsed)
            else:
                metodo_pago = metodo_pago_raw
        except:
            metodo_pago = metodo_pago_raw
        # Dynamic tipo: detect if mixed
        tipos_encontrados = set(it.get("tipo","producto") for it in items)
        tipo_venta = "mixta" if len(tipos_encontrados) > 1 else next(iter(tipos_encontrados))
        cur = database.execute(conn,"INSERT INTO ventas (id_cliente,id_animal,fecha,total,tipo,estado,tipo_comprobante,tipo_pago,metodo_pago,cliente_dni,cliente_nombre,descuento,igv,subtotal,serie,numero) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
            (int(request.form.get("id_cliente",0)) if request.form.get("id_cliente") else None,
             None, str(date.today()), total_final, tipo_venta, "completada",
             tipo_comp, "contado", metodo_pago,
             dni_cliente, nombre_cliente, descuento_global, igv, subtotal_sin_desc, serie, numero))
        if database._using_pg():
            conn.commit()
            venta_id = cur.fetchone()[0] if cur.description else None
            if not venta_id: venta_id = database.fetchone(conn, "SELECT MAX(id) as id FROM ventas")["id"]
        else: venta_id = cur.lastrowid
        for it in items:
            tipo_item = it.get("tipo","producto")
            ref_id = it.get("id")
            kilo_mode = it.get("kiloMode", False)
            cant = float(it.get("cant", 1)) if kilo_mode else int(it.get("cant", 1))
            database.execute(conn,"INSERT INTO venta_items (id_venta,tipo_item,referencia_id,nombre,cantidad,precio_unitario,subtotal) VALUES (?,?,?,?,?,?,?)",
                (venta_id, tipo_item, ref_id, it.get("nombre",""), cant,
                 float(it.get("precio",0)), float(it.get("subtotal",0))))
            if tipo_item == "producto" and ref_id and not kilo_mode:
                database.execute(conn,"UPDATE productos SET stock=stock-? WHERE id=? AND stock>=?",(int(it.get("cant",1)),ref_id,int(it.get("cant",1))))
                registrar_movimiento(conn, ref_id, 'salida', int(it.get("cant",1)), 'venta', venta_id, 'Venta #{}'.format(venta_id))
            if tipo_item == "servicio_medico" and ref_id:
                insumos = database.fetchall(conn, "SELECT si.*, p.stock, p.nombre as pnom FROM servicio_insumos si JOIN productos p ON si.id_producto=p.id WHERE si.id_servicio=?", (ref_id,))
                for ins in insumos:
                    if ins["stock"] < ins["cantidad"]:
                        logger.warning("Stock insuficiente para insumo '%s' en venta #%s: %d disp, %d req", ins["pnom"], venta_id, ins["stock"], ins["cantidad"])
                        flash(f"Stock insuficiente para '{ins['pnom']}' ({ins['stock']} disp, {ins['cantidad']} req). No se descont\u00f3.", "warning")
                    else:
                        database.execute(conn, "INSERT INTO insumos_utilizados (id_registro,id_producto,cantidad) VALUES (?,?,?)", (venta_id, ins["id_producto"], ins["cantidad"]))
                        database.execute(conn, "UPDATE productos SET stock=stock-? WHERE id=?", (ins["cantidad"], ins["id_producto"]))
                        registrar_movimiento(conn, ins["id_producto"], "salida", ins["cantidad"], "servicio", venta_id, "Venta #{} (servicio)".format(venta_id))
            database.execute(conn,"INSERT INTO caja (fecha,tipo,concepto,monto,referencia_tipo,referencia_id) VALUES (?,?,?,?,?,?)",
                (str(date.today()),"ingreso","Venta #"+str(venta_id),float(it.get("subtotal",0)),"venta",venta_id))
        # Save split payments if provided
        pagos_json = request.form.get("pagos","[]")
        try: pagos = json.loads(pagos_json)
        except: pagos = []
        if pagos:
            for p in pagos:
                database.execute(conn, "INSERT INTO pagos (id_venta,monto,metodo,fecha) VALUES (?,?,?,?)", (venta_id, float(p.get("monto",0)), p.get("metodo","efectivo"), str(date.today())))
        conn.commit()
        conn.close()
        return redirect(url_for("ticket", venta_id=venta_id))
    except Exception as e:
        conn.rollback()
        conn.close()
        logger.error("Error al guardar venta: %s", e)
        return f"Error al procesar la venta: {e}", 500

@app.route("/ventas/<int:id>")
@login_required
@role_required('admin', 'recepcion')
def ver_venta(id):
    conn = database.get_db()
    v = database.fetchone(conn,"SELECT * FROM ventas WHERE id=?",(id,))
    items = database.fetchall(conn,"SELECT * FROM venta_items WHERE id_venta=?",(id,))
    pagos = database.fetchall(conn,"SELECT * FROM pagos WHERE id_venta=? ORDER BY id",(id,))
    nc = None
    if v["estado"] == "anulada":
        nc = database.fetchone(conn, "SELECT id FROM notas_credito WHERE id_venta=?", (id,))
    conn.close()
    return render_template("admin_venta_detail.html", v=v, items=items, pagos=pagos, nc=nc, title="Venta #"+str(id), active="ventas", **_ctx())

@app.route("/ventas/<int:id>/anular", methods=["POST"])
@login_required
@role_required('admin', 'recepcion')
def anular_venta(id):
    conn = database.get_db()
    v = database.fetchone(conn,"SELECT * FROM ventas WHERE id=?",(id,))
    if not v or v["estado"] != "completada":
        conn.close()
        return redirect(url_for("listar_ventas"))
    try:
        # Revert stock for product items
        items = database.fetchall(conn,"SELECT * FROM venta_items WHERE id_venta=?",(id,))
        for it in items:
            if it["tipo_item"] == "producto" and it["referencia_id"]:
                database.execute(conn,"UPDATE productos SET stock=stock+? WHERE id=?",(it["cantidad"],it["referencia_id"]))
                registrar_movimiento(conn, it["referencia_id"], 'entrada', it["cantidad"], 'anulacion', id, 'Anulación Venta #{}'.format(id))
        # Generate Nota de Crédito
        serie, numero = _generar_comprobante(conn, 'Nota Credito')
        motivo = request.form.get("motivo_anulacion", "Anulación voluntaria")
        nc_subtotal = v["subtotal"] or v["total"]
        nc_igv = v["igv"] or 0
        nc_total = v["total"]
        database.execute(conn,
            "INSERT INTO notas_credito (id_venta,serie,numero,fecha,subtotal,igv,total,motivo) VALUES (?,?,?,?,?,?,?,?)",
            (id, serie, numero, str(date.today()), nc_subtotal, nc_igv, nc_total, motivo))
        nc_id = database.fetchone(conn, "SELECT MAX(id) as id FROM notas_credito")["id"]
        database.execute(conn,"UPDATE ventas SET estado='anulada' WHERE id=?",(id,))
        conn.commit()
        conn.close()
        return redirect(url_for("ver_nota_credito", id=nc_id))
    except Exception as e:
        conn.rollback()
        conn.close()
        logger.error("Error al anular venta: %s", e)
        return f"Error al anular venta: {e}", 500



@app.route("/ventas/<int:id>/emitir-sunat", methods=["POST"])
@login_required
@role_required('admin')
def emitir_sunat(id):
    conn = database.get_db()
    v = database.fetchone(conn, "SELECT * FROM ventas WHERE id=?", (id,))
    if not v or v["estado"] != "completada":
        conn.close()
        flash("Venta no encontrada o no completada", "error")
        return redirect(url_for("ver_venta", id=id))
    sunat_estado = v.get("sunat_estado", "pendiente") or "pendiente"
    if sunat_estado in ("aceptado", "pendiente"):
        flash("Ya se envio un comprobante a SUNAT (estado: %s)" % sunat_estado, "warning")
        conn.close()
        return redirect(url_for("ver_venta", id=id))
    items = database.fetchall(conn, "SELECT * FROM venta_items WHERE id_venta=?", (id,))
    conn.close()
    data = sunat_api.armar_comprobante(v, items)
    res = sunat_api.emitir_comprobante(data)
    if res.get("success") or res.get("ticket"):
        ticket = res.get("ticket", "")
        hash_val = res.get("hash", "")
        conn = database.get_db()
        database.execute(conn, "UPDATE ventas SET sunat_estado=?, sunat_ticket=?, sunat_hash=?, sunat_fecha_envio=? WHERE id=?",
            ("pendiente", ticket, hash_val, str(date.today()), id))
        conn.commit()
        conn.close()
        flash("Comprobante enviado a SUNAT. Ticket: " + str(ticket), "success")
    else:
        flash("Error al emitir: " + res.get("message", "Error desconocido"), "error")
    return redirect(url_for("ver_venta", id=id))

@app.route("/ventas/<int:id>/check-sunat", methods=["POST"])
@login_required
@role_required('admin')
def check_sunat(id):
    conn = database.get_db()
    v = database.fetchone(conn, "SELECT * FROM ventas WHERE id=?", (id,))
    if not v or not v.get("sunat_ticket"):
        conn.close()
        flash("No hay ticket pendiente", "warning")
        return redirect(url_for("ver_venta", id=id))
    conn.close()
    res = sunat_api.consultar_estado(v["sunat_ticket"])
    if res.get("success"):
        estado_sunat = res.get("payload", {}).get("estado", "pendiente")
        cdr = res.get("payload", {}).get("cdr", "")
        conn = database.get_db()
        database.execute(conn, "UPDATE ventas SET sunat_estado=?, sunat_cdr_xml=? WHERE id=?",
            (estado_sunat, cdr, id))
        conn.commit()
        conn.close()
        flash("Estado actualizado: " + estado_sunat, "success")
    else:
        flash("Error al consultar SUNAT: " + res.get("message", "Error desconocido"), "error")
    return redirect(url_for("ver_venta", id=id))

@app.route("/api/consultar-ruc", methods=["POST"])
def api_consultar_ruc():
    data = request.get_json()
    if not data or not data.get("ruc"):
        return jsonify({"success": False, "message": "RUC requerido"})
    res = sunat_api.consultar_ruc(data["ruc"])
    return jsonify(res)

@app.route("/api/consultar-dni", methods=["POST"])
def api_consultar_dni():
    data = request.get_json()
    if not data or not data.get("dni"):
        return jsonify({"success": False, "message": "DNI requerido"})
    res = sunat_api.consultar_dni(data["dni"])
    return jsonify(res)

@app.route("/nota-credito/<int:id>")
@login_required
def ver_nota_credito(id):
    conn = database.get_db()
    nc = database.fetchone(conn, "SELECT nc.*, v.tipo_comprobante, v.cliente_nombre, v.cliente_dni FROM notas_credito nc JOIN ventas v ON nc.id_venta=v.id WHERE nc.id=?", (id,))
    if not nc:
        conn.close()
        return redirect(url_for("listar_ventas"))
    items = database.fetchall(conn, "SELECT * FROM venta_items WHERE id_venta=?", (nc["id_venta"],))
    conn.close()
    return render_template("admin_nota_credito.html", nc=nc, items=items, title="Nota de Crédito", active="ventas", **_ctx())

@app.route("/consulta/<int:animal_id>")
def consulta(animal_id):
    if "doctor" not in session:
        return redirect(url_for("index"))
    conn = database.get_db()
    p = database.fetchone(conn,
        "SELECT a.*, d.nombre as dueno_nombre, d.telefono, d.direccion "
        "FROM animales a JOIN duenos d ON a.id_dueno = d.id WHERE a.id=?", (animal_id,))
    if not p:
        conn.close()
        return "Paciente no encontrado", 404
    historial = database.fetchall(conn,
        "SELECT id, fecha, hora, diagnostico, tratamiento, observaciones, peso, doctor, proximo_control "
        "FROM registros_medicos WHERE id_animal=? ORDER BY fecha DESC, hora DESC LIMIT 20",
        (animal_id,))
    sv = None
    if historial:
        ultimo_id = historial[0]["id"]
        sv = database.fetchone(conn, "SELECT * FROM signos_vitales WHERE id_registro=?", (ultimo_id,))
    vacunas = database.fetchall(conn,
        "SELECT * FROM vacunas WHERE id_animal=? ORDER BY fecha DESC", (animal_id,))
    alergias = database.fetchall(conn,
        "SELECT * FROM alergias WHERE id_animal=? ORDER BY id", (animal_id,))
    medicacion_activa = database.fetchall(conn,
        "SELECT * FROM medicacion WHERE id_animal=? AND activo=1 ORDER BY fecha_inicio DESC",
        (animal_id,))
    servicios = database.fetchall(conn,
        "SELECT id, nombre, precio FROM servicios_medicos ORDER BY nombre")
    servicios_grooming = database.fetchall(conn,
        "SELECT id, nombre, precio FROM servicios_grooming ORDER BY nombre")
    conn.close()
    precios_cita = {}
    for s in servicios:
        precios_cita[s["nombre"].lower()] = s["precio"]
    for s in servicios_grooming:
        precios_cita[s["nombre"].lower()] = s["precio"]
    now = datetime.now()
    return render_template("web_consulta.html",
                           paciente=p, dueno_nombre=p["dueno_nombre"],
                           dueno_telefono=p["telefono"] or "",
                           dueno_direccion=p["direccion"] or "",
                           doctor=session["doctor"],
                           hoy=now.strftime("%Y-%m-%d"),
                           hora=now.strftime("%H:%M"),
                            servicios=servicios,
                            precios_cita=json.dumps(precios_cita),
                            historial=historial,
                           signos_vitales=sv,
                           vacunas=vacunas,
                           alergias=alergias,
                            medicacion=medicacion_activa)

@app.route("/historial_completo/<int:animal_id>")
def web_historial_completo(animal_id):
    if "doctor" not in session:
        return redirect(url_for("index"))
    conn = database.get_db()
    p = database.fetchone(conn,
        "SELECT a.*, d.nombre as dueno_nombre, d.telefono, d.direccion "
        "FROM animales a JOIN duenos d ON a.id_dueno = d.id WHERE a.id=?", (animal_id,))
    if not p:
        conn.close()
        return "Paciente no encontrado", 404
    registros = database.fetchall(conn,
        "SELECT * FROM registros_medicos WHERE id_animal=? ORDER BY fecha DESC, hora DESC",
        (animal_id,))
    conn.close()
    return render_template("web_historial_completo.html",
        paciente=p, dueno=p, registros=registros, doctor=session["doctor"])

@app.route("/guardar_consulta/<int:animal_id>", methods=["POST"])
def guardar_consulta(animal_id):
    if "doctor" not in session:
        return redirect(url_for("index"))
    doctor = request.form.get("doctor", session.get("doctor", ""))
    fecha = request.form.get("fecha", str(date.today()))
    hora = request.form.get("hora", "")
    peso = request.form.get("peso", "").strip()
    diagnostico = request.form.get("diagnostico", "").strip()
    tratamiento = request.form.get("tratamiento", "").strip()
    observaciones = request.form.get("observaciones", "").strip()
    anamnesis = request.form.get("anamnesis", "").strip()
    diagnostico_presuntivo = request.form.get("diagnostico_presuntivo", "").strip()
    diagnostico_definitivo = request.form.get("diagnostico_definitivo", "").strip()
    cita_fecha = request.form.get("cita_fecha","").strip()
    conn = database.get_db()
    cur = database.execute(conn,
        "INSERT INTO registros_medicos (id_animal, fecha, hora, peso, doctor, diagnostico, tratamiento, "
        "observaciones, anamnesis, diagnostico_presuntivo, diagnostico_definitivo, proximo_control) "
        "VALUES (?,?,?,?,?,?,?,?,?,?,?,?)",
        (animal_id, fecha, hora, float(peso) if peso else None, doctor,
         diagnostico, tratamiento, observaciones, anamnesis,
         diagnostico_presuntivo, diagnostico_definitivo,
         cita_fecha if cita_fecha else None))
    if database._using_pg():
        conn.commit()
        reg_id = cur.fetchone()[0] if cur.description else None
        if not reg_id:
            reg_id = database.fetchone(conn, "SELECT MAX(id) as id FROM registros_medicos WHERE id_animal=?", (animal_id,))["id"]
    else:
        reg_id = cur.lastrowid

    temp = request.form.get("temperatura", "").strip()
    fc = request.form.get("fc", "").strip()
    fr = request.form.get("fr", "").strip()
    ps = request.form.get("presion_sistolica", "").strip()
    pd = request.form.get("presion_diastolica", "").strip()
    cm = request.form.get("color_mucosa", "").strip()
    if temp or fc or fr or ps or pd or cm:
        database.execute(conn,
            "INSERT INTO signos_vitales (id_registro, temperatura, frecuencia_cardiaca, "
            "frecuencia_respiratoria, presion_sistolica, presion_diastolica, color_mucosa) VALUES (?,?,?,?,?,?,?)",
            (reg_id, float(temp) if temp else None, int(fc) if fc else None,
             int(fr) if fr else None, int(ps) if ps else None, int(pd) if pd else None,
             cm if cm else None))

    vac_nombre = request.form.get("vacuna_nombre", "").strip()
    vac_fecha = request.form.get("vacuna_fecha", "").strip()
    if vac_nombre and vac_fecha:
        database.execute(conn,
            "INSERT INTO vacunas (id_animal, tipo, nombre, fecha, doctor) VALUES (?,?,?,?,?)",
            (animal_id, 'vacuna', vac_nombre, vac_fecha, doctor))

    alg_alergeno = request.form.get("alergeno", "").strip()
    alg_severidad = request.form.get("severidad", "leve")
    if alg_alergeno:
        database.execute(conn,
            "INSERT INTO alergias (id_animal, alergeno, severidad) VALUES (?,?,?)",
            (animal_id, alg_alergeno, alg_severidad))

    med_nombre = request.form.get("med_nombre", "").strip()
    med_dosis = request.form.get("med_dosis", "").strip()
    med_frecuencia = request.form.get("med_frecuencia", "").strip()
    med_via = request.form.get("med_via", "").strip()
    med_duracion = request.form.get("med_duracion", "").strip()
    if med_nombre and med_dosis:
        database.execute(conn,
            "INSERT INTO medicacion (id_animal, medicamento, dosis, frecuencia, via, fecha_inicio, observaciones) "
            "VALUES (?,?,?,?,?,?,?)",
            (animal_id, med_nombre, med_dosis, med_frecuencia, med_via, fecha, med_duracion))

    sv_selected = request.form.getlist("sv_sel")
    if sv_selected:
        servicios_list = []
        total = 0.0
        for sv_id in sv_selected:
            precio_str = request.form.get(f"sv_precio_{sv_id}", "0")
            conn2 = database.get_db()
            sv = database.fetchone(conn2, "SELECT id, nombre, precio FROM servicios_medicos WHERE id=?", (sv_id,))
            conn2.close()
            if sv:
                try:
                    p = float(precio_str)
                except:
                    p = float(sv["precio"])
                servicios_list.append({"id": sv["id"], "nombre": sv["nombre"], "precio": p})
                total += p
        if servicios_list:
            database.execute(conn,
                "INSERT INTO cobros_pendientes (id_registro, id_animal, servicios, total, created_by, fecha) "
                "VALUES (?,?,?,?,?,?)",
                (reg_id, animal_id, json.dumps(servicios_list), total, None, fecha))

    if cita_fecha:
        a_info = database.fetchone(conn, "SELECT id_dueno FROM animales WHERE id=?", (animal_id,))
        if a_info:
            database.execute(conn,
                "INSERT INTO citas (id_animal,id_dueno,fecha,hora,motivo,tipo,precio) VALUES (?,?,?,?,?,?,?)",
                (animal_id, a_info["id_dueno"], cita_fecha, request.form.get("cita_hora",""),
                 request.form.get("cita_motivo",""),
                 request.form.get("cita_tipo","consulta"),
                 float(request.form.get("cita_precio",0))))
    conn.commit()
    conn.close()
    conn2 = database.get_db()
    p = database.fetchone(conn2, "SELECT nombre FROM animales WHERE id=?", (animal_id,))
    conn2.close()
    return render_template("web_exito.html", paciente=p["nombre"] if p else "",
                           paciente_id=animal_id, doctor=doctor,
                           fecha=fecha, hora=hora, diagnostico=diagnostico,
                           registro_id=reg_id)

@app.route("/api/ia_sugerir", methods=["POST"])
def ia_sugerir():
    api_key = AI_CONFIG.get("gemini_api_key", "")
    if not api_key:
        return {"error":"IA no configurada. Agrega gemini_api_key en config.json"}, 400
    data = request.get_json() or {}
    especie = data.get("especie","")
    raza = data.get("raza","")
    edad = data.get("edad","")
    sexo = data.get("sexo","")
    peso = data.get("peso","")
    anamnesis = data.get("anamnesis","")
    temperatura = data.get("temperatura","")
    fc = data.get("fc","")
    fr = data.get("fr","")
    sintomas = data.get("sintomas","")

    prompt = f"""Eres un veterinario experto. Con los siguientes datos del paciente, sugiere un diagnóstico presuntivo y un tratamiento.

Datos del paciente:
- Especie: {especie or "No especificada"}
- Raza: {raza or "No especificada"}
- Edad: {edad or "No especificada"} años
- Sexo: {sexo or "No especificado"}
- Peso: {peso or "No especificado"} kg

Motivo de consulta (anamnesis): {anamnesis or "No especificado"}

Signos vitales:
- Temperatura: {temperatura or "No registrada"} °C
- FC: {fc or "No registrada"} lpm
- FR: {fr or "No registrada"} rpm

Síntomas adicionales: {sintomas or "No especificados"}

Responde ÚNICAMENTE en el siguiente formato JSON (sin explicaciones adicionales):
{{"diagnostico_presuntivo": "texto del diagnóstico presuntivo", "tratamiento": "texto del tratamiento recomendado con medicamentos, dosis y duración"}}"""

    try:
        resp = requests.post(
            f"https://generativelanguage.googleapis.com/v1beta/models/gemini-2.5-flash:generateContent?key={api_key}",
            json={"contents":[{"role":"user","parts":[{"text":prompt}]}]},
            timeout=30
        )
        result = resp.json()
        if "candidates" not in result or not result["candidates"]:
            err = result.get("error",{}).get("message","") or str(result)
            return {"error":f"Error de IA: {err}"}, 500
        text = result["candidates"][0]["content"]["parts"][0]["text"]
        text = text.strip().removeprefix("```json").removesuffix("```").strip()
        suggestion = json.loads(text)
        return {"diagnostico_presuntivo": suggestion.get("diagnostico_presuntivo",""),
                "tratamiento": suggestion.get("tratamiento","")}
    except Exception as e:
        return {"error":f"Error al consultar IA: {str(e)}"}, 500

@app.route("/agregar_insumos/<int:registro_id>")
def web_agregar_insumos(registro_id):
    if "doctor" not in session:
        return redirect(url_for("index"))
    conn = database.get_db()
    r = database.fetchone(conn, """
        SELECT r.*, a.nombre as animal_nombre, d.nombre as dueno_nombre
        FROM registros_medicos r
        JOIN animales a ON r.id_animal = a.id
        JOIN duenos d ON a.id_dueno = d.id
        WHERE r.id=?
    """, (registro_id,))
    productos = database.fetchall(conn,
        "SELECT id, nombre, stock, precio_venta FROM productos WHERE activo=1 ORDER BY nombre")
    conn.close()
    if not r:
        return "Registro no encontrado", 404
    return render_template("web_insumos.html", registro=r, productos=productos,
                           doctor=session["doctor"])

@app.route("/guardar_insumos/<int:registro_id>", methods=["POST"])
def guardar_insumos(registro_id):
    if "doctor" not in session:
        return redirect(url_for("index"))
    conn = database.get_db()
    insumo_ids = request.form.getlist("insumo_id")
    insumo_cants = request.form.getlist("insumo_cant")
    try:
        for iid, cant in zip(insumo_ids, insumo_cants):
            q = int(cant)
            pid = int(iid)
            if q > 0 and pid > 0:
                prod = database.fetchone(conn, "SELECT stock FROM productos WHERE id=? AND activo=1", (pid,))
                if not prod or prod["stock"] < q:
                    conn.rollback()
                    conn.close()
                    return f"Stock insuficiente para producto ID {pid}", 400
                database.execute(conn,
                    "INSERT INTO insumos_utilizados (id_registro, id_producto, cantidad) VALUES (?,?,?)",
                    (registro_id, pid, q))
                database.execute(conn,
                    "UPDATE productos SET stock = stock - ? WHERE id = ? AND stock >= ?", (q, pid, q))
                registrar_movimiento(conn, pid, 'salida', q, 'insumo', registro_id, 'Uso en registro médico #{}'.format(registro_id))
        conn.commit()
        conn.close()
        return redirect(url_for("web_pendientes_insumos", ok=1))
    except Exception as e:
        conn.rollback()
        conn.close()
        logger.error("Error al guardar insumos: %s", e)
        return f"Error al guardar insumos: {e}", 500

@app.route("/pendientes_insumos")
def web_pendientes_insumos():
    if "doctor" not in session:
        return redirect(url_for("index"))
    conn = database.get_db()
    doctorname = session["doctor"]
    hoy_str = date.today().isoformat()
    rows = database.fetchall(conn, """
        SELECT r.id, r.fecha, r.diagnostico, a.nombre as animal_nombre,
               d.nombre as dueno_nombre,
               (SELECT COUNT(*) FROM insumos_utilizados iu WHERE iu.id_registro = r.id) as insumos_count
        FROM registros_medicos r
        JOIN animales a ON r.id_animal = a.id
        JOIN duenos d ON a.id_dueno = d.id
        WHERE r.doctor=? AND r.fecha=?
        ORDER BY r.hora DESC
    """, (doctorname, hoy_str))
    conn.close()
    return render_template("web_pendientes_insumos.html",
                           registros=rows, doctor=doctorname, hoy=hoy_str)

@app.route("/subir_foto/<int:animal_id>", methods=["GET", "POST"])
def web_subir_foto(animal_id):
    if "doctor" not in session:
        return redirect(url_for("index"))
    conn = database.get_db()
    a = database.fetchone(conn, "SELECT id, nombre FROM animales WHERE id=?", (animal_id,))
    conn.close()
    if not a:
        return "Animal no encontrado", 404
    error = None
    if request.method == "POST":
        foto = request.files.get("foto")
        if foto and foto.filename:
            try:
                fname, _ = storage.save_file(foto, 'photos')
                conn2 = database.get_db()
                database.execute(conn2, "UPDATE animales SET foto=? WHERE id=?", (fname, animal_id))
                conn2.commit()
                conn2.close()
                return redirect(url_for("consulta", animal_id=animal_id))
            except Exception as e:
                error = "Error al guardar la foto: {}".format(str(e))
        else:
            error = "No se recibi\u00f3 ninguna foto"
    return render_template("web_subir_foto.html", animal=a, error=error)

# ===== GROOMING =====
@app.route("/grooming")
@login_required
@role_required('admin', 'recepcion')
def listar_grooming():
    conn = database.get_db()
    registros = database.fetchall(conn, """
        SELECT h.*, a.nombre as animal_nombre, sg.nombre as servicio_nombre
        FROM historial_grooming h
        JOIN animales a ON h.id_animal=a.id
        LEFT JOIN servicios_grooming sg ON h.id_servicio=sg.id
        ORDER BY h.fecha DESC, h.id DESC LIMIT 100
    """)
    animales = database.fetchall(conn, "SELECT a.*, d.nombre as dueno_nombre FROM animales a JOIN duenos d ON a.id_dueno=d.id ORDER BY a.nombre")
    servicios = database.fetchall(conn, "SELECT * FROM servicios_grooming ORDER BY nombre")
    conn.close()
    return render_template("admin_grooming.html", registros=registros, animales=animales,
                           servicios=servicios, title="Grooming", active="grooming", **_ctx())

@app.route("/grooming/nuevo", methods=["POST"])
@login_required
@role_required('admin', 'recepcion')
def crear_grooming():
    conn = database.get_db()
    database.execute(conn,
        "INSERT INTO historial_grooming (id_animal,id_servicio,fecha,observaciones,precio,estilista) VALUES (?,?,?,?,?,?)",
        (request.form["id_animal"], request.form.get("id_servicio") or None,
         request.form.get("fecha", str(date.today())), request.form.get("observaciones",""),
         float(request.form.get("precio",0)), request.form.get("estilista","")))
    conn.commit(); conn.close()
    return redirect(url_for("listar_grooming"))

@app.route("/grooming/<int:id>/eliminar")
@login_required
@role_required('admin')
def eliminar_grooming(id):
    conn = database.get_db()
    database.execute(conn, "DELETE FROM historial_grooming WHERE id=?", (id,))
    conn.commit(); conn.close()
    return redirect(url_for("listar_grooming"))

# ===== SERVICIOS GROOMING CRUD =====
@app.route("/servicios-grooming")
@login_required
def listar_servicios_grooming():
    conn = database.get_db()
    items = database.fetchall(conn, "SELECT * FROM servicios_grooming ORDER BY nombre")
    conn.close()
    return render_template("admin_servicios_grooming.html", items=items,
        title="Servicios de Grooming", active="servicios_grooming", **_ctx())

@app.route("/servicios-grooming/nuevo", methods=["POST"])
@login_required
def crear_servicio_grooming():
    conn = database.get_db()
    database.execute(conn,
        "INSERT INTO servicios_grooming (nombre,descripcion,precio,duracion_minutos,tipo) VALUES (?,?,?,?,?)",
        (request.form["nombre"], request.form.get("descripcion",""),
         float(request.form.get("precio",0)), int(request.form.get("duracion_minutos",30)),
         request.form.get("tipo","Baño")))
    conn.commit(); conn.close()
    return redirect(url_for("listar_servicios_grooming"))

@app.route("/servicios-grooming/<int:id>/editar", methods=["POST"])
@login_required
def editar_servicio_grooming(id):
    conn = database.get_db()
    database.execute(conn,
        "UPDATE servicios_grooming SET nombre=?,descripcion=?,precio=?,duracion_minutos=?,tipo=? WHERE id=?",
        (request.form["nombre"], request.form.get("descripcion",""),
         float(request.form.get("precio",0)), int(request.form.get("duracion_minutos",30)),
         request.form.get("tipo","Baño"), id))
    conn.commit(); conn.close()
    return redirect(url_for("listar_servicios_grooming"))

@app.route("/servicios-grooming/<int:id>/eliminar")
@login_required
def eliminar_servicio_grooming(id):
    conn = database.get_db()
    database.execute(conn, "DELETE FROM servicios_grooming WHERE id=?", (id,))
    conn.commit(); conn.close()
    return redirect(url_for("listar_servicios_grooming"))

@app.route("/servicios-grooming/exportar")
@login_required
def exportar_servicios_grooming():
    conn = database.get_db()
    items = database.fetchall(conn, "SELECT * FROM servicios_grooming ORDER BY nombre")
    conn.close()
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Border, Side
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Grooming"
    hdr_fill = PatternFill(start_color="1abc9c", end_color="1abc9c", fill_type="solid")
    hdr_font = Font(color="ffffff", bold=True, size=11)
    thin = Side(style="thin", color="cccccc"); border = Border(top=thin, left=thin, right=thin, bottom=thin)
    for c, h in enumerate(["ID", "Nombre", "Descripción", "Precio", "Duración (min)", "Tipo"], 1):
        cell = ws.cell(row=1, column=c, value=h); cell.fill = hdr_fill; cell.font = hdr_font; cell.border = border
    for r, row in enumerate(items, 2):
        for c, v in enumerate([row["id"], row["nombre"], row["descripcion"] or "", row["precio"], row["duracion_minutos"], row["tipo"]], 1):
            cell = ws.cell(row=r, column=c, value=v); cell.border = border
            if c == 4: cell.number_format = '#,##0.00'
    ws.column_dimensions["A"].width = 6; ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 35; ws.column_dimensions["D"].width = 10
    ws.column_dimensions["E"].width = 14; ws.column_dimensions["F"].width = 18
    out = io.BytesIO(); wb.save(out); out.seek(0)
    return Response(out.getvalue(), mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=servicios_grooming.xlsx"})

# ===== SERVICIOS MEDICOS =====
@app.route("/servicios")
@login_required
def listar_servicios():
    conn = database.get_db()
    servicios = database.fetchall(conn, """
        SELECT sm.*,
               (SELECT COUNT(*) FROM servicio_insumos si WHERE si.id_servicio=sm.id) as insumos_count,
               (SELECT COALESCE(SUM(si.cantidad * p.precio_compra), 0) FROM servicio_insumos si JOIN productos p ON si.id_producto=p.id WHERE si.id_servicio=sm.id) as costo_insumos
        FROM servicios_medicos sm ORDER BY sm.id DESC""")
    insumos_map = {}
    historial_map = {}
    for s in servicios:
        ins = database.fetchall(conn,
            "SELECT si.*, p.nombre as producto_nombre, p.stock FROM servicio_insumos si JOIN productos p ON si.id_producto=p.id WHERE si.id_servicio=?",
            (s["id"],))
        insumos_map[s["id"]] = ins
        hist = database.fetchall(conn, """
            SELECT iu.id, iu.cantidad, iu.id_producto, p.nombre as producto_nombre,
                   v.id as venta_id, v.fecha, v.cliente_nombre
            FROM insumos_utilizados iu
            JOIN productos p ON iu.id_producto=p.id
            JOIN ventas v ON iu.id_registro=v.id
            JOIN venta_items vi ON vi.id_venta=v.id AND vi.referencia_id=? AND vi.tipo_item='servicio_medico'
            WHERE iu.id_registro IN (SELECT id FROM ventas)
            ORDER BY v.fecha DESC LIMIT 20""", (s["id"],))
        historial_map[s["id"]] = hist
    productos = database.fetchall(conn, "SELECT id,nombre,stock FROM productos WHERE activo=1 ORDER BY nombre")
    conn.close()
    return render_template("admin_servicios.html", servicios=servicios, insumos_map=insumos_map,
                           historial_map=historial_map, productos=productos,
                           title="Servicios", active="servicios", **_ctx())

@app.route("/servicios/nuevo", methods=["POST"])
@login_required
def crear_servicio():
    conn = database.get_db()
    database.execute(conn,
        "INSERT INTO servicios_medicos (nombre,descripcion,precio,tipo) VALUES (?,?,?,?)",
        (request.form["nombre"], request.form.get("descripcion",""),
         float(request.form.get("precio",0) or 0), request.form.get("tipo","consulta")))
    conn.commit(); conn.close()
    return redirect(url_for("listar_servicios"))

@app.route("/servicios/<int:id>/editar", methods=["GET", "POST"])
@login_required
def editar_servicio(id):
    conn = database.get_db()
    if request.method == "POST":
        database.execute(conn,
            "UPDATE servicios_medicos SET nombre=?,descripcion=?,precio=?,tipo=? WHERE id=?",
            (request.form["nombre"], request.form.get("descripcion",""),
             float(request.form.get("precio",0) or 0), request.form.get("tipo","consulta"), id))
        conn.commit(); conn.close()
        return redirect(url_for("listar_servicios"))
    item = database.fetchone(conn, "SELECT * FROM servicios_medicos WHERE id=?", (id,))
    conn.close()
    return render_template("admin_form.html", fields=[
        ("nombre","Nombre",True),("descripcion","Descripci\u00f3n","textarea"),
        ("precio","Precio","number"),("tipo","Tipo")
    ], item=item, title="Editar Servicio", active="servicios", **_ctx())

@app.route("/servicios/<int:id>/eliminar")
@login_required
def eliminar_servicio(id):
    conn = database.get_db()
    used = database.fetchone(conn,
        "SELECT COUNT(*) as c FROM venta_items WHERE tipo_item='servicio_medico' AND referencia_id=?", (id,))
    if used and used["c"] > 0:
        conn.close()
        return "No se puede eliminar: el servicio tiene ventas asociadas", 400
    database.execute(conn, "DELETE FROM servicio_insumos WHERE id_servicio=?", (id,))
    database.execute(conn, "DELETE FROM servicios_medicos WHERE id=?", (id,))
    conn.commit(); conn.close()
    return redirect(url_for("listar_servicios"))

# ===== INSUMOS POR SERVICIO =====
@app.route("/servicios/<int:id>/insumos", methods=["GET", "POST"])
@login_required
def servicio_insumos(id):
    conn = database.get_db()
    servicio = database.fetchone(conn, "SELECT * FROM servicios_medicos WHERE id=?", (id,))
    if not servicio: conn.close(); return redirect(url_for("listar_servicios"))
    if request.method == "POST":
        database.execute(conn, "DELETE FROM servicio_insumos WHERE id_servicio=?", (id,))
        for pid, cant in zip(request.form.getlist("producto_id"), request.form.getlist("cantidad[]")):
            if int(cant) > 0:
                database.execute(conn, "INSERT INTO servicio_insumos (id_servicio,id_producto,cantidad) VALUES (?,?,?)", (id, int(pid), int(cant)))
        conn.commit(); conn.close()
        return redirect(url_for("servicio_insumos", id=id))
    insumos = database.fetchall(conn,
        "SELECT si.*, p.nombre as producto_nombre, p.stock FROM servicio_insumos si JOIN productos p ON si.id_producto=p.id WHERE si.id_servicio=? ORDER BY p.nombre", (id,))
    productos = database.fetchall(conn, "SELECT id,nombre,stock FROM productos WHERE activo=1 ORDER BY nombre")
    conn.close()
    return render_template("admin_servicio_insumos.html", servicio=servicio, insumos=insumos, productos=productos,
        title="Insumos: "+servicio["nombre"], active="servicios", **_ctx())

@app.route("/servicios/<int:id>/insumos/agregar", methods=["POST"])
@login_required
def servicio_insumo_agregar(id):
    conn = database.get_db()
    pid = int(request.form["producto_id"])
    cant = int(request.form.get("cantidad", 1))
    existing = database.fetchone(conn, "SELECT id,cantidad FROM servicio_insumos WHERE id_servicio=? AND id_producto=?", (id, pid))
    if existing:
        database.execute(conn, "UPDATE servicio_insumos SET cantidad=cantidad+? WHERE id=?", (cant, existing["id"]))
    else:
        database.execute(conn, "INSERT INTO servicio_insumos (id_servicio,id_producto,cantidad) VALUES (?,?,?)", (id, pid, cant))
    conn.commit(); conn.close()
    return "", 200

@app.route("/servicios/<int:id>/insumos/agregar-nuevo", methods=["POST"])
@login_required
def servicio_insumo_agregar_nuevo(id):
    conn = database.get_db()
    nombre = request.form["nombre"]
    precio_compra = float(request.form.get("precio_compra", 0) or 0)
    stock = int(request.form.get("stock", 0) or 0)
    cant = int(request.form.get("cantidad", 1) or 1)
    # Check if product already exists with same name
    existing = database.fetchone(conn, "SELECT id FROM productos WHERE nombre=? AND activo=1", (nombre,))
    if existing:
        pid = existing["id"]
    else:
        cur = database.execute(conn,
            "INSERT INTO productos (nombre,precio_compra,precio_venta,stock,activo,categoria) VALUES (?,?,0,?,1,'Insumo')",
            (nombre, precio_compra, stock))
        pid = cur.lastrowid
    # Link as insumo
    linked = database.fetchone(conn, "SELECT id,cantidad FROM servicio_insumos WHERE id_servicio=? AND id_producto=?", (id, pid))
    if linked:
        database.execute(conn, "UPDATE servicio_insumos SET cantidad=cantidad+? WHERE id=?", (cant, linked["id"]))
    else:
        database.execute(conn, "INSERT INTO servicio_insumos (id_servicio,id_producto,cantidad) VALUES (?,?,?)", (id, pid, cant))
    conn.commit(); conn.close()
    return "", 200

@app.route("/servicios/<int:id>/insumos/eliminar", methods=["POST"])
@login_required
def servicio_insumo_eliminar(id):
    conn = database.get_db()
    pid = int(request.form["producto_id"])
    database.execute(conn, "DELETE FROM servicio_insumos WHERE id_servicio=? AND id_producto=?", (id, pid))
    conn.commit(); conn.close()
    return "", 200

@app.route("/servicios/<int:id>/insumos/actualizar-cantidad", methods=["POST"])
@login_required
def servicio_insumo_actualizar_cantidad(id):
    conn = database.get_db()
    pid = int(request.form["producto_id"])
    cantidad = int(request.form.get("cantidad", 1))
    database.execute(conn, "UPDATE servicio_insumos SET cantidad=? WHERE id_servicio=? AND id_producto=?", (cantidad, id, pid))
    conn.commit(); conn.close()
    return "", 200

@app.route("/servicios/<int:id>/usar-insumos", methods=["POST"])
@login_required
def usar_insumos_servicio(id):
    conn = database.get_db()
    registro_id = request.form.get("registro_id")
    insumos = database.fetchall(conn, "SELECT si.*, p.stock FROM servicio_insumos si JOIN productos p ON si.id_producto=p.id WHERE si.id_servicio=?", (id,))
    for ins in insumos:
        if ins["stock"] < ins["cantidad"]:
            conn.close()
            return f"Stock insuficiente para {ins['producto_nombre'] if 'producto_nombre' in ins else 'producto'}", 400
        database.execute(conn, "INSERT INTO insumos_utilizados (id_registro,id_producto,cantidad) VALUES (?,?,?)", (registro_id, ins["id_producto"], ins["cantidad"]))
        database.execute(conn, "UPDATE productos SET stock=stock-? WHERE id=?", (ins["cantidad"], ins["id_producto"]))
        registrar_movimiento(conn, ins["id_producto"], 'salida', ins["cantidad"], 'servicio', registro_id, 'Uso automático por servicio #{}'.format(id))
    conn.commit(); conn.close()
    return redirect(request.referrer or url_for("listar_servicios"))

# ===== COMBOS =====
@app.route("/combos")
@login_required
def listar_combos():
    conn = database.get_db()
    rows = database.fetchall(conn, "SELECT * FROM combos WHERE activo=1 ORDER BY nombre")
    combos = []
    for c in rows:
        item = dict(c)
        item["combo_items"] = database.fetchall(conn, "SELECT * FROM combo_items WHERE id_combo=?", (c["id"],))
        combos.append(item)
    conn.close()
    return render_template("admin_combos.html", combos=combos, title="Combos/Paquetes", active="combos", **_ctx())

@app.route("/combos/nuevo", methods=["GET", "POST"])
@login_required
@role_required('admin')
def crear_combo():
    conn = database.get_db()
    if request.method == "POST":
        nombre = request.form["nombre"]
        descripcion = request.form.get("descripcion","")
        database.execute(conn, "INSERT INTO combos (nombre,descripcion,precio_total) VALUES (?,?,?)", (nombre, descripcion, 0))
        combo_id = database.fetchone(conn, "SELECT MAX(id) as id FROM combos")["id"]
        precios = request.form.getlist("precio_item")
        total = 0
        for i, (tid, rid, nom, cant, pu) in enumerate(zip(
            request.form.getlist("tipo_item"), request.form.getlist("referencia_id"),
            request.form.getlist("nombre_item"), request.form.getlist("cantidad_item"),
            precios)):
            c = int(cant) if cant else 1
            p = float(pu) if pu else 0
            database.execute(conn, "INSERT INTO combo_items (id_combo,tipo_item,referencia_id,nombre,cantidad,precio_unitario) VALUES (?,?,?,?,?,?)",
                (combo_id, tid, int(rid) if rid else None, nom, c, p))
            total += c * p
        database.execute(conn, "UPDATE combos SET precio_total=? WHERE id=?", (total, combo_id))
        conn.commit(); conn.close()
        return redirect(url_for("listar_combos"))
    productos = database.fetchall(conn, "SELECT id,nombre,precio_venta FROM productos WHERE activo=1 ORDER BY nombre")
    servicios_med = database.fetchall(conn, "SELECT id,nombre,precio FROM servicios_medicos ORDER BY nombre")
    servicios_groom = database.fetchall(conn, "SELECT id,nombre,precio FROM servicios_grooming ORDER BY nombre")
    conn.close()
    return render_template("admin_combo_form.html", productos=productos, servicios_med=servicios_med,
        servicios_groom=servicios_groom, title="Nuevo Combo", active="combos", **_ctx())

@app.route("/combos/<int:id>/editar", methods=["GET", "POST"])
@login_required
@role_required('admin')
def editar_combo(id):
    conn = database.get_db()
    if request.method == "POST":
        database.execute(conn, "UPDATE combos SET nombre=?,descripcion=? WHERE id=?", (request.form["nombre"], request.form.get("descripcion",""), id))
        database.execute(conn, "DELETE FROM combo_items WHERE id_combo=?", (id,))
        precios = request.form.getlist("precio_item")
        total = 0
        for tid, rid, nom, cant, pu in zip(
            request.form.getlist("tipo_item"), request.form.getlist("referencia_id"),
            request.form.getlist("nombre_item"), request.form.getlist("cantidad_item"), precios):
            c = int(cant) if cant else 1; p = float(pu) if pu else 0
            database.execute(conn, "INSERT INTO combo_items (id_combo,tipo_item,referencia_id,nombre,cantidad,precio_unitario) VALUES (?,?,?,?,?,?)",
                (id, tid, int(rid) if rid else None, nom, c, p))
            total += c * p
        database.execute(conn, "UPDATE combos SET precio_total=? WHERE id=?", (total, id))
        conn.commit(); conn.close()
        return redirect(url_for("listar_combos"))
    c = database.fetchone(conn, "SELECT * FROM combos WHERE id=?", (id,))
    if not c: conn.close(); return redirect(url_for("listar_combos"))
    c["combo_items"] = database.fetchall(conn, "SELECT * FROM combo_items WHERE id_combo=?", (id,))
    productos = database.fetchall(conn, "SELECT id,nombre,precio_venta FROM productos WHERE activo=1 ORDER BY nombre")
    servicios_med = database.fetchall(conn, "SELECT id,nombre,precio FROM servicios_medicos ORDER BY nombre")
    servicios_groom = database.fetchall(conn, "SELECT id,nombre,precio FROM servicios_grooming ORDER BY nombre")
    conn.close()
    return render_template("admin_combo_form.html", item=c, productos=productos, servicios_med=servicios_med,
        servicios_groom=servicios_groom, title="Editar Combo", active="combos", **_ctx())

@app.route("/combos/exportar")
@login_required
@role_required('admin')
def exportar_combos():
    conn = database.get_db()
    rows = database.fetchall(conn, "SELECT * FROM combos WHERE activo=1 ORDER BY nombre")
    combos = []
    for c in rows:
        d = dict(c); d["combo_items"] = database.fetchall(conn, "SELECT * FROM combo_items WHERE id_combo=?", (c["id"],))
        combos.append(d)
    conn.close()
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Border, Side
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Combos"
    hdr_fill = PatternFill(start_color="1abc9c", end_color="1abc9c", fill_type="solid")
    hdr_font = Font(color="ffffff", bold=True, size=11)
    thin = Side(style="thin", color="cccccc"); border = Border(top=thin, left=thin, right=thin, bottom=thin)
    for c, h in enumerate(["ID", "Nombre", "Descripción", "Items", "Precio Total"], 1):
        cell = ws.cell(row=1, column=c, value=h); cell.fill = hdr_fill; cell.font = hdr_font; cell.border = border
    for r, co in enumerate(combos, 2):
        items_str = ", ".join(f"{it['nombre']} x{it['cantidad']}" for it in co["combo_items"])
        for col, v in enumerate([co["id"], co["nombre"], co["descripcion"] or "", items_str, co["precio_total"]], 1):
            cell = ws.cell(row=r, column=col, value=v); cell.border = border
            if col == 5: cell.number_format = '#,##0.00'
    ws.column_dimensions["A"].width = 6; ws.column_dimensions["B"].width = 28; ws.column_dimensions["C"].width = 35
    ws.column_dimensions["D"].width = 50; ws.column_dimensions["E"].width = 12
    out = io.BytesIO(); wb.save(out); out.seek(0)
    return Response(out.getvalue(), mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=combos.xlsx"})

@app.route("/combos/<int:id>/eliminar")
@login_required
@role_required('admin')
def eliminar_combo(id):
    conn = database.get_db()
    database.execute(conn, "DELETE FROM combo_items WHERE id_combo=?", (id,))
    database.execute(conn, "DELETE FROM combos WHERE id=?", (id,))
    conn.commit(); conn.close()
    return redirect(url_for("listar_combos"))

# ===== PROVEEDORES =====
@app.route("/proveedores")
@login_required
@role_required('admin')
def listar_proveedores():
    conn = database.get_db()
    q = request.args.get("q","").strip()
    where = " WHERE activo=1"
    params = []
    if q:
        where += " AND (nombre LIKE ? OR contacto LIKE ? OR ruc LIKE ?)"
        params.extend([f"%{q}%", f"%{q}%", f"%{q}%"])
    proveedores = database.fetchall(conn, "SELECT * FROM proveedores"+where+" ORDER BY nombre", params)
    conn.close()
    return render_template("admin_proveedores.html", proveedores=proveedores, q=q, title="Proveedores", active="proveedores", **_ctx())

@app.route("/proveedores/nuevo", methods=["GET", "POST"])
@login_required
@role_required('admin')
def crear_proveedor():
    conn = database.get_db()
    if request.method == "POST":
        database.execute(conn, "INSERT INTO proveedores (nombre,contacto,telefono,email,direccion,ruc) VALUES (?,?,?,?,?,?)",
            (request.form["nombre"], request.form.get("contacto",""), request.form.get("telefono",""),
             request.form.get("email",""), request.form.get("direccion",""), request.form.get("ruc","")))
        conn.commit(); conn.close()
        return redirect(url_for("listar_proveedores"))
    conn.close()
    return redirect(url_for("listar_proveedores"))

@app.route("/proveedores/<int:id>/editar", methods=["GET", "POST"])
@login_required
@role_required('admin')
def editar_proveedor(id):
    conn = database.get_db()
    if request.method == "POST":
        database.execute(conn, "UPDATE proveedores SET nombre=?,contacto=?,telefono=?,email=?,direccion=?,ruc=? WHERE id=?",
            (request.form["nombre"], request.form.get("contacto",""), request.form.get("telefono",""),
             request.form.get("email",""), request.form.get("direccion",""), request.form.get("ruc",""), id))
        conn.commit(); conn.close()
        return redirect(url_for("listar_proveedores"))
    conn.close()
    return redirect(url_for("listar_proveedores"))

@app.route("/proveedores/<int:id>/eliminar")
@login_required
@role_required('admin')
def eliminar_proveedor(id):
    conn = database.get_db()
    database.execute(conn, "UPDATE proveedores SET activo=0 WHERE id=?", (id,))
    conn.commit(); conn.close()
    return redirect(url_for("listar_proveedores"))

@app.route("/proveedores/exportar")
@login_required
@role_required('admin')
def exportar_proveedores():
    conn = database.get_db()
    q = request.args.get("q","").strip()
    where = " WHERE activo=1"; params = []
    if q:
        where += " AND (nombre LIKE ? OR contacto LIKE ? OR ruc LIKE ?)"
        params.extend([f"%{q}%", f"%{q}%", f"%{q}%"])
    proveedores = database.fetchall(conn, "SELECT * FROM proveedores"+where+" ORDER BY nombre", params) if params else database.fetchall(conn, "SELECT * FROM proveedores WHERE activo=1 ORDER BY nombre")
    conn.close()
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Border, Side
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Proveedores"
    hdr_fill = PatternFill(start_color="1abc9c", end_color="1abc9c", fill_type="solid")
    hdr_font = Font(color="ffffff", bold=True, size=11)
    thin = Side(style="thin", color="cccccc"); border = Border(top=thin, left=thin, right=thin, bottom=thin)
    for c, h in enumerate(["ID", "Nombre", "Contacto", "Teléfono", "RUC", "Email", "Dirección"], 1):
        cell = ws.cell(row=1, column=c, value=h); cell.fill = hdr_fill; cell.font = hdr_font; cell.border = border
    for r, p in enumerate(proveedores, 2):
        for col, v in enumerate([p["id"], p["nombre"], p["contacto"] or "", p["telefono"] or "", p["ruc"] or "", p["email"] or "", p["direccion"] or ""], 1):
            cell = ws.cell(row=r, column=col, value=v); cell.border = border
    ws.column_dimensions["A"].width = 6; ws.column_dimensions["B"].width = 28; ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 14; ws.column_dimensions["E"].width = 16; ws.column_dimensions["F"].width = 28
    ws.column_dimensions["G"].width = 35
    out = io.BytesIO(); wb.save(out); out.seek(0)
    return Response(out.getvalue(), mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=proveedores.xlsx"})

# ===== ORDENES DE COMPRA =====
@app.route("/ordenes-compra")
@login_required
@role_required('admin')
def listar_ordenes_compra():
    conn = database.get_db()
    estado = request.args.get("estado","")
    q = request.args.get("q","").strip()
    where = ""; params = []
    if estado:
        where = " WHERE o.estado=?"
        params.append(estado)
    if q:
        where = (where + " AND" if where else " WHERE") + " p.nombre LIKE ?"
        params.append(f"%{q}%")
    ordenes = database.fetchall(conn, """
        SELECT o.*, p.nombre as proveedor_nombre
        FROM ordenes_compra o JOIN proveedores p ON o.id_proveedor=p.id
    """+where+" ORDER BY o.fecha DESC LIMIT 100", params)
    conn.close()
    return render_template("admin_ordenes_compra.html", ordenes=ordenes, estado=estado, q=q,
        title="&Oacute;rdenes de Compra", active="ordenes_compra", **_ctx())

@app.route("/ordenes-compra/exportar")
@login_required
@role_required('admin')
def exportar_ordenes_compra():
    conn = database.get_db()
    estado = request.args.get("estado","")
    q = request.args.get("q","").strip()
    where = ""; params = []
    if estado:
        where = " WHERE o.estado=?"
        params.append(estado)
    if q:
        where = (where + " AND" if where else " WHERE") + " p.nombre LIKE ?"
        params.append(f"%{q}%")
    ordenes = database.fetchall(conn, """
        SELECT o.*, p.nombre as proveedor_nombre
        FROM ordenes_compra o JOIN proveedores p ON o.id_proveedor=p.id
    """+where+" ORDER BY o.fecha DESC LIMIT 100", params)
    conn.close()
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Border, Side
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Órdenes Compra"
    hdr_fill = PatternFill(start_color="1abc9c", end_color="1abc9c", fill_type="solid")
    hdr_font = Font(color="ffffff", bold=True, size=11)
    thin = Side(style="thin", color="cccccc"); border = Border(top=thin, left=thin, right=thin, bottom=thin)
    for c, h in enumerate(["# OC", "Proveedor", "Fecha", "Total", "Estado"], 1):
        cell = ws.cell(row=1, column=c, value=h); cell.fill = hdr_fill; cell.font = hdr_font; cell.border = border
    for r, o in enumerate(ordenes, 2):
        for col, v in enumerate([o["id"], o["proveedor_nombre"], o["fecha"], o["total"], o["estado"].capitalize()], 1):
            cell = ws.cell(row=r, column=col, value=v); cell.border = border
            if col == 4: cell.number_format = '#,##0.00'
    ws.column_dimensions["A"].width = 8; ws.column_dimensions["B"].width = 28; ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 12; ws.column_dimensions["E"].width = 14
    out = io.BytesIO(); wb.save(out); out.seek(0)
    return Response(out.getvalue(), mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=ordenes_compra.xlsx"})

@app.route("/ordenes-compra/nuevo", methods=["GET", "POST"])
@login_required
@role_required('admin')
def crear_orden_compra():
    conn = database.get_db()
    if request.method == "POST":
        proveedor_id = int(request.form["id_proveedor"])
        database.execute(conn, "INSERT INTO ordenes_compra (id_proveedor,fecha,estado,total,observaciones) VALUES (?,?,?,?,?)",
            (proveedor_id, request.form.get("fecha",str(date.today())), "pendiente", 0, request.form.get("observaciones","")))
        orden_id = database.fetchone(conn, "SELECT MAX(id) as id FROM ordenes_compra")["id"]
        total = 0
        for pid, cant, pu in zip(request.form.getlist("producto_id"), request.form.getlist("cantidad"), request.form.getlist("precio_unitario")):
            c = int(cant) if cant else 0
            p = float(pu) if pu else 0
            if c > 0 and pid:
                subtotal = c * p
                database.execute(conn, "INSERT INTO orden_compra_items (id_orden,id_producto,cantidad,precio_unitario,subtotal) VALUES (?,?,?,?,?)", (orden_id, int(pid), c, p, subtotal))
                total += subtotal
        database.execute(conn, "UPDATE ordenes_compra SET total=? WHERE id=?", (total, orden_id))
        conn.commit(); conn.close()
        return redirect(url_for("listar_ordenes_compra"))
    proveedores = database.fetchall(conn, "SELECT * FROM proveedores WHERE activo=1 ORDER BY nombre")
    productos = database.fetchall(conn, "SELECT id,nombre,precio_compra,precio_venta FROM productos WHERE activo=1 ORDER BY nombre")
    conn.close()
    return render_template("admin_orden_compra_form.html", proveedores=proveedores, productos=productos,
        hoy=str(date.today()), title="Nueva Orden de Compra", active="ordenes_compra", **_ctx())

@app.route("/ordenes-compra/<int:id>")
@login_required
@role_required('admin')
def ver_orden_compra(id):
    conn = database.get_db()
    orden = database.fetchone(conn, """
        SELECT o.*, p.nombre as proveedor_nombre, p.contacto, p.telefono, p.ruc
        FROM ordenes_compra o JOIN proveedores p ON o.id_proveedor=p.id WHERE o.id=?
    """, (id,))
    if not orden: conn.close(); return redirect(url_for("listar_ordenes_compra"))
    items = database.fetchall(conn, """
        SELECT oi.*, pr.nombre as producto_nombre, pr.stock
        FROM orden_compra_items oi JOIN productos pr ON oi.id_producto=pr.id WHERE oi.id_orden=? ORDER BY pr.nombre
    """, (id,))
    conn.close()
    return render_template("admin_orden_compra_detail.html", orden=orden, items=items,
        title="OC #"+str(orden["id"]), active="ordenes_compra", **_ctx())

@app.route("/ordenes-compra/<int:id>/recibir", methods=["POST"])
@login_required
@role_required('admin')
def recibir_orden_compra(id):
    conn = database.get_db()
    items = database.fetchall(conn, "SELECT * FROM orden_compra_items WHERE id_orden=?", (id,))
    for it in items:
        database.execute(conn, "UPDATE productos SET stock=stock+?, precio_compra=? WHERE id=?", (it["cantidad"], it["precio_unitario"], it["id_producto"]))
        registrar_movimiento(conn, it["id_producto"], 'entrada', it["cantidad"], 'orden_compra', id, 'OC #{} recibida'.format(id))
    database.execute(conn, "UPDATE ordenes_compra SET estado='recibida' WHERE id=?", (id,))
    conn.commit(); conn.close()
    return redirect(url_for("ver_orden_compra", id=id))

@app.route("/ordenes-compra/<int:id>/anular")
@login_required
@role_required('admin')
def anular_orden_compra(id):
    conn = database.get_db()
    database.execute(conn, "UPDATE ordenes_compra SET estado='anulada' WHERE id=?", (id,))
    conn.commit(); conn.close()
    return redirect(url_for("listar_ordenes_compra"))

def _detectar_codigo(linea):
    """Extrae el código de producto de una línea. Retorna el código limpio o None."""
    if not linea or len(linea) < 2:
        return None
    # Rechazar lineas con palabras de factura/documento que no son productos
    bajos_full = linea.lower()
    palabras_rechazo = ['lic.conducir', 'const. inscrip', 'ruc transportista', 'marcwplaca', 'marcawplaca',
                        'p. partida', 'p. llegada', 'factura electronica', 'condicion', 'cliente:',
                        'vendedor', 'transportista', 'zona:', 'licencia', 'pedido', 'vehiculo']
    if any(p in bajos_full for p in palabras_rechazo):
        return None

    # Quitar TODO lo que no sea alfanumérico al inicio (OCR noise: ©, °, ), etc.)
    limpia = re.sub(r'^[^A-Za-zÁÉÍÓÚÑáéíóúñ0-9]+', '', linea)
    if not limpia:
        return None
    # Intentar código alfanumérico (ej. HER2500119)
    m = re.match(r'([A-ZÁÉÍÓÚÑ]{2,}[A-ZÁÉÍÓÚÑ0-9/\-.*]*\d[A-ZÁÉÍÓÚÑ0-9/\-.*]*)', limpia)
    if m:
        cand = m.group(1)
        if 4 <= len(cand) <= 25 and not re.search(r'\d+\.\d+|\d+\*\d+', cand):
            bajos = cand.lower()
            if not any(p in bajos for p in ['page', 'pagina', 'fecha', 'telefono', 'ruc', 'email', 'web', 'www', 'http', 'total', 'subtotal', 'igv', 'direccion']):
                return cand
    # Fallback: código puramente numérico (ej. 064268) - buscar en toda la línea
    limpia2 = re.sub(r'^[^\d]+', '', linea.strip())
    if not limpia2:
        return None
    # Buscar el primer grupo de 5-10 dígitos consecutivos (saltando el número de fila inicial si es corto)
    mn = re.search(r'(?<!\d)(\d{5,10})(?!\d)', limpia2)
    if mn:
        cand = mn.group(1)
        # Rechazar si parece numero de documento/licencia/etc
        if any(p in bajos_full for p in ['lic.conducir', 'const. inscrip', 'const.inscrip',
                                          'ruc transportista', 'marcwplaca', 'licencia',
                                          'dnl', 'dni:', 'ruc:', 'celular:', 'telefono:', 'page',
                                          'pagina', 'página', 'www.', 'http', '@',
                                          'subtotal', 'igv ', 'total ', 'son:',
                                          'pedido', 'condicion', 'cliente', 'vendedor',
                                          'transportista', 'marca', 'placa',
                                          'partida', 'llegada']):
            return None
        return cand
    return None

def _parse_text_productos(texto_crudo):
    """Parse product data from raw text. Returns (productos, texto_crudo con debug)."""

    def _fix_decimal(s):
        """Recover decimal points eaten by OCR (e.g., '16801' -> '168.01')."""
        s = s.strip().replace('O', '0').replace('o', '0').replace('l', '1')
        s = s.replace(',', '.')
        # Already has a decimal
        if re.match(r'^\d+\.\d+$', s):
            return s
        # Pure digits - try to insert decimal before last 2 digits if 5+ digits
        m = re.match(r'^(\d{3,})(\d{2})$', s)
        if m:
            candidate = f"{m.group(1)}.{m.group(2)}"
            # Only if it looks like a reasonable price (< 100000)
            try:
                if float(candidate) < 100000:
                    return candidate
            except ValueError:
                pass
        return s

    def _try_num(parte):
        """Try to parse a string as a number, handling OCR noise. Returns float or None."""
        if not parte or not re.search(r'\d', parte):
            return None
        # Strip leading/trailing OCR noise (parentheses, quotes, symbols)
        parte = re.sub(r'^[^A-Za-zÁÉÍÓÚÑ0-9.,\-]+|[^A-Za-zÁÉÍÓÚÑ0-9.,\-]+$', '', parte)
        if not parte or not re.search(r'\d', parte):
            return None
        # If it contains letters other than OCR-confusable ones, skip
        cleaned_basic = re.sub(r'[\d.,\-eE+]', '', parte)
        if cleaned_basic:
            remaining = re.sub(r'[Ool]', '', cleaned_basic)
            if remaining:
                return None
        cleaned = parte.replace('O', '0').replace('o', '0').replace('l', '1')
        cleaned = _fix_decimal(re.sub(r'[^\d.,\-eE+]', '', cleaned))
        if not cleaned:
            return None
        try:
            return float(cleaned.replace(',', '.'))
        except ValueError:
            return None

    productos = []
    lineas_filtradas_todas = []
    texto_debug = ""
    lineas = [l.strip() for l in texto_crudo.split('\n') if l.strip()]
    texto_debug += f"--- Total lineas obtenidas ({len(lineas)}) ---\n"
    for li, l in enumerate(lineas):
        filtrada = ""
        baja = l.lower()
        if any(p in baja for p in ['page ', 'pagina ', 'página ', 'www.', 'http', '@',
                                   'ruc:', 'celular:', 'telefono:', 'direccion:',
                                   'subtotal', 'igv ', 'total ', 'son:']):
            filtrada = " [FILTRADA]"
        texto_debug += f"[{li:3d}]|{l}|{filtrada}\n"
        if not filtrada:
            lineas_filtradas_todas.append((l, 0))
    texto_debug += f"\n=== LINEAS FILTRADAS ({len(lineas_filtradas_todas)}) ===\n"
    for idx, (l, pn) in enumerate(lineas_filtradas_todas):
        cod_test = _detectar_codigo(l)
        marc = f" <-- CODIGO? -> {cod_test!r}" if cod_test else ""
        texto_debug += f"[{idx:3d}] p{pn} |{l}|{marc}\n"
    indices_codigos = [idx for idx, (l, pn) in enumerate(lineas_filtradas_todas) if _detectar_codigo(l)]
    texto_debug += f"\n=== CODIGOS DETECTADOS: {len(indices_codigos)} ===\n"
    for idx in indices_codigos:
        texto_debug += f"  idx={idx}: {lineas_filtradas_todas[idx][0]!r} -> {_detectar_codigo(lineas_filtradas_todas[idx][0])!r}\n"
    for ci, start in enumerate(indices_codigos):
        end = indices_codigos[ci + 1] if ci + 1 < len(indices_codigos) else len(lineas_filtradas_todas)
        bloque = lineas_filtradas_todas[start:end]
        linea_code, page_num = bloque[0]
        codigo = _detectar_codigo(linea_code)
        if not codigo:
            continue
        # Try to extract inline values from the same line as the code
        # Format: CODE DESCRIPTION QTY P1 P2 TOTAL
        resto = linea_code[len(codigo):].strip()
        inline_nums = []
        for parte in resto.split():
            n = _try_num(parte)
            if n is not None:
                inline_nums.append(n)
        # If we got 3-5 numbers inline, try to parse them
        if len(inline_nums) >= 3:
            desc = resto
            nums_strs = [str(int(n)) if n == int(n) else f'{n:.2f}' for n in inline_nums]
            for ns in reversed(nums_strs):
                idx = desc.rfind(' ' + ns)
                if idx >= 0:
                    desc = desc[:idx]
                else:
                    idx = desc.rfind(' ' + ns.rstrip('0').rstrip('.'))
                    if idx >= 0:
                        desc = desc[:idx]
            desc = desc.strip()
            cantidad = int(inline_nums[0])
            p1 = inline_nums[1]
            p2 = inline_nums[2]
            total = inline_nums[3] if len(inline_nums) >= 4 else cantidad * p2
            precio_venta = math.ceil(p2 * 2)
            # Validate consistency
            expected = cantidad * p2
            if abs(expected - total) / max(total, 0.01) > 0.15:
                texto_debug += f"[PARSE-INLINE-WARN] #{len(productos)+1}: inconsistencia qty*p2={expected:.2f} vs total={total:.2f} (cod={codigo!r})\n"
            productos.append(dict(
                codigo=codigo, descripcion=desc.strip(), cantidad=cantidad,
                precio_compra=p2, precio_venta=precio_venta, total=total,
                page=page_num
            ))
            texto_debug += f"[PARSE-INLINE] #{len(productos)}: cod={codigo!r} desc={desc!r} qty={cantidad} p1={p1} p2={p2} tot={total}\n"
            continue
        # Strategy 2: format tipo factura - precios en la(s) linea(s) siguiente(s)
        # Pattern: "N QTY CODE DESCRIPTION" then next line "REF QTY_UNIT UNIT_PRICE DISCOUNT TOTAL"
        desc_lineas_2 = []
        # Extraer qty y desc de la misma linea del codigo
        # Formato: (row_num) (qty) (CODE) (description...)
        limpia_linea = re.sub(r'^[^A-Za-zÁÉÍÓÚÑáéíóúñ0-9]+', '', linea_code)
        parts_code = limpia_linea.split()
        qty_from_line = None
        desc_start_idx = None
        for pi, p in enumerate(parts_code):
            if p == codigo:
                desc_start_idx = pi + 1
                # Buscar cantidad antes del código: ultimo numero entero pequeño antes del codigo
                for pj in range(pi - 1, -1, -1):
                    try:
                        n = int(parts_code[pj])
                        if 1 <= n <= 9999:
                            qty_from_line = n
                            break
                    except ValueError:
                        break
                break
        desc_from_line = " ".join(parts_code[desc_start_idx:]) if desc_start_idx else ""
        # Buscar precios en las siguientes 1-2 lineas
        precios_encontrados = None
        qty_from_price_line = None
        for bi in range(1, min(3, len(bloque))):
            val = bloque[bi][0]
            # Limpiar y buscar numeros (precios)
            nums_linea = []
            for parte in val.split():
                n = _try_num(parte)
                if n is not None:
                    nums_linea.append(n)
            # Buscar patron: (qty_unit) (unit_price) (discount) (total)
            # O saltar ref inicial y buscar (unit_price) (total)
            if len(nums_linea) >= 3:
                # Intentar: el ultimo numero es total, el penultimo es discount, el antepenultimo es unit_price
                # Pero si hay un numero extra al inicio (ref), saltarlo
                candidates = []
                for offset in range(0, len(nums_linea) - 2):
                    cand_qty = None
                    cand_up = nums_linea[offset]
                    cand_disc = nums_linea[offset + 1]
                    cand_tot = nums_linea[offset + 2]
                    # unit_price debe ser razonable (< 10000)
                    if cand_up < 1 or cand_up > 10000:
                        continue
                    # discount puede ser 0 o un porcentaje
                    if cand_disc < 0 or cand_disc > cand_up:
                        continue
                    # total debe ser >= unit_price
                    if cand_tot < cand_up * 0.5:
                        continue
                    candidates.append((cand_qty, cand_up, cand_disc, cand_tot))
                if candidates:
                    precios_encontrados = candidates[0]
                    qty_from_price_line = candidates[0][0]
                    break
            # Si hay exactamente 2 numeros grandes (unit_price y total)
            if len(nums_linea) >= 2:
                candidates_2 = []
                for offset in range(0, len(nums_linea) - 1):
                    cand_up = nums_linea[offset]
                    cand_tot = nums_linea[offset + 1]
                    # Reject if up looks like a quantity/reference (< 5) not a price
                    if cand_up < 5 or cand_up > 10000:
                        continue
                    if cand_tot > cand_up * 0.5:
                        cand_qty = None
                        for q_offset in range(offset - 1, -1, -1):
                            if nums_linea[q_offset] == int(nums_linea[q_offset]) and 1 <= nums_linea[q_offset] <= 999:
                                cand_qty = int(nums_linea[q_offset])
                                break
                        candidates_2.append((cand_qty, cand_up, 0, cand_tot))
                if candidates_2:
                    precios_encontrados = candidates_2[0]
                    qty_from_price_line = candidates_2[0][0]
                    break
        if precios_encontrados and desc_from_line:
            _qty, p1, _disc, total = precios_encontrados
            cantidad = qty_from_line or qty_from_price_line or 1
            p2 = p1
            precio_venta = math.ceil(p2 * 2)
            if total > 0 and p2 > 0:
                expected = cantidad * p2
                if abs(expected - total) / max(total, 0.01) > 0.15:
                    texto_debug += f"[PARSE-FACTURA-WARN] #{len(productos)+1}: inconsistencia qty*p2={expected:.2f} vs total={total:.2f} (cod={codigo!r})\n"
                productos.append(dict(
                    codigo=codigo, descripcion=desc_from_line.strip(), cantidad=cantidad,
                    precio_compra=p2, precio_venta=precio_venta, total=total,
                    page=page_num
                ))
                texto_debug += f"[PARSE-FACTURA] #{len(productos)}: cod={codigo!r} desc={desc_from_line!r} qty={cantidad} p1={p1} p2={p2} tot={total}\n"
                continue
        # Fallback: collect ALL lines as multi-line description, scan for prices
        desc_lineas = []
        all_nums = []  # (value, is_int, original_text)
        for bi in range(1, len(bloque)):
            val = bloque[bi][0]
            # Collect all numbers from this line
            partes = val.split()
            nums_en_linea = []
            for parte in partes:
                n = _try_num(parte)
                if n is not None:
                    nums_en_linea.append(n)
            if nums_en_linea:
                all_nums.extend(nums_en_linea)
                # Still add text to description (price lines are also description context)
            # Add to description regardless (we'll trim later)
            desc_lineas.append(val)
        # Try to match (qty, unit_price, total) or (unit_price, total) patterns from collected numbers
        cantidad = None
        p2 = None
        total = None
        # Strategy A: look for pattern [qty, price, price, total] with qty being integer
        for i in range(len(all_nums) - 2):
            qty_cand = all_nums[i]
            if qty_cand == int(qty_cand) and 1 <= qty_cand <= 99999:
                up_cand = all_nums[i + 1]
                tot_cand = all_nums[i + 2] if i + 2 < len(all_nums) else None
                if 1 <= up_cand <= 100000 and tot_cand and tot_cand >= up_cand * 0.8:
                    # Validate: qty * up ≈ total
                    if abs(qty_cand * up_cand - tot_cand) / max(tot_cand, 0.01) < 0.15:
                        cantidad = int(qty_cand)
                        p2 = up_cand
                        total = tot_cand
                        break
        # Strategy B: [qty, price, total] or [price, total] if A failed
        if cantidad is None:
            for i in range(len(all_nums) - 1):
                a, b = all_nums[i], all_nums[i + 1]
                if 5 <= a <= 100000 and b >= a * 0.8:
                    # Could be [unit_price, total] — find qty before
                    qty_before = None
                    for j in range(i - 1, -1, -1):
                        if all_nums[j] == int(all_nums[j]) and 1 <= all_nums[j] <= 9999:
                            qty_before = int(all_nums[j])
                            break
                    if qty_before and abs(qty_before * a - b) / max(b, 0.01) < 0.15:
                        cantidad = qty_before
                        p2 = a
                        total = b
                        break
                    elif not qty_before:
                        cantidad = 1
                        p2 = a
                        total = b
                        break
        # Strategy C: just find the most likely unit_price (middle value)
        if cantidad is None and len(all_nums) >= 2:
            sorted_nums = sorted(all_nums)
            median = sorted_nums[len(sorted_nums) // 2]
            if 1 <= median <= 100000:
                p2 = median
                cantidad = 1
                total = max(all_nums)
        if cantidad is None or p2 is None:
            # Last resort: register with price 0 (bonus/editar después)
            desc = " ".join(desc_lineas).strip() if desc_lineas else (desc_from_line.strip() or "Sin descripcion")
            productos.append(dict(
                codigo=codigo, descripcion=desc, cantidad=qty_from_line or 1,
                precio_compra=0, precio_venta=0, total=0,
                page=page_num
            ))
            texto_debug += f"[PARSE-BONUS] #{len(productos)}: cod={codigo!r} desc={desc!r} qty={qty_from_line or 1} PRECIO 0 (editar despues)\n"
            continue
        if p2 > 100000:
            texto_debug += f"[PARSE] SALTANDO idx={start}: cod={codigo!r} precio muy alto ({p2})\n"
            continue
        desc = " ".join(desc_lineas).strip() if desc_lineas else (desc_from_line.strip() or "")
        precio_venta = math.ceil(p2 * 2)
        # Validate consistency
        if total and cantidad and p2:
            expected = cantidad * p2
            if abs(expected - total) / max(total, 0.01) > 0.15:
                texto_debug += f"[PARSE-WARN] #{len(productos)+1}: inconsistencia qty*up={expected:.2f} vs total={total:.2f} (cod={codigo!r})\n"
        productos.append(dict(
            codigo=codigo, descripcion=desc, cantidad=cantidad,
            precio_compra=p2, precio_venta=precio_venta, total=total or (cantidad * p2),
            page=page_num
        ))
        texto_debug += f"[PARSE] #{len(productos)}: cod={codigo!r} desc={desc!r} qty={cantidad} p2={p2} tot={total} (multi-line)\n"
    try:
        debug_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "_debug_parser.txt")
        with open(debug_path, "w", encoding="utf-8") as f:
            f.write(texto_debug)
    except:
        pass
    return productos, texto_debug

def _parse_pdf_productos(pdf_path):
    doc = fitz.open(pdf_path)
    imagenes = []
    texto_crudo = ""
    for page_num, pagina in enumerate(doc):
        blocks = pagina.get_text("blocks")
        page_h = pagina.rect.height
        for b in blocks:
            if len(b) < 5 or not b[4].strip():
                continue
            x0, y0, x1, y1, text = b[:5]
            # Skip header (top 12%) and footer (bottom 8%) using position
            if y0 < page_h * 0.12 or y1 > page_h * 0.92:
                continue
            # Skip very small blocks (page numbers, stray chars)
            if len(text.strip()) < 3:
                continue
            texto_crudo += text + "\n"
    doc.close()
    productos, texto_debug = _parse_text_productos(texto_crudo)
    return productos, imagenes, texto_debug

def _ocr_image_productos(image_path):
    """Extract products from an image using OCR. Returns (productos, [], texto_crudo)."""
    import pytesseract
    pytesseract.pytesseract.tesseract_cmd = r"C:\Program Files\Tesseract-OCR\tesseract.exe"
    import cv2
    import numpy as np
    from PIL import Image as PILImage, ImageEnhance, ImageFilter, ImageOps

    img = PILImage.open(image_path)
    # 1) Escalar si muy pequeña (< 1600px de ancho)
    w, h = img.size
    if w < 1600:
        ratio = 1600 / w
        img = img.resize((1600, int(h * ratio)), PILImage.LANCZOS)

    # 2) Convertir a grises + contraste moderado
    img = img.convert('L')
    enhancer = ImageEnhance.Contrast(img)
    img = enhancer.enhance(1.5)

    # 3) Suavizado ligero para reducir ruido de camara
    img = img.filter(ImageFilter.MedianFilter(size=3))

    # 4) OCR con PSM 6 (bloque uniforme) primero, fallback a PSM 4
    text = pytesseract.image_to_string(img, lang='spa+eng', config='--psm 6 --oem 3')
    if not text.strip():
        text = pytesseract.image_to_string(img, lang='spa', config='--psm 4 --oem 3')

    # 5) Si aun sin texto, intentar con binarizacion OTSU
    if not text.strip():
        arr = np.array(img)
        _, arr_bin = cv2.threshold(arr, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)
        if np.mean(arr_bin) > 127:
            arr_bin = cv2.bitwise_not(arr_bin)
        img_bin = PILImage.fromarray(arr_bin)
        text = pytesseract.image_to_string(img_bin, lang='spa+eng', config='--psm 6 --oem 3')

    productos, texto_debug = _parse_text_productos(text)
    return productos, [], texto_debug

_TEMP_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "productos_photos")

@app.route("/compras/carga-rapida", methods=["GET", "POST"])
@login_required
@role_required('admin')
def carga_rapida():
    conn = database.get_db()
    categorias_list = database.fetchall(conn, "SELECT nombre FROM categorias ORDER BY nombre")
    categorias_list = [c["nombre"] for c in categorias_list]
    conn.close()
    # Limpiar temp files viejos (>1 hora)
    if os.path.isdir(_TEMP_DIR):
        ahora = time.time()
        for f in os.listdir(_TEMP_DIR):
            if f.startswith("_tmp_"):
                fp = os.path.join(_TEMP_DIR, f)
                if ahora - os.path.getmtime(fp) > 3600:
                    try: os.remove(fp)
                    except: pass
    token = request.args.get("token") or request.form.get("token")
    if request.method == "POST" and request.form.get("confirmar"):
        try:
            datos = json.loads(request.form["datos_json"])
        except (json.JSONDecodeError, KeyError) as e:
            return render_template("admin_carga_rapida.html",
                title="Carga Rápida", active="carga_rapida",
                categorias=categorias_list,
                error=f"Error al leer datos del formulario: {e}. Intenta recargar la página y subir el PDF nuevamente.", **_ctx())
        if not isinstance(datos, list):
            return render_template("admin_carga_rapida.html",
                title="Carga Rápida", active="carga_rapida",
                categorias=categorias_list,
                error="Error: datos_json no es una lista válida.", **_ctx())
        cantidades = request.form.getlist("cantidad_edit")
        precios_compra = request.form.getlist("precio_compra_edit")
        precios_venta = request.form.getlist("precio_venta_edit")
        img_checked = request.form.getlist("img_check")
        creados = 0
        img_idx = 0
        productos_ids = []
        proveedores = []
        conn2 = database.get_db()
        try:
            proveedores = database.fetchall(conn2, "SELECT id, nombre FROM proveedores WHERE activo=1 ORDER BY nombre")
            total_compra_calculado = 0.0
            for i, p in enumerate(datos):
                cant = int(cantidades[i]) if i < len(cantidades) and cantidades[i] else p.get("cantidad", 1)
                pc = float(precios_compra[i]) if i < len(precios_compra) and precios_compra[i] else p.get("precio_compra", 0)
                pv = float(precios_venta[i]) if i < len(precios_venta) and precios_venta[i] else p.get("precio_venta", 0)
                total_compra_calculado += cant * pc
                cod_aux = (p.get("codigo") or "").strip()
                desc_aux = (p.get("descripcion") or "").strip()
                nombre = (f"{cod_aux} - {desc_aux}" if cod_aux else desc_aux)[:200]
                existe = database.fetchone(conn2, "SELECT id FROM productos WHERE nombre=? OR codigo_barras=?", (nombre, cod_aux)) if cod_aux else database.fetchone(conn2, "SELECT id FROM productos WHERE nombre=?", (nombre,))
                if not existe and cod_aux and desc_aux:
                    existe = database.fetchone(conn2, "SELECT id FROM productos WHERE nombre=?", (desc_aux,))
                if existe:
                    pid = existe["id"]
                    desc_mayor = float(request.form.get("descuento_mayorista","0"))
                    cod_prod = (p.get("codigo") or "")[:50]
                    database.execute(conn2, """UPDATE productos SET
                        stock=stock+?, precio_compra=?, precio_venta=?,
                        por_mayor=?, descuento_mayorista=?,
                        nombre=?, codigo_barras=COALESCE(codigo_barras,?),
                        categoria=?, activo=1
                        WHERE id=?""",
                        (cant, pc, pv,
                         1 if desc_mayor > 0 else 0, desc_mayor,
                         nombre, cod_prod if cod_prod else None,
                         request.form.get("categoria","General"), pid))
                else:
                    desc_mayor = float(request.form.get("descuento_mayorista","0"))
                    cod_prod = (p.get("codigo") or "")[:50]
                    database.execute(conn2, """INSERT INTO productos
                        (nombre, descripcion, precio_compra, precio_venta, stock, categoria,
                         por_mayor, descuento_mayorista, codigo_barras)
                        VALUES (?,?,?,?,?,?,?,?,?)""",
                        (nombre, p.get("descripcion",""), pc, pv,
                         cant, request.form.get("categoria","General"),
                         1 if desc_mayor > 0 else 0, desc_mayor,
                         cod_prod if cod_prod else None))
                    pid = database.fetchone(conn2, "SELECT MAX(id) as id FROM productos")["id"]
                productos_ids.append({"id": pid, "cantidad": cant, "precio_unitario": pc})
                registrar_movimiento(conn2, pid, 'entrada', cant, 'carga_rapida', None, 'Importación rápida desde PDF')
                if token and img_idx < len(img_checked):
                    img_real_idx = int(img_checked[img_idx])
                    tmp_name = f"_tmp_{token}_{img_real_idx}.png"
                    tmp_path = os.path.join(_TEMP_DIR, tmp_name)
                    raw_name = str(p.get('codigo','prod'))
                    safe_name = re.sub('[<>:"/\\\\|?*\x00-\x1f\xa0]', '_', raw_name).strip('._ ')
                    final_name = f"{safe_name}_{pid}.png"
                    if os.path.exists(tmp_path):
                        from rembg import remove
                        from PIL import Image
                        import io as _io
                        with open(tmp_path, "rb") as _f:
                            img_data = _f.read()
                        try:
                            result = remove(img_data)
                            img = Image.open(_io.BytesIO(result)).convert('RGBA')
                            img.save(os.path.join(_TEMP_DIR, final_name), 'PNG')
                        except Exception:
                            os.rename(tmp_path, os.path.join(_TEMP_DIR, final_name))
                        os.remove(tmp_path)
                        database.execute(conn2, "UPDATE productos SET foto=? WHERE id=?", (final_name, pid))
                    img_idx += 1
                creados += 1
            # Registrar compra total en Caja como egreso si hay comprobante
            comprobante_file = request.form.get("comprobante_file", "")
            if comprobante_file:
                total_compra = total_compra_calculado
                concepto_compra = f"Compra proveedor - {creados} producto(s)"
                database.execute(conn2,
                    "INSERT INTO caja (fecha,tipo,concepto,monto,metodo_pago,observacion,comprobante,referencia_tipo) VALUES (?,?,?,?,?,?,?,?)",
                    (str(date.today()), "egreso", concepto_compra, total_compra,
                     request.form.get("metodo_pago_carga","efectivo"), f"Importación rápida desde PDF",
                     comprobante_file, "carga_rapida"))
            # Crear Orden de Compra si se seleccionó proveedor
            id_proveedor = request.form.get("id_proveedor", "")
            if id_proveedor and int(id_proveedor) > 0 and productos_ids:
                obs = f"Importación rápida desde PDF - {creados} producto(s)"
                database.execute(conn2,
                    "INSERT INTO ordenes_compra (id_proveedor,fecha,estado,total,observaciones) VALUES (?,?,?,?,?)",
                    (int(id_proveedor), str(date.today()), "recibida", total_compra_calculado, obs))
                orden_id = database.fetchone(conn2, "SELECT MAX(id) as id FROM ordenes_compra")["id"]
                for item in productos_ids:
                    subtotal = item["cantidad"] * item["precio_unitario"]
                    database.execute(conn2,
                        "INSERT INTO orden_compra_items (id_orden,id_producto,cantidad,precio_unitario,subtotal) VALUES (?,?,?,?,?)",
                        (orden_id, item["id"], item["cantidad"], item["precio_unitario"], subtotal))
                database.execute(conn2, "UPDATE ordenes_compra SET total=? WHERE id=?", (total_compra_calculado, orden_id))
            conn2.commit(); conn2.close()
        except Exception as e:
            import traceback
            traceback.print_exc()
            if 'conn2' in dir(): conn2.close()
            try:
                datos = json.loads(request.form.get("datos_json", "[]"))
                cantidades_err = request.form.getlist("cantidad_edit")
                precios_compra_err = request.form.getlist("precio_compra_edit")
                precios_venta_err = request.form.getlist("precio_venta_edit")
                for i, p in enumerate(datos):
                    if i < len(cantidades_err) and cantidades_err[i]: p["cantidad"] = int(cantidades_err[i])
                    if i < len(precios_compra_err) and precios_compra_err[i]: p["precio_compra"] = float(precios_compra_err[i])
                    if i < len(precios_venta_err) and precios_venta_err[i]: p["precio_venta"] = float(precios_venta_err[i])
                    p["total"] = p["cantidad"] * p["precio_compra"]
                return render_template("admin_carga_rapida.html",
                    title="Carga Rápida", active="carga_rapida",
                    categorias=categorias_list,
                    productos=datos, token=request.form.get("token",""),
                    texto_crudo="", descuento_mayorista=request.form.get("descuento_mayorista","0"),
                    comprobante_file=request.form.get("comprobante_file",""),
                    datos_json=json.dumps(datos),
                    error=f"Error al importar productos: {e}",
                    proveedores=proveedores, debug_imagenes=[], **_ctx())
            except:
                return render_template("admin_carga_rapida.html",
                    title="Carga Rápida", active="carga_rapida",
                    categorias=categorias_list,
                    error=f"Error al importar productos: {e}", **_ctx())
        if token:
            for f in os.listdir(_TEMP_DIR):
                if f.startswith(f"_tmp_{token}_"):
                    os.remove(os.path.join(_TEMP_DIR, f))
        return render_template("admin_carga_rapida.html",
            title="Carga Rápida", active="carga_rapida",
            categorias=categorias_list,
            msg=f"¡Importación completada! {creados} producto(s) procesado(s).", **_ctx())
    if request.method == "POST":
        archivo = request.files.get("pdf")
        if not archivo or not archivo.filename:
            return render_template("admin_carga_rapida.html",
                title="Carga Rápida", active="carga_rapida",
                categorias=categorias_list,
                error="Debes seleccionar un archivo PDF o imagen.", **_ctx())
        fname_lower = archivo.filename.lower()
        es_pdf = fname_lower.endswith(".pdf")
        es_imagen = any(fname_lower.endswith(ext) for ext in [".png",".jpg",".jpeg",".webp",".bmp",".tiff",".tif"])
        if not es_pdf and not es_imagen:
            return render_template("admin_carga_rapida.html",
                title="Carga Rápida", active="carga_rapida",
                categorias=categorias_list,
                error="Formato no soportado. Sube un PDF o imagen (PNG, JPG, WEBP).", **_ctx())
        comprobante_file = ""
        if "comprobante" in request.files and request.files["comprobante"] and request.files["comprobante"].filename:
            fname, _ = storage.save_file(request.files["comprobante"], "comprobantes")
            comprobante_file = fname
        ext = ".pdf" if es_pdf else ".png"
        temp_path = os.path.join(_TEMP_DIR, f"_temp_upload{ext}")
        archivo.save(temp_path)
        try:
            try:
                if es_pdf:
                    productos, imagenes, texto_crudo = _parse_pdf_productos(temp_path)
                else:
                    productos, imagenes, texto_crudo = _ocr_image_productos(temp_path)
            except Exception as e:
                import traceback
                traceback.print_exc()
                return render_template("admin_carga_rapida.html",
                    title="Carga Rápida", active="carga_rapida",
                    categorias=categorias_list,
                    error=f"Error al procesar el archivo: {e}. Verifica que no esté dañado.",
                    debug_texto="", **_ctx())
        finally:
            if os.path.exists(temp_path): os.remove(temp_path)
        if not productos:
            import logging as _lg
            _lg.getLogger(__name__).error("PDF parse failed. Raw text preview:\n%s", texto_crudo[:2000])
            preview = "\n".join(texto_crudo.split("\n")[:30])
            return render_template("admin_carga_rapida.html",
                title="Carga Rápida", active="carga_rapida",
                categorias=categorias_list,
                error="No se pudieron extraer productos del PDF. Verifica que el formato sea el esperado.",
                debug_texto=preview, **_ctx())
        token = uuid.uuid4().hex[:12]
        os.makedirs(_TEMP_DIR, exist_ok=True)
        # Guardar imágenes como temp files
        for idx, img_data in enumerate(imagenes):
            tmp_name = f"_tmp_{token}_{idx}.png"
            with open(os.path.join(_TEMP_DIR, tmp_name), "wb") as f:
                f.write(img_data["bytes"])
        # (imagen_idx ya fue asignado por _parse_pdf_productos)
        # Verificar productos existentes y obtener proveedores
        conn_check = database.get_db()
        proveedores = database.fetchall(conn_check, "SELECT id, nombre FROM proveedores WHERE activo=1 ORDER BY nombre")
        for p in productos:
            cod = (p.get("codigo") or "").strip()
            desc = (p.get("descripcion") or "").strip()
            nombre = (f"{cod} - {desc}" if cod else desc)[:200]
            row = database.fetchone(conn_check, "SELECT id, stock FROM productos WHERE nombre=? OR codigo_barras=?", (nombre, cod)) if cod else database.fetchone(conn_check, "SELECT id, stock FROM productos WHERE nombre=?", (nombre,))
            if not row and cod and desc:
                row = database.fetchone(conn_check, "SELECT id, stock FROM productos WHERE nombre=?", (desc,))
            p["existe"] = row is not None
            p["stock_actual"] = row["stock"] if row else 0
        conn_check.close()
        desc_mayor = request.form.get("descuento_mayorista","0")
        return render_template("admin_carga_rapida.html",
            title="Carga Rápida", active="carga_rapida",
            categorias=categorias_list,
            productos=productos, token=token,
            debug_imagenes=imagenes,
            texto_crudo=texto_crudo,
            descuento_mayorista=desc_mayor,
            comprobante_file=comprobante_file,
            datos_json=json.dumps(productos),
            proveedores=proveedores, **_ctx())
    return render_template("admin_carga_rapida.html",
        title="Carga Rápida", active="carga_rapida",
        categorias=categorias_list, **_ctx())

# ===== API items unificados =====
@app.route("/api/ventas/<int:id>/items")
def api_venta_items(id):
    conn = database.get_db()
    items = database.fetchall(conn, "SELECT * FROM venta_items WHERE id_venta=?", (id,))
    conn.close()
    return Response(json.dumps([dict(r) for r in items]), mimetype="application/json")

@app.route("/api/items/unificados")
@login_required
@role_required('admin', 'recepcion')
def api_items_unificados():
    conn = database.get_db()
    items = []
    for p in database.fetchall(conn, "SELECT id,nombre,precio_venta as precio,'producto' as tipo FROM productos WHERE activo=1 ORDER BY nombre"):
        items.append(dict(p))
    for s in database.fetchall(conn, "SELECT id,nombre,precio,'servicio_medico' as tipo FROM servicios_medicos ORDER BY nombre"):
        items.append(dict(s))
    for s in database.fetchall(conn, "SELECT id,nombre,precio,'servicio_grooming' as tipo FROM servicios_grooming ORDER BY nombre"):
        items.append(dict(s))
    conn.close()
    return Response(json.dumps(items), mimetype="application/json")

# ===== CREDITOS =====
@app.route("/creditos")
@login_required
@role_required('admin')
def listar_creditos():
    conn = database.get_db()
    creditos = database.fetchall(conn, """
        SELECT c.*, d.nombre as cliente_nombre
        FROM creditos c
        LEFT JOIN duenos d ON c.id_cliente=d.id
        ORDER BY c.fecha_venta DESC
    """)
    total_pendiente = database.fetchone(conn, "SELECT COALESCE(SUM(saldo),0) as tot FROM creditos WHERE estado='pendiente'")["tot"]
    conn.close()
    return render_template("admin_creditos.html", creditos=creditos, total_pendiente=total_pendiente,
                           title="Cr\u00e9ditos", active="creditos", **_ctx())

@app.route("/creditos/<int:id>/pagar", methods=["POST"])
@login_required
@role_required('admin')
def pagar_credito(id):
    conn = database.get_db()
    monto = float(request.form.get("monto", 0))
    metodo = request.form.get("metodo", "efectivo")
    fecha = str(date.today())
    database.execute(conn,
        "INSERT INTO pagos_credito (id_credito,monto,fecha,metodo) VALUES (?,?,?,?)",
        (id, monto, fecha, metodo))
    database.execute(conn,
        "UPDATE creditos SET saldo=saldo-? WHERE id=? AND saldo>=?", (monto, id, monto))
    database.execute(conn,
        "UPDATE creditos SET estado='pagado' WHERE id=? AND saldo<=0", (id,))
    database.execute(conn,
        "INSERT INTO caja (fecha,tipo,concepto,monto,referencia_tipo,referencia_id) VALUES (?,?,?,?,?,?)",
        (fecha, "ingreso", "Pago credito #"+str(id), monto, "credito", id))
    conn.commit(); conn.close()
    return redirect(url_for("listar_creditos"))

# ===== REPORTES =====
def _query_reporte(desde, hasta, doctor=None):
    conn = database.get_db()
    params_base = [desde, hasta]
    where_doc = ""
    params_doc = list(params_base)
    if doctor:
        where_doc = " AND doctor=?"
        params_doc.append(doctor)
    total_ventas = database.fetchone(conn, "SELECT COALESCE(SUM(total),0) as tot FROM ventas WHERE fecha>=? AND fecha<=?", params_base)["tot"]
    total_egresos = database.fetchone(conn, "SELECT COALESCE(SUM(monto),0) as tot FROM caja WHERE tipo='egreso' AND fecha>=? AND fecha<=?", params_base)["tot"]
    total_consultas = database.fetchone(conn, f"SELECT COUNT(*) as cnt FROM registros_medicos WHERE fecha>=? AND fecha<=?{where_doc}", params_doc)["cnt"]
    where_doc_hg = where_doc.replace("doctor","estilista")
    total_grooming = database.fetchone(conn, f"SELECT COUNT(*) as cnt FROM historial_grooming WHERE fecha>=? AND fecha<=?{where_doc_hg}", params_doc)["cnt"]
    total_creditos = database.fetchone(conn, "SELECT COALESCE(SUM(saldo),0) as tot FROM creditos WHERE estado='pendiente'")["tot"]
    total_costo = database.fetchone(conn, """SELECT COALESCE(SUM(vi.cantidad*COALESCE(p.precio_compra,0)),0) as tot FROM venta_items vi JOIN ventas v ON vi.id_venta=v.id LEFT JOIN productos p ON vi.id_producto=p.id WHERE v.fecha>=? AND v.fecha<=? AND vi.tipo_item='producto'""", params_base)["tot"]
    margen_bruto = round(total_ventas - total_costo, 2)
    ventas_por_dia = database.fetchall(conn, "SELECT fecha,SUM(total) as total FROM ventas WHERE fecha>=? AND fecha<=? GROUP BY fecha ORDER BY fecha", params_base)
    top_productos = database.fetchall(conn, """SELECT p.nombre,SUM(vi.cantidad) as cantidad,SUM(vi.subtotal) as total FROM venta_items vi JOIN ventas v ON vi.id_venta=v.id LEFT JOIN productos p ON vi.id_producto=p.id WHERE v.fecha>=? AND v.fecha<=? AND vi.tipo_item='producto' GROUP BY vi.id_producto ORDER BY cantidad DESC LIMIT 10""", params_base)
    top_doctores = database.fetchall(conn, "SELECT doctor,COUNT(*) as cnt FROM registros_medicos WHERE fecha>=? AND fecha<=? GROUP BY doctor ORDER BY cnt DESC", params_base) if not doctor else [{"doctor":doctor,"cnt":total_consultas}]
    grooming_servicios = database.fetchall(conn, "SELECT COALESCE(sg.nombre,'Grooming') as servicio,COUNT(*) as cantidad FROM historial_grooming hg LEFT JOIN servicios_grooming sg ON hg.id_servicio=sg.id WHERE hg.fecha>=? AND hg.fecha<=? GROUP BY hg.id_servicio ORDER BY cantidad DESC", params_base)
    ventas_list = database.fetchall(conn, "SELECT * FROM ventas WHERE fecha>=? AND fecha<=? ORDER BY fecha,id", params_base)
    ventas_items = {}
    for v in ventas_list:
        ventas_items[v["id"]] = database.fetchall(conn, "SELECT vi.*,p.nombre as pnombre FROM venta_items vi LEFT JOIN productos p ON vi.id_producto=p.id WHERE vi.id_venta=? ORDER BY vi.id", (v["id"],))
    try:
        dias = (datetime.strptime(hasta,"%Y-%m-%d") - datetime.strptime(desde,"%Y-%m-%d")).days
        ant_desde = (datetime.strptime(desde,"%Y-%m-%d") - timedelta(days=dias+1)).strftime("%Y-%m-%d")
        ant_hasta = (datetime.strptime(desde,"%Y-%m-%d") - timedelta(days=1)).strftime("%Y-%m-%d")
        ventas_anterior = database.fetchone(conn, "SELECT COALESCE(SUM(total),0) as tot FROM ventas WHERE fecha>=? AND fecha<=?",(ant_desde,ant_hasta))["tot"]
        variacion = round(((total_ventas-ventas_anterior)/ventas_anterior*100) if ventas_anterior>0 else 0, 1)
    except: ventas_anterior=0; variacion=0
    conn.close()
    return dict(titulo=f"Reporte {desde} a {hasta}", desde=desde, hasta=hasta,
        total_ventas=total_ventas, total_egresos=total_egresos, total_consultas=total_consultas,
        total_grooming=total_grooming, total_creditos=total_creditos,
        margen_bruto=margen_bruto, comparativa=dict(ventas_anterior=ventas_anterior, variacion=variacion),
        ventas_por_dia=ventas_por_dia, top_productos=top_productos, top_doctores=top_doctores,
        grooming_servicios=grooming_servicios, ventas_list=ventas_list, ventas_items=ventas_items)

@app.route("/reportes")
@login_required
def reportes():
    conn = database.get_db()
    doctores = database.fetchall(conn, "SELECT nombre FROM doctores ORDER BY nombre")
    conn.close()
    hoy_dt = date.today()
    hace7 = (hoy_dt - timedelta(days=7)).isoformat()
    hace30 = (hoy_dt - timedelta(days=30)).isoformat()
    return render_template("admin_reportes.html", reporte=None, doctores=doctores,
                           hace7=hace7, hace30=hace30,
                           title="Reportes", active="reportes", **_ctx())

@app.route("/reportes/rango")
@login_required
def reporte_rango():
    desde = request.args.get("desde", str(date.today()))
    hasta = request.args.get("hasta", str(date.today()))
    doctor = request.args.get("doctor", "")
    r = _query_reporte(desde, hasta, doctor or None)
    conn = database.get_db()
    doctores = database.fetchall(conn, "SELECT nombre FROM doctores ORDER BY nombre")
    conn.close()
    hoy_dt = date.today()
    hace7 = (hoy_dt - timedelta(days=7)).isoformat()
    hace30 = (hoy_dt - timedelta(days=30)).isoformat()
    return render_template("admin_reportes.html", reporte=r, doctores=doctores,
                           hace7=hace7, hace30=hace30,
                           title="Reportes", active="reportes", **_ctx())

@app.route("/reportes/pdf")
@login_required
def reporte_pdf():
    desde = request.args.get("desde", str(date.today()))
    hasta = request.args.get("hasta", str(date.today()))
    doctor = request.args.get("doctor", "")
    r = _query_reporte(desde, hasta, doctor or None)
    html = render_template("admin_reportes_pdf.html", reporte=r, hoy=str(date.today()),
                           title="Reporte PDF", active="reportes", **_ctx())
    options = {"page-size":"A4","margin-top":"10mm","margin-right":"10mm","margin-bottom":"10mm","margin-left":"10mm","encoding":"UTF-8"}
    pdf = pdfkit.from_string(html, False, options=options)
    return Response(pdf, mimetype="application/pdf",
        headers={"Content-Disposition": f"attachment;filename=reporte_{desde}_{hasta}.pdf"})

@app.route("/reportes/csv/<tipo>")
@login_required
def reporte_csv(tipo):
    conn = database.get_db()
    titles_map = {"ventas":"Ventas","productos":"Inventario","animales":"Pacientes","citas":"Citas","creditos":"Cr\u00e9ditos"}
    if tipo == "ventas":
        rows = database.fetchall(conn, "SELECT * FROM ventas ORDER BY fecha DESC")
        cols = ["ID","Fecha","Cliente","Total","Tipo","Estado"]
        keys = ["id","fecha","id_cliente","total","tipo","estado"]
    elif tipo == "productos":
        rows = database.fetchall(conn, "SELECT * FROM productos WHERE activo=1 ORDER BY nombre")
        cols = ["ID","Nombre","Categor\u00eda","Stock","Compra","Venta"]
        keys = ["id","nombre","categoria","stock","precio_compra","precio_venta"]
    elif tipo == "animales":
        rows = database.fetchall(conn, "SELECT a.*, d.nombre as dueno_nombre FROM animales a JOIN duenos d ON a.id_dueno=d.id ORDER BY a.nombre")
        cols = ["ID","Nombre","Especie","Raza","Edad","Due\u00f1o"]
        keys = ["id","nombre","especie","raza","edad","dueno_nombre"]
    elif tipo == "citas":
        rows = database.fetchall(conn, "SELECT c.*, a.nombre as animal_nombre, d.nombre as dueno_nombre FROM citas c JOIN animales a ON c.id_animal=a.id JOIN duenos d ON c.id_dueno=d.id ORDER BY c.fecha DESC")
        cols = ["ID","Fecha","Mascota","Due\u00f1o","Motivo","Estado"]
        keys = ["id","fecha","animal_nombre","dueno_nombre","motivo","estado"]
    elif tipo == "creditos":
        rows = database.fetchall(conn, "SELECT c.*, d.nombre as cliente_nombre FROM creditos c LEFT JOIN duenos d ON c.id_cliente=d.id ORDER BY c.fecha_venta DESC")
        cols = ["ID","Venta","Cliente","Total","Saldo","Vencimiento","Estado"]
        keys = ["id","id_venta","cliente_nombre","total","saldo","fecha_vencimiento","estado"]
    else:
        conn.close()
        return "Tipo no v\u00e1lido", 404
    conn.close()
    title = titles_map.get(tipo, tipo.capitalize())
    xldata = _excel_file(title, cols, rows, keys, col_widths={"A":8,"B":16,"C":24,"D":14,"E":14,"F":14,"G":14})
    return Response(xldata, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    headers={"Content-Disposition": f"attachment;filename={tipo}.xlsx"})

# ===== CAJA =====
@app.route("/caja")
@login_required
@role_required('admin')
def listar_caja():
    conn = database.get_db()
    fecha = request.args.get("fecha", "")
    q = request.args.get("q", "").strip()
    tf = request.args.get("tipo", "")
    page = max(1, int(request.args.get("page", 1)))
    per_page = 50
    wheres = []; params = []
    if fecha: wheres.append("fecha=?"); params.append(fecha)
    if q: wheres.append("concepto LIKE ?"); params.append(f"%{q}%")
    if tf: wheres.append("tipo=?"); params.append(tf)
    where_sql = (" WHERE " + " AND ".join(wheres)) if wheres else ""
    cnt = database.fetchone(conn, f"SELECT COUNT(*) as cnt FROM caja{where_sql}", params)["cnt"]
    movimientos = database.fetchall(conn, f"SELECT * FROM caja{where_sql} ORDER BY fecha DESC, id DESC LIMIT ? OFFSET ?", params+[per_page, (page-1)*per_page])
    sum_where = where_sql + (" AND tipo=?" if wheres else " WHERE tipo=?")
    total_ingresos = database.fetchone(conn, f"SELECT COALESCE(SUM(monto),0) as tot FROM caja{sum_where}", params+["ingreso"])["tot"]
    total_egresos = database.fetchone(conn, f"SELECT COALESCE(SUM(monto),0) as tot FROM caja{sum_where}", params+["egreso"])["tot"]
    total_pages = max(1, (cnt + per_page - 1) // per_page)
    page = min(page, total_pages)
    conn.close()
    return render_template("admin_caja.html", movimientos=movimientos,
                           total_ingresos=total_ingresos, total_egresos=total_egresos,
                           hoy_selected=fecha, q=q, tipo_f=tf, page=page, total_pages=total_pages,
                           title="Caja", active="caja", **_ctx())

@app.route("/caja/nuevo", methods=["POST"])
@login_required
@role_required('admin')
def crear_movimiento_caja():
    conn = database.get_db()
    comprobante = ""
    if "comprobante" in request.files and request.files["comprobante"] and request.files["comprobante"].filename:
        fname, _ = storage.save_file(request.files["comprobante"], "comprobantes")
        comprobante = fname
    database.execute(conn,
        "INSERT INTO caja (fecha,tipo,concepto,monto,metodo_pago,observacion,comprobante) VALUES (?,?,?,?,?,?,?)",
        (str(date.today()), request.form["tipo"], request.form.get("concepto",""),
         float(request.form.get("monto",0)), request.form.get("metodo_pago","efectivo"),
         request.form.get("observacion",""), comprobante))
    conn.commit(); conn.close()
    return redirect(url_for("listar_caja"))

@app.route("/caja/<int:id>/editar", methods=["POST"])
@login_required
@role_required('admin')
def editar_movimiento_caja(id):
    conn = database.get_db()
    m = database.fetchone(conn, "SELECT * FROM caja WHERE id=?", (id,))
    if not m:
        conn.close()
        return redirect(url_for("listar_caja"))
    comprobante = m["comprobante"] or ""
    if "comprobante" in request.files and request.files["comprobante"] and request.files["comprobante"].filename:
        if comprobante:
            storage.delete_file(comprobante, "comprobantes")
        fname, _ = storage.save_file(request.files["comprobante"], "comprobantes")
        comprobante = fname
    database.execute(conn,
        "UPDATE caja SET tipo=?, concepto=?, monto=?, metodo_pago=?, observacion=?, comprobante=? WHERE id=?",
        (request.form["tipo"], request.form.get("concepto",""),
         float(request.form.get("monto",0)),
         request.form.get("metodo_pago","efectivo"),
         request.form.get("observacion",""),
         comprobante, id))
    conn.commit(); conn.close()
    return redirect(url_for("listar_caja"))

@app.route("/caja/<int:id>/eliminar", methods=["POST"])
@login_required
@role_required('admin')
def eliminar_movimiento_caja(id):
    conn = database.get_db()
    m = database.fetchone(conn, "SELECT comprobante FROM caja WHERE id=?", (id,))
    if m and m.get("comprobante"):
        storage.delete_file(m["comprobante"], "comprobantes")
    database.execute(conn, "DELETE FROM caja WHERE id=?", (id,))
    conn.commit(); conn.close()
    return redirect(url_for("listar_caja"))

@app.route("/caja/cerrar", methods=["POST"])
@login_required
@role_required('admin')
def cerrar_caja():
    conn = database.get_db()
    hoy = str(date.today())
    ing = database.fetchone(conn, "SELECT COALESCE(SUM(monto),0) as tot FROM caja WHERE fecha=? AND tipo='ingreso'", (hoy,))["tot"]
    egr = database.fetchone(conn, "SELECT COALESCE(SUM(monto),0) as tot FROM caja WHERE fecha=? AND tipo='egreso'", (hoy,))["tot"]
    ap = database.fetchone(conn, "SELECT COALESCE(SUM(monto),0) as tot FROM caja WHERE fecha=? AND tipo='apertura'", (hoy,))["tot"]
    saldo = ap + ing - egr
    database.execute(conn,
        "INSERT INTO caja (fecha,tipo,concepto,monto) VALUES (?,?,?,?)",
        (hoy, "cierre", f"Cierre de caja - {session.get('doctor','')} - Ing:{ing:.2f} Egr:{egr:.2f} Saldo:{saldo:.2f}", saldo))
    conn.commit(); conn.close()
    return redirect(url_for("listar_caja"))


@app.route("/caja/exportar")
@login_required
@role_required('admin')
def exportar_caja():
    import openpyxl, io
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, Color
    conn = database.get_db()
    fecha = request.args.get("fecha", "")
    if fecha:
        movs = database.fetchall(conn, "SELECT * FROM caja WHERE fecha=? ORDER BY id DESC", (fecha,))
        ti = sum(r["monto"] for r in movs if r["tipo"]=="ingreso")
        te = sum(r["monto"] for r in movs if r["tipo"]=="egreso")
    else:
        movs = database.fetchall(conn, "SELECT * FROM caja ORDER BY fecha DESC, id DESC")
        ti = database.fetchone(conn, "SELECT COALESCE(SUM(monto),0) as tot FROM caja WHERE tipo='ingreso'")["tot"]
        te = database.fetchone(conn, "SELECT COALESCE(SUM(monto),0) as tot FROM caja WHERE tipo='egreso'")["tot"]
    conn.close()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Caja"
    ws.sheet_properties.tabColor = "1abc9c"
    hf = Font(bold=True, color="FFFFFF", size=11)
    hfill = PatternFill(start_color="1abc9c", end_color="1abc9c", fill_type="solid")
    ha = Alignment(horizontal="center", vertical="center")
    tbc = Color(rgb="CCCCCC"); tb = Border(left=Side(style="thin",color=tbc),right=Side(style="thin",color=tbc),top=Side(style="thin",color=tbc),bottom=Side(style="thin",color=tbc))
    ws.merge_cells("A1:F1")
    c = ws.cell(row=1, column=1, value=f"Reporte de Caja{f' - {fecha}' if fecha else ''}")
    c.font = Font(bold=True, size=14, color="1abc9c"); c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30
    ws.merge_cells("A2:B2")
    ws.cell(row=2, column=1, value=f"Ingresos: S/ {ti:.2f}").font = Font(bold=True, color="27ae60", size=11)
    ws.merge_cells("C2:D2")
    ws.cell(row=2, column=3, value=f"Egresos: S/ {te:.2f}").font = Font(bold=True, color="e74c3c", size=11)
    ws.merge_cells("E2:F2")
    bal = ti - te
    ws.cell(row=2, column=5, value=f"Balance: S/ {bal:.2f}").font = Font(bold=True, color="27ae60" if bal>=0 else "e74c3c", size=11)
    cols = ["ID","Fecha","Tipo","Concepto","Monto (S/)","Referencia"]
    for ci, cn in enumerate(cols, 1):
        c = ws.cell(row=4, column=ci, value=cn); c.font = hf; c.fill = hfill; c.alignment = ha; c.border = tb
    ws.row_dimensions[4].height = 22
    for ri, r in enumerate(movs, 5):
        vals = [r["id"], r["fecha"], r["tipo"].upper(), r["concepto"] or "-", r["monto"],
                f"{r['referencia_tipo'] or '-'} {r['referencia_id'] or ''}".strip()]
        for ci, v in enumerate(vals, 1):
            c = ws.cell(row=ri, column=ci, value=v); c.border = tb; c.font = Font(size=10)
            if ci == 3:
                if v == "INGRESO": c.font = Font(size=10, color="27ae60", bold=True)
                elif v == "EGRESO": c.font = Font(size=10, color="e74c3c", bold=True)
                elif v == "APERTURA": c.font = Font(size=10, color="3498db", bold=True)
                elif v == "CIERRE": c.font = Font(size=10, color="f39c12", bold=True)
            if ci == 5:
                c.number_format = "#,##0.00"
    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 40
    ws.column_dimensions["E"].width = 16
    ws.column_dimensions["F"].width = 20
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return Response(output.getvalue(), mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    headers={"Content-Disposition": f"attachment;filename=caja{f'_{fecha}' if fecha else ''}.xlsx"})


# ===== COBROS (Admin) =====
@app.route("/cobros")
@login_required
@role_required('admin', 'recepcion')
def listar_cobros():
    conn = database.get_db()
    filtro = request.args.get("filtro","pendientes")
    q = request.args.get("q","").strip()
    page = int(request.args.get("page","1"))
    per_page = 30

    where = " WHERE 1=1"
    params = []
    if filtro == "pendientes":
        where += " AND cp.cobrado=0"
    elif filtro == "cobrados":
        where += " AND cp.cobrado=1"
    if q:
        where += " AND (a.nombre LIKE ? OR d.nombre LIKE ?)"
        params.extend([f"%{q}%", f"%{q}%"])

    total = database.fetchone(conn,
        "SELECT COUNT(*) as c FROM cobros_pendientes cp JOIN animales a ON cp.id_animal=a.id JOIN duenos d ON a.id_dueno=d.id"+where,
        params)["c"]
    total_pages = max(1, (total + per_page - 1) // per_page)
    page = min(page, total_pages)
    offset = (page - 1) * per_page

    rows = database.fetchall(conn, """
        SELECT cp.*, a.nombre as animal_nombre, d.nombre as dueno_nombre
        FROM cobros_pendientes cp
        JOIN animales a ON cp.id_animal = a.id
        JOIN duenos d ON a.id_dueno = d.id
        """+where+"""
        ORDER BY cp.fecha DESC
        LIMIT ? OFFSET ?
    """, params+[per_page, offset])
    conn.close()
    cobros = []
    total_pendiente = 0
    for r in rows:
        sv = json.loads(r["servicios"]) if isinstance(r["servicios"], str) else (r["servicios"] or [])
        if not r["cobrado"]:
            total_pendiente += r["total"]
        cobros.append({
            "id": r["id"], "animal_nombre": r["animal_nombre"],
            "dueno_nombre": r["dueno_nombre"], "fecha": r["fecha"],
            "total": r["total"], "servicios_list": sv,
            "cobrado": r["cobrado"], "cobrado_en": r["cobrado_en"] or "",
            "metodo_pago": r["metodo_pago"] or "",
            "observacion": r["observacion"] or "",
            "comprobante": r["comprobante"] or ""
        })
    msg = "Cobro registrado exitosamente" if request.args.get("cobrado") else None
    return render_template("admin_cobros.html", cobros=cobros, total_pendiente=total_pendiente,
                           title="Cobros", active="cobros", msg=msg, filtro=filtro, q=q,
                           page=page, total_pages=total_pages, total=total, **_ctx())

@app.route("/cobros/<int:id>/cobrar", methods=["POST"])
@login_required
@role_required('admin', 'recepcion')
def cobrar_pendiente(id):
    conn = database.get_db()
    try:
        cp = database.fetchone(conn, "SELECT * FROM cobros_pendientes WHERE id=?", (id,))
        if not cp or cp["cobrado"]:
            conn.close()
            return redirect(url_for("listar_cobros"))
        metodo_pago = request.form.get("metodo_pago","efectivo")
        observacion = request.form.get("observacion","")
        fecha_cobro = request.form.get("fecha_cobro", str(date.today()))
        comprobante = ''
        if 'comprobante' in request.files and request.files['comprobante'] and request.files['comprobante'].filename:
            try:
                fname, _ = storage.save_file(request.files['comprobante'], 'comprobantes')
                comprobante = fname
            except Exception as e:
                logger.warning("Error al guardar comprobante cobro: %s", e)
        database.execute(conn,
            "INSERT INTO caja (fecha,tipo,concepto,monto,metodo_pago,observacion,comprobante,referencia_tipo,referencia_id) VALUES (?,?,?,?,?,?,?,?,?)",
            (fecha_cobro, "ingreso", "Cobro registro #"+str(cp["id_registro"]),
             cp["total"], metodo_pago, observacion, comprobante, "cobro", cp["id_registro"]))
        database.execute(conn,
            "UPDATE cobros_pendientes SET cobrado=1, cobrado_en=?, metodo_pago=?, observacion=?, comprobante=? WHERE id=?",
            (fecha_cobro, metodo_pago, observacion, comprobante, id))
        conn.commit()
        conn.close()
        return redirect(url_for("listar_cobros", cobrado=1))
    except Exception as e:
        conn.rollback()
        conn.close()
        logger.error("Error al cobrar: %s", e)
        return f"Error al procesar el cobro: {e}", 500

@app.route("/cobros/exportar")
@login_required
@role_required('admin', 'recepcion')
def exportar_cobros():
    conn = database.get_db()
    filtro = request.args.get("filtro", "pendientes")
    q = request.args.get("q", "").strip()
    where = "WHERE rc.id_registro IS NOT NULL"
    if filtro == "cobrados":
        where += " AND rc.cobrado=1"
    else:
        where += " AND (rc.cobrado IS NULL OR rc.cobrado=0)"
    if q:
        where += " AND (a.nombre LIKE ? OR d.nombre LIKE ?)"
    params = ["%" + q + "%", "%" + q + "%"] if q else []
    sql = f"""
        SELECT rc.id, rc.monto, a.nombre as animal_nombre, d.nombre as dueno_nombre,
               rc.fecha, rc.cobrado, rc.cobrado_en, rc.metodo_pago, rc.observacion, sv.nombre as servicio_nombre
        FROM registros_cobro rc
        LEFT JOIN registros_medicos rm ON rc.id_registro = rm.id
        LEFT JOIN animales a ON rm.id_animal = a.id
        LEFT JOIN duenos d ON a.id_dueno = d.id
        LEFT JOIN servicios_medicos sv ON rm.id_servicio = sv.id
        {where}
        ORDER BY rc.fecha DESC
    """
    rows = database.fetchall(conn, sql, params) if params else database.fetchall(conn, sql)
    conn.close()
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Border, Side
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Cobros"
    hdr_fill = PatternFill(start_color="1abc9c", end_color="1abc9c", fill_type="solid")
    hdr_font = Font(color="ffffff", bold=True, size=11)
    thin = Side(style="thin", color="cccccc")
    border = Border(top=thin, left=thin, right=thin, bottom=thin)
    headers = ["ID", "Paciente", "Due\u00f1o", "Servicio", "Monto", "Fecha", "Estado", "Cobrado En", "M\u00e9todo Pago", "Observaci\u00f3n"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.fill = hdr_fill; cell.font = hdr_font; cell.border = border
    for r, row in enumerate(rows, 2):
        vals = [row["id"], row["animal_nombre"] or "", row["dueno_nombre"] or "",
                row["servicio_nombre"] or "", row["monto"], row["fecha"] or "",
                "Cobrado" if row["cobrado"] else "Pendiente",
                row["cobrado_en"] or "", row["metodo_pago"] or "", row["observacion"] or ""]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.border = border
            if c == 5: cell.number_format = '#,##0.00'
    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 22
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 14
    ws.column_dimensions["G"].width = 12
    ws.column_dimensions["H"].width = 16
    ws.column_dimensions["I"].width = 14
    ws.column_dimensions["J"].width = 30
    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return Response(out.getvalue(), mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    headers={"Content-Disposition": "attachment;filename=cobros.xlsx"})

@app.route("/mis_consultas")
def web_mis_consultas():
    if "doctor" not in session:
        return redirect(url_for("index"))
    conn = database.get_db()
    doctorname = session["doctor"]
    rows = database.fetchall(conn, """
        SELECT r.id, r.fecha, r.hora, r.diagnostico, r.proximo_control,
               r.id_animal, a.nombre as animal_nombre, a.especie, d.nombre as dueno_nombre
        FROM registros_medicos r
        JOIN animales a ON r.id_animal = a.id
        JOIN duenos d ON a.id_dueno = d.id
        WHERE r.doctor=?
        ORDER BY r.fecha DESC, r.hora DESC
        LIMIT 100
    """, (doctorname,))
    conn.close()
    return render_template("web_mis_consultas.html", consultas=rows, doctor=doctorname)

@app.route("/mis_citas")
def web_mis_citas():
    if "doctor" not in session:
        return redirect(url_for("index"))
    conn = database.get_db()
    doctorname = session["doctor"]
    hoy = date.today().isoformat()
    citas_hoy = database.fetchall(conn, """
        SELECT c.*, a.nombre as animal_nombre, d.nombre as dueno_nombre, d.telefono as dueno_telefono
        FROM citas c JOIN animales a ON c.id_animal=a.id JOIN duenos d ON c.id_dueno=d.id
        WHERE c.fecha=? ORDER BY c.id ASC
    """, (hoy,))
    citas_futuras = database.fetchall(conn, """
        SELECT c.*, a.nombre as animal_nombre, d.nombre as dueno_nombre, d.telefono as dueno_telefono
        FROM citas c JOIN animales a ON c.id_animal=a.id JOIN duenos d ON c.id_dueno=d.id
        WHERE c.fecha>? ORDER BY c.fecha ASC, c.id ASC
    """, (hoy,))
    conn.close()
    return render_template("web_mis_citas.html",
        citas_hoy=citas_hoy, citas_futuras=citas_futuras, doctor=doctorname, hoy=hoy)

# ===== USUARIOS =====
@app.route("/usuarios")
@login_required
@role_required('admin')
def listar_usuarios():
    conn = database.get_db()
    usuarios = database.fetchall(conn, "SELECT * FROM usuarios ORDER BY nombre")
    conn.close()
    return render_template("admin_usuarios.html", items=usuarios, title="Usuarios", active="usuarios", **_ctx())

@app.route("/usuarios/nuevo", methods=["GET","POST"])
@login_required
@role_required('admin')
def crear_usuario():
    conn = database.get_db()
    if request.method == "POST":
        database.execute(conn,
            "INSERT INTO usuarios (username,password,nombre,rol) VALUES (?,?,?,?)",
            (request.form["username"], request.form["password"],
             request.form["nombre"], request.form.get("rol","empleado")))
        conn.commit(); conn.close()
        return redirect(url_for("listar_usuarios"))
    conn.close()
    return redirect(url_for("listar_usuarios"))

@app.route("/usuarios/<int:id>/editar", methods=["GET","POST"])
@login_required
@role_required('admin')
def editar_usuario(id):
    conn = database.get_db()
    if request.method == "POST":
        pwd = request.form.get("password","")
        if pwd:
            database.execute(conn,
                "UPDATE usuarios SET username=?,password=?,nombre=?,rol=?,activo=? WHERE id=?",
                (request.form["username"], pwd, request.form["nombre"],
                 request.form.get("rol","empleado"),
                 int(request.form.get("activo",1)), id))
        else:
            database.execute(conn,
                "UPDATE usuarios SET username=?,nombre=?,rol=?,activo=? WHERE id=?",
                (request.form["username"], request.form["nombre"],
                 request.form.get("rol","empleado"),
                 int(request.form.get("activo",1)), id))
        conn.commit(); conn.close()
        return redirect(url_for("listar_usuarios"))
    conn.close()
    return redirect(url_for("listar_usuarios"))

@app.route("/usuarios/<int:id>/eliminar")
@login_required
@role_required('admin')
def eliminar_usuario(id):
    conn = database.get_db()
    database.execute(conn, "DELETE FROM usuarios WHERE id=?", (id,))
    conn.commit(); conn.close()
    return redirect(url_for("listar_usuarios"))

# ===== DOCTORES =====
@app.route("/doctores")
@login_required
def listar_doctores():
    conn = database.get_db()
    doctores = database.fetchall(conn, "SELECT * FROM doctores ORDER BY nombre")
    conn.close()
    return render_template("admin_doctores.html", items=doctores, title="Doctores", active="doctores", **_ctx())

@app.route("/doctores/nuevo", methods=["GET","POST"])
@login_required
def crear_doctor():
    conn = database.get_db()
    if request.method == "POST":
        database.execute(conn, "INSERT INTO doctores (nombre) VALUES (?)",
            (request.form["nombre"],))
        conn.commit(); conn.close()
        return redirect(url_for("listar_doctores"))
    conn.close()
    return redirect(url_for("listar_doctores"))

@app.route("/doctores/<int:id>/editar", methods=["GET","POST"])
@login_required
def editar_doctor(id):
    conn = database.get_db()
    if request.method == "POST":
        database.execute(conn, "UPDATE doctores SET nombre=? WHERE id=?",
            (request.form["nombre"], id))
        conn.commit(); conn.close()
        return redirect(url_for("listar_doctores"))
    conn.close()
    return redirect(url_for("listar_doctores"))

@app.route("/doctores/<int:id>/eliminar")
@login_required
def eliminar_doctor(id):
    conn = database.get_db()
    database.execute(conn, "DELETE FROM doctores WHERE id=?", (id,))
    conn.commit(); conn.close()
    return redirect(url_for("listar_doctores"))

# ===== EDIT / DELETE REGISTRO MEDICO =====
@app.route("/registros/<int:id>/editar", methods=["GET","POST"])
@login_required
def editar_registro(id):
    conn = database.get_db()
    if request.method == "POST":
        database.execute(conn,
            "UPDATE registros_medicos SET fecha=?,hora=?,peso=?,doctor=?,diagnostico=?,tratamiento=?,"
            "observaciones=?,anamnesis=?,diagnostico_presuntivo=?,diagnostico_definitivo=?,proximo_control=? WHERE id=?",
            (request.form.get("fecha",""), request.form.get("hora",""),
             float(request.form["peso"]) if request.form.get("peso") else None,
             request.form.get("doctor",""), request.form.get("diagnostico",""),
             request.form.get("tratamiento",""), request.form.get("observaciones",""),
             request.form.get("anamnesis",""), request.form.get("diagnostico_presuntivo",""),
             request.form.get("diagnostico_definitivo",""), request.form.get("proximo_control",""), id))
        conn.commit(); conn.close()
        return redirect(request.referrer or url_for("dashboard"))
    r = database.fetchone(conn,
        "SELECT r.*, a.nombre as animal_nombre, a.id as animal_id FROM registros_medicos r "
        "JOIN animales a ON r.id_animal=a.id WHERE r.id=?", (id,))
    conn.close()
    if not r: return "No encontrado", 404
    return render_template("admin_edit_registro.html", r=r, title="Editar Registro", active="", **_ctx())

@app.route("/registros/<int:id>/eliminar")
@login_required
def eliminar_registro(id):
    conn = database.get_db()
    database.execute(conn, "DELETE FROM signos_vitales WHERE id_registro=?", (id,))
    database.execute(conn, "DELETE FROM cobros_pendientes WHERE id_registro=?", (id,))
    database.execute(conn, "DELETE FROM insumos_utilizados WHERE id_registro=?", (id,))
    database.execute(conn, "DELETE FROM registros_medicos WHERE id=?", (id,))
    conn.commit(); conn.close()
    return redirect(request.referrer or url_for("dashboard"))

# ===== DOCUMENTOS ADJUNTOS =====
@app.route("/animales/<int:id>/documentos")
@login_required
def listar_documentos(id):
    conn = database.get_db()
    a = database.fetchone(conn, "SELECT id,nombre FROM animales WHERE id=?", (id,))
    docs = database.fetchall(conn, "SELECT * FROM documentos_adjuntos WHERE id_animal=? ORDER BY fecha DESC", (id,))
    conn.close()
    if not a: return "No encontrado", 404
    return render_template("admin_documentos.html", animal=a, documentos=docs, title="Documentos", active="animales", **_ctx())

@app.route("/animales/<int:id>/documentos/nuevo", methods=["POST"])
@login_required
def crear_documento(id):
    conn = database.get_db()
    archivo = request.files.get("archivo")
    fname = ""
    if archivo and archivo.filename:
        fname, _ = storage.save_file(archivo, 'examenes')
    database.execute(conn,
        "INSERT INTO documentos_adjuntos (id_animal,nombre,tipo,fecha,archivo,observaciones) VALUES (?,?,?,?,?,?)",
        (id, request.form.get("nombre","Documento"), request.form.get("tipo","otro"),
         str(date.today()), fname, request.form.get("observaciones","")))
    conn.commit(); conn.close()
    return redirect(url_for("listar_documentos", id=id))

@app.route("/documentos/<int:doc_id>/eliminar")
@login_required
def eliminar_documento(doc_id):
    conn = database.get_db()
    doc = database.fetchone(conn, "SELECT id_animal, archivo FROM documentos_adjuntos WHERE id=?", (doc_id,))
    if doc:
        database.execute(conn, "DELETE FROM documentos_adjuntos WHERE id=?", (doc_id,))
        conn.commit()
    conn.close()
    animal_id = doc["id_animal"] if doc else 0
    return redirect(url_for("listar_documentos", id=animal_id))

# ===== HISTORIAL CLÍNICO COMPLETO =====
@app.route("/animales/<int:id>/historial")
@login_required
def historial_clinico(id):
    conn = database.get_db()
    a = database.fetchone(conn,
        "SELECT a.*, d.nombre as dueno_nombre, d.telefono as dueno_telefono, d.dni as dueno_dni, d.direccion as dueno_direccion, d.email as dueno_email "
        "FROM animales a JOIN duenos d ON a.id_dueno=d.id WHERE a.id=?", (id,))
    if not a: conn.close(); return "No encontrado", 404

    # Filtros de fecha para consultas
    desde = request.args.get("desde", "")
    hasta = request.args.get("hasta", "")
    sql_reg = ("SELECT r.*, sv.temperatura, sv.frecuencia_cardiaca, sv.frecuencia_respiratoria, "
        "sv.presion_sistolica, sv.presion_diastolica, sv.color_mucosa "
        "FROM registros_medicos r LEFT JOIN signos_vitales sv ON sv.id_registro=r.id "
        "WHERE r.id_animal=?")
    params_reg = [id]
    if desde:
        sql_reg += " AND r.fecha>=?"
        params_reg.append(desde)
    if hasta:
        sql_reg += " AND r.fecha<=?"
        params_reg.append(hasta)
    sql_reg += " ORDER BY r.fecha DESC, r.hora DESC"

    # Paginacion
    page = request.args.get("page", 1, type=int)
    per_page = 10
    # Count total for pagination
    count_sql = "SELECT COUNT(*) as cnt FROM registros_medicos WHERE id_animal=?"
    count_params = [id]
    if desde:
        count_sql += " AND fecha>=?"
        count_params.append(desde)
    if hasta:
        count_sql += " AND fecha<=?"
        count_params.append(hasta)
    total_reg = database.fetchone(conn, count_sql, count_params)["cnt"]
    total_pages = max(1, (total_reg + per_page - 1) // per_page)
    page = min(page, total_pages)
    offset = (page - 1) * per_page
    sql_reg += " LIMIT ? OFFSET ?"
    params_reg.extend([per_page, offset])

    registros = database.fetchall(conn, sql_reg, params_reg)
    vacunas = database.fetchall(conn,
        "SELECT * FROM vacunas WHERE id_animal=? ORDER BY fecha DESC", (id,))
    alergias = database.fetchall(conn,
        "SELECT * FROM alergias WHERE id_animal=? ORDER BY id", (id,))
    medicacion = database.fetchall(conn,
        "SELECT * FROM medicacion WHERE id_animal=? AND activo=1 ORDER BY fecha_inicio DESC", (id,))
    examenes = database.fetchall(conn,
        "SELECT * FROM examenes_auxiliares WHERE id_animal=? ORDER BY fecha DESC", (id,))
    citas = database.fetchall(conn,
        "SELECT * FROM citas WHERE id_animal=? ORDER BY fecha DESC LIMIT 10", (id,))

    # Otras mascotas del mismo dueno
    hermanos = database.fetchall(conn,
        "SELECT id, nombre, especie, raza, foto FROM animales WHERE id_dueno=? AND id!=? ORDER BY nombre", (a["id_dueno"], id))

    # Resumen rapido
    ultima_consulta = database.fetchone(conn,
        "SELECT fecha, diagnostico, doctor FROM registros_medicos WHERE id_animal=? ORDER BY fecha DESC, hora DESC LIMIT 1", (id,))
    ultimo_peso = database.fetchone(conn,
        "SELECT fecha, peso FROM registros_medicos WHERE id_animal=? AND peso IS NOT NULL ORDER BY fecha DESC LIMIT 1", (id,))
    ultima_vacuna = database.fetchone(conn,
        "SELECT fecha, nombre, proxima_dosis FROM vacunas WHERE id_animal=? ORDER BY fecha DESC LIMIT 1", (id,))

    # Alertas de proxima dosis
    today = str(date.today())
    prox_2 = str(date.today() + timedelta(days=2))
    vacunas_alerta = database.fetchall(conn,
        "SELECT * FROM vacunas WHERE id_animal=? AND proxima_dosis IS NOT NULL AND proxima_dosis!='' "
        "AND proxima_dosis<=? ORDER BY proxima_dosis ASC", (id, prox_2))

    conn.close()
    return render_template("admin_historial_clinico.html", a=a,
        registros=registros, vacunas=vacunas, alergias=alergias,
        medicacion=medicacion, examenes=examenes, citas=citas,
        hermanos=hermanos,
        ultima_consulta=ultima_consulta, ultimo_peso=ultimo_peso,
        ultima_vacuna=ultima_vacuna, vacunas_alerta=vacunas_alerta,
        page=page, total_pages=total_pages, desde=desde, hasta=hasta,
        total_reg=total_reg,
        title="Historial Cl\u00ednico: "+a["nombre"], active="animales", **_ctx())

@app.route("/animales/<int:id>/historial/nuevo", methods=["GET","POST"])
@login_required
def nuevo_historial(id):
    conn = database.get_db()
    a = database.fetchone(conn, "SELECT * FROM animales WHERE id=?", (id,))
    if not a: conn.close(); return "No encontrado", 404
    now = datetime.now()
    if request.method == "POST":
        cur = database.execute(conn,
            "INSERT INTO registros_medicos (id_animal,fecha,hora,peso,doctor,diagnostico,tratamiento,"
            "observaciones,anamnesis,diagnostico_presuntivo,diagnostico_definitivo,proximo_control) "
            "VALUES (?,?,?,?,?,?,?,?,?,?,?,?)",
            (id, request.form.get("fecha",str(date.today())), request.form.get("hora",""),
             float(request.form["peso"]) if request.form.get("peso") else None,
             request.form.get("doctor",session.get("doctor","")),
             request.form.get("diagnostico",""), request.form.get("tratamiento",""),
             request.form.get("observaciones",""), request.form.get("anamnesis",""),
             request.form.get("diagnostico_presuntivo",""),
             request.form.get("diagnostico_definitivo",""),
             request.form.get("cita_fecha","")))
        if database._using_pg():
            conn.commit()
            reg_id = cur.fetchone()[0] if cur.description else None
            if not reg_id: reg_id = database.fetchone(conn, "SELECT MAX(id) as id FROM registros_medicos WHERE id_animal=?", (id,))["id"]
        else: reg_id = cur.lastrowid
        # Signos vitales
        temp = request.form.get("temperatura","").strip()
        fc = request.form.get("fc","").strip()
        fr = request.form.get("fr","").strip()
        ps = request.form.get("presion_sistolica","").strip()
        pd = request.form.get("presion_diastolica","").strip()
        cm = request.form.get("color_mucosa","").strip()
        if temp or fc or fr or ps or pd or cm:
            database.execute(conn,
                "INSERT INTO signos_vitales (id_registro,temperatura,frecuencia_cardiaca,"
                "frecuencia_respiratoria,presion_sistolica,presion_diastolica,color_mucosa) VALUES (?,?,?,?,?,?,?)",
                (reg_id, float(temp) if temp else None, int(fc) if fc else None,
                 int(fr) if fr else None, int(ps) if ps else None, int(pd) if pd else None,
                 cm if cm else None))
        # Save exam files
        exam_types = [("ecografia","Ecograf\u00eda"),("radiografia","Radiograf\u00eda"),("hemograma","Hemograma"),("bioquimico","Bioqu\u00edmico"),("orina","Orina"),("otros","Otros")]
        for field, label in exam_types:
            file_obj = request.files.get("exam_"+field)
            if file_obj and file_obj.filename:
                fname, _ = storage.save_file(file_obj, 'examenes')
                database.execute(conn,
                    "INSERT INTO examenes_auxiliares (id_animal,tipo,nombre,fecha,archivo,resultados,observaciones) VALUES (?,?,?,?,?,?,?)",
                    (id, label, label+" - "+str(date.today()), str(date.today()), fname,
                     request.form.get("exam_"+field+"_resultados",""),
                     request.form.get("exam_"+field+"_obs","")))
        cita_fecha = request.form.get("cita_fecha","").strip()
        if cita_fecha:
            database.execute(conn,
                "INSERT INTO citas (id_animal,id_dueno,fecha,hora,motivo,tipo,precio) VALUES (?,?,?,?,?,?,?)",
                (id, a["id_dueno"], cita_fecha, request.form.get("cita_hora",""),
                 request.form.get("cita_motivo",""),
                 request.form.get("cita_tipo","consulta"),
                 float(request.form.get("cita_precio",0))))
        conn.commit(); conn.close()
        return redirect(url_for("historial_clinico", id=id))
    conn.close()
    return render_template("admin_historial_form.html", a=a, now=now,
        title="Nuevo Registro: "+a["nombre"], active="animales", **_ctx())

@app.route("/animales/<int:id>/historial/vacunas/nuevo", methods=["POST"])
@login_required
def nueva_vacuna(id):
    conn = database.get_db()
    database.execute(conn,
        "INSERT INTO vacunas (id_animal,tipo,nombre,fecha,lote,dosis,proxima_dosis,doctor,observaciones) "
        "VALUES (?,?,?,?,?,?,?,?,?)",
        (id, request.form.get("tipo","vacuna"), request.form.get("nombre",""),
         request.form.get("fecha",""), request.form.get("lote",""),
         request.form.get("dosis",""), request.form.get("proxima_dosis",""),
         request.form.get("doctor",""), request.form.get("observaciones","")))
    conn.commit(); conn.close()
    return redirect(url_for("historial_clinico", id=id))

@app.route("/animales/<int:id>/historial/alergias/nuevo", methods=["POST"])
@login_required
def nueva_alergia(id):
    conn = database.get_db()
    database.execute(conn,
        "INSERT INTO alergias (id_animal,alergeno,tipo,severidad,observaciones) VALUES (?,?,?,?,?)",
        (id, request.form.get("alergeno",""), request.form.get("tipo",""),
         request.form.get("severidad","leve"), request.form.get("observaciones","")))
    conn.commit(); conn.close()
    return redirect(url_for("historial_clinico", id=id))

@app.route("/animales/<int:id>/historial/medicacion/nuevo", methods=["POST"])
@login_required
def nueva_medicacion(id):
    conn = database.get_db()
    database.execute(conn,
        "INSERT INTO medicacion (id_animal, medicamento, dosis, frecuencia, via, fecha_inicio, observaciones) "
        "VALUES (?,?,?,?,?,?,?)",
        (id, request.form.get("med_nombre",""), request.form.get("med_dosis",""),
         request.form.get("med_frecuencia",""), request.form.get("med_via",""),
         request.form.get("med_fecha_inicio",""), request.form.get("med_observaciones","")))
    conn.commit(); conn.close()
    return redirect(url_for("historial_clinico", id=id))

@app.route("/animales/<int:id>/historial/examenes/nuevo", methods=["POST"])
@login_required
def nuevo_examen(id):
    conn = database.get_db()
    archivo = None
    file_obj = request.files.get("archivo")
    if file_obj and file_obj.filename:
        from storage import storage
        archivo, _ = storage.save_file(file_obj, 'examenes')
    database.execute(conn,
        "INSERT INTO examenes_auxiliares (id_animal,tipo,nombre,fecha,archivo,resultados,observaciones) VALUES (?,?,?,?,?,?,?)",
        (id, request.form.get("tipo",""), request.form.get("nombre",""),
         str(date.today()), archivo, request.form.get("resultados",""),
         request.form.get("observaciones","")))
    conn.commit(); conn.close()
    return redirect(url_for("historial_clinico", id=id))

@app.route("/vacunas/<int:id>/editar", methods=["POST"])
@login_required
def editar_vacuna(id):
    conn = database.get_db()
    database.execute(conn,
        "UPDATE vacunas SET nombre=?,fecha=?,lote=?,dosis=?,proxima_dosis=?,doctor=?,observaciones=? WHERE id=?",
        (request.form.get("nombre",""), request.form.get("fecha",""),
         request.form.get("lote",""), request.form.get("dosis",""),
         request.form.get("proxima_dosis",""), request.form.get("doctor",""),
         request.form.get("observaciones",""), id))
    conn.commit()
    animal_id = database.fetchone(conn, "SELECT id_animal FROM vacunas WHERE id=?", (id,))["id_animal"]
    conn.close()
    return redirect(url_for("historial_clinico", id=animal_id))

@app.route("/vacunas/<int:id>/eliminar", methods=["POST"])
@login_required
def eliminar_vacuna(id):
    conn = database.get_db()
    v = database.fetchone(conn, "SELECT id_animal FROM vacunas WHERE id=?", (id,))
    if v:
        database.execute(conn, "DELETE FROM vacunas WHERE id=?", (id,))
        conn.commit()
        animal_id = v["id_animal"]
    else:
        animal_id = 0
    conn.close()
    return redirect(url_for("historial_clinico", id=animal_id))

@app.route("/alergias/<int:id>/eliminar", methods=["POST"])
@login_required
def eliminar_alergia(id):
    conn = database.get_db()
    al = database.fetchone(conn, "SELECT id_animal FROM alergias WHERE id=?", (id,))
    if al:
        database.execute(conn, "DELETE FROM alergias WHERE id=?", (id,))
        conn.commit()
        animal_id = al["id_animal"]
    else:
        animal_id = 0
    conn.close()
    return redirect(url_for("historial_clinico", id=animal_id))

@app.route("/animales/<int:id>/historial/exportar")
@login_required
def exportar_historial(id):
    conn = database.get_db()
    a = database.fetchone(conn,
        "SELECT a.*, d.nombre as dueno_nombre, d.telefono as dueno_telefono, d.dni as dueno_dni "
        "FROM animales a JOIN duenos d ON a.id_dueno=d.id WHERE a.id=?", (id,))
    if not a: conn.close(); return "No encontrado", 404
    registros = database.fetchall(conn,
        "SELECT * FROM registros_medicos WHERE id_animal=? ORDER BY fecha DESC", (id,))
    vacunas = database.fetchall(conn,
        "SELECT * FROM vacunas WHERE id_animal=? ORDER BY fecha DESC", (id,))
    alergias = database.fetchall(conn,
        "SELECT * FROM alergias WHERE id_animal=? ORDER BY id", (id,))
    medicacion = database.fetchall(conn,
        "SELECT * FROM medicacion WHERE id_animal=? AND activo=1 ORDER BY fecha_inicio DESC", (id,))
    examenes = database.fetchall(conn,
        "SELECT * FROM examenes_auxiliares WHERE id_animal=? ORDER BY fecha DESC", (id,))
    conn.close()
    return render_template("admin_historial_print.html", a=a,
        registros=registros, vacunas=vacunas, alergias=alergias,
        medicacion=medicacion, examenes=examenes,
        **_ctx())
# ===== WEIGHT CHART DATA (JSON) =====
@app.route("/api/peso/<int:animal_id>")
@login_required
def api_peso(animal_id):
    conn = database.get_db()
    data = database.fetchall(conn,
        "SELECT fecha, peso FROM registros_medicos WHERE id_animal=? AND peso IS NOT NULL ORDER BY fecha ASC", (animal_id,))
    conn.close()
    return Response(json.dumps([{"fecha": d["fecha"], "peso": d["peso"]} for d in data]),
                    mimetype="application/json")

# ===== RECURRENT DIAGNOSES =====
@app.route("/api/diagnosticos/<int:animal_id>")
@login_required
def api_diagnosticos(animal_id):
    conn = database.get_db()
    data = database.fetchall(conn,
        "SELECT diagnostico, COUNT(*) as cnt FROM registros_medicos WHERE id_animal=? AND diagnostico!='' GROUP BY diagnostico ORDER BY cnt DESC LIMIT 10", (animal_id,))
    conn.close()
    return Response(json.dumps([{"diagnostico": d["diagnostico"], "count": d["cnt"]} for d in data]),
                    mimetype="application/json")

# ===== PDF RECETA =====
@app.route("/receta/<int:registro_id>")
@login_required
def receta(registro_id):
    conn = database.get_db()
    r = database.fetchone(conn,
        "SELECT r.*, a.nombre as animal_nombre, a.especie, a.edad, d.nombre as dueno_nombre, d.telefono, d.direccion "
        "FROM registros_medicos r JOIN animales a ON r.id_animal=a.id "
        "JOIN duenos d ON a.id_dueno=d.id WHERE r.id=?", (registro_id,))
    conn.close()
    if not r: return "No encontrado", 404
    return render_template("admin_receta.html", r=r, title="Receta", active="", **_ctx())

# ===== FICHA CLINICA =====
@app.route("/ficha/<int:animal_id>")
@login_required
def ficha_clinica(animal_id):
    conn = database.get_db()
    a = database.fetchone(conn,
        "SELECT a.*, d.nombre as dueno_nombre, d.telefono, d.direccion FROM animales a "
        "JOIN duenos d ON a.id_dueno=d.id WHERE a.id=?", (animal_id,))
    if not a: conn.close(); return "No encontrado", 404
    historial = database.fetchall(conn,
        "SELECT fecha, diagnostico, tratamiento, doctor, peso FROM registros_medicos WHERE id_animal=? ORDER BY fecha DESC", (animal_id,))
    vacunas = database.fetchall(conn,
        "SELECT fecha, nombre, doctor FROM vacunas WHERE id_animal=? ORDER BY fecha DESC", (animal_id,))
    conn.close()
    return render_template("admin_ficha.html", a=a, historial=historial, vacunas=vacunas, title="Ficha Cl\u00ednica", active="", **_ctx())

# ===== HISTORIAL CLÍNICO GLOBAL =====
@app.route("/historial")
@login_required
def historial_global():
    conn = database.get_db()
    page = request.args.get("page", 1, type=int)
    per_page = 25
    q = request.args.get("q","").strip()

    count_sql = ("SELECT COUNT(*) as cnt FROM registros_medicos r "
                 "JOIN animales a ON r.id_animal=a.id "
                 "JOIN duenos d ON a.id_dueno=d.id WHERE 1=1")
    sql = ("SELECT r.*, a.nombre as animal_nombre, a.especie, a.id_dueno, d.nombre as dueno_nombre "
           "FROM registros_medicos r JOIN animales a ON r.id_animal=a.id "
           "JOIN duenos d ON a.id_dueno=d.id WHERE 1=1")
    params = []
    if q:
        where = " AND (a.nombre LIKE ? OR d.nombre LIKE ? OR r.diagnostico LIKE ? OR r.doctor LIKE ?)"
        count_sql += where
        sql += where
        params.extend([f"%{q}%", f"%{q}%", f"%{q}%", f"%{q}%"])
    total = database.fetchone(conn, count_sql, params)["cnt"]
    total_pages = max(1, (total + per_page - 1) // per_page)
    if page < 1: page = 1
    if page > total_pages: page = total_pages
    offset = (page - 1) * per_page
    sql += " ORDER BY r.fecha DESC, r.hora DESC LIMIT ? OFFSET ?"
    registros = database.fetchall(conn, sql, params + [per_page, offset])
    edited_animal = request.args.get("edited_animal", type=int)
    edited_dueno = request.args.get("edited_dueno", type=int)
    conn.close()
    return render_template("admin_historial_global.html", registros=registros, q=q,
        page=page, total_pages=total_pages, total=total,
        edited_animal=edited_animal, edited_dueno=edited_dueno,
        title="Historial Cl\u00ednico", active="historial", **_ctx())

# ===== TICKET VENTA =====
@app.route("/ticket/<int:venta_id>")
@login_required
def ticket(venta_id):
    conn = database.get_db()
    v = database.fetchone(conn, "SELECT * FROM ventas WHERE id=?", (venta_id,))
    if not v: conn.close(); return "No encontrado", 404
    items = database.fetchall(conn, "SELECT * FROM venta_items WHERE id_venta=?", (venta_id,))
    pagos = database.fetchall(conn, "SELECT * FROM pagos WHERE id_venta=?", (venta_id,))
    conn.close()
    # Build WhatsApp link
    tel = (v.get("cliente_telefono") or "").strip()
    msg = "Ticket #" + str(venta_id) + " - S/ " + "%.2f" % v["total"]
    wa_link = "https://wa.me/51" + tel + "?text=" + urllib.parse.quote(msg) if tel else "#"
    return render_template("admin_ticket.html", v=v, items=items, pagos=pagos, wa_link=wa_link, title="Ticket", active="", **_ctx())

# ===== EDITAR VENTA =====
@app.route("/ventas/<int:id>/editar", methods=["GET","POST"])
@login_required
@role_required('admin', 'recepcion')
def editar_venta(id):
    conn = database.get_db()
    if request.method == "POST":
        try:
            items_json = request.form.get("items","[]")
            try: new_items = json.loads(items_json)
            except: new_items = []
            # Revert stock for old product items
            old_items = database.fetchall(conn, "SELECT * FROM venta_items WHERE id_venta=?", (id,))
            for oit in old_items:
                if oit["tipo_item"] == "producto" and oit["referencia_id"]:
                    database.execute(conn, "UPDATE productos SET stock=stock+? WHERE id=?", (oit["cantidad"], oit["referencia_id"]))
                    registrar_movimiento(conn, oit["referencia_id"], 'entrada', oit["cantidad"], 'edit_venta', id, 'Reversión edición Venta #{}'.format(id))
            # Delete old items
            database.execute(conn, "DELETE FROM venta_items WHERE id_venta=?", (id,))
            # Insert new/updated items
            total = 0
            for it in new_items:
                tipo_item = it.get("tipo","producto")
                ref_id = it.get("id")
                nombre = it.get("nombre","")
                cant = int(it.get("cant",1))
                precio = float(it.get("precio",0))
                subt = float(it.get("subtotal", cant * precio))
                database.execute(conn,"INSERT INTO venta_items (id_venta,tipo_item,referencia_id,nombre,cantidad,precio_unitario,subtotal) VALUES (?,?,?,?,?,?,?)",
                    (id, tipo_item, ref_id, nombre, cant, precio, subt))
                total += subt
                if tipo_item == "producto" and ref_id:
                    database.execute(conn,"UPDATE productos SET stock=stock-? WHERE id=? AND stock>=?", (cant, ref_id, cant))
                    registrar_movimiento(conn, ref_id, 'salida', cant, 'edit_venta', id, 'Edición Venta #{}'.format(id))
            # Update venta header
            nuevo_tipo = request.form.get("tipo_comprobante","Boleta")
            if nuevo_tipo not in ('Boleta', 'Factura', 'Nota de Venta'):
                nuevo_tipo = 'Boleta'
            v_actual = database.fetchone(conn, "SELECT tipo_comprobante FROM ventas WHERE id=?", (id,))
            if v_actual and v_actual["tipo_comprobante"] != nuevo_tipo:
                serie, numero = _generar_comprobante(conn, nuevo_tipo)
                database.execute(conn, "UPDATE ventas SET total=?,cliente_dni=?,cliente_nombre=?,tipo_comprobante=?,metodo_pago=?,serie=?,numero=? WHERE id=?",
                    (total, request.form.get("cliente_dni",""), request.form.get("cliente_nombre",""),
                     nuevo_tipo, request.form.get("metodo_pago","efectivo"), serie, numero, id))
            else:
                database.execute(conn, "UPDATE ventas SET total=?,cliente_dni=?,cliente_nombre=?,tipo_comprobante=?,metodo_pago=? WHERE id=?",
                    (total, request.form.get("cliente_dni",""), request.form.get("cliente_nombre",""),
                     nuevo_tipo, request.form.get("metodo_pago","efectivo"), id))
            conn.commit()
            conn.close()
            return redirect(url_for("ver_venta", id=id))
        except Exception as e:
            conn.rollback()
            conn.close()
            logger.error("Error al editar venta: %s", e)
            return f"Error al editar venta: {e}", 500
    v = database.fetchone(conn, "SELECT * FROM ventas WHERE id=?", (id,))
    items = database.fetchall(conn, "SELECT * FROM venta_items WHERE id_venta=?", (id,))
    productos = database.fetchall(conn,"SELECT id,nombre,precio_venta,stock,codigo_barras,por_mayor,descuento_mayorista FROM productos WHERE activo=1 ORDER BY nombre")
    conn.close()
    return render_template("admin_edit_venta.html", v=v, items=items, productos=productos, title="Editar Venta", active="ventas", **_ctx())

# ===== PAGAR VENTA AL CRÉDITO =====
@app.route("/ventas/<int:id>/pagar", methods=["GET","POST"])
@login_required
@role_required('admin', 'recepcion')
def pagar_venta(id):
    conn = database.get_db()
    v = database.fetchone(conn, "SELECT * FROM ventas WHERE id=?", (id,))
    if not v: conn.close(); return redirect(url_for("listar_ventas"))
    if request.method == "POST":
        monto = float(request.form.get("monto",0))
        metodo = request.form.get("metodo_pago","efectivo")
        if monto > 0:
            database.execute(conn, "INSERT INTO pagos (id_venta,monto,metodo,fecha) VALUES (?,?,?,?)", (id, monto, metodo, str(date.today())))
            nuevo_saldo = (v.get("saldo_pendiente") or v["total"]) - monto
            if nuevo_saldo <= 0:
                database.execute(conn, "UPDATE ventas SET saldo_pendiente=0, estado_pago='pagado' WHERE id=?", (id,))
            else:
                database.execute(conn, "UPDATE ventas SET saldo_pendiente=?, estado_pago='pendiente' WHERE id=?", (nuevo_saldo, id))
            conn.commit()
        conn.close()
        return redirect(url_for("ver_venta", id=id))
    conn.close()
    return redirect(url_for("ver_venta", id=id))

# ===== EXPORT EXCEL =====
@app.route("/exportar/excel/<tipo>")
@login_required
def exportar_excel(tipo):
    conn = database.get_db()
    title_map = {"animales":"Pacientes","ventas":"Ventas","productos":"Inventario","creditos":"Cr\u00e9ditos"}
    if tipo == "animales":
        data = database.fetchall(conn, "SELECT a.*, d.nombre as dueno_nombre FROM animales a JOIN duenos d ON a.id_dueno=d.id ORDER BY a.nombre")
        cols = ["ID","Nombre","Especie","Raza","Edad","Peso","Sexo","Color","Due\u00f1o"]
        keys = ["id","nombre","especie","raza","edad","peso","sexo","color","dueno_nombre"]
    elif tipo == "ventas":
        data = database.fetchall(conn, "SELECT * FROM ventas ORDER BY fecha DESC")
        cols = ["ID","Fecha","Cliente","Total","Tipo","Estado"]
        keys = ["id","fecha","id_cliente","total","tipo","estado"]
    elif tipo == "productos":
        data = database.fetchall(conn, "SELECT * FROM productos WHERE activo=1 ORDER BY nombre")
        cols = ["ID","Nombre","Categor\u00eda","Stock","Compra","Venta"]
        keys = ["id","nombre","categoria","stock","precio_compra","precio_venta"]
    elif tipo == "creditos":
        data = database.fetchall(conn, "SELECT c.*, d.nombre as cliente_nombre FROM creditos c LEFT JOIN duenos d ON c.id_cliente=d.id ORDER BY c.fecha_venta DESC")
        cols = ["ID","Cliente","Total","Saldo","Venta","Vencimiento","Estado"]
        keys = ["id","cliente_nombre","total","saldo","id_venta","fecha_vencimiento","estado"]
    else:
        conn.close(); return "Tipo no v\u00e1lido", 404
    conn.close()
    xldata = _excel_file(title_map.get(tipo, tipo.capitalize()), cols, data, keys, col_widths={"A":8,"B":28,"C":14,"D":14,"E":14,"F":14,"G":14,"H":14,"I":28})
    return Response(xldata, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    headers={"Content-Disposition": f"attachment;filename={tipo}.xlsx"})

# ===== GLOBAL SEARCH =====
@app.route("/buscar")
@login_required
def buscar():
    q = request.args.get("q", "").strip()
    results = {"duenos": [], "animales": [], "citas": [], "ventas": [], "productos": []}
    if q:
        conn = database.get_db()
        results["duenos"] = database.fetchall(conn,
            "SELECT * FROM duenos WHERE nombre LIKE ? OR telefono LIKE ? OR dni LIKE ? LIMIT 20",
            (f"%{q}%", f"%{q}%", f"%{q}%"))
        results["animales"] = database.fetchall(conn,
            "SELECT a.*, d.nombre as dueno_nombre FROM animales a JOIN duenos d ON a.id_dueno=d.id "
            "WHERE a.nombre LIKE ? OR a.especie LIKE ? OR a.raza LIKE ? LIMIT 20",
            (f"%{q}%", f"%{q}%", f"%{q}%"))
        results["citas"] = database.fetchall(conn,
            "SELECT c.*, a.nombre as animal_nombre FROM citas c JOIN animales a ON c.id_animal=a.id "
            "WHERE c.motivo LIKE ? OR a.nombre LIKE ? LIMIT 20",
            (f"%{q}%", f"%{q}%"))
        results["ventas"] = database.fetchall(conn,
            "SELECT * FROM ventas WHERE CAST(id AS TEXT) LIKE ? LIMIT 20",
            (f"%{q}%",))
        results["productos"] = database.fetchall(conn,
            "SELECT * FROM productos WHERE activo=1 AND (nombre LIKE ? OR categoria LIKE ?) LIMIT 20",
            (f"%{q}%", f"%{q}%"))
        conn.close()
    return render_template("admin_search.html", q=q, results=results, title="B\u00fasqueda", active="", **_ctx())

# ===== CAJA APERTURA =====
@app.route("/caja/aperturar", methods=["POST"])
@login_required
@role_required('admin')
def aperturar_caja():
    conn = database.get_db()
    database.execute(conn,
        "INSERT INTO caja (fecha,tipo,concepto,monto) VALUES (?,?,?,?)",
        (str(date.today()), "apertura", "Apertura de caja - "+session.get("doctor",""),
         float(request.form.get("monto_inicial",0))))
    conn.commit(); conn.close()
    return redirect(url_for("listar_caja"))

@app.route("/api/caja/semanal")
@login_required
@role_required('admin')
def api_caja_semanal():
    conn = database.get_db()
    from datetime import timedelta
    rows = []
    for i in range(6, -1, -1):
        d = str(date.today() - timedelta(days=i))
        ing = database.fetchone(conn, "SELECT COALESCE(SUM(monto),0) as tot FROM caja WHERE fecha=? AND tipo='ingreso'", (d,))["tot"]
        egr = database.fetchone(conn, "SELECT COALESCE(SUM(monto),0) as tot FROM caja WHERE fecha=? AND tipo='egreso'", (d,))["tot"]
        rows.append({"fecha": d, "ingresos": ing, "egresos": egr})
    conn.close()
    return {"dias": rows}

@app.route("/api/caja/resumen")
@login_required
@role_required('admin')
def api_caja_resumen():
    conn = database.get_db()
    hoy = str(date.today())
    apertura = database.fetchone(conn, "SELECT COALESCE(SUM(monto),0) as tot FROM caja WHERE fecha=? AND tipo='apertura'", (hoy,))["tot"]
    ingresos = database.fetchone(conn, "SELECT COALESCE(SUM(monto),0) as tot FROM caja WHERE fecha=? AND tipo='ingreso'", (hoy,))["tot"]
    egresos = database.fetchone(conn, "SELECT COALESCE(SUM(monto),0) as tot FROM caja WHERE fecha=? AND tipo='egreso'", (hoy,))["tot"]
    conn.close()
    return {"apertura": apertura, "ingresos": ingresos, "egresos": egresos}

# ===== BACKUP =====
@app.route("/api/backup/crear", methods=["POST"])
@login_required
def api_crear_backup():
    if session.get("rol") != "admin":
        return {"ok": False, "error": "Solo admin"}, 403
    database.backup_database()
    return {"ok": True, "message": "Backup creado exitosamente"}

@app.route("/backup/descargar")
@login_required
def descargar_backup():
    if session.get("rol") != "admin":
        return redirect(url_for("dashboard"))
    backup_dir = os.path.join(os.path.dirname(database.DB_PATH) if not database._using_pg() else ".", 'backups')
    if not os.path.exists(backup_dir):
        return "No hay backups", 404
    backups = sorted([f for f in os.listdir(backup_dir) if f.endswith('.db')])
    if not backups:
        return "No hay backups", 404
    latest = os.path.join(backup_dir, backups[-1])
    return send_from_directory(backup_dir, backups[-1], as_attachment=True)

# ===== API: NOTIFICACIONES =====
@app.route("/api/notificaciones")
@login_required
def api_notificaciones():
    conn = database.get_db()
    hoy = str(date.today())
    citas_hoy = database.fetchone(conn, "SELECT COUNT(*) as cnt FROM citas WHERE fecha=? AND estado='pendiente'", (hoy,))["cnt"]
    cobros_pend = database.fetchone(conn, "SELECT COUNT(*) as cnt FROM cobros_pendientes WHERE cobrado=0")["cnt"]
    stock_bajo = database.fetchone(conn, "SELECT COUNT(*) as cnt FROM productos WHERE activo=1 AND ((stock_minimo IS NOT NULL AND stock<=stock_minimo) OR (stock_minimo IS NULL AND stock<=5))")["cnt"]
    conn.close()
    return Response(json.dumps({"citas_hoy": citas_hoy, "cobros_pend": cobros_pend, "stock_bajo": stock_bajo}), mimetype="application/json")

# ===== PWA =====
@app.route("/manifest.json")
def manifest():
    return Response(json.dumps({
        "name": "SCRIPTYFY - Veterinaria",
        "short_name": "SCRIPTYFY",
        "start_url": "/",
        "display": "standalone",
        "background_color": "#1B4332",
        "theme_color": "#1B4332",
        "icons": [{"src": "/photos/logo.png", "sizes": "192x192", "type": "image/png"}]
    }), mimetype="application/json")

@app.route("/sw.js")
def service_worker():
    return Response("""
self.addEventListener('install',e=>self.skipWaiting());
self.addEventListener('fetch',e=>e.respondWith(fetch(e.request).catch(()=>new Response('Offline',{status:503}))));
""", mimetype="application/javascript")

# ===== SUSCRIPCI\u00d3N MENSUAL =====
@app.route("/suscripcion")
@login_required
@role_required('admin')
def suscripcion():
    conn = database.get_db()
    plan = database.fetchone(conn, "SELECT * FROM suscripcion_plan WHERE id=1")
    if not plan:
        database.execute(conn, "INSERT INTO suscripcion_plan (id,estado,plan,precio,fecha_creacion) VALUES (1,'inactivo','mensual',99,?)", (str(date.today()),))
        conn.commit()
        plan = database.fetchone(conn, "SELECT * FROM suscripcion_plan WHERE id=1")
    pagos = database.fetchall(conn, "SELECT * FROM suscripcion_pagos ORDER BY fecha_pago DESC LIMIT 50")
    # Auto-vencer si pas\u00f3 la fecha
    if plan and plan["estado"] == "activo" and plan.get("fecha_vencimiento"):
        try:
            v = datetime.strptime(plan["fecha_vencimiento"], "%Y-%m-%d").date()
            if v < date.today():
                database.execute(conn, "UPDATE suscripcion_plan SET estado='vencido' WHERE id=1")
                conn.commit()
                plan["estado"] = "vencido"
        except:
            pass
    conn.close()
    return render_template("admin_suscripcion.html", plan=plan, pagos=pagos,
                           title="Suscripci\u00f3n", active="suscripcion", **_ctx())

@app.route("/suscripcion/activar", methods=["POST"])
@login_required
@role_required('admin')
def suscripcion_activar():
    conn = database.get_db()
    plan = database.fetchone(conn, "SELECT * FROM suscripcion_plan WHERE id=1")
    hoy = date.today()
    if plan and plan["estado"] == "activo":
        venc = plan.get("fecha_vencimiento")
        if venc:
            v = datetime.strptime(venc, "%Y-%m-%d").date()
            nuevo_venc = v + timedelta(days=30)
        else:
            nuevo_venc = hoy + timedelta(days=30)
    else:
        nuevo_venc = hoy + timedelta(days=30)
    database.execute(conn, "UPDATE suscripcion_plan SET estado='activo',fecha_inicio=COALESCE(fecha_inicio,?),fecha_vencimiento=?,fecha_creacion=COALESCE(fecha_creacion,?) WHERE id=1",
                     (str(hoy), str(nuevo_venc), str(hoy)))
    database.execute(conn, "INSERT INTO suscripcion_pagos (monto,fecha_pago,metodo,notas,registrado_por) VALUES (?,?,?,?,?)",
                     (plan["precio"] if plan else 99, str(hoy), "manual", "Activaci\u00f3n manual", session.get("doctor","")))
    conn.commit(); conn.close()
    flash("Suscripci\u00f3n activada por 30 d\u00edas", "success")
    return redirect(url_for("suscripcion"))

@app.route("/suscripcion/configurar", methods=["POST"])
@login_required
@role_required('admin')
def suscripcion_configurar():
    precio = float(request.form.get("precio", 99))
    conn = database.get_db()
    database.execute(conn, "UPDATE suscripcion_plan SET precio=? WHERE id=1", (precio,))
    conn.commit(); conn.close()
    flash("Precio actualizado", "success")
    return redirect(url_for("suscripcion"))

@app.route("/suscripcion/desactivar", methods=["POST"])
@login_required
@role_required('admin')
def suscripcion_desactivar():
    conn = database.get_db()
    database.execute(conn, "UPDATE suscripcion_plan SET estado='inactivo' WHERE id=1")
    conn.commit(); conn.close()
    flash("Suscripci\u00f3n desactivada", "warning")
    return redirect(url_for("suscripcion"))

@app.route("/api/suscripcion/estado")
@login_required
def api_suscripcion_estado():
    conn = database.get_db()
    plan = database.fetchone(conn, "SELECT * FROM suscripcion_plan WHERE id=1")
    conn.close()
    if not plan or plan["estado"] != "activo":
        return {"estado": plan["estado"] if plan else "inactivo"}
    venc = plan.get("fecha_vencimiento")
    if not venc:
        return {"estado": "activo", "dias_restantes": 999}
    try:
        v = datetime.strptime(venc, "%Y-%m-%d").date()
        dias = (v - date.today()).days
        if dias < 0:
            conn = database.get_db()
            database.execute(conn, "UPDATE suscripcion_plan SET estado='vencido' WHERE id=1")
            conn.commit(); conn.close()
            return {"estado": "vencido", "dias_restantes": dias}
        return {"estado": "activo", "dias_restantes": dias, "vencimiento": venc}
    except:
        return {"estado": "activo", "dias_restantes": 999}

# ===== Test DNI page =====
@app.route("/testdni")
def test_dni():
    return render_template("test_dni.html")

@app.route("/testdni2")
def test_dni2():
    return render_template("test_dni2.html")

# ===== DNI LOOKUP (SUNAT) =====
@app.route("/api/dni/<numero>")
def api_dni(numero):
    if len(numero) != 8:
        return Response(json.dumps({"error": "DNI debe tener 8 d\u00edgitos"}), mimetype="application/json", status=400)
    conn = database.get_db()
    local = database.fetchone(conn, "SELECT nombre, dni, telefono, direccion FROM duenos WHERE dni=? AND dni!=''", (numero,))
    conn.close()
    if local:
        return Response(json.dumps({
            "nombre": local["nombre"], "apellido_paterno": "", "apellido_materno": "",
            "direccion": local["direccion"] or ""
        }), mimetype="application/json")
    # Try free RENIEC API (no key needed)
    sources = [
        ("RENIEC", f"https://graphperu.daustinn.com/api/query/{numero}"),
        ("SUNAT", f"https://ww1.sunat.gob.pe/ol-ti-itfisdenreg/itfisdenreg.htm?accion=obtenerDatosDni&numDocumento={numero}"),
    ]
    for source_name, url in sources:
        try:
            resp = requests.get(url, timeout=10)
            if resp.status_code != 200:
                continue
            data = resp.json()
            if source_name == "RENIEC" and data.get("fullName"):
                return Response(json.dumps({
                    "nombre": data["names"] or "", "apellido_paterno": data.get("paternalLastName","") or "",
                    "apellido_materno": data.get("maternalLastName","") or "", "direccion": ""
                }), mimetype="application/json")
            if source_name == "SUNAT" and data.get("message") == "success" and data.get("lista"):
                item = data["lista"][0]
                nombre_completo = (item.get("nombresapellidos") or item.get("apenomdenunciado") or "").strip()
                if "," in nombre_completo:
                    parts = nombre_completo.split(",")
                    apellidos = parts[0].strip()
                    nombre = parts[1].strip()
                    return Response(json.dumps({
                        "nombre": nombre, "apellido_paterno": apellidos, "apellido_materno": "", "direccion": ""
                    }), mimetype="application/json")
                return Response(json.dumps({
                    "nombre": nombre_completo, "apellido_paterno": "", "apellido_materno": "", "direccion": ""
                }), mimetype="application/json")
        except Exception:
            continue
    return Response(json.dumps({"error": "No se encontr\u00f3 DNI en RENIEC. Escribe los datos manualmente."}), mimetype="application/json")

# ===== RUC LOOKUP (OpenRUC + local fallback) =====
@app.route("/api/ruc/<numero>")
def api_ruc(numero):
    if len(numero) != 11:
        return Response(json.dumps({"error": "RUC debe tener 11 d\u00edgitos"}), mimetype="application/json", status=400)
    conn = database.get_db()
    local = database.fetchone(conn, "SELECT nombre, dni, telefono, direccion FROM duenos WHERE dni=? AND dni!=''", (numero,))
    conn.close()
    if local:
        return Response(json.dumps({
            "nombre": local["nombre"], "direccion": local["direccion"] or ""
        }), mimetype="application/json")
    try:
        resp = requests.get(f"https://openruc.com/api/ruc/{numero}", timeout=10)
        if resp.status_code == 200:
            data = resp.json()
            if data.get("razon_social"):
                return Response(json.dumps({
                    "nombre": data.get("razon_social", ""),
                    "direccion": data.get("direccion", "") or ""
                }), mimetype="application/json")
        return Response(json.dumps({"error": "No se encontr\u00f3 RUC. Escribe los datos manualmente."}), mimetype="application/json")
    except Exception as e:
        return Response(json.dumps({"error": "Error de conexi\u00f3n. Escribe los datos manualmente."}), mimetype="application/json")

# ===== API: Categorias =====
@app.route("/api/categorias")
def api_categorias():
    conn = database.get_db()
    cats = database.fetchall(conn, "SELECT * FROM categorias ORDER BY nombre")
    conn.close()
    return Response(json.dumps([dict(c) for c in cats]), mimetype="application/json")

@app.route("/api/categorias", methods=["POST"])
def api_categorias_add():
    try:
        nombre = request.form.get("nombre","").strip()
        if not nombre:
            return Response(json.dumps({"error":"Nombre requerido"}), mimetype="application/json", status=400)
        conn = database.get_db()
        try:
            database.execute(conn, "INSERT INTO categorias (nombre) VALUES (?)", (nombre,))
            conn.commit()
            cid = database.fetchone(conn, "SELECT MAX(id) as id FROM categorias")["id"]
            conn.close()
            return Response(json.dumps({"id":cid,"nombre":nombre}), mimetype="application/json")
        except:
            conn.close()
            return Response(json.dumps({"error":"Esa categor\u00eda ya existe"}), mimetype="application/json", status=400)
    except Exception as e:
        return Response(json.dumps({"error":str(e)}), mimetype="application/json", status=500)

# ===== API: Buscar dueño por DNI =====
@app.route("/api/duenos/buscar")
def api_duenos_buscar():
    try:
        dni = request.args.get("dni","").strip()
        if not dni: return Response(json.dumps({"error":"Falta DNI"}), mimetype="application/json", status=400)
        conn = database.get_db()
        d = database.fetchone(conn, "SELECT id,nombre,telefono,direccion,email FROM duenos WHERE dni=? LIMIT 1", (dni,))
        conn.close()
        if d: return Response(json.dumps({"id":d["id"],"nombre":d["nombre"] or "","telefono":d["telefono"] or "","direccion":d["direccion"] or "","email":d["email"] or ""}), mimetype="application/json")
        return Response(json.dumps({}), mimetype="application/json")
    except Exception as e:
        return Response(json.dumps({"error":str(e)}), mimetype="application/json", status=500)

# ===== API: Crear dueño desde venta =====
@app.route("/api/duenos/crear", methods=["POST"])
def api_duenos_crear():
    dni = request.form.get("dni","").strip()
    nombre = request.form.get("nombre","").strip()
    if not nombre: return Response(json.dumps({"error":"Nombre requerido"}), mimetype="application/json", status=400)
    conn = database.get_db()
    # Check if already exists
    exist = database.fetchone(conn, "SELECT id FROM duenos WHERE dni=? AND dni!=''", (dni,)) if dni else None
    if exist:
        conn.close()
        return Response(json.dumps({"error":"Ya existe un cliente con ese DNI","id":exist["id"]}), mimetype="application/json", status=400)
    database.execute(conn,"INSERT INTO duenos (nombre,dni,telefono,direccion,email) VALUES (?,?,?,?,?)",
        (nombre, dni, request.form.get("telefono",""), request.form.get("direccion",""), request.form.get("email","")))
    conn.commit()
    nuevo_id = database.fetchone(conn, "SELECT MAX(id) as id FROM duenos")["id"]
    conn.close()
    return Response(json.dumps({"id":nuevo_id,"nombre":nombre,"dni":dni}), mimetype="application/json")

@app.route("/api/pacientes/crear", methods=["POST"])
@login_required
def api_pacientes_crear():
    conn = database.get_db()
    nombre = request.form.get("nombre","").strip()
    if not nombre: conn.close(); return Response(json.dumps({"error":"Nombre requerido"}), mimetype="application/json", status=400)
    # Find or create owner
    dni = request.form.get("dni","").strip()
    dueno_nombre = request.form.get("dueno_nombre","").strip()
    telefono = request.form.get("telefono","").strip()
    dueno_id = None
    if dni:
        exist = database.fetchone(conn, "SELECT id FROM duenos WHERE dni=? AND dni!=''", (dni,))
        if exist: dueno_id = exist["id"]
    if not dueno_id and dueno_nombre:
        direccion = request.form.get("direccion","")
        database.execute(conn,"INSERT INTO duenos (nombre,dni,telefono,direccion) VALUES (?,?,?,?)", (dueno_nombre, dni, telefono, direccion))
        conn.commit()
        dueno_id = database.fetchone(conn, "SELECT MAX(id) as id FROM duenos")["id"]
    if not dueno_id:
        conn.close()
        return Response(json.dumps({"error":"Debes ingresar nombre del due\u00f1o o DNI"}), mimetype="application/json", status=400)
    esterilizado = 1 if request.form.get("esterilizado") == "1" else 0
    database.execute(conn,"INSERT INTO animales (nombre,especie,raza,sexo,esterilizado,id_dueno) VALUES (?,?,?,?,?,?)",
        (nombre, request.form.get("especie","Perro"), request.form.get("raza",""),
         request.form.get("sexo",""), esterilizado, dueno_id))
    conn.commit()
    nuevo_id = database.fetchone(conn, "SELECT MAX(id) as id FROM animales")["id"]
    if 'foto' in request.files and request.files['foto'] and request.files['foto'].filename:
        try:
            fname, _ = storage.save_file(request.files['foto'], 'photos')
            database.execute(conn, "UPDATE animales SET foto=? WHERE id=?", (fname, nuevo_id))
            conn.commit()
        except Exception as e:
            logger.warning("Error al guardar foto: %s", e)
    conn.close()
    return Response(json.dumps({"id":nuevo_id,"nombre":nombre}), mimetype="application/json")

@app.route("/api/historial/<dni>")
def api_historial(dni):
    conn = database.get_db()
    rows = database.fetchall(conn, "SELECT v.id,v.fecha,v.total,v.tipo_comprobante,v.tipo FROM ventas v WHERE v.cliente_dni=? ORDER BY v.fecha DESC LIMIT 10", (dni,))
    conn.close()
    return Response(json.dumps([dict(r) for r in rows]), mimetype="application/json")

@app.errorhandler(404)
def api_404(e):
    if request.path.startswith("/api/"):
        return Response(json.dumps({"error": "Endpoint no encontrado"}), mimetype="application/json", status=404)
    return e

def start_server():
    database.init_db()
    port = int(os.environ.get('PORT', 5000))
    logger.info("Servidor web iniciado en puerto %s", port)
    app.run(host="0.0.0.0", port=port, debug=(os.environ.get('FLASK_DEBUG', '0') == '1'))

if __name__ == "__main__":
    start_server()
